"""Tests for the CAT-style review grid workflow: confirm_segment (Ctrl+Enter),
interactive propagation (apply_propagation_matches/undo_propagation_batch),
TM learning, and the simplified apply_manual_edits materializer.

Imports `app` directly (proven safe headlessly — module-level code only
instantiates engines, no Streamlit runtime calls at import time) and
monkeypatches the DB-touching functions app.py calls at module scope, so
these tests never touch the real sqlite file.
"""

import io
import sys
import time

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, "/Users/User/Documents/Home24demo")
import app  # noqa: E402


def _row(id, column, source, translated_text, severity=app.SEVERITY_MEDIUM, row=None, **extra):
    return {
        "id": id, "column": column, "original_text": source, "translated_text": translated_text,
        "severity": severity, "row": row if row is not None else int(id.split("|")[1]),
        "category": "Manual review recommended", "reason": "Manual review recommended", **extra,
    }


@pytest.fixture(autouse=True)
def _no_db(monkeypatch):
    """confirm_segment/apply_propagation_matches only touch the DB through
    these two calls — stub them so tests never hit sqlite."""
    monkeypatch.setattr(app, "db_log_manual_correction", lambda *a, **k: None)
    monkeypatch.setattr(app, "db_save_glossary_suggestions", lambda *a, **k: None)


def _empty_tm():
    return {"entries": {}}


# =============================================================================
# 1. Exact source propagation (through confirm_segment, not just the engine)
# =============================================================================

class TestExactSourcePropagation:
    def test_confirm_propagates_to_identical_sources(self):
        items = [
            _row("s|2|name", "name", "Armlehnenstuhl ohne Dekoration", "Fauteuil sans décoration"),
            _row("s|3|name", "name", "Armlehnenstuhl ohne Dekoration", "Fauteuil sans décoration"),
            _row("s|4|name", "name", "Table basse", "Table basse"),
        ]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise a accoudoirs sans decoration", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester@home24.com", tm,
        )
        assert result["ok"], result.get("reason")
        assert {p["id"] for p in result["propagated"]} == {"s|3|name"}
        assert "s|2|name" in confirmed_ids and "s|3|name" in confirmed_ids
        assert "s|4|name" not in confirmed_ids


# =============================================================================
# 2. Model-safe pattern propagation
# =============================================================================

class TestModelSafePropagation:
    def test_evira_nova_preserved_through_confirm(self):
        items = [
            _row("s|2|name", "name", "Drehbarer Armlehnenstuhl EVIRA 2er-Set", "Fauteuil pivotant EVIRA 2er-Set"),
            _row("s|3|name", "name", "Drehbarer Armlehnenstuhl NOVA 2er-Set", "Fauteuil pivotant NOVA 2er-Set"),
        ]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise pivotante EVIRA 2er-Set", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")
        assert len(result["propagated"]) == 1
        assert result["propagated"][0]["id"] == "s|3|name"
        assert "NOVA" in manual_edits["s|3|name"]
        assert "EVIRA" not in manual_edits["s|3|name"]


# =============================================================================
# 3. Quantity safety
# =============================================================================

class TestQuantitySafety:
    def test_2er_set_never_becomes_4er_set(self):
        items = [
            _row("s|2|name", "name", "Kissenbezug 2er-Set", "Housse de coussin 2er-Set"),
            _row("s|3|name", "name", "Kissenbezug 4er-Set", "Housse de coussin 4er-Set"),
        ]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Taie d'oreiller 2er-Set", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")
        assert manual_edits["s|3|name"] == "Taie d'oreiller 4er-Set"
        assert "2er-Set" not in manual_edits["s|3|name"]


# =============================================================================
# 4. Capitalization propagation
# =============================================================================

class TestCapitalizationPropagation:
    def test_propagated_text_is_capitalization_correct(self):
        items = [
            _row("s|2|colorDetail", "colorDetail", "Bezug: Samt", "Revêtement : Velours"),
            _row("s|3|colorDetail", "colorDetail", "Bezug: Samt", "Revêtement : Velours"),
        ]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Revêtement : Velours côtelé", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")
        # colorDetail profile lowercases the value after ':' — the finalize
        # pass must run on the propagated text too, not just copy it.
        assert manual_edits["s|3|colorDetail"] == "Revêtement : velours côtelé"


# =============================================================================
# 5. Glossary learning
# =============================================================================

class TestGlossaryLearning:
    def test_clean_term_correction_surfaces_via_confirm(self, monkeypatch):
        captured = []
        monkeypatch.setattr(
            app, "db_save_glossary_suggestions",
            lambda suggestions, **k: captured.extend(suggestions),
        )
        items = [_row("s|2|name", "name", "Armlehnenstuhl", "Fauteuil avec accoudoirs")]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise à accoudoirs", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")
        assert captured and captured[0]["term"] == "Armlehnenstuhl"


# =============================================================================
# 6. TM learning
# =============================================================================

class TestTMLearning:
    def test_confirm_writes_high_trust_tm_entry(self):
        items = [_row("s|2|name", "name", "Armlehnenstuhl", "Fauteuil avec accoudoirs")]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise à accoudoirs", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "reviewer@home24.com", tm,
        )
        assert result["ok"], result.get("reason")
        key = app._tm_key("Armlehnenstuhl", app._tm_col_type("name"), "French")
        entry = tm["entries"][key]
        assert entry["source_type"] == "HUMAN_REVIEW"
        assert entry["confidence"] == 100
        assert entry["confirmed"] is True
        assert entry["corrected_by"] == "reviewer@home24.com"
        assert entry["previous_translation"] == "Fauteuil avec accoudoirs"
        assert entry["translation"] == "Chaise à accoudoirs"


# =============================================================================
# 7. Undo safety
# =============================================================================

class TestUndoSafety:
    def test_undo_restores_prior_state(self):
        items = [
            _row("s|2|name", "name", "Armlehnenstuhl EVIRA", "Fauteuil EVIRA"),
            _row("s|3|name", "name", "Armlehnenstuhl NOVA", "Fauteuil NOVA"),
        ]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise EVIRA", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")
        assert "s|3|name" in manual_edits and "s|3|name" in confirmed_ids

        app.undo_propagation_batch(result["undo_batch"], manual_edits, confirmed_ids)

        assert "s|2|name" not in manual_edits
        assert "s|2|name" not in confirmed_ids
        assert "s|3|name" not in manual_edits
        assert "s|3|name" not in confirmed_ids

    def test_undo_restores_previously_confirmed_row_instead_of_deleting_it(self):
        """A row that was already independently confirmed before the
        propagation touched it must come back confirmed, not disappear."""
        items = [
            _row("s|2|name", "name", "Armlehnenstuhl EVIRA", "Fauteuil EVIRA"),
            _row("s|3|name", "name", "Armlehnenstuhl NOVA", "Fauteuil NOVA"),
        ]
        manual_edits = {"s|3|name": "Ancienne correction NOVA"}
        confirmed_ids = {"s|3|name"}
        tm = _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise EVIRA", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")
        # confirmed rows are never overwritten by propagation
        assert manual_edits["s|3|name"] == "Ancienne correction NOVA"
        assert "s|3|name" in confirmed_ids
        assert result["propagated"] == []


# =============================================================================
# 8. 200-row performance smoke
# =============================================================================

class TestGridPerformance:
    def test_grid_rows_build_quickly_for_200_rows(self):
        items = [
            _row(f"s|{i}|name", "name", f"Quelle {i}", f"Cible {i}", row=i)
            for i in range(2, 202)
        ]
        manual_edits, confirmed_ids = {}, set()
        start = time.perf_counter()
        rows = app._build_cat_grid_rows(items, manual_edits, confirmed_ids)
        elapsed = time.perf_counter() - start
        assert elapsed < 2.0
        assert len(rows) == 200
        assert {r["id"] for r in rows} == {f"s|{i}|name" for i in range(2, 202)}


# =============================================================================
# 9. Automatic download remains unchanged / materializer writes confirmed edits
# =============================================================================

def _make_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name"])
    for r in rows:
        ws.append([r])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestMaterializer:
    def test_automatic_bytes_untouched_and_confirmed_edit_written(self, monkeypatch):
        monkeypatch.setattr(app, "load_translation_memory", lambda: _empty_tm())
        monkeypatch.setattr(app, "save_translation_memory", lambda tm: None)

        original_bytes = _make_workbook(["Drehbarer Armlehnenstuhl EVIRA 2er-Set"])
        items = [_row("Sheet1|2|name", "name", "Drehbarer Armlehnenstuhl EVIRA 2er-Set",
                       "Fauteuil pivotant EVIRA 2er-Set")]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise pivotante EVIRA 2er-Set", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")

        export = app.apply_manual_edits(
            original_bytes, "Sheet1", 1, items, manual_edits, None, "French", "test.xlsx",
            corrected_by="tester",
        )
        assert export["applied"] and export["applied"][0]["new"] == "Chaise pivotante EVIRA 2er-Set"

        # the automatic bytes we started from are never mutated in place
        wb_orig = load_workbook(io.BytesIO(original_bytes))
        assert wb_orig["Sheet1"].cell(row=2, column=1).value == "Drehbarer Armlehnenstuhl EVIRA 2er-Set"

        wb_corrected = load_workbook(io.BytesIO(export["corrected_workbook_bytes"]))
        assert wb_corrected["Sheet1"].cell(row=2, column=1).value == "Chaise pivotante EVIRA 2er-Set"


# =============================================================================
# 10. Corrected export contains the confirm + its auto-propagated sibling
# =============================================================================

class TestExportIncludesPropagation:
    def test_export_contains_direct_and_propagated_rows(self, monkeypatch):
        monkeypatch.setattr(app, "load_translation_memory", lambda: _empty_tm())
        monkeypatch.setattr(app, "save_translation_memory", lambda tm: None)

        original_bytes = _make_workbook([
            "Drehbarer Armlehnenstuhl EVIRA 2er-Set",
            "Drehbarer Armlehnenstuhl NOVA 2er-Set",
        ])
        items = [
            _row("Sheet1|2|name", "name", "Drehbarer Armlehnenstuhl EVIRA 2er-Set",
                 "Fauteuil pivotant EVIRA 2er-Set"),
            _row("Sheet1|3|name", "name", "Drehbarer Armlehnenstuhl NOVA 2er-Set",
                 "Fauteuil pivotant NOVA 2er-Set"),
        ]
        manual_edits, confirmed_ids, tm = {}, set(), _empty_tm()
        result = app.confirm_segment(
            items[0], "Chaise pivotante EVIRA 2er-Set", items, manual_edits, confirmed_ids,
            None, "French", "Sheet1", "tester", tm,
        )
        assert result["ok"], result.get("reason")
        assert "Sheet1|3|name" in manual_edits  # auto-propagated before export even runs

        export = app.apply_manual_edits(
            original_bytes, "Sheet1", 1, items, manual_edits, None, "French", "test.xlsx",
            corrected_by="tester",
        )
        wb_corrected = load_workbook(io.BytesIO(export["corrected_workbook_bytes"]))
        ws = wb_corrected["Sheet1"]
        assert ws.cell(row=2, column=1).value == "Chaise pivotante EVIRA 2er-Set"
        assert ws.cell(row=3, column=1).value == "Chaise pivotante NOVA 2er-Set"
