"""
German to French E-commerce Translator
AI-powered product localization tool for e-commerce platforms.

Author: Yves Koulle Banga
"""

import streamlit as st
import os
import re
import csv
import json
import uuid
import copy
import hmac
import string
import tempfile
import threading
import time
import logging
import smtplib
import unicodedata
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO, StringIO
from pathlib import Path
from datetime import datetime, timedelta

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openai import OpenAI

from database import (
    init_db,
    db_load_history,
    db_load_history_for_user,
    db_save_history_record,
    db_save_warnings,
    db_load_translation_memory,
    db_save_translation_memory,
    db_load_glossary,
    db_save_glossary,
    db_get_status,
    db_log_login,
    db_get_login_activity,
    db_get_admin_stats,
    db_save_glossary_suggestions,
    db_load_glossary_suggestions,
    db_update_suggestion_status,
    db_load_forbidden_patterns,
    db_save_forbidden_pattern,
    db_delete_forbidden_pattern,
    db_load_corpus_entries,
    db_add_corpus_entry,
    db_get_corpus_count,
    db_get_forbidden_count,
    db_save_issue_report,
    db_load_issue_reports,
    db_update_issue_report_status,
    db_get_issue_report_counts,
    db_update_jira_metadata,
    db_get_jira_stats,
)
from intelligence import (
    normalize_text,
    dedup_api_queue,
    try_glossary_only,
    try_pattern_translation,
    semantic_tm_match,
    detect_product_type,
    get_product_type_hint,
    extract_glossary_suggestions,
    apply_furniture_terms,
    auto_learn_glossary_from_source,
    run_local_consistency_pass,
    apply_french_semantic_normalization,
    apply_french_typography_rules,
    FURNITURE_TERM_MAP_FR,
    FURNITURE_TERM_MAP_NL,
    build_row_context,
    apply_context_terminology_fr,
    get_context_prompt_hint,
    get_corpus_style_hint,
    apply_forbidden_patterns,
)
from nl_engine import (
    Home24DutchCorpusEngine,
    DutchWorkbookConsistencyMemory,
    parse_trados_xlsx,
    nl_post_process,
    nl_qa_check,
    detect_nl_german_residue,
    apply_nl_post_colon_lowercase,
    safe_shorten_product_name_nl,
    segment_cell_for_tm_matching,
)
from database import (
    db_nl_trados_import,
    db_nl_trados_load_all,
    db_nl_trados_count,
)
from pipeline import (
    LargeFileModeConfig,
    detect_large_file_mode,
    SemanticRowClusterer,
    WorkbookConsistencyMemory,
    build_clustered_batches,
    qa_cell_needs_ai_fix,
    SheetDebugMetrics,
    LARGE_FILE_ROW_THRESHOLD,
)

load_dotenv()

try:
    from jira_client import get_jira_client, jira_configured
    _JIRA_AVAILABLE = True
except ImportError:
    _JIRA_AVAILABLE = False

    def get_jira_client():  # type: ignore[misc]
        return None, "jira_client module not installed."

    def jira_configured() -> bool:  # type: ignore[misc]
        return False


class _JiraFileProxy:
    """
    Minimal proxy that mimics st.file_uploader's result interface.
    Lets Jira-downloaded Excel bytes flow through the existing upload pipeline
    without any changes to the detection or translation logic.
    """

    def __init__(self, name: str, data: bytes) -> None:
        self.name = name
        self.size = len(data)
        self._data = data

    def getvalue(self) -> bytes:
        return self._data

    def read(self) -> bytes:
        return self._data


# =============================================================================
# NL CORPUS ENGINE — singleton loader
# =============================================================================

_NL_TM_AUTO_PATH = Path(__file__).parent.parent / "Downloads" / "Translation_Memory_Export_NL.xlsx"


def _get_nl_corpus_engine() -> "Home24DutchCorpusEngine | None":
    """
    Return the cached Dutch corpus engine.
    Loaded from DB once per session.
    If DB is empty and the Trados XLSX is found locally, auto-imports it.
    """
    if "nl_corpus_engine" not in st.session_state:
        entries = db_nl_trados_load_all()

        # Auto-import from known local path if DB is empty
        if not entries and _NL_TM_AUTO_PATH.exists():
            try:
                raw = parse_trados_xlsx(str(_NL_TM_AUTO_PATH))
                db_nl_trados_import(raw)
                entries = db_nl_trados_load_all()
            except Exception:
                entries = []

        if entries:
            st.session_state["nl_corpus_engine"] = Home24DutchCorpusEngine(entries)
        else:
            st.session_state["nl_corpus_engine"] = None
    return st.session_state["nl_corpus_engine"]


def _reload_nl_corpus_engine() -> "Home24DutchCorpusEngine | None":
    """Force a reload of the corpus engine (call after import)."""
    st.session_state.pop("nl_corpus_engine", None)
    return _get_nl_corpus_engine()


# =============================================================================
# WORKBOOK CONSISTENCY MEMORY — session-state singleton
# =============================================================================

def _get_consistency_memory(file_key: str) -> WorkbookConsistencyMemory:
    """
    Return the per-upload WorkbookConsistencyMemory singleton.
    One memory instance per uploaded file (keyed by filename + size).
    A new upload clears the old memory automatically.
    """
    mem_key = f"_wcm_{file_key}"
    if mem_key not in st.session_state:
        # Clear any stale instances from a previous file
        for k in list(st.session_state.keys()):
            if k.startswith("_wcm_"):
                del st.session_state[k]
        st.session_state[mem_key] = WorkbookConsistencyMemory()
    return st.session_state[mem_key]


# =============================================================================
# CONFIGURATION
# =============================================================================

HISTORY_FILE  = Path(__file__).parent / "translation_history.json"
TM_FILE       = Path(__file__).parent / "translation_memory.json"
GLOSSARY_FILE = Path(__file__).parent / "glossary.json"

CANDIDATE_SHEETS = ["Tabelle1", "Translations", "Sheet1"]

COLUMNS_TO_TRANSLATE = [
    "name", "colorDetail", "deliveryScope", "materialDetail",
    "otherMeasurements", "qualityDetail", "textileCompositionCover1", "variantName",
]

OPENAI_MODEL           = "gpt-4o-mini"
_INPUT_COST_PER_TOKEN  = 0.15 / 1_000_000
_OUTPUT_COST_PER_TOKEN = 0.60 / 1_000_000
MANUAL_SECONDS_PER_CELL = 45

DEFAULT_BATCH_SIZE      = 15
DEFAULT_MAX_CONCURRENT  = 3
MAX_BATCH_RETRIES       = 2
API_TIMEOUT_SECONDS = 45
MAX_API_RETRIES     = 3
RETRY_BASE_DELAY    = 1.0   # seconds; doubles on each retry

# Premium AI Refinement
REFINEMENT_COLUMNS   = {"name", "materialDetail", "qualityDetail", "deliveryScope", "variantName"}
REFINEMENT_MIN_CHARS = 20   # skip very short/obvious texts
REFINEMENT_BATCH_SIZE = 10  # smaller batches — refinement prompts are longer

REVIEW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

# Warning severity levels
SEVERITY_CRITICAL = "Critical"
SEVERITY_HIGH     = "High"
SEVERITY_MEDIUM   = "Medium"
SEVERITY_LOW      = "Low"

SEVERITY_DEDUCTION = {
    SEVERITY_CRITICAL: 10,
    SEVERITY_HIGH:      5,
    SEVERITY_MEDIUM:    2,
    SEVERITY_LOW:       1,
}

SEVERITY_ORDER = [SEVERITY_CRITICAL, SEVERITY_HIGH, SEVERITY_MEDIUM, SEVERITY_LOW]

DEFAULT_GLOSSARY_TERMS = {
    "Bezug":         "Revêtement",
    "Gestell":       "Structure",
    "Füße":          "Pieds",
    "Bettwäsche":    "linge de lit",
    "Webstoff":      "tissu tissé",
    "Strukturstoff": "tissu structuré",
    "Samtstoff":     "tissu velours",
    "Velours":       "velours",
    "Eiche":         "chêne",
    "Buche":         "hêtre",
    "Kiefer":        "pin",
    "Nussbaum":      "noyer",
    "Ahorn":         "érable",
    "Birke":         "bouleau",
    "Massiv":        "massif",
    "Furnier":       "plaqué",
    "lackiert":      "verni",
    "geölt":         "huilé",
    "gebeizt":       "teinté",
    "dunkelgrau":    "gris foncé",
    "hellgrau":      "gris clair",
    "dunkelbraun":   "brun foncé",
    "hellbraun":     "brun clair",
    "dunkelblau":    "bleu foncé",
    "hellblau":      "bleu clair",
    "dunkelgrün":    "vert foncé",
    "hellgrün":      "vert clair",
    "Anthrazit":     "anthracite",
    "Sandbeige":     "beige sable",
    "Baumwolle":     "coton",
    "Leinen":        "lin",
    "Wolle":         "laine",
    "Sofa":               "Canapé",
    "Sessel":             "Fauteuil",
    "Ecksofa":            "Canapé d'angle",
    "Schlafsofa":         "Canapé-lit",
    "Tisch":              "Table",
    "Stuhl":              "Chaise",
    "Schrank":            "Armoire",
    "Kommode":            "Commode",
    "Regal":              "Étagère",
    "inkl.":              "inclus",
    # Outdoor / lounge / garden
    "Loungeset":          "salon de jardin",
    "Sofaelement":        "module de canapé",
    "Gartenessgruppe":    "ensemble de jardin",
    "Gartengruppe":       "salon de jardin",
    "Gartenset":          "salon de jardin",
    "Gartenstuhl":        "chaise de jardin",
    "Gartentisch":        "table de jardin",
    "Gartenliege":        "chaise longue de jardin",
    "Gartenmöbel":        "mobilier de jardin",
    "Terrassenmöbel":     "mobilier de terrasse",
    # Materials / finishes
    "pulverbeschichtet":  "thermolaqué",
    "Geflecht":           "résine tressée",
    "Polyrattan":         "résine tressée",
    "Rattan":             "rotin",
    "Tischgestell":       "piètement de table",
    # Phrases
    "bestehend aus":      "composé de",
    "Set bestehend aus":  "ensemble composé de",
    "ohne Dekoration":    "sans décoration",
    "Absetzung":          "bordure contrastante",
    # Mattress
    "Matratze":                        "matelas",
    "Taschenfederkernmatratze":        "matelas à ressorts ensachés",
    "Taschenfederkern":                "ressorts ensachés",
    "Kokosmatte":                      "couche de coco",
    "Einseitige Kokosmatte":           "couche de coco sur une face",
    "Doppeltuch":                      "coutil double",
    "Reißverschluss":                  "fermeture éclair",
    "4-seitiger Reißverschluss":       "fermeture éclair sur 4 côtés",
    "Abnehmbarer Bezug":               "revêtement amovible",
}

DEFAULT_NL_GLOSSARY_TERMS = {
    "Bezug":         "bekleding",
    "Gestell":       "onderstel",
    "Füße":          "poten",
    "Bettwäsche":    "beddengoed",
    "Webstoff":      "geweven stof",
    "Strukturstoff": "structuurstof",
    "Samtstoff":     "fluwelen stof",
    "Velours":       "velours",
    "Eiche":         "eiken",
    "Buche":         "beuk",
    "Kiefer":        "grenen",
    "Nussbaum":      "notelaar",
    "Ahorn":         "esdoorn",
    "Birke":         "berk",
    "Massiv":        "massief",
    "Furnier":       "fineer",
    "lackiert":      "gelakt",
    "geölt":         "geolieerd",
    "gebeizt":       "gebeitst",
    "dunkelgrau":    "donkergrijs",
    "hellgrau":      "lichtgrijs",
    "dunkelbraun":   "donkerbruin",
    "hellbraun":     "lichtbruin",
    "dunkelblau":    "donkerblauw",
    "hellblau":      "lichtblauw",
    "dunkelgrün":    "donkergroen",
    "hellgrün":      "lichtgroen",
    "Anthrazit":     "antraciet",
    "Sandbeige":     "zandbeige",
    "Baumwolle":     "katoen",
    "Leinen":        "linnen",
    "Wolle":         "wol",
    "Sofa":              "bank",
    "Sessel":            "fauteuil",
    "Ecksofa":           "Hoekbank",      # TM: capital H in product names
    "Wohnlandschaft":    "Zithoek",       # TM: "Wohnlandschaft Fardah" → "Zithoek Fardah"
    "Ottomane":          "ottomane",      # TM: "mit Ottomane" → "met ottomane"
    "Ottoman":           "ottomane",
    "Schlafsofa":        "slaapbank",
    "Kombi":             "combi",         # TM: "Kombi A" → "combi A"
    "Variante":          "variant",       # TM: "Variante A" → "variant A"
    "Tisch":             "tafel",
    "Stuhl":             "stoel",
    "Schrank":           "kast",
    "Kommode":           "Kast",          # TM: "Kommode Weallup" → "Kast Weallup"
    "Regal":             "boekenrek",
    "inkl.":         "incl.",
    "Schublade":     "lade",
    "Türen":         "deuren",
    "Korpus":        "romp",
    "Maße":          "afmetingen",
    "Breite":        "breedte",
    "Höhe":          "hoogte",
    "Tiefe":         "diepte",
    "Länge":         "lengte",
    "Lieferumfang":  "leveringsomvang",
    "Lieferung":     "levering",
    "Holzwerkstoff": "houtmateriaal",
    "Spanplatte":    "spaanplaat",
    "Massivholz":    "massief hout",
    # Mattress
    "Matratze":                        "matras",
    "Taschenfederkernmatratze":        "pocketveringmatras",
    "Taschenfederkern":                "pocketveringkern",
    "Kokosmatte":                      "kokoslaag",
    "Einseitige Kokosmatte":           "kokoslaag aan één zijde",
    "Doppeltuch":                      "dubbeldoek",
    "Reißverschluss":                  "ritssluiting",
    "4-seitiger Reißverschluss":       "ritssluiting aan 4 zijden",
    "Abnehmbarer Bezug":               "afneembare hoes",
    # Dishwasher / GSP — TM: GSP-Blende → vaatwasserpaneel
    "GSP-Blende":                      "vaatwasserpaneel",
    "GSP Blende":                      "vaatwasserpaneel",
    "Geschirrspüler-Blende":           "vaatwasserpaneel",
    "Geschirrspülerblende":            "vaatwasserpaneel",
    "Geschirrspüler":                  "vaatwasser",
    # Dimensions — TM canonical: no spaces (BxHxD)
    "BHT":                             "BxHxD",
    "BxHxT":                           "BxHxD",
    "B x H x T":                       "BxHxD",
    "B/H/T":                           "BxHxD",
    "Breite x Höhe x Tiefe":           "breedte x hoogte x diepte",
    # Handles
    "Grifflos":                        "greeploos",
    "grifflos":                        "greeploos",
    # Drawer runners
    "Unterflurauszug":                 "onderliggende ladegeleider",
    "Unterflur-Auszug":                "onderliggende ladegeleider",
    "Unterflurführung":                "onderliggende ladegeleider",
    "Auszug":                          "lade",
    "Schubkasten":                     "lade",
    "Schubladen":                      "lades",
    # Kitchen furniture
    "Singleküche":                     "Mini keuken",      # TM: exact
    "Pantryküche":                     "Pantrykeuken",
    "Küchenleerblock":                 "Keukenblok",       # never "Keukenleerblok"
    "Kücheninsel":                     "Kookeiland",       # TM: exact
    "Küchenzeile":                     "Keukenblok",       # TM: exact
    "Einbauküche":                     "inbouwkeuken",
    "Arbeitsplatte":                   "werkblad",
    "Spüle":                           "spoelbak",
    "Spülenschrank":                   "spoelkast",
    "Unterschrank":                    "onderkast",
    "Hängeschrank":                    "hangkast",
    "Oberschrank":                     "bovenkast",
    "Hochschrank":                     "hoge kast",
    "Apothekerschrank":                "apothekerskast",
    "Blende":                          "frontpaneel",
    "Sockel":                          "plint",
    "Griff":                           "greep",
    "Griffe":                          "grepen",
    # Bathroom
    "Einzelwaschtisch":                "enkele wastafel",
    "Doppelwaschtisch":                "dubbele wastafel",
    "Waschtisch":                      "wastafelmeubel",
    "Waschbecken":                     "wastafel",
    "Waschbeckenunterschrank":         "wastafelonderkast",
    "Ablage":                          "legplank",
    "Armatur":                         "kraan",
    "Siphon":                          "sifon",
    "Überlauf":                        "overloop",
    "Soft-Close":                      "soft-close",
    "Softclose":                       "soft-close",
    "Dämpfung":                        "demping",
}

GERMAN_RESIDUE_WORDS = [
    "mit", "ohne", "und", "oder", "für", "aus", "inkl", "inklusive",
    "bei", "zur", "zum", "vom", "von", "samt", "sowie",
    "Maße", "Masse", "Breite", "Höhe", "Hoehe", "Tiefe", "Länge", "Laenge",
    "Bezug", "Gestell", "Füße", "Fuesse", "Fuße", "Beine",
    "Sitz", "Rücken", "Ruecken", "Armlehne", "Armlehnen",
    "Schublade", "Schubladen", "Türen", "Tueren", "Tür", "Tuer",
    "Polster", "Polsterung", "Lehne",
    "Holz", "Metall", "Kunststoff", "Stoff", "Leder", "Glas",
    "Webstoff", "Strukturstoff", "Samtstoff", "Velours",
    "Eiche", "Buche", "Kiefer", "Nussbaum", "Ahorn", "Birke",
    "Massiv", "massiv", "Furnier",
    "lackiert", "geölt", "geoelt", "gebeizt", "natur",
    "dunkelgrau", "hellgrau", "dunkelbraun", "hellbraun",
    "dunkelblau", "hellblau", "dunkelgrün", "hellgrün",
    "Sand", "sand", "Sandbeige", "Anthrazit",
    "weiß", "weiss", "schwarz", "grün", "gruen", "blau", "rot", "gelb",
    "grau", "braun", "creme", "Creme",
    "Lindgrün", "Moosgrün", "Mintgrün",
    "Bettwäsche", "Bettwasche", "Kissen", "Decke", "Matratze",
    "Baumwolle", "Leinen", "Wolle", "Seide", "Polyester",
    "Sofa", "Couch", "Sessel", "Tisch", "Stuhl", "Stühle", "Stuehle",
    "Schrank", "Bett", "Regal", "Lampe", "Kommode",
    "Sitzer", "Ecksofa", "Schlafsofa", "Longchair",
    "teilig", "Stück", "Stueck", "Set",
    "groß", "gross", "klein", "hoch", "niedrig", "breit", "schmal",
    "höhenverstellbar", "hoehenverstellbar", "ausziehbar", "klappbar",
    "Lieferumfang", "Hinweis", "Achtung", "Wichtig",
    "Zierkissen", "Dekoration",
    # Outdoor / lounge / garden furniture
    "Loungeset", "Loungesofa", "Loungesessel",
    "Sofaelement", "Sofamodul",
    "Gartenessgruppe", "Gartengruppe", "Gartenset",
    "Gartenstuhl", "Gartentisch", "Gartenbank", "Gartenliege", "Gartensofa",
    "Gartenmöbel", "Terrassenmöbel", "Terrassenmoebel", "Terrassenset",
    # Materials / finishes
    "pulverbeschichtet", "Pulverbeschichtung", "thermobeschichtet",
    "Geflecht", "Kunststoffgeflecht", "Flechtwerk", "Polyrattan",
    "Rattan",
    # Frame / structural parts
    "Tischgestell", "Untergestell", "Zargen", "Zarge",
    # Descriptors / phrases
    "bestehend", "Absetzung", "Abhebung",
    # Mattress / bedding
    "Taschenfederkern", "Taschenfederkernmatratze", "Bonellfeder",
    "Kaltschaummatratze", "Latexmatratze",
    "Kokosmatte", "Kokosschicht",
    "Doppeltuch",
    "Reißverschluss", "Reissverschluss",
    # Kitchen / GSP / abbreviations
    "GSP", "Geschirrspüler", "Geschirrspülerblende",
    "Grifflos", "grifflos",
    "Unterflurauszug", "Unterflurführung",
    "Küchenzeile", "Einbauküche", "Arbeitsplatte",
    "Spülenschrank", "Apothekerschrank",
    "Einzelwaschtisch", "Doppelwaschtisch", "Waschbeckenunterschrank",
    "Softclose", "Dämpfung",
    "BHT", "BxHxT",
    # Carpet / rug types
    "Fußmatte", "Fussmatte", "Läufer", "Laufer",
    "Hochflorteppich", "Kurzflorteppich", "Teppichläufer",
    "Teppich", "Kuhfellteppich", "Sisalteppich", "Juteteppich",
    "Schaffell", "Kunstfell",
    # Colors (German)
    "Elfenbein", "Puderrosa",
    # Textile composition materials
    "Polypropylen", "Polyamid", "Modacryl",
    "Kokosfaser", "Kokos", "Gummi", "Mikrofaser",
]

FRENCH_ACCEPTABLE_WORDS = [
    # Colors identical in German and French
    "beige", "taupe",
    # E-commerce / color terms used unchanged
    "multi", "multicolore",
    # Textile materials — same word or close cognate acceptable in French output
    "polyester", "polyamide", "viscose", "modacrylique", "polypropylène",
    "coco", "caoutchouc", "latex", "nylon", "sisal", "jute", "chenille",
    "microfibre", "coton", "laine", "lin", "soie",
    # Style/product terms used unchanged
    "set", "bouclé", "boucle",
]

# Dutch words that overlap with GERMAN_RESIDUE_WORDS — suppress false positives
DUTCH_ACCEPTABLE_WORDS = [
    "beige", "taupe", "polyester", "set", "velours", "glas",
    "creme", "bouclé", "boucle", "klein",
]

# Dutch words that must never appear in French output — with their French equivalents.
# Longest entries first so multi-word patterns replace before single words.
DUTCH_IN_FRENCH_MAP: list[tuple[str, str]] = [
    # Compound Dutch color descriptors
    ("donkergrijs",       "gris foncé"),
    ("lichtgrijs",        "gris clair"),
    ("donkerbruin",       "marron foncé"),
    ("lichtbruin",        "marron clair"),
    ("donkerblauw",       "bleu foncé"),
    ("lichtblauw",        "bleu clair"),
    ("donkergroen",       "vert foncé"),
    ("lichtgroen",        "vert clair"),
    ("antracietkleurig",  "anthracite"),
    ("grafietkleurig",    "graphite"),
    ("zilverkleurig",     "argenté"),
    ("goudkleurig",       "doré"),
    ("crèmekleurig",      "crème"),
    # Dutch fabric / texture names from TM
    ("fijnbever",         "castorette fine"),
    ("fijnbiber",         "castorette fine"),
    ("kerstdeken",        "couverture de Noël"),
    # Basic Dutch color words (title-case and lower-case)
    ("Donkergrijs",       "Gris foncé"),
    ("Lichtgrijs",        "Gris clair"),
    ("Donkerbruin",       "Marron foncé"),
    ("Lichtbruin",        "Marron clair"),
    ("Antraciet",         "Anthracite"),
    ("Bever",             "Castor"),
    ("Bruin",             "Marron"),
    ("Grijs",             "Gris"),
    ("Groen",             "Vert"),
    ("Rood",              "Rouge"),
    ("Geel",              "Jaune"),
    ("Blauw",             "Bleu"),
    ("Zwart",             "Noir"),
    ("Wit",               "Blanc"),
    ("Roze",              "Rose"),
    ("Oranje",            "Orange"),
    ("Beige",             "Beige"),
    # lower-case forms
    ("bever",             "castor"),
    ("bruin",             "marron"),
    ("grijs",             "gris"),
    ("groen",             "vert"),
    ("rood",              "rouge"),
    ("geel",              "jaune"),
    ("blauw",             "bleu"),
    ("zwart",             "noir"),
    ("wit",               "blanc"),
    ("roze",              "rose"),
    ("oranje",            "orange"),
    # Dutch furniture / product terms
    ("hoekbank",          "canapé d'angle"),
    ("zithoek",           "salon d'angle"),
    ("dressoir",          "buffet"),
    ("kookeiland",        "îlot de cuisine"),
    ("vaatwasserpaneel",  "panneau lave-vaisselle"),
    ("ottomane",          "méridienne"),
    ("eikenlook",         "décor chêne"),
    ("betonlook",         "béton ciré"),
    ("notenlook",         "décor noyer"),
    ("marmerlook",        "marbre"),
    # Dutch structural words that leak
    ("bestaande uit",     "composé de"),
    ("inclusief",         "inclus"),
    ("zonder",            "sans"),
    ("met",               "avec"),
]

# Compiled regex patterns for Dutch-in-French detection (word-boundary aware)
_DUTCH_FR_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r'\b' + re.escape(nl) + r'\b', re.IGNORECASE | re.UNICODE), fr)
    for nl, fr in DUTCH_IN_FRENCH_MAP
]

# Set of Dutch words (lowercase) for fast containment check
_DUTCH_WORD_SET: set[str] = {nl.lower() for nl, _ in DUTCH_IN_FRENCH_MAP}


def detect_dutch_in_french(text: str) -> list[str]:
    """Return list of Dutch words found in text (for French QA)."""
    if not text:
        return []
    found = []
    tl = text.lower()
    for nl, _ in DUTCH_IN_FRENCH_MAP:
        nl_lower = nl.lower()
        pattern = re.compile(r'\b' + re.escape(nl_lower) + r'\b', re.IGNORECASE | re.UNICODE)
        if pattern.search(tl):
            found.append(nl)
    return found


def apply_dutch_to_french_fixes(text: str) -> tuple[str, int]:
    """Replace Dutch words in text with French equivalents. Returns (fixed_text, count)."""
    count = 0
    for pattern, fr_word in _DUTCH_FR_PATTERNS:
        new = pattern.sub(fr_word, text)
        if new != text:
            count += 1
            text = new
    return text, count

# Protected column detection — two tiers to avoid false positives on short words
PROTECTED_SUBSTRINGS = [
    "articlenumber", "article_number", "articlenr", "artnummer", "artnr",
    "artikelnummer", "artikelnr", "sku", "productid", "product_id",
    "produktid", "gtin", "barcode", "ean13", "ean8",
]
PROTECTED_EXACT = {
    "id", "ean", "article", "artikel", "ref", "reference", "reference id",
}
# Keep old name for backward compat with any callers
PROTECTED_KEYWORDS = PROTECTED_SUBSTRINGS

IMPORTANT_KEYWORDS = [
    "name", "color", "colour", "delivery", "measurement",
    "quality", "textile", "composition", "material", "variant",
    "farbe", "liefer", "masse", "qualitat", "beschreibung",
]

# Keywords used to score candidate header rows
HEADER_SCORE_KEYWORDS = {
    "article", "artikelnummer", "sku", "product", "name", "color", "colour",
    "farbe", "material", "composition", "textile", "textil", "delivery",
    "lieferumfang", "scope", "quality", "detail", "measurement", "masse",
    "dimensions", "variant", "beschreibung", "nummer", "id", "ean", "gtin",
}

# T1: exact match against _normalize_col_header() output (spaces-collapsed form also tried)
TRANSLATE_ALIASES_T1 = {
    "name": [
        "name", "productname", "product name", "produktname", "produkt name",
        "artikelname", "artikel name", "artikelbezeichnung", "artikel bezeichnung",
        "produktbezeichnung", "produkt bezeichnung", "bezeichnung", "titelname",
        "title", "product title", "producttitle",
        # additional real-world variations
        "long name", "longname", "long description", "longdescription",
        "kurzbeschreibung", "produkttitel", "item name", "itemname",
    ],
    "colorDetail": [
        "colordetail", "color detail", "colourdetail", "colour detail",
        "farbe", "farbdetail", "farb detail", "farbbezeichnung",
        "colorname", "color name", "colourname", "colour name",
        "variantcolor", "variant color", "couleur", "detailcouleur",
        "detail couleur", "couleurdetail", "couleur detail",
        "colorangabe", "color angabe", "couleurduproduit",
        # additional real-world variations
        "farbton", "farbvariante", "color description", "colour description",
        "kleur", "kleurdetail", "kleur detail",
    ],
    "deliveryScope": [
        "deliveryscope", "delivery scope", "delivery_scope", "lieferumfang",
        "delivery contents", "deliverycontents", "lieferinhalt",
        "lieferung", "inhaltsangabe", "contenulivraison", "contenu livraison",
        "perimetre livraison", "perimetre de livraison", "scope de livraison",
        "leveringsomvang", "leveromfang",
        # additional real-world variations
        "im lieferumfang", "lieferumfang enthalt", "was ist enthalten",
        "included items", "included accessories", "inbegrepen",
    ],
    "materialDetail": [
        "materialdetail", "material detail", "materialdetails", "material details",
        "materialinfo", "material info", "materialbeschreibung", "material beschreibung",
        "materialangaben", "material angaben", "werkstoffe", "werkstoff",
        "materiaux", "detail matiere", "detailmatiere", "compositionmatiere",
        "composition matiere", "matiere",
        # additional real-world variations
        "materials", "material description", "rohstoffe", "material overview",
        "materiaaldetail", "materiaal detail", "materialen",
    ],
    "otherMeasurements": [
        "othermeasurements", "other measurements", "other measurement",
        "masse", "abmessungen", "abmessung", "measurements", "measurement",
        "dimensions", "dimension", "maßangaben", "maß angaben",
        "gesamtmasse", "gesamt masse", "produktmasse", "produkt masse",
        "dimensionen", "ausmasse", "aus masse", "autresmesures", "autres mesures",
        "mesures", "groesse", "breite hohe tiefe",
        # additional real-world variations
        "product dimensions", "product size", "produktgrosse", "grosse",
        "size", "sizes", "maten", "afmetingen", "breedte hoogte diepte",
        "b x h x t", "width height depth", "breite x hohe x tiefe",
    ],
    "qualityDetail": [
        "qualitydetail", "quality detail", "qualitydetails", "quality details",
        "qualitatsdetail", "qualitats detail", "qualite", "detail qualite",
        "detailqualite", "qualitaet", "pflegehinweise", "pflege hinweise",
        "eigenschaften", "produkteigenschaften",
        # additional real-world variations
        "care instructions", "care info", "pflegeanleitung", "pflegeempfehlung",
        "features", "product features", "kwaliteit", "kwaliteitsdetail",
        "onderhoud", "onderhoudsinstructies",
    ],
    "textileCompositionCover1": [
        "textilecompositioncover1", "textile composition cover 1",
        "textilecompositioncover",  "textile composition cover",
        "textilecomposition",       "textile composition",
        "textecomposition",         "texte composition",
        "compositioncover",         "composition cover",
        "compositioncover1",        "composition cover 1",
        "textecompositioncover1",   "texte composition cover 1",
        "compositiontextile",       "composition textile",
        "textilzusammensetzung",    "textil zusammensetzung",
        "zusammensetzung",
        "materialzusammensetzung",  "material zusammensetzung",
        "textilescomposition",      "textiles composition",
        "bezugzusammensetzung",     "bezug zusammensetzung",
        # additional real-world variations
        "fabric composition", "fabric content", "stoffzusammensetzung",
        "stof samenstelling", "textielesamenstelling", "samenstellingbekleding",
    ],
    "variantName": [
        "variantname", "variant name", "variantenname", "varianten name",
        "variantbezeichnung", "variant bezeichnung",
        "ausfuhrung", "ausfuehrung", "ausführung", "variante",
        "variantennamen", "varianten namen",
        # additional real-world variations
        "option", "option name", "optionname", "model", "modell",
        "uitvoering", "variantbenaming",
    ],
}

# T2: substring match against normalized header (word-boundary safe substrings)
TRANSLATE_ALIASES_T2 = {
    "textileCompositionCover1": [
        "textile", "textil", "composition", "zusammensetzung", "compositiontextile",
    ],
    "materialDetail":    ["material", "matiere", "werkstoff", "materiaux"],
    "colorDetail":       ["color", "colour", "farbe", "couleur"],
    "deliveryScope":     ["delivery", "lieferumfang", "livraison", "lieferinhalt"],
    "otherMeasurements": ["measurement", "dimension", "abmessung", "mesure"],
    "qualityDetail":     ["quality", "qualite", "qualitat", "qualitaet", "pflege"],
    "variantName":       ["variante", "ausfuhrung"],
    "name":              ["designation", "bezeichnung"],
}

# T3: word-set matching for compound/scrambled headers (≥2 matching words required)
CANONICAL_WORD_SETS: dict[str, set[str]] = {
    "name": {
        "name", "product", "produkt", "artikel", "bezeichnung", "title",
    },
    "colorDetail": {
        "color", "colour", "farbe", "couleur", "detail", "variant",
    },
    "deliveryScope": {
        "delivery", "scope", "lieferumfang", "lieferung", "inhalt", "contenu", "livraison",
    },
    "materialDetail": {
        "material", "detail", "matiere", "werkstoff", "werkstoffe", "materiaux",
    },
    "otherMeasurements": {
        "measurement", "dimension", "masse", "abmessung", "mesure", "groesse",
        "breite", "hohe", "tiefe", "laenge",
    },
    "qualityDetail": {
        "quality", "qualite", "qualitat", "detail", "pflege", "eigenschaften",
    },
    "textileCompositionCover1": {
        "textile", "composition", "cover", "zusammensetzung", "textil", "bezug",
    },
    "variantName": {
        "variant", "ausfuhrung", "ausfuehrung", "variante", "ausführung",
    },
}

# =============================================================================
# HOME24 KNOWN HEADER REGISTRY
# Maps normalized (lowercase, space-separated) column names to their canonical.
# This is the fast-path: if ANY of these appear in a file the file is accepted
# immediately — no density check, no row-count check.
# =============================================================================

HOME24_TRANSLATABLE_CANONICAL: dict[str, str] = {
    # name variants
    "name":                              "name",
    "productname":                       "name",
    "product name":                      "name",
    "artikelname":                       "name",
    "bezeichnung":                       "name",
    # materialDetail variants
    "material detail":                   "materialDetail",
    "materialdetail":                    "materialDetail",
    "material details":                  "materialDetail",
    "materialdetails":                   "materialDetail",
    # colorDetail variants
    "color detail":                      "colorDetail",
    "colordetail":                       "colorDetail",
    "colour detail":                     "colorDetail",
    "colourdetail":                      "colorDetail",
    "farbe":                             "colorDetail",
    "farbdetail":                        "colorDetail",
    # textileCompositionCover1 variants
    "textile composition":               "textileCompositionCover1",
    "textilecomposition":                "textileCompositionCover1",
    "textile composition cover":         "textileCompositionCover1",
    "textilecompositioncover":           "textileCompositionCover1",
    "textile composition cover 1":       "textileCompositionCover1",
    "textilecompositioncover 1":         "textileCompositionCover1",
    "textilecompositioncover1":          "textileCompositionCover1",
    "zusammensetzung":                   "textileCompositionCover1",
    "textzusammensetzung":               "textileCompositionCover1",
    "textilzusammensetzung":             "textileCompositionCover1",
    # qualityDetail variants
    "quality detail":                    "qualityDetail",
    "qualitydetail":                     "qualityDetail",
    "quality details":                   "qualityDetail",
    "qualitydetails":                    "qualityDetail",
    "qualitatsdetail":                   "qualityDetail",
    "qualitat detail":                   "qualityDetail",
    "qualitaet":                         "qualityDetail",
    # deliveryScope variants
    "delivery scope":                    "deliveryScope",
    "deliveryscope":                     "deliveryScope",
    "lieferumfang":                      "deliveryScope",
    "deliverycontents":                  "deliveryScope",
    "delivery contents":                 "deliveryScope",
    # variantName variants
    "variant name":                      "variantName",
    "variantname":                       "variantName",
    "variantenname":                     "variantName",
    "varianten name":                    "variantName",
    "ausfuhrung":                        "variantName",
    "ausfuehrung":                       "variantName",
    "ausführung":                        "variantName",
}

# Set of all known normalized Home24 translatable headers for O(1) lookup
HOME24_TRANSLATABLE_NORMALIZED: frozenset[str] = frozenset(HOME24_TRANSLATABLE_CANONICAL.keys())

# Detection logger — logs decisions at DEBUG level so they stay silent by default
# but can be surfaced with `logging.basicConfig(level=logging.DEBUG)`.
_DETECT_LOG = logging.getLogger("home24.detect")


# =============================================================================
# AUTHENTICATION
# =============================================================================

def _get_admin_credentials() -> tuple[str, str]:
    # I try secrets first, then env vars; support both ADMIN_* and legacy APP_USER_* names
    try:
        email    = st.secrets.get("ADMIN_EMAIL") or st.secrets.get("APP_USER_EMAIL", "")
        password = st.secrets.get("ADMIN_PASSWORD") or st.secrets.get("APP_USER_PASSWORD", "")
        if email and password:
            return str(email), str(password)
    except Exception:
        pass
    email    = os.environ.get("ADMIN_EMAIL") or os.environ.get("APP_USER_EMAIL", "")
    password = os.environ.get("ADMIN_PASSWORD") or os.environ.get("APP_USER_PASSWORD", "")
    return email or "", password or ""


def _get_guest_credentials() -> tuple[str, str]:
    try:
        email    = st.secrets.get("GUEST_EMAIL", "")
        password = st.secrets.get("GUEST_PASSWORD", "")
        if email and password:
            return str(email), str(password)
    except Exception:
        pass
    return os.environ.get("GUEST_EMAIL", ""), os.environ.get("GUEST_PASSWORD", "")


def _get_justus_credentials() -> tuple[str, str]:
    try:
        email    = st.secrets.get("JUSTUS_EMAIL", "")
        password = st.secrets.get("JUSTUS_PASSWORD", "")
        if email and password:
            return str(email), str(password)
    except Exception:
        pass
    return os.environ.get("JUSTUS_EMAIL", ""), os.environ.get("JUSTUS_PASSWORD", "")


def _get_secret(key: str, default: str = "") -> str:
    """Read a value from Streamlit secrets first, then env vars."""
    try:
        val = st.secrets.get(key, "")
        if val:
            return str(val)
    except Exception:
        pass
    return os.environ.get(key, default)


def verify_credentials(input_email: str, input_password: str) -> str | None:
    # Returns "admin", "standard_user", "guest", or None on failure
    email = input_email.strip().lower()

    admin_email, admin_password = _get_admin_credentials()
    if admin_email and admin_password:
        if (hmac.compare_digest(email, admin_email.strip().lower()) and
                hmac.compare_digest(input_password, admin_password)):
            return "admin"

    justus_email, justus_password = _get_justus_credentials()
    if justus_email and justus_password:
        if (hmac.compare_digest(email, justus_email.strip().lower()) and
                hmac.compare_digest(input_password, justus_password)):
            return "standard_user"

    guest_email, guest_password = _get_guest_credentials()
    if guest_email and guest_password:
        if (hmac.compare_digest(email, guest_email.strip().lower()) and
                hmac.compare_digest(input_password, guest_password)):
            return "guest"

    return None


# =============================================================================
# HISTORY
# =============================================================================

def load_history() -> list:
    return db_load_history()


def save_history_record(record: dict) -> None:
    db_save_history_record(record)


# =============================================================================
# ISSUE REPORTING — EMAIL
# =============================================================================

_ISSUE_CATEGORIES = [
    "Translation error",
    "German residue",
    "Wrong terminology",
    "Product name issue",
    "Column detection issue",
    "Excel export issue",
    "CSV export issue",
    "Login / access issue",
    "UI/UX bug",
    "Other",
]

_ISSUE_SEVERITIES = ["Low", "Medium", "High", "Critical"]

_ISSUE_STATUSES = ["open", "in progress", "resolved", "ignored"]

_ISSUE_LANGUAGES = ["French", "Dutch", "Both", "Not language-related"]


def _send_issue_report_email(report: dict) -> bool:
    """
    Send issue report notification to the admin address.
    Returns True if the email was sent successfully, False otherwise.
    Missing SMTP config is not an error — the report is always saved to DB.
    """
    smtp_host = _get_secret("SMTP_HOST")
    smtp_port = int(_get_secret("SMTP_PORT", "587"))
    smtp_user = _get_secret("SMTP_USER")
    smtp_pass = _get_secret("SMTP_PASSWORD")
    smtp_from = _get_secret("SMTP_FROM") or smtp_user
    report_to = _get_secret("REPORT_EMAIL_TO", "yves.banga@home24.de")

    if not (smtp_host and smtp_user and smtp_pass):
        return False

    category = report.get("category", "Other")
    severity = report.get("severity", "Medium")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"[Localization Platform] New {severity} issue: {category}"
    msg["From"]    = smtp_from
    msg["To"]      = report_to

    body_lines = [
        "A new issue report has been submitted on the Home24 Localization Platform.",
        "",
        f"Reporter:           {report.get('user_email', '—')}",
        f"Role:               {report.get('user_role', '—')}",
        f"Category:           {category}",
        f"Severity:           {severity}",
        f"Target language:    {report.get('target_language', '—')}",
        f"File:               {report.get('filename', '—') or '—'}",
        f"Row reference:      {report.get('row_reference', '—') or '—'}",
        f"Column reference:   {report.get('column_reference', '—') or '—'}",
        f"Submitted at:       {report.get('created_at', '—')}",
        "",
        "─── Description ────────────────────────────────────────────────────",
        report.get("description", ""),
        "",
    ]
    correction = report.get("expected_correction", "").strip()
    if correction:
        body_lines += [
            "─── Expected correction ─────────────────────────────────────────────",
            correction,
            "",
        ]
    body_lines += [
        "────────────────────────────────────────────────────────────────────",
        "Home24 AI Localization Platform — internal tool",
    ]

    msg.attach(MIMEText("\n".join(body_lines), "plain", "utf-8"))

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        return True
    except Exception:
        return False


# =============================================================================
# TRANSLATION MEMORY
# =============================================================================

def _tm_col_type(canonical: str) -> str:
    if canonical == "name":
        return "name"
    if canonical == "materialDetail":
        return "materialDetail"
    return "other"


def _tm_key(text: str, col_type: str, target_language: str = "French") -> str:
    lang = "nl" if target_language == "Dutch" else "fr"
    return f"{lang}:{col_type}:{' '.join(text.strip().split())}"


def load_translation_memory() -> dict:
    return db_load_translation_memory()


def save_translation_memory(tm: dict) -> None:
    db_save_translation_memory(tm)


def tm_get(tm: dict, text: str, col_type: str, target_language: str = "French") -> str | None:
    key   = _tm_key(text, col_type, target_language)
    entry = tm["entries"].get(key)
    if entry is not None:
        entry["hit_count"] = entry.get("hit_count", 0) + 1
        return entry["translation"]
    return None


def tm_put(tm: dict, source: str, translation: str, col_type: str, target_language: str = "French") -> None:
    key = _tm_key(source, col_type, target_language)
    if key not in tm["entries"]:
        tm["entries"][key] = {
            "translation": translation,
            "col_type":    col_type,
            "created_at":  datetime.now().isoformat(timespec="seconds"),
            "hit_count":   0,
        }


# =============================================================================
# GLOSSARY SYSTEM
# =============================================================================

def load_glossary(target_language: str = "French") -> dict:
    data = db_load_glossary(target_language)
    if data is not None:
        return data
    default = DEFAULT_NL_GLOSSARY_TERMS if target_language == "Dutch" else DEFAULT_GLOSSARY_TERMS
    return {
        "terms":           default.copy(),
        "target_language": target_language,
        "stats":           {"total_hits": 0, "term_counts": {}},
    }


def save_glossary(glossary: dict, target_language: str = "French") -> None:
    db_save_glossary(glossary, target_language)


def _glossary_prompt_block(glossary: dict) -> str:
    terms = glossary.get("terms", {})
    if not terms:
        return ""
    lines = [f"- {de} → {fr}" for de, fr in list(terms.items())[:25]]
    return "\nAlways use these standard terms consistently:\n" + "\n".join(lines)


def count_glossary_hits(text: str, glossary: dict) -> dict:
    hits      = {}
    terms     = glossary.get("terms", {})
    text_lower = text.lower()
    for de_term in terms:
        pattern = r'\b' + re.escape(de_term.lower()) + r'\b'
        if re.search(pattern, text_lower):
            hits[de_term] = hits.get(de_term, 0) + 1
    return hits


def update_glossary_stats(glossary: dict, term_counts: dict) -> None:
    s  = glossary.setdefault("stats", {"total_hits": 0, "term_counts": {}})
    tc = s.setdefault("term_counts", {})
    for term, count in term_counts.items():
        tc[term]           = tc.get(term, 0) + count
        s["total_hits"]    = s.get("total_hits", 0) + count


def _issue_meta(issue_type: str, source: str, translation: str) -> dict:
    """Return severity, category, suggested_fix for a given issue type."""
    if issue_type == "german_residue":
        return {
            "severity":      SEVERITY_CRITICAL,
            "category":      "German residue",
            "suggested_fix": "Retranslate manually or apply glossary corrections",
        }
    if issue_type == "identical_output":
        return {
            "severity":      SEVERITY_CRITICAL,
            "category":      "Missing translation",
            "suggested_fix": "Output is identical to source — retranslate this cell manually",
        }
    if issue_type == "api_failure":
        return {
            "severity":      SEVERITY_CRITICAL,
            "category":      "API failure",
            "suggested_fix": "API failed for this cell — retranslate manually",
        }
    if issue_type == "batch_mismatch":
        return {
            "severity":      SEVERITY_CRITICAL,
            "category":      "Batch output mismatch",
            "suggested_fix": "Batch returned wrong count — retranslate this cell manually",
        }
    if issue_type == "lost_br_tags":
        missing = source.count("<br>") - translation.count("<br>")
        return {
            "severity":      SEVERITY_HIGH,
            "category":      "Formatting issue",
            "suggested_fix": f"Reinsert {missing} missing <br> tag(s) at the correct position(s)",
        }
    if issue_type == "name_too_long":
        return {
            "severity":      SEVERITY_HIGH,
            "category":      "Product name rule violation",
            "suggested_fix": f"Shorten to ≤40 chars (currently {len(translation)}); remove adjectives if needed",
        }
    if issue_type == "name_has_comma":
        return {
            "severity":      SEVERITY_HIGH,
            "category":      "Product name rule violation",
            "suggested_fix": "Remove the comma from the product name",
        }
    if issue_type == "glossary_inconsistency":
        return {
            "severity":      SEVERITY_HIGH,
            "category":      "Glossary inconsistency",
            "suggested_fix": "Apply the standard glossary term listed in the reason field",
        }
    if issue_type == "too_short":
        return {
            "severity":      SEVERITY_MEDIUM,
            "category":      "Suspicious translation length",
            "suggested_fix": "Verify the translation is complete; re-run if truncated",
        }
    if issue_type == "too_long":
        return {
            "severity":      SEVERITY_MEDIUM,
            "category":      "Suspicious translation length",
            "suggested_fix": "Check for duplicated or hallucinated content; translation should not exceed ~2.5× source length",
        }
    if issue_type == "possible_hallucination":
        return {
            "severity":      SEVERITY_MEDIUM,
            "category":      "Possible hallucination",
            "suggested_fix": "Review carefully — unexpected content may have been generated",
        }
    return {
        "severity":      SEVERITY_LOW,
        "category":      "Manual review recommended",
        "suggested_fix": "Review this cell manually",
    }


def _check_glossary_inconsistency(source: str, translation: str, glossary: dict) -> list[str]:
    """Check first 25 glossary terms (those in the model prompt) for missed applications."""
    terms = glossary.get("terms", {})
    prompt_terms = dict(list(terms.items())[:25])
    src_lower = source.lower()
    tr_lower  = translation.lower()
    violations = []
    for de, fr in prompt_terms.items():
        if len(de) <= 5:
            continue
        if re.search(r'\b' + re.escape(de.lower()) + r'\b', src_lower):
            if fr.lower() not in tr_lower:
                violations.append(f"{de} → {fr}")
    return violations[:2]


def compute_quality_score(warnings: list) -> int:
    """Score 0–100: start at 100, deduct per warning severity, floor at 0."""
    score = 100
    for w in warnings:
        score -= SEVERITY_DEDUCTION.get(w.get("severity", SEVERITY_LOW), 1)
    return max(0, score)


def quality_verdict(score: int) -> tuple[str, str]:
    """Return (verdict_text, color_token) for the score."""
    if score >= 95:
        return "Excellent — ready to use", "#22C55E"
    if score >= 85:
        return "Good — minor review recommended", "#0369A1"
    if score >= 70:
        return "Needs review before use", "#f59e0b"
    return "Do not publish without manual review", "#ef4444"


def _normalize_header(header: str) -> set[str]:
    """Split camelCase / PascalCase / snake_case header into lowercase word tokens."""
    no_under = header.replace("_", " ")
    spaced   = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', no_under)
    spaced   = re.sub(r'(?<=[A-Z])(?=[A-Z][a-z])', ' ', spaced)
    return {w.lower() for w in spaced.split() if w} | {header.strip().lower()}


def _normalize_col_header(raw) -> str:
    """
    Full normalization of a raw Excel header for alias matching.
    Handles: camelCase, snake_case, line-breaks, invisible chars, accents.
    Returns a lowercase space-separated string.
    """
    text = str(raw)
    # Replace line breaks, tabs, non-breaking spaces, zero-width chars
    text = re.sub(r'[\n\r\t\xa0​‌‍﻿]', ' ', text)
    # Remove remaining control/format characters (keep spaces)
    text = ''.join(
        c for c in text
        if unicodedata.category(c) not in ('Cc', 'Cf') or c == ' '
    )
    # Expand camelCase / PascalCase: insert space before uppercase run
    text = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)
    text = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', text)
    # Insert space between letter and digit (e.g. Cover1 → Cover 1)
    text = re.sub(r'([a-zA-Z])(\d)', r'\1 \2', text)
    text = re.sub(r'(\d)([a-zA-Z])', r'\1 \2', text)
    # Replace underscores, hyphens, dots with spaces
    text = re.sub(r'[_\-\.]', ' ', text)
    # Normalize unicode — NFKD decomposes characters; then drop combining marks
    # so ä→a, ö→o, ü→u, é→e, etc. (good for cross-language matching)
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    # Collapse whitespace, lowercase
    return ' '.join(text.split()).lower()


# =============================================================================
# DESIGN SYSTEM — CSS
# =============================================================================

def inject_custom_css():
    # ── Background ──────────────────────────────────────────────────────────────
    bg_app    = "#F5F7FA"
    bg_sb     = "#FFFFFF"
    bg_card   = "#FFFFFF"
    bg_input  = "#F8FAFC"
    bg_hover  = "rgba(15,61,158,0.025)"
    bg_subtle = "rgba(15,61,158,0.04)"
    # ── Primary accent ───────────────────────────────────────────────────────────
    primary   = "#0F3D9E"
    primary_d = "#0B2D7E"
    primary_lt= "#EAF2FF"
    # ── Status ────────────────────────────────────────────────────────────────────
    green     = "#12A150"
    green_d   = "#0E8A42"
    green_bg  = "#F0FBF5"
    amber     = "#E6A23C"
    amber_lt  = "#FEF3C7"
    red       = "#EF4444"
    # ── Borders / dividers ────────────────────────────────────────────────────────
    divider   = "#E8ECF2"
    divider_s = "#F0F3F8"
    border    = "#E8ECF2"
    border_sm = "#E8ECF2"
    border_md = "#D1D9E6"
    border_dsh= "#C8D3E6"
    border_hv = "#B0BDD6"
    hover_rb  = "rgba(15,61,158,0.05)"
    # ── Text ──────────────────────────────────────────────────────────────────────
    text      = "#1A2035"
    text2     = "#2D3A52"
    text2b    = "#4A566D"
    text3     = "#6B7A99"
    text4     = "#9BA8BE"
    text5     = "#C3CBDB"
    text6     = "#9BA8BE"
    # ── Component tokens ──────────────────────────────────────────────────────────
    sb_btn    = "#6B7A99"
    code_bg   = "rgba(15,61,158,0.06)"
    code_clr  = "#0F3D9E"
    prog_trk  = "#E8ECF2"
    chip_bg   = "rgba(15,61,158,0.05)"
    chip_bdr  = "#E8ECF2"

    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ── Reset & Base ─────────────────────────────────────────── */
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

    .stApp {{
        background-color: {bg_app} !important;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        color: {text};
        -webkit-font-smoothing: antialiased;
    }}

    [data-testid="stAppViewContainer"],
    [data-testid="stMain"],
    .main {{
        background-color: {bg_app} !important;
    }}

    .main .block-container {{
        padding: 2.5rem 2.5rem 4rem !important;
        max-width: 1080px !important;
    }}

    #MainMenu, footer, header,
    div[data-testid="stDecoration"],
    [data-testid="stToolbar"] {{ display: none !important; visibility: hidden !important; }}

    /* ── Animations ───────────────────────────────────────────── */
    @keyframes fadeUp {{
        from {{ opacity: 0; transform: translateY(14px); }}
        to   {{ opacity: 1; transform: translateY(0); }}
    }}
    @keyframes glow-pulse {{
        0%, 100% {{ box-shadow: 0 0 0 0 rgba(15,61,158,0.40); }}
        50%       {{ box-shadow: 0 0 0 8px rgba(15,61,158,0); }}
    }}
    @keyframes dot-pulse {{
        0%, 100% {{ opacity: 1; transform: scale(1); }}
        50%       {{ opacity: 0.4; transform: scale(0.75); }}
    }}
    @keyframes slide-in {{
        from {{ opacity: 0; transform: translateX(-6px); }}
        to   {{ opacity: 1; transform: translateX(0); }}
    }}

    /* ── Sidebar — permanently visible, never collapsible ──── */
    [data-testid="stSidebar"] {{
        background-color: {bg_sb} !important;
        border-right: 1px solid {divider} !important;
        min-width: 240px !important;
        width: 240px !important;
        transform: translateX(0) !important;
        visibility: visible !important;
        display: flex !important;
        box-shadow: 2px 0 14px rgba(15,61,158,0.04) !important;
    }}

    [data-testid="stSidebarCollapseButton"],
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] {{
        display: none !important;
        visibility: hidden !important;
        pointer-events: none !important;
    }}
    [data-testid="stSidebarContent"] {{ padding: 24px 16px !important; }}

    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] label {{ color: {text2} !important; }}

    [data-testid="stSidebar"] hr {{
        border: none !important;
        border-top: 1px solid {divider} !important;
        margin: 12px 0 !important;
    }}

    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] {{
        font-size: 13.5px !important;
        font-weight: 500 !important;
        padding: 9px 12px !important;
        border-radius: 10px !important;
        cursor: pointer !important;
        transition: background 0.18s !important;
        display: flex !important;
        align-items: center !important;
        margin-bottom: 2px !important;
        background: transparent !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] p {{
        font-size: 13.5px !important;
        font-weight: 500 !important;
        color: {text3} !important;
        transition: color 0.18s !important;
        margin: 0 !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:hover {{
        background: {hover_rb} !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:hover p {{
        color: {text} !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] > div:first-child {{
        display: none !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:has(input:checked) {{
        background: {primary_lt} !important;
        border-radius: 10px !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:has(input:checked) p {{
        color: {primary} !important;
        font-weight: 700 !important;
    }}

    [data-testid="stSidebar"] .stButton > button {{
        background: transparent !important;
        color: {sb_btn} !important;
        border: 1px solid {border} !important;
        border-radius: 8px !important;
        font-size: 12px !important;
        font-weight: 500 !important;
        padding: 7px 14px !important;
        box-shadow: none !important;
        letter-spacing: 0.01em !important;
        transition: color 0.15s, border-color 0.15s, background 0.15s !important;
    }}
    [data-testid="stSidebar"] .stButton > button:hover {{
        color: #EF4444 !important;
        border-color: rgba(239,68,68,0.25) !important;
        background: rgba(239,68,68,0.06) !important;
        transform: none !important;
        box-shadow: none !important;
    }}

    .sb-brand {{ padding: 0 0 20px; }}
    .sb-wordmark {{
        display: flex; align-items: center; gap: 10px;
        font-size: 15px; font-weight: 700; letter-spacing: -0.02em;
        color: {text} !important;
    }}
    .sb-dot {{
        width: 8px; height: 8px; border-radius: 50%;
        background: {green}; flex-shrink: 0;
        box-shadow: 0 0 0 3px rgba(18,161,80,0.15);
    }}
    .sb-org {{
        font-size: 10.5px; font-weight: 600; text-transform: uppercase;
        letter-spacing: 0.10em; color: {text4} !important;
        margin-top: 4px; padding-left: 18px;
    }}
    .sb-nav-label {{
        font-size: 10px; font-weight: 700; text-transform: uppercase;
        letter-spacing: 0.12em; color: {text5} !important;
        padding: 0 12px; margin-bottom: 6px; display: block;
    }}
    .sb-user {{
        background: {bg_subtle};
        border: 1px solid {border};
        border-radius: 10px; padding: 11px 13px; margin: 6px 0;
    }}
    .sb-user-label {{
        font-size: 10px; font-weight: 700; text-transform: uppercase;
        letter-spacing: 0.09em; color: {text5} !important; display: block;
    }}
    .sb-user-email {{
        font-size: 11.5px; color: {text2b} !important;
        margin-top: 4px; display: block; word-break: break-all;
        font-family: Menlo, Monaco, 'Cascadia Code', monospace;
    }}

    /* ── Login page ───────────────────────────────────────────── */
    .login-hero {{
        text-align: center; padding: 56px 0 36px;
        animation: fadeUp 0.5s ease;
    }}
    .login-lockup {{
        display: inline-flex; align-items: center; gap: 9px;
        font-size: 13px; font-weight: 600; color: {text3};
        letter-spacing: 0.07em; text-transform: uppercase;
        margin-bottom: 36px;
    }}
    .login-lockup-dot {{ width: 7px; height: 7px; border-radius: 50%; background: {green}; }}
    .login-title {{
        font-size: 32px; font-weight: 800; color: {primary};
        letter-spacing: -0.04em; margin: 0 0 10px; line-height: 1.15;
    }}
    .login-subtitle {{ font-size: 14px; color: {text3}; font-weight: 400; }}
    .login-footer {{
        text-align: center; font-size: 11px; color: {text6};
        margin-top: 18px; font-weight: 500;
    }}
    .login-form-title {{
        font-size: 22px; font-weight: 800; color: {text};
        letter-spacing: -0.03em;
    }}
    .login-form-sub {{
        font-size: 13px; color: {text4}; margin-top: 5px; font-weight: 400;
    }}

    [data-testid="stForm"] {{
        background: {bg_card} !important;
        border: 1px solid {border_sm} !important;
        border-radius: 20px !important;
        padding: 40px 44px !important;
        box-shadow: 0 4px 28px rgba(15,61,158,0.07) !important;
        animation: fadeUp 0.45s ease 0.08s both;
    }}

    /* ── Page header ──────────────────────────────────────────── */
    .page-hd {{
        padding: 4px 0 28px;
        border-bottom: 1px solid {divider};
        margin-bottom: 36px;
        animation: fadeUp 0.3s ease;
    }}
    .page-hd-title {{
        font-size: 28px; font-weight: 800; color: {primary};
        letter-spacing: -0.04em; line-height: 1.2;
    }}
    .page-hd-sub {{ font-size: 14px; color: {text3}; margin-top: 7px; font-weight: 400; }}

    /* ── Section label ────────────────────────────────────────── */
    .section-label {{
        font-size: 10.5px; font-weight: 700; color: {text4};
        text-transform: uppercase; letter-spacing: 0.12em;
        margin: 36px 0 14px;
    }}

    /* ── Cards ────────────────────────────────────────────────── */
    .card {{
        background: {bg_card};
        border: 1px solid {border};
        border-radius: 20px; padding: 32px; margin: 14px 0;
        box-shadow: 0 2px 14px rgba(15,61,158,0.05);
        animation: fadeUp 0.3s ease;
        transition: border-color 0.2s, box-shadow 0.2s;
    }}
    .card:hover {{ border-color: {border_hv}; box-shadow: 0 6px 24px rgba(15,61,158,0.09); }}
    .card-title {{
        font-size: 10.5px; font-weight: 700; color: {text4};
        text-transform: uppercase; letter-spacing: 0.12em;
        margin-bottom: 20px; padding-bottom: 16px;
        border-bottom: 1px solid {divider};
    }}

    /* ── Alert / message blocks ───────────────────────────────── */
    .alert {{
        display: flex; gap: 11px; align-items: flex-start;
        padding: 13px 16px; border-radius: 9px; margin: 10px 0;
        font-size: 13px; line-height: 1.55;
        animation: fadeUp 0.3s ease;
    }}
    .alert-icon {{ font-size: 14px; flex-shrink: 0; margin-top: 1px; }}
    .alert-info  {{ background: {primary_lt}; border: 1px solid #C0D8FB; color: {primary}; }}
    .alert-success {{ background: {green_bg}; border: 1px solid rgba(18,161,80,0.22); color: {green_d}; }}
    .alert-warn  {{ background: {amber_lt}; border: 1px solid rgba(230,162,60,0.28); color: #7A4A0F; }}
    .alert strong {{ color: {text}; font-weight: 600; }}
    .alert code {{
        font-family: Menlo, Monaco, monospace; font-size: 11px;
        background: {code_bg}; padding: 1px 5px; border-radius: 4px;
        color: {code_clr};
    }}

    /* ── Stat result cards ────────────────────────────────────── */
    .result-grid {{ display: grid; grid-template-columns: repeat(4,1fr); gap: 14px; margin: 20px 0; }}
    .result-card {{
        background: {bg_card}; border: 1px solid {border};
        border-radius: 16px; padding: 24px 22px;
        box-shadow: 0 2px 12px rgba(15,61,158,0.05);
        transition: border-color 0.2s, transform 0.2s, box-shadow 0.2s;
        animation: fadeUp 0.35s ease;
    }}
    .result-card:hover {{ border-color: {border_hv}; transform: translateY(-2px); box-shadow: 0 8px 24px rgba(15,61,158,0.09); }}
    .result-card-label {{
        font-size: 10.5px; font-weight: 600; color: {text4};
        text-transform: uppercase; letter-spacing: 0.09em; margin-bottom: 13px;
    }}
    .result-card-value {{
        font-size: 32px; font-weight: 800; letter-spacing: -0.04em;
        font-variant-numeric: tabular-nums; color: {text};
    }}
    .result-card-value.accent  {{ color: {primary}; }}
    .result-card-value.success {{ color: {green}; }}
    .result-card-value.warn    {{ color: {amber}; }}

    /* ── Column chips ─────────────────────────────────────────── */
    .chip {{
        display: inline-flex; align-items: center; gap: 5px;
        padding: 3px 10px; border-radius: 6px;
        font-size: 11px; font-weight: 600;
        font-family: Menlo, Monaco, 'Cascadia Code', monospace;
        margin: 3px 3px 3px 0;
    }}
    .chip-accent {{
        background: #DCFCE7;
        border: 1px solid #BBF7D0;
        color: #16A34A;
    }}
    .chip-muted {{
        background: {chip_bg};
        border: 1px solid {chip_bdr};
        color: {text3};
    }}
    .chip-arrow {{ color: {text3}; font-family: sans-serif; font-weight: 400; }}

    /* ── File chip ────────────────────────────────────────────── */
    .file-chip {{
        display: inline-flex; align-items: center; gap: 8px;
        background: {primary_lt};
        border: 1px solid #C0D8FB;
        color: {primary}; padding: 6px 16px; border-radius: 20px;
        font-size: 12px; font-weight: 600; margin: 8px 0;
        font-family: Menlo, Monaco, 'Cascadia Code', monospace;
        animation: slide-in 0.25s ease;
    }}

    /* ── Progress shell ───────────────────────────────────────── */
    .prog-shell {{
        background: {bg_card};
        border: 1px solid {border};
        border-radius: 20px; padding: 34px 36px; margin: 16px 0;
        animation: fadeUp 0.3s ease;
        box-shadow: 0 2px 16px rgba(15,61,158,0.05);
    }}
    .prog-head {{
        display: flex; align-items: center;
        justify-content: space-between; margin-bottom: 24px;
    }}
    .prog-phase {{
        font-size: 13px; font-weight: 700; color: {text};
        text-transform: uppercase; letter-spacing: 0.08em;
    }}
    .prog-sheet {{ font-size: 12px; color: {text4}; margin-top: 4px; }}
    .prog-badge {{
        display: inline-flex; align-items: center; gap: 6px;
        padding: 4px 12px; border-radius: 20px;
        font-size: 10.5px; font-weight: 700; letter-spacing: 0.06em;
        background: {green_bg};
        border: 1px solid rgba(18,161,80,0.22);
        color: {green_d};
    }}
    .prog-badge-dot {{
        width: 6px; height: 6px; border-radius: 50%; background: {green};
        animation: dot-pulse 1.4s ease infinite;
    }}
    .prog-track {{
        width: 100%; height: 4px;
        background: {prog_trk};
        border-radius: 3px; overflow: hidden; margin: 18px 0;
        position: relative;
    }}
    .prog-bar {{
        height: 4px; border-radius: 3px;
        background: linear-gradient(90deg, {primary} 0%, #3B6EDE 100%);
        transition: width 0.45s cubic-bezier(0.4,0,0.2,1); position: relative;
    }}
    .prog-bar::after {{
        content: ''; position: absolute; right: -1px; top: -3px;
        width: 10px; height: 10px; background: {primary};
        border-radius: 50%; animation: glow-pulse 1.6s ease infinite;
        box-shadow: 0 0 0 3px {primary_lt};
    }}
    .prog-item {{
        display: flex; align-items: center; gap: 9px;
        font-size: 12.5px; color: {text3}; margin: 11px 0;
        font-family: Menlo, Monaco, 'Cascadia Code', monospace;
    }}
    .prog-item-dot {{
        width: 6px; height: 6px; border-radius: 50%; background: {primary}; flex-shrink: 0;
        animation: dot-pulse 1.4s ease infinite;
    }}
    .prog-item-col {{ color: {primary}; }}
    .prog-item-row {{ color: {text4}; margin-left: 6px; }}
    .prog-stats {{
        display: flex; gap: 32px; margin-top: 22px; padding-top: 18px;
        border-top: 1px solid {divider_s};
        flex-wrap: wrap;
    }}
    .prog-stat-val {{
        font-size: 16px; font-weight: 800; color: {text};
        font-variant-numeric: tabular-nums; display: block;
    }}
    .prog-stat-lbl {{
        font-size: 9.5px; font-weight: 700; color: {text4};
        text-transform: uppercase; letter-spacing: 0.10em;
        margin-top: 3px; display: block;
    }}

    /* ── Quality gate ─────────────────────────────────────────── */
    .qg {{
        background: {bg_card}; border: 1px solid {border};
        border-radius: 14px; overflow: hidden; margin: 14px 0;
        animation: fadeUp 0.35s ease;
        box-shadow: 0 1px 8px rgba(15,61,158,0.04);
    }}
    .qg-row {{
        display: flex; align-items: center; gap: 16px;
        padding: 14px 22px;
        border-bottom: 1px solid {divider_s};
        font-size: 13px;
        transition: background 0.15s;
    }}
    .qg-row:last-child {{ border-bottom: none; }}
    .qg-row:hover {{ background: {bg_hover}; }}
    .qg-icon {{ flex-shrink: 0; font-size: 13px; }}
    .qg-label {{ font-weight: 600; color: {text2}; min-width: 160px; font-size: 12.5px; }}
    .qg-value {{ color: {text3}; font-size: 12.5px; font-family: Menlo, Monaco, monospace; }}

    /* ── Warning detail ───────────────────────────────────────── */
    .warn-detail {{
        display: flex; gap: 12px; align-items: flex-start;
        padding: 14px 18px; margin: 8px 0;
        background: rgba(230,162,60,0.04);
        border: 1px solid rgba(230,162,60,0.12);
        border-radius: 11px; font-size: 12.5px; color: {text2};
        animation: fadeUp 0.3s ease;
    }}
    .warn-detail-dot {{
        width: 5px; height: 5px; border-radius: 50%;
        background: {amber}; margin-top: 5px; flex-shrink: 0;
    }}
    .warn-detail strong {{ color: #b07020; }}

    /* ── Success / completion banner ──────────────────────────── */
    .success-banner {{
        display: flex; align-items: center; gap: 18px;
        padding: 22px 28px;
        background: {green_bg};
        border: 1px solid rgba(18,161,80,0.22);
        border-radius: 16px; margin: 18px 0;
        animation: fadeUp 0.3s ease;
    }}
    .success-banner-icon {{
        width: 44px; height: 44px; border-radius: 50%;
        background: rgba(18,161,80,0.12);
        border: 1px solid rgba(18,161,80,0.22);
        display: flex; align-items: center; justify-content: center;
        font-size: 20px; flex-shrink: 0;
    }}
    .success-banner-title {{ font-size: 15px; font-weight: 700; color: {green_d}; }}
    .success-banner-sub   {{ font-size: 12.5px; color: rgba(14,138,66,0.75); margin-top: 4px; }}

    .warn-banner {{
        padding: 18px 24px;
        background: rgba(230,162,60,0.04);
        border: 1px solid rgba(230,162,60,0.14);
        border-radius: 13px; margin: 18px 0;
        animation: fadeUp 0.3s ease;
    }}
    .warn-banner-title {{ font-size: 14px; font-weight: 700; color: {amber}; }}
    .warn-banner-sub   {{ font-size: 11.5px; color: {text2}; margin-top: 4px; }}

    /* ── Metric cards ─────────────────────────────────────────── */
    .kpi-row   {{ display: grid; grid-template-columns: repeat(4,1fr); gap: 14px; margin: 22px 0; }}
    .kpi-row-3 {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 14px; margin: 22px 0; }}
    .kpi {{
        background: {bg_card};
        border: 1px solid {border};
        border-radius: 16px; padding: 26px 24px;
        box-shadow: 0 2px 12px rgba(15,61,158,0.05);
        transition: border-color 0.2s, transform 0.2s, box-shadow 0.2s;
        animation: fadeUp 0.35s ease;
    }}
    .kpi:hover {{ border-color: {border_hv}; transform: translateY(-2px); box-shadow: 0 8px 24px rgba(15,61,158,0.09); }}
    .kpi-label {{
        font-size: 10.5px; font-weight: 600; color: {text4};
        text-transform: uppercase; letter-spacing: 0.09em; margin-bottom: 13px;
    }}
    .kpi-value {{
        font-size: 32px; font-weight: 800; letter-spacing: -0.04em;
        color: {text}; font-variant-numeric: tabular-nums;
    }}
    .kpi-value.accent  {{ color: {primary}; }}
    .kpi-value.success {{ color: {green}; }}
    .kpi-value.warn    {{ color: {amber}; }}
    .kpi-value.muted   {{ color: {text4}; }}
    .kpi-sub {{ font-size: 11.5px; color: {text4}; margin-top: 7px; }}

    /* ── Hero metric ──────────────────────────────────────────── */
    .hero-kpi {{
        text-align: center; padding: 60px 32px; border-radius: 20px;
        background: linear-gradient(135deg, {primary_lt} 0%, {green_bg} 100%);
        border: 1px solid #C0D8FB; margin: 22px 0;
        animation: fadeUp 0.4s ease;
    }}
    .hero-kpi-value {{
        font-size: 80px; font-weight: 800; letter-spacing: -0.05em; line-height: 1;
        background: linear-gradient(135deg, {primary} 0%, #2563EB 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text;
        margin: 0;
    }}
    .hero-kpi-label {{ font-size: 15px; color: {text3}; margin: 16px 0 0; font-weight: 500; }}
    .hero-kpi-sub   {{ font-size: 12px; color: {text4}; margin: 7px 0 0; }}

    /* ── History ──────────────────────────────────────────────── */
    .history-empty {{
        text-align: center; padding: 70px 20px;
        color: {text4}; font-size: 14px; font-weight: 500;
    }}
    .history-empty-sub {{ font-size: 12px; color: {text5}; }}
    .history-empty-sub strong {{ color: {text4}; }}
    .cloud-note {{
        padding: 12px 16px; border-radius: 11px; margin: 14px 0;
        font-size: 11.5px; line-height: 1.55;
        background: {primary_lt};
        border: 1px solid #C0D8FB;
        color: {primary};
    }}

    /* ── Footer ───────────────────────────────────────────────── */
    .footer-author {{ color: {text3} !important; }}
    .footer-version {{ color: {text6} !important; }}

    /* ── Streamlit native overrides ───────────────────────────── */
    [data-testid="stTextInput"] input {{
        background: {bg_input} !important;
        border: 1px solid {border_md} !important;
        border-radius: 10px !important;
        color: {text} !important;
        font-size: 13.5px !important;
        padding: 11px 14px !important;
        caret-color: {primary} !important;
    }}
    [data-testid="stTextInput"] input:focus {{
        border-color: rgba(15,61,158,0.45) !important;
        box-shadow: 0 0 0 3px rgba(15,61,158,0.10) !important;
        outline: none !important;
    }}
    [data-testid="stTextInput"] label p {{ color: {text2b} !important; font-size: 12.5px !important; font-weight: 500 !important; }}

    [data-testid="stNumberInput"] input {{
        background: {bg_input} !important;
        border: 1px solid {border_md} !important;
        border-radius: 9px !important;
        color: {text} !important;
        font-size: 13px !important;
    }}
    [data-testid="stNumberInput"] label p {{ color: {text2b} !important; font-size: 12.5px !important; }}

    [data-testid="stSelectbox"] > div > div {{
        background: {bg_input} !important;
        border: 1px solid {border_md} !important;
        border-radius: 9px !important;
        color: {text} !important;
        font-size: 13px !important;
    }}
    [data-testid="stSelectbox"] label p {{ color: {text2b} !important; font-size: 12.5px !important; }}

    [data-testid="stTextarea"] textarea {{
        background: {bg_input} !important;
        border: 1px solid {border_md} !important;
        border-radius: 10px !important;
        color: {text} !important;
        font-size: 13.5px !important;
        line-height: 1.6 !important;
        caret-color: {primary} !important;
    }}
    [data-testid="stTextarea"] textarea:focus {{
        border-color: rgba(15,61,158,0.40) !important;
        box-shadow: 0 0 0 3px rgba(15,61,158,0.08) !important;
        outline: none !important;
    }}
    [data-testid="stTextarea"] label p {{ color: {text2b} !important; font-size: 12.5px !important; font-weight: 500 !important; }}

    .stProgress {{ padding: 6px 0 !important; }}
    .stProgress > div > div > div {{
        background: linear-gradient(90deg, {primary} 0%, #3B6EDE 100%) !important;
        border-radius: 3px !important;
    }}
    .stProgress > div > div {{
        border-radius: 3px !important;
        background: {prog_trk} !important;
    }}

    .stButton > button {{
        background: {primary} !important;
        color: #fff !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 12px 28px !important;
        font-weight: 600 !important;
        font-size: 13.5px !important;
        letter-spacing: 0.01em !important;
        transition: background 0.15s, transform 0.15s, box-shadow 0.15s !important;
        box-shadow: 0 2px 14px rgba(15,61,158,0.22) !important;
    }}
    .stButton > button:hover {{
        background: {primary_d} !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 20px rgba(15,61,158,0.32) !important;
    }}
    .stButton > button:active {{ transform: translateY(0) !important; }}

    .stDownloadButton > button {{
        background: #FFFFFF !important;
        color: {primary} !important;
        border: 1.5px solid {primary} !important;
        border-radius: 12px !important;
        padding: 12px 28px !important;
        font-weight: 600 !important;
        font-size: 13.5px !important;
        box-shadow: none !important;
        transition: background 0.15s, border-color 0.15s, transform 0.15s !important;
    }}
    .stDownloadButton > button:hover {{
        background: {primary_lt} !important;
        border-color: {primary_d} !important;
        transform: translateY(-1px) !important;
    }}

    [data-testid="stFileUploader"] section,
    [data-testid="stFileUploader"] > div {{
        background: {bg_hover} !important;
        border: 1.5px dashed {border_dsh} !important;
        border-radius: 14px !important;
        transition: border-color 0.2s, background 0.2s !important;
        padding: 24px !important;
    }}
    [data-testid="stFileUploader"] section:hover,
    [data-testid="stFileUploader"] > div:hover {{
        border-color: rgba(15,61,158,0.35) !important;
        background: rgba(15,61,158,0.03) !important;
    }}
    [data-testid="stFileUploader"] span,
    [data-testid="stFileUploader"] p {{ color: {text3} !important; font-size: 13px !important; }}
    [data-testid="stFileUploader"] small {{ color: {text5} !important; }}
    [data-testid="stFileUploader"] button {{
        background: {primary_lt} !important;
        color: {primary} !important;
        border: 1px solid #C0D8FB !important;
        border-radius: 9px !important;
        font-size: 12.5px !important;
        font-weight: 600 !important;
    }}

    [data-testid="stExpander"] {{
        background: {bg_card} !important;
        border: 1px solid {border} !important;
        border-radius: 12px !important;
    }}
    [data-testid="stExpander"] summary {{
        color: {text3} !important;
        font-size: 12.5px !important;
        font-weight: 600 !important;
        padding: 12px 16px !important;
    }}

    [data-testid="stDataFrame"] {{
        border-radius: 13px !important;
        overflow: hidden !important;
        border: 1px solid {border} !important;
    }}

    .site-footer {{
        margin-top: 80px; padding: 22px 0;
        border-top: 1px solid {divider_s};
        display: flex; align-items: center; justify-content: space-between;
        font-size: 11px; color: {text6};
        animation: fadeUp 0.4s ease;
    }}

    /* ── Severity badges ─────────────────────────────────── */
    .sev-badge {{
        display: inline-flex; align-items: center;
        padding: 2px 8px; border-radius: 4px;
        font-size: 11px; font-weight: 700;
        letter-spacing: 0.04em; text-transform: uppercase;
    }}
    .sev-critical {{
        background: rgba(239,68,68,0.12);
        border: 1px solid rgba(239,68,68,0.22);
        color: #ef4444;
    }}
    .sev-high {{
        background: rgba(245,158,11,0.12);
        border: 1px solid rgba(245,158,11,0.22);
        color: #f59e0b;
    }}
    .sev-medium {{
        background: rgba(250,204,21,0.10);
        border: 1px solid rgba(250,204,21,0.20);
        color: #ca8a04;
    }}
    .sev-low {{
        background: rgba(129,140,248,0.10);
        border: 1px solid rgba(129,140,248,0.20);
        color: #818cf8;
    }}
    .alert-success {{
        background: {green_bg};
        border: 1px solid rgba(18,161,80,0.22);
        color: {green_d};
    }}

    /* ── Translator hero card ────────────────────────────────── */
    .tr-hero {{
        background: {bg_card};
        border: 1px solid {border};
        border-radius: 24px; padding: 44px 48px; margin-bottom: 36px;
        box-shadow: 0 4px 28px rgba(15,61,158,0.06);
        animation: fadeUp 0.35s ease;
        position: relative; overflow: hidden;
    }}
    .tr-hero::before {{
        content: ''; position: absolute;
        top: -70px; right: -70px;
        width: 220px; height: 220px; border-radius: 50%;
        background: radial-gradient(circle, {primary_lt} 0%, transparent 70%);
        pointer-events: none;
    }}
    .tr-hero-tag {{
        font-size: 10.5px; font-weight: 700; text-transform: uppercase;
        letter-spacing: 0.12em; color: {primary}; margin-bottom: 12px;
    }}
    .tr-hero-title {{
        font-size: 30px; font-weight: 800; color: {primary};
        letter-spacing: -0.03em; margin: 0 0 12px; line-height: 1.2;
    }}
    .tr-hero-sub {{
        font-size: 15px; color: {text3}; margin: 0; font-weight: 400;
        max-width: 520px; line-height: 1.65;
    }}

    @media (max-width: 780px) {{
        .kpi-row, .result-grid {{ grid-template-columns: repeat(2,1fr) !important; }}
        .hero-kpi-value {{ font-size: 52px !important; }}
        .main .block-container {{ padding: 1.5rem 1.4rem 3rem !important; }}
    }}
    </style>
    """, unsafe_allow_html=True)


# =============================================================================
# UI COMPONENTS
# =============================================================================

def render_header():
    st.markdown("""
    <div class="login-hero">
        <div class="login-lockup">
            <div class="login-lockup-dot"></div>
            Home24 AI Localization
        </div>
        <h1 class="login-title">Welcome back</h1>
        <p class="login-subtitle">Sign in to access your localization workspace</p>
    </div>
    """, unsafe_allow_html=True)


def render_page_header(title: str, subtitle: str = ""):
    sub = f'<div class="page-hd-sub">{subtitle}</div>' if subtitle else ""
    st.markdown(f"""
    <div class="page-hd">
        <div class="page-hd-title">{title}</div>
        {sub}
    </div>
    """, unsafe_allow_html=True)


def render_stats(stats: dict):
    warn = "warn" if stats["unresolved_warnings"] > 0 else "success"
    st.markdown(f"""
    <div class="result-grid">
        <div class="result-card">
            <div class="result-card-label">Cells Translated</div>
            <div class="result-card-value accent">{stats["cells_translated"]}</div>
        </div>
        <div class="result-card">
            <div class="result-card-label">Cells Skipped</div>
            <div class="result-card-value">{stats["cells_skipped"]}</div>
        </div>
        <div class="result-card">
            <div class="result-card-label">Residue Fixed</div>
            <div class="result-card-value success">{stats["residue_corrections"]}</div>
        </div>
        <div class="result-card">
            <div class="result-card-label">Warnings</div>
            <div class="result-card-value {warn}">{stats["unresolved_warnings"]}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_intelligence_stats(stats: dict):
    tm_hits      = stats.get("tm_hits", 0)
    batch_count  = stats.get("batch_count", 0)
    avg_batch    = stats.get("avg_batch_size", 0.0)
    gloss_hits   = stats.get("glossary_hits", 0)
    api_reduced  = stats.get("api_calls_reduced", 0)
    concurrency  = stats.get("max_concurrent_used", 1)
    avg_dur      = stats.get("avg_batch_duration", 0.0)
    failed_b     = stats.get("failed_batches", 0)
    refined      = stats.get("cells_refined", 0)
    refine_calls = stats.get("refinement_api_calls", 0)
    refine_on    = stats.get("refinement_enabled", False)

    conc_sub    = f"{concurrency}× parallel" if concurrency > 1 else "sequential mode"
    refine_sub  = f"{refine_calls} refinement batch(es)" if refine_on and refine_calls else ("disabled" if not refine_on else "—")

    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-label">TM Cache Hits</div>
            <div class="kpi-value success">{tm_hits}</div>
            <div class="kpi-sub">Served from memory</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">API Batches Sent</div>
            <div class="kpi-value accent">{batch_count}</div>
            <div class="kpi-sub">~{api_reduced} calls saved</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Avg Batch Size</div>
            <div class="kpi-value">{avg_batch}</div>
            <div class="kpi-sub">cells per request</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Glossary Matches</div>
            <div class="kpi-value">{gloss_hits}</div>
            <div class="kpi-sub">Consistent terms enforced</div>
        </div>
    </div>
    <div class="kpi-row" style="margin-top:8px;">
        <div class="kpi">
            <div class="kpi-label">Concurrency</div>
            <div class="kpi-value accent">{concurrency}</div>
            <div class="kpi-sub">{conc_sub}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Failed Batches</div>
            <div class="kpi-value {'warn' if failed_b else 'success'}">{failed_b}</div>
            <div class="kpi-sub">{'Fell back to single-cell' if failed_b else 'All batches succeeded'}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Premium Refinement</div>
            <div class="kpi-value {'accent' if refined else 'muted'}">{refined}</div>
            <div class="kpi-sub">{refine_sub}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Retries</div>
            <div class="kpi-value">{stats.get("retry_count", 0)}</div>
            <div class="kpi-sub">API retry events</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    forbidden = stats.get("forbidden_corrections", 0)
    ctx_recs  = stats.get("context_reconstructions", 0)

    if forbidden > 0 or ctx_recs > 0:
        st.markdown(f"""
        <div class="kpi-row" style="margin-top:8px;">
            <div class="kpi">
                <div class="kpi-label">Forbidden Patterns Fixed</div>
                <div class="kpi-value {'accent' if forbidden else 'muted'}">{forbidden}</div>
                <div class="kpi-sub">Quality corrections applied</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Context Reconstructions</div>
                <div class="kpi-value">{ctx_recs}</div>
                <div class="kpi-sub">Row contexts built</div>
            </div>
        </div>
        """, unsafe_allow_html=True)


def format_time(seconds: float) -> str:
    return str(timedelta(seconds=int(seconds)))[2:7]


def render_column_report(classification: dict):
    to_translate    = classification["to_translate"]
    protected       = classification["protected"]
    ignored         = classification["ignored"]
    possible_missed = classification["possible_missed"]
    normalized_map  = classification.get("normalized_map", {})
    header_row      = classification.get("header_row", 1)

    if to_translate:
        chips = ""
        for header, (_, canonical) in to_translate.items():
            if header == canonical:
                chips += f'<span class="chip chip-accent">{header}</span>'
            else:
                chips += (
                    f'<span class="chip chip-accent">{header}'
                    f'<span class="chip-arrow"> → {canonical}</span></span>'
                )
        will_translate_html = f"""
        <div style="margin-bottom:16px;">
            <div class="section-label" style="margin-top:0;">Will translate — {len(to_translate)} column(s)</div>
            <div>{chips}</div>
        </div>"""
    else:
        will_translate_html = """
        <div class="alert alert-warn">
            <span class="alert-icon">⚠</span>
            <span>No translatable columns detected automatically — see debug info below.</span>
        </div>"""

    prot_html = ""
    if protected:
        prot_chips = "".join(f'<span class="chip chip-muted">{h}</span>' for h in protected)
        prot_html = f"""
        <div style="margin-top:14px;">
            <div class="section-label" style="margin-top:0;">Protected — never modified</div>
            <div>{prot_chips}</div>
        </div>"""

    missed_html = ""
    if possible_missed:
        missed_chips = "".join(f'<span class="chip chip-muted">{h}</span>' for h in possible_missed)
        missed_html = f"""
        <div class="alert alert-warn" style="margin-top:14px;">
            <span class="alert-icon">⚠</span>
            <div>
                <strong>Possible missed columns:</strong><br>
                <div style="margin-top:6px;">{missed_chips}</div>
                <div style="font-size:11px;color:#92400E;margin-top:5px;">
                    These look like translatable columns but didn't match any known pattern.
                </div>
            </div>
        </div>"""

    row_note = f' <span style="font-size:11px;color:#64748B;">(headers detected in row {header_row})</span>' if header_row != 1 else ""

    st.markdown(f"""
    <div class="card">
        <div class="card-title">Column detection{row_note}</div>
        {will_translate_html}
        {prot_html}
        {missed_html}
    </div>
    """, unsafe_allow_html=True)

    if ignored:
        with st.expander(f"Ignored columns ({len(ignored)})"):
            st.markdown(", ".join(f"`{h}`" for h in ignored))

    # Debug panel — shown only when automatic detection found nothing
    if not to_translate:
        with st.expander("🔍 Detection debug report", expanded=True):
            ws_info = classification.get("ws_info", {})

            real_rows = ws_info.get("real_max_row") or ws_info.get("max_row", "—")
            real_cols = ws_info.get("real_max_col") or ws_info.get("max_column", "—")
            opl_rows  = ws_info.get("openpyxl_max_row", ws_info.get("max_row", "—"))
            opl_cols  = ws_info.get("openpyxl_max_col", ws_info.get("max_column", "—"))

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Header row", header_row)
            c2.metric("Cols detected", ws_info.get("columns_found", len(normalized_map)))
            c3.metric("Real rows", real_rows)
            c4.metric("Real cols", real_cols)
            c5.metric("openpyxl rows", opl_rows, delta=None,
                      help="openpyxl's max_row from XML metadata — can be wrong if file metadata is stale")

            if real_rows != opl_rows or real_cols != opl_cols:
                st.info(
                    f"openpyxl reported {opl_rows} rows × {opl_cols} cols, "
                    f"but the actual data scan found {real_rows} rows × {real_cols} cols. "
                    "This is normal for files with stale workbook metadata."
                )

            all_raw = list(normalized_map.keys())
            if all_raw:
                all_col_idx = {}
                all_col_idx.update(protected)
                all_col_idx.update(ignored)
                match_tiers = classification.get("match_tiers", {})
                rows = [
                    {
                        "Col #":        all_col_idx.get(h, ""),
                        "Raw header":   h,
                        "Normalized":   normalized_map.get(h, ""),
                        "Status":       "Protected" if h in protected else "Ignored",
                        "Match tier":   match_tiers.get(h, ""),
                        "Reason":       "Protected column" if h in protected
                                        else classification.get("ignored_reasons", {}).get(h, ""),
                    }
                    for h in all_raw
                ]
                st.dataframe(rows, use_container_width=True, hide_index=True)
                st.caption(
                    "H24-exact / H24-collapsed = Home24 registry match. "
                    "T1-exact = alias match. T2-substring = keyword match. "
                    "T3-wordset = word-overlap match."
                )
            else:
                st.warning(
                    f"No headers found in row {header_row}. "
                    f"Real data scan found {real_rows} rows × {real_cols} cols. "
                    "Try the manual column selector below."
                )


_EXCEL_EXTENSIONS = (".xlsx", ".xls", ".xlsm", ".xlsb")


def jira_tickets_page() -> None:
    role = st.session_state.get("user_role", "")
    if role == "guest":
        st.markdown(
            '<div class="alert alert-warn">'
            '<span class="alert-icon">⚠</span>'
            '<span>Jira integration is not available for Guest accounts.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    render_page_header(
        "Jira Tickets",
        "Search tickets · download source files · upload translations",
    )

    # ── Connection status ─────────────────────────────────────────────────────
    _conn_key = "_jira_conn_status"
    if st.button("Test connection", key="jira_test_conn_btn"):
        st.session_state.pop(_conn_key, None)

    if _conn_key not in st.session_state:
        if not jira_configured():
            st.session_state[_conn_key] = {
                "ok": False, "name": "", "email": "",
                "error": (
                    "Jira credentials not configured. "
                    "Add JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN "
                    "to your Streamlit secrets or .env file."
                ),
            }
        else:
            _jc0, _jerr0 = get_jira_client()
            st.session_state[_conn_key] = (
                _jc0.test_connection() if _jc0
                else {"ok": False, "name": "", "email": "", "error": _jerr0}
            )

    _conn = st.session_state[_conn_key]
    if _conn["ok"]:
        st.markdown(
            f'<div class="alert alert-info">'
            f'<span class="alert-icon">✓</span>'
            f'<span>Connected as <strong>{_conn["name"]}</strong> '
            f'({_conn["email"]})</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div class="alert alert-warn">'
            f'<span class="alert-icon">⚠</span>'
            f'<span><strong>Not connected:</strong> {_conn["error"]}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
        with st.expander("Setup instructions", expanded=True):
            st.markdown("""
**Local development** — add to `.env`:
```
JIRA_BASE_URL=https://your-company.atlassian.net
JIRA_EMAIL=your-email@home24.de
JIRA_API_TOKEN=your-api-token-here
```
**Streamlit Cloud** — add to Secrets panel:
```toml
JIRA_BASE_URL  = "https://your-company.atlassian.net"
JIRA_EMAIL     = "your-email@home24.de"
JIRA_API_TOKEN = "your-api-token-here"
```
Get your API token at https://id.atlassian.com/manage-profile/security/api-tokens
            """)
        return

    st.markdown("---")

    # Helper: run a JQL search and store results in session_state
    def _run_search(jql: str, append: bool = False) -> None:
        _jc_s, _jerr_s = get_jira_client()
        if not _jc_s:
            st.error(f"Connection failed: {_jerr_s}")
            return
        _start  = 0
        _token  = None
        if append and "_jira_search_results" in st.session_state:
            _prev = st.session_state["_jira_search_results"]
            _start = _prev.get("start_at", 0)
            _token = _prev.get("next_page_token")
        with st.spinner("Searching Jira..."):
            _res = _jc_s.search_issues(
                jql,
                max_results=50,
                start_at=_start,
                next_page_token=_token,
            )
        _res["_jql"] = jql
        if append and "_jira_search_results" in st.session_state:
            _prev_issues = st.session_state["_jira_search_results"].get("issues", [])
            _res["issues"] = _prev_issues + _res.get("issues", [])
        st.session_state["_jira_search_results"] = _res
        st.session_state["_jira_jql_saved"] = jql

    # Helper: load a saved filter and search with its JQL
    def _run_filter_search(filter_id: str, filter_name: str, append: bool = False) -> None:
        _jc_f, _jerr_f = get_jira_client()
        if not _jc_f:
            st.error(f"Connection failed: {_jerr_f}")
            return
        _f = _jc_f.get_filter(filter_id)
        if not _f:
            st.error(f"Filter '{filter_name}' (ID {filter_id}) not found or inaccessible.")
            return
        _run_search(_f["jql"], append=append)

    # ── Home24 saved filters ──────────────────────────────────────────────────
    st.markdown(
        '<div class="section-label">Home24 Jira Filters</div>',
        unsafe_allow_html=True,
    )

    _filter_ready_id = _get_secret("JIRA_FILTER_READY_FR", "")
    _filter_inprog_id = _get_secret("JIRA_FILTER_IN_PROGRESS_FR", "")

    _FILTER_READY_NAME    = "FR product translation - ready for translation"
    _FILTER_INPROG_NAME   = "FR product translation - in progress"

    _fb1, _fb2, _fb3 = st.columns(3)

    with _fb1:
        if st.button(
            "Load: Ready for translation",
            key="jira_filter_ready_btn",
            use_container_width=True,
        ):
            st.session_state.pop("_jira_search_results", None)
            if _filter_ready_id:
                _run_filter_search(_filter_ready_id, _FILTER_READY_NAME)
            else:
                # Search saved filters by name
                _jc_f2, _ = get_jira_client()
                if _jc_f2:
                    _matches = _jc_f2.search_filters(_FILTER_READY_NAME)
                    if _matches:
                        _run_search(_matches[0]["jql"])
                        st.session_state["_jira_filter_label"] = _matches[0]["name"]
                    else:
                        # Fallback JQL for Home24 FR translation workflow
                        _run_search('summary ~ "FR -" AND status != Done ORDER BY updated DESC')
                        st.session_state["_jira_filter_label"] = "JQL fallback (filter not found)"

    with _fb2:
        if st.button(
            "Load: In progress",
            key="jira_filter_inprog_btn",
            use_container_width=True,
        ):
            st.session_state.pop("_jira_search_results", None)
            if _filter_inprog_id:
                _run_filter_search(_filter_inprog_id, _FILTER_INPROG_NAME)
            else:
                _jc_f3, _ = get_jira_client()
                if _jc_f3:
                    _matches2 = _jc_f3.search_filters(_FILTER_INPROG_NAME)
                    if _matches2:
                        _run_search(_matches2[0]["jql"])
                        st.session_state["_jira_filter_label"] = _matches2[0]["name"]
                    else:
                        _run_search('summary ~ "FR -" AND status ~ "progress" ORDER BY updated DESC')
                        st.session_state["_jira_filter_label"] = "JQL fallback (filter not found)"

    with _fb3:
        if st.button(
            "Search filters by name…",
            key="jira_filter_search_btn",
            use_container_width=True,
        ):
            st.session_state["_jira_show_filter_search"] = True

    # Filter name search (shown when user clicks "Search filters by name…")
    if st.session_state.get("_jira_show_filter_search"):
        _fname_q = st.text_input(
            "Filter name to search:",
            value="FR product translation",
            key="jira_filter_name_input",
        )
        if st.button("Find filters", key="jira_find_filters_btn"):
            _jc_fs, _ = get_jira_client()
            if _jc_fs:
                _found_filters = _jc_fs.search_filters(_fname_q, max_results=20)
                st.session_state["_jira_found_filters"] = _found_filters
        _found = st.session_state.get("_jira_found_filters", [])
        if _found:
            _filt_options = {f["name"]: f for f in _found}
            _chosen_fname = st.selectbox(
                "Select filter:", list(_filt_options.keys()), key="jira_filter_pick"
            )
            _chosen_f = _filt_options.get(_chosen_fname)
            if _chosen_f and st.button("Use this filter", key="jira_use_filter_btn"):
                st.session_state.pop("_jira_search_results", None)
                _run_search(_chosen_f["jql"])
                st.session_state["_jira_filter_label"] = _chosen_f["name"]
                st.session_state["_jira_show_filter_search"] = False
                st.rerun()
        elif "jira_found_filters" in st.session_state:
            st.info("No matching filters found.")

    st.markdown("---")

    # ── Manual JQL search ─────────────────────────────────────────────────────
    st.markdown('<div class="section-label">Search by JQL</div>', unsafe_allow_html=True)

    _default_jql = _get_secret(
        "JIRA_TRANSLATION_JQL",
        'summary ~ "FR -" ORDER BY updated DESC',
    )
    _jql_input = st.text_input(
        "JQL query:",
        value=st.session_state.get("_jira_jql_saved", _default_jql),
        key="jira_jql_input",
        help=(
            'Examples:\n'
            '  summary ~ "FR -" ORDER BY updated DESC\n'
            '  summary ~ "FR" AND status ~ "ready" ORDER BY updated DESC\n'
            '  key in (PC-20820, PC-20815)'
        ),
    )

    _sc1, _sc2, _sc3 = st.columns([2, 1, 1])
    with _sc1:
        _do_search = st.button("Search", key="jira_search_btn", use_container_width=True)
    with _sc2:
        if st.button("Clear", key="jira_clear_results_btn", use_container_width=True):
            for _k in ["_jira_search_results", "_jira_filter_label", "_jira_jql_saved"]:
                st.session_state.pop(_k, None)
            st.rerun()
    with _sc3:
        _can_load_more = (
            "_jira_search_results" in st.session_state
            and len(st.session_state["_jira_search_results"].get("issues", []))
            < st.session_state["_jira_search_results"].get("total", 0)
        )
        if st.button(
            "Load more",
            key="jira_load_more_btn",
            disabled=not _can_load_more,
            use_container_width=True,
        ):
            _prev_jql = st.session_state["_jira_search_results"].get("_jql", _jql_input)
            _run_search(_prev_jql, append=True)
            st.rerun()

    if _do_search:
        st.session_state.pop("_jira_search_results", None)
        st.session_state.pop("_jira_filter_label", None)
        _run_search(_jql_input)

    # ── Results ───────────────────────────────────────────────────────────────
    _results = st.session_state.get("_jira_search_results")
    if _results is None:
        st.info(
            "Use one of the Home24 filter buttons above, or enter a JQL query and click Search."
        )
        return

    # Show the active filter / JQL label
    _active_label = st.session_state.get("_jira_filter_label", "")
    _active_jql   = _results.get("_jql", "")
    if _active_label:
        st.caption(f"Filter: {_active_label}")
    if _active_jql:
        st.caption(f"JQL: `{_active_jql}`")

    if _results.get("error"):
        _ep = _results.get("endpoint", "")
        st.markdown(
            f'<div class="alert alert-warn">'
            f'<span class="alert-icon">⚠</span>'
            f'<div><strong>Search failed</strong> (endpoint: {_ep})<br>'
            f'<code>{_results["error"]}</code><br>'
            f'Try using a saved filter above, or check your JQL syntax.</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        return

    _issues = _results.get("issues", [])
    if not _issues:
        st.info("No tickets found. Try adjusting your JQL or use a saved filter.")
        return

    import pandas as _pd_jira
    _total   = _results.get("total", len(_issues))
    _showing = len(_issues)
    _ep_used = _results.get("endpoint", "")
    st.markdown(
        f'<div class="alert alert-info">'
        f'<span class="alert-icon">ℹ</span>'
        f'<span>Showing {_showing} of {_total} tickets'
        f'{" — via " + _ep_used if _ep_used else ""}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    _table_rows = [
        {
            "Key":         i["key"],
            "Summary":     i["summary"][:80] + ("…" if len(i["summary"]) > 80 else ""),
            "Status":      i["status"],
            "Assignee":    i["assignee"],
            "Updated":     i["updated"],
            "Attachments": i["attachment_count"],
        }
        for i in _issues
    ]
    st.dataframe(
        _pd_jira.DataFrame(_table_rows),
        hide_index=True,
        use_container_width=True,
    )

    if _showing < _total:
        st.caption(
            f"{_total - _showing} more tickets available — click **Load more** above to fetch them."
        )

    st.markdown("---")

    # ── Ticket selector ───────────────────────────────────────────────────────
    st.markdown('<div class="section-label">Select ticket</div>', unsafe_allow_html=True)

    _ticket_display = [f'{i["key"]} — {i["summary"][:60]}' for i in _issues]
    _ticket_keys    = [i["key"] for i in _issues]
    _sel_idx = st.selectbox(
        "Ticket:",
        range(len(_ticket_display)),
        format_func=lambda x: _ticket_display[x],
        key="jira_ticket_selector",
    )
    _selected_issue = _issues[_sel_idx] if _issues else None
    _selected_key   = _ticket_keys[_sel_idx] if _issues else ""

    if not _selected_issue:
        return

    st.markdown(
        f'<div class="alert alert-info">'
        f'<span class="alert-icon">ℹ</span>'
        f'<span><strong>{_selected_key}</strong>: {_selected_issue["summary"]}'
        f' — Status: <strong>{_selected_issue["status"]}</strong>'
        f' · Assignee: {_selected_issue["assignee"]}'
        f'</span></div>',
        unsafe_allow_html=True,
    )

    # ── Attachment list ───────────────────────────────────────────────────────
    _all_atts   = _selected_issue.get("attachments", [])
    _excel_atts = [
        a for a in _all_atts
        if a["filename"].lower().endswith(_EXCEL_EXTENSIONS)
    ]
    _other_atts = [a for a in _all_atts if a not in _excel_atts]

    st.markdown('<div class="section-label">Attachments</div>', unsafe_allow_html=True)

    if not _all_atts:
        st.info("No attachments on this ticket.")
        return

    if _other_atts:
        with st.expander(f"Non-Excel attachments ({len(_other_atts)})", expanded=False):
            for _att in _other_atts:
                _sz = round(_att["size"] / 1024, 1)
                st.markdown(f"- `{_att['filename']}` ({_sz} KB)")

    if not _excel_atts:
        st.markdown(
            '<div class="alert alert-warn">'
            '<span class="alert-icon">⚠</span>'
            '<span>No Excel attachments (.xlsx / .xls / .xlsm / .xlsb) on this ticket.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    _att_options     = [a["filename"] for a in _excel_atts]
    _selected_att_fn = st.selectbox(
        "Excel attachment to translate:", _att_options, key="jira_att_selector"
    )
    _selected_att = next((a for a in _excel_atts if a["filename"] == _selected_att_fn), None)

    if _selected_att:
        _sz_kb = round(_selected_att["size"] / 1024, 1)
        st.markdown(
            f'<div class="alert alert-info">'
            f'<span class="alert-icon">📄</span>'
            f'<span><strong>{_selected_att["filename"]}</strong>'
            f' — {_sz_kb} KB · uploaded {_selected_att["created"]}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("---")

    # ── Download and route to Translator ─────────────────────────────────────
    if st.button(
        f"⬇ Download and translate: {_selected_att_fn}",
        key="jira_dl_translate_btn",
        use_container_width=True,
    ):
        _jc3, _jerr3 = get_jira_client()
        if not _jc3:
            st.error(f"Connection failed: {_jerr3}")
        elif _selected_att:
            with st.spinner(f"Downloading {_selected_att_fn}..."):
                _file_bytes = _jc3.download_attachment(_selected_att["content_url"])
            if _file_bytes is None:
                st.error(
                    "Download failed. Verify that your API token has attachment read access."
                )
            else:
                st.session_state["_jira_upload"]              = _JiraFileProxy(
                    _selected_att_fn, _file_bytes
                )
                st.session_state["_jira_ticket_key"]          = _selected_key
                st.session_state["_jira_ticket_summary"]      = _selected_issue["summary"]
                st.session_state["_jira_attachment_id"]       = _selected_att["id"]
                st.session_state["_jira_attachment_filename"] = _selected_att_fn
                # Clear any stale translation results and sheet caches
                st.session_state.pop("_tr_result", None)
                for _sk in [k for k in st.session_state if k.startswith("_sheet_")]:
                    del st.session_state[_sk]
                st.session_state["_tr_result_file"] = _selected_att_fn
                # Navigate to Translator
                st.session_state["nav_radio"] = "Translator"
                st.success(f"Downloaded {_selected_att_fn}. Opening Translator…")
                st.rerun()


def render_sidebar() -> str:
    with st.sidebar:
        target_language = st.session_state.get("target_language", "French")
        lang_code       = "NL" if target_language == "Dutch" else "FR"
        lang_flag       = "🇳🇱" if target_language == "Dutch" else "🇫🇷"

        st.markdown(f"""
        <div class="sb-brand">
            <div class="sb-wordmark"><div class="sb-dot"></div>DE→{lang_code} Localization</div>
            <div class="sb-org">Home24 Internal</div>
        </div>
        """, unsafe_allow_html=True)

        # Target language badge
        badge_bg    = "#FEF9C3" if target_language == "Dutch" else "#DCFCE7"
        badge_bdr   = "#FDE047" if target_language == "Dutch" else "#BBF7D0"
        badge_color = "#713F12" if target_language == "Dutch" else "#15803D"
        st.markdown(
            f'<div style="display:inline-flex;align-items:center;gap:6px;'
            f'padding:4px 12px;border-radius:20px;background:{badge_bg};'
            f'border:1px solid {badge_bdr};'
            f'font-size:11px;font-weight:700;color:{badge_color};margin-bottom:8px;">'
            f'{lang_flag} {target_language} ({lang_code})'
            f'</div>',
            unsafe_allow_html=True,
        )

        if st.button("Switch language", key="switch_lang_btn", use_container_width=True):
            st.session_state["language_selected"] = False
            st.session_state.pop("target_language", None)
            st.session_state.pop("_tr_result", None)
            st.session_state.pop("_tr_result_file", None)
            st.session_state.pop("nl_corpus_engine", None)
            st.rerun()

        st.markdown("---")
        st.markdown('<span class="sb-nav-label">Navigation</span>', unsafe_allow_html=True)

        role = st.session_state.get("user_role", "")
        nav_options = ["Translator", "Translation History", "Analytics", "Glossary", "Translation Memory"]
        if role == "admin":
            nav_options.append("Admin Dashboard")
            nav_options.append("Issue Reports")
        nav_options.append("Report an Issue")
        # Jira integration temporarily disabled — code preserved in jira_client.py

        page = st.radio(
            "Navigation",
            nav_options,
            key="nav_radio",
            label_visibility="collapsed",
        )

        st.markdown("---")

        email = st.session_state.get("user_email", "")
        if role == "guest":
            role_label = "Guest Demo"
        elif role == "admin":
            role_label = "Administrator"
        elif role == "standard_user":
            role_label = "Standard User"
        else:
            role_label = ""

        st.markdown(f"""
        <div class="sb-user">
            <span class="sb-user-label">Signed in as</span>
            <span class="sb-user-email">{email}</span>
            {"<span style='display:block;font-size:0.68rem;color:#22C55E;margin-top:2px;'>" + role_label + "</span>" if role_label else ""}
        </div>
        """, unsafe_allow_html=True)

        if st.button("Sign out", key="logout_btn", use_container_width=True):
            for key in ["authenticated", "user_role", "user_email", "language_selected",
                        "target_language", "session_id", "_tr_result", "_tr_result_file"]:
                st.session_state.pop(key, None)
            st.session_state["authenticated"] = False
            st.rerun()

        st.markdown("---")
        db_status = db_get_status()
        dot   = "🟢" if db_status["connected"] else "🔴"
        label = "SQLite · connected" if db_status["connected"] else "SQLite · error"
        st.markdown(
            f'<div style="font-size:0.72rem;color:#94A3B8;margin-top:2px;">'
            f'{dot} <strong style="color:#64748B;">{label}</strong><br>'
            f'Jobs: {db_status["jobs"]} &nbsp;·&nbsp; '
            f'TM: {db_status["tm_entries"]} &nbsp;·&nbsp; '
            f'Glossary: {db_status["glossary_terms"]}'
            f"</div>",
            unsafe_allow_html=True,
        )

    return page


def render_footer():
    st.markdown("""
    <div class="site-footer">
        <span>Built by <strong class="footer-author">Yves Koulle Banga</strong></span>
        <span class="footer-version">DE Multilingual Translator · v6.0</span>
    </div>
    """, unsafe_allow_html=True)


# =============================================================================
# REPORT AN ISSUE — user-facing form
# =============================================================================

def report_issue_page():
    render_page_header("Report an Issue", "Report a translation error, bug, or quality problem.")

    user_email = st.session_state.get("user_email", "")
    user_role  = st.session_state.get("user_role", "")

    # Pre-fill filename from the last translated file if available
    last_file = st.session_state.get("_tr_result_file", "")

    st.markdown("""
    <div class="alert alert-info" style="margin-bottom:20px;">
        <span class="alert-icon">ℹ</span>
        <span>Use this form to report any issue you encounter. Your report is saved immediately
        and the admin will be notified by email.</span>
    </div>
    """, unsafe_allow_html=True)

    with st.form("issue_report_form", clear_on_submit=True):
        col_cat, col_sev = st.columns(2)
        with col_cat:
            category = st.selectbox(
                "Issue category",
                _ISSUE_CATEGORIES,
                index=0,
            )
        with col_sev:
            severity = st.selectbox(
                "Severity",
                _ISSUE_SEVERITIES,
                index=1,
            )

        col_lang, col_file = st.columns(2)
        with col_lang:
            target_language = st.selectbox(
                "Target language",
                _ISSUE_LANGUAGES,
                index=0,
            )
        with col_file:
            filename = st.text_input(
                "File name (optional)",
                value=last_file,
                placeholder="e.g. Translation_PC-20815.xlsx",
            )

        col_row, col_col = st.columns(2)
        with col_row:
            row_reference = st.text_input(
                "Row reference (optional)",
                placeholder="e.g. Row 14",
            )
        with col_col:
            column_reference = st.text_input(
                "Column reference (optional)",
                placeholder="e.g. materialDetail",
            )

        description = st.text_area(
            "Description",
            placeholder="Describe the issue clearly. What happened? What did you expect?",
            height=130,
        )

        expected_correction = st.text_area(
            "Expected correction (optional)",
            placeholder="If you know the correct translation or fix, write it here.",
            height=80,
        )

        submitted = st.form_submit_button("Send Report", use_container_width=True, type="primary")

    if submitted:
        if not description.strip():
            st.error("Please add a description before submitting.")
            return

        from datetime import datetime as _dt
        report = {
            "user_email":          user_email,
            "user_role":           user_role,
            "category":            category,
            "severity":            severity,
            "target_language":     target_language,
            "filename":            filename.strip(),
            "row_reference":       row_reference.strip(),
            "column_reference":    column_reference.strip(),
            "description":         description.strip(),
            "expected_correction": expected_correction.strip(),
            "created_at":          _dt.now().isoformat(timespec="seconds"),
        }

        db_save_issue_report(report)
        email_sent = _send_issue_report_email(report)

        if email_sent:
            st.success("Thank you. Your report has been sent to the admin.")
        else:
            st.success("Report saved. Email notification is not configured — the admin will review it in the dashboard.")


# =============================================================================
# ADMIN ISSUE REPORTS — admin-only dashboard
# =============================================================================

_SEVERITY_BADGE: dict[str, str] = {
    "Critical": "background:#FEE2E2;color:#991B1B;border:1px solid #FECACA;",
    "High":     "background:#FEF3C7;color:#92400E;border:1px solid #FDE68A;",
    "Medium":   "background:#DBEAFE;color:#1E40AF;border:1px solid #BFDBFE;",
    "Low":      "background:#F0FDF4;color:#166534;border:1px solid #BBF7D0;",
}

_STATUS_BADGE: dict[str, str] = {
    "open":        "background:#FEE2E2;color:#991B1B;border:1px solid #FECACA;",
    "in progress": "background:#FEF3C7;color:#92400E;border:1px solid #FDE68A;",
    "resolved":    "background:#DCFCE7;color:#166534;border:1px solid #BBF7D0;",
    "ignored":     "background:#F1F5F9;color:#64748B;border:1px solid #E2E8F0;",
}


def admin_issue_reports_page():
    if st.session_state.get("user_role") != "admin":
        st.error("Access restricted to administrators.")
        return

    render_page_header("Issue Reports", "Bug reports and translation quality issues from users.")

    counts = db_get_issue_report_counts()
    total  = sum(counts.values())
    open_  = counts.get("open", 0)
    in_progress = counts.get("in progress", 0)
    resolved    = counts.get("resolved", 0)
    ignored     = counts.get("ignored", 0)

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total", total)
    c2.metric("Open", open_)
    c3.metric("In progress", in_progress)
    c4.metric("Resolved", resolved)
    c5.metric("Ignored", ignored)

    st.markdown("---")
    st.markdown('<div class="section-label">Filters</div>', unsafe_allow_html=True)

    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        f_status = st.selectbox("Status", ["All"] + _ISSUE_STATUSES, key="ir_f_status")
    with fc2:
        f_severity = st.selectbox("Severity", ["All"] + _ISSUE_SEVERITIES, key="ir_f_severity")
    with fc3:
        f_category = st.selectbox("Category", ["All"] + _ISSUE_CATEGORIES, key="ir_f_category")
    with fc4:
        f_lang = st.selectbox("Language", ["All"] + _ISSUE_LANGUAGES, key="ir_f_lang")

    reports = db_load_issue_reports(
        status_filter=f_status,
        severity_filter=f_severity,
        category_filter=f_category,
        lang_filter=f_lang,
    )

    st.markdown(f'<div class="section-label" style="margin-top:16px;">Reports — {len(reports)} found</div>', unsafe_allow_html=True)

    if not reports:
        st.markdown("""
        <div class="alert alert-info">
            <span class="alert-icon">ℹ</span>
            <span>No reports match the current filters.</span>
        </div>
        """, unsafe_allow_html=True)
        return

    for rep in reports:
        sev   = rep.get("severity", "Medium")
        stat  = rep.get("status", "open")
        sev_style  = _SEVERITY_BADGE.get(sev,  "background:#F1F5F9;color:#334155;")
        stat_style = _STATUS_BADGE.get(stat, "background:#F1F5F9;color:#334155;")

        with st.expander(
            f"[{sev}] {rep.get('category','?')}  ·  {rep.get('user_email','?')}  ·  {rep.get('created_at','')[:16]}",
            expanded=False,
        ):
            m1, m2, m3 = st.columns(3)
            m1.markdown(
                f'<span style="font-size:11px;font-weight:600;padding:2px 8px;border-radius:10px;{sev_style}">{sev}</span>',
                unsafe_allow_html=True,
            )
            m2.markdown(
                f'<span style="font-size:11px;font-weight:600;padding:2px 8px;border-radius:10px;{stat_style}">{stat}</span>',
                unsafe_allow_html=True,
            )
            m3.write("")

            info_rows = [
                ("Reporter",        rep.get("user_email", "—")),
                ("Role",            rep.get("user_role", "—")),
                ("Category",        rep.get("category", "—")),
                ("Target language", rep.get("target_language", "—")),
                ("File",            rep.get("filename", "—") or "—"),
                ("Row",             rep.get("row_reference", "—") or "—"),
                ("Column",          rep.get("column_reference", "—") or "—"),
                ("Submitted",       rep.get("created_at", "—")),
            ]
            for label, val in info_rows:
                st.markdown(
                    f'<div style="display:flex;gap:12px;font-size:13px;padding:2px 0;">'
                    f'<span style="min-width:130px;color:#64748B;font-weight:500;">{label}</span>'
                    f'<span style="color:#0F172A;">{val}</span></div>',
                    unsafe_allow_html=True,
                )

            st.markdown("**Description**")
            st.markdown(
                f'<div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:6px;'
                f'padding:10px 14px;font-size:13px;color:#334155;white-space:pre-wrap;">'
                f'{rep.get("description","")}</div>',
                unsafe_allow_html=True,
            )

            correction = rep.get("expected_correction", "").strip()
            if correction:
                st.markdown("**Expected correction**")
                st.markdown(
                    f'<div style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:6px;'
                    f'padding:10px 14px;font-size:13px;color:#166534;white-space:pre-wrap;">'
                    f'{correction}</div>',
                    unsafe_allow_html=True,
                )

            st.markdown("**Update status**")
            new_status = st.selectbox(
                "New status",
                _ISSUE_STATUSES,
                index=_ISSUE_STATUSES.index(stat) if stat in _ISSUE_STATUSES else 0,
                key=f"status_sel_{rep['id']}",
                label_visibility="collapsed",
            )
            if st.button("Save status", key=f"status_btn_{rep['id']}"):
                db_update_issue_report_status(rep["id"], new_status)
                st.success(f"Status updated to '{new_status}'.")
                st.rerun()


# =============================================================================
# VALIDATION
# =============================================================================

def validate_product_name(name: str) -> str:
    if not name:
        return name
    name = " ".join(name.split())
    name = name.replace(",", "")
    name = re.sub(r'\([^)]*\)', '', name)
    name = name.replace("(", "").replace(")", "")
    name = re.sub(r'\[[^\]]*\]', '', name)
    name = name.replace("[", "").replace("]", "")
    name = " ".join(name.split())
    if len(name) > 40:
        truncated = name[:40]
        last_space = truncated.rfind(" ")
        name = truncated[:last_space] if last_space > 20 else truncated
    return name.strip()


def detect_german_residue(text: str, target_language: str = "French") -> list[str]:
    if not text:
        return []
    acceptable = DUTCH_ACCEPTABLE_WORDS if target_language == "Dutch" else FRENCH_ACCEPTABLE_WORDS
    detected = []
    masked = text.lower()
    for word in acceptable:
        masked = masked.replace(word.lower(), "X" * len(word))
    for word in GERMAN_RESIDUE_WORDS:
        word_lower = word.lower()
        if word_lower in [w.lower() for w in acceptable]:
            continue
        if re.compile(r'\b' + re.escape(word_lower) + r'\b', re.IGNORECASE).search(masked):
            detected.append(word)
    return detected


def analyze_translation_quality(
    source: str, translation: str, canonical: str, glossary: dict | None = None
) -> list[dict]:
    """Return enriched quality issues. Each dict has: type, reason, severity, category, suggested_fix.

    Excel highlighting signals:
      - Critical/High issues → REVIEW_FILL only when checkbox enabled
      - original_excel_highlights → always preserved
      - glossary_hits → analytics only, never affect cell color
    """
    issues   = []
    _glossary = glossary or {}

    # Identical output (Critical) — check before residue so I don't double-flag
    if translation.strip() == source.strip():
        src_lower      = source.strip().lower()
        glossary_terms = _glossary.get("terms", {})
        is_expected    = (
            src_lower in {w.lower() for w in FRENCH_ACCEPTABLE_WORDS}
            or any(
                src_lower == de.lower() and translation.strip().lower() == fr.lower()
                for de, fr in glossary_terms.items()
            )
        )
        if not is_expected:
            meta = _issue_meta("identical_output", source, translation)
            issues.append({
                "type":   "identical_output",
                "reason": "Translation matches source text — likely untranslated",
                **meta,
            })

    # German residue (recorded here; final verdict via warning_details after all fix passes)
    residue = detect_german_residue(translation)
    if residue:
        meta = _issue_meta("german_residue", source, translation)
        issues.append({
            "type":   "german_residue",
            "reason": f"German words detected: {', '.join(residue[:3])}",
            **meta,
        })

    # Too short (Medium)
    if len(source) > 6 and len(translation) < max(3, len(source) * 0.4):
        meta = _issue_meta("too_short", source, translation)
        issues.append({
            "type":   "too_short",
            "reason": f"Very short ({len(translation)} chars vs {len(source)} source)",
            **meta,
        })

    # Too long (Medium)
    if len(source) > 20 and len(translation) > len(source) * 2.5:
        meta = _issue_meta("too_long", source, translation)
        issues.append({
            "type":   "too_long",
            "reason": f"Unusually long ({len(translation)} chars vs {len(source)} source)",
            **meta,
        })

    # Lost <br> tags (High)
    src_br = source.count("<br>")
    if src_br > 0 and translation.count("<br>") < src_br:
        meta = _issue_meta("lost_br_tags", source, translation)
        issues.append({
            "type":   "lost_br_tags",
            "reason": f"Lost {src_br - translation.count('<br>')} <br> tag(s)",
            **meta,
        })

    # Product name rules (High)
    if canonical == "name":
        if len(translation) > 40:
            meta = _issue_meta("name_too_long", source, translation)
            issues.append({
                "type":   "name_too_long",
                "reason": f"Exceeds 40 chars ({len(translation)})",
                **meta,
            })
        if "," in translation:
            meta = _issue_meta("name_has_comma", source, translation)
            issues.append({
                "type":   "name_has_comma",
                "reason": "Contains a comma",
                **meta,
            })

    # Glossary inconsistency (High) — only prompt-block terms, long ones only
    violations = _check_glossary_inconsistency(source, translation, _glossary)
    if violations:
        meta = _issue_meta("glossary_inconsistency", source, translation)
        issues.append({
            "type":   "glossary_inconsistency",
            "reason": f"Glossary term(s) not applied: {' | '.join(violations)}",
            **meta,
        })

    return issues


def _api_call_with_retry(fn, *, max_retries: int = MAX_API_RETRIES, notify_fn=None):
    """Call fn() with exponential backoff on failure. Re-raises if all attempts fail."""
    delay    = RETRY_BASE_DELAY
    last_exc = None
    for attempt in range(max_retries + 1):
        try:
            return fn()
        except Exception as exc:
            last_exc = exc
            if attempt >= max_retries:
                break
            msg  = str(exc).lower()
            wait = delay * 4 if ("rate limit" in msg or "429" in msg) else delay
            if notify_fn:
                notify_fn(f"API error (attempt {attempt + 1}/{max_retries + 1}), retrying in {wait:.0f}s…")
            time.sleep(wait)
            delay *= 2
    raise last_exc


# =============================================================================
# API KEY + CLIENT
# =============================================================================

def _get_api_key() -> str | None:
    try:
        key = st.secrets.get("OPENAI_API_KEY")
        if key:
            return key
    except Exception:
        pass
    return os.environ.get("OPENAI_API_KEY")


def get_openai_client():
    api_key = _get_api_key()
    if not api_key:
        st.error(
            "**OPENAI_API_KEY not configured.**\n\n"
            "- **Locally:** add `OPENAI_API_KEY=sk-...` to your `.env` file.\n"
            "- **Streamlit Cloud:** App settings → Secrets."
        )
        st.stop()
    return OpenAI(api_key=api_key)


# =============================================================================
# TRANSLATION FUNCTIONS
# =============================================================================

_FR_SHARED_RULES = (
    "- All German words MUST be translated — zero German residue allowed\n"
    "- French typography: always write a space before and after ':' (e.g. 'Structure : métal')\n"
    "- Spaces around '/' in color/material combinations (e.g. 'Noir / Gris', 'Aluminium / Polyester')\n"
    "- \"pulverbeschichtet\" → \"thermolaqué\" (NEVER 'revêtu de poudre' or 'revêtement de poudre')\n"
    "- \"Geflecht\" / \"Polyrattan\" / \"Kunststoffgeflecht\" → \"résine tressée\"\n"
    "- \"bestehend aus\" → \"composé de\", \"Set bestehend aus\" → \"Ensemble composé de\"\n"
    "- Preserve <br> tags exactly — NEVER replace with semicolons\n"
    "- Preserve numbers, dimensions and percentages exactly\n"
    "- Write real home24.fr wording — not literal German structure"
)


def _build_system_prompt(
    canonical: str,
    glossary_block: str,
    target_language: str = "French",
    tm_guidance: str = "",
) -> str:
    if target_language == "Dutch":
        return _build_nl_system_prompt(canonical, glossary_block, tm_guidance=tm_guidance)
    # French prompts
    if canonical == "name":
        return (
            "You are a premium French copywriter for Home24 France furniture e-commerce.\n"
            "Translate the German product name to natural, elegant French — real home24.fr wording.\n\n"
            "STRICT rules:\n"
            "- Maximum 40 characters total\n"
            "- No commas, no brackets, no parentheses\n"
            "- All German words MUST be translated — zero German residue allowed\n"
            "- \"Sofa\" → \"Canapé\"\n"
            "- \"Sessel\" → \"Fauteuil\"\n"
            "- \"Ecksofa\" → \"Canapé d'angle\"\n"
            "- \"Schlafsofa\" → \"Canapé convertible\"\n"
            "- \"Sitzer\" → \"places\" (\"3-Sitzer\" = \"3 places\")\n"
            "- \"Loungeset\" → \"Salon de jardin\"\n"
            "- \"Gartenessgruppe\" / \"Gartengruppe\" → \"Ensemble de jardin\" / \"Salon de jardin\"\n"
            "- \"Sofaelement\" → \"Module de canapé\"\n"
            "- \"Gartenstuhl\" → \"Chaise de jardin\"\n"
            "- \"Gartentisch\" → \"Table de jardin\"\n"
            "- \"Chaiselongue\" / \"Chaise longue\" → ALWAYS two words: \"Chaise longue\"\n"
            "- \"Fußhocker\" / \"Fusshocker\" in seating/lounge context → \"Repose-pieds\"\n"
            "- \"Esstisch\" → \"Table\" (NOT 'table à manger' unless clearly a dining table category)\n"
            "- Ausziehbar table → \"Table extensible\"\n"
            "- \"Artisan Eiche Dekor\" / \"Eiche Artisan Dekor\" → \"Décor chêne artisan\"\n"
            "- \"Eiche Dekor\" → \"Décor chêne\" (décor FIRST, then wood name)\n"
            "- Colors: \"Schlamm\" / \"mud\" → \"argile\"; \"Terra\" (color) → \"Terracotta\"\n"
            "- \"boue\" is FORBIDDEN in customer-facing copy — use \"argile\" or \"taupe\"\n"
            "- Pluralise item names when source indicates multiple items:\n"
            "  6-teilig / Set / Lot / x6 / 6 Stück → pluralise the item noun\n"
            "  e.g. 6 plates → \"Assiettes de service\" not \"Assiette de service\"\n"
            "  Do NOT pluralise model or collection names\n"
            "- Carpet / rug product types (CRITICAL — do not confuse these):\n"
            "  • \"Fußmatte\" → \"Tapis d'entrée\" (entry/door mat)\n"
            "  • \"Hochflorteppich\" → \"Tapis à poils longs\" (high-pile rug)\n"
            "  • \"Kurzflorteppich\" → \"Tapis à poils courts\" (low-pile rug)\n"
            "  • \"Läufer\" → \"Chemin de couloir\" (NEVER \"Couverture\" — that means blanket)\n"
            "  • \"Teppich\" → \"Tapis\", \"Teppichläufer\" → \"Tapis de couloir\"\n"
            "- Preserve product series names exactly as-is (e.g. Entry, Natural Guard, Cosy, Firenze)\n"
            "  Do NOT translate 'Entry' as 'd'accueil' — it is a product series name\n"
            "- Preserve model/collection names exactly (e.g. Vedene, Arin, Bocca, Level36)\n"
            "- Preserve dimensions and numbers exactly\n"
            "- Write elegant, commercial French — not literal word-for-word translation"
            f"{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "materialDetail":
        return (
            "You are a premium French copywriter for Home24 France furniture e-commerce.\n"
            "Translate the German material description to natural, precise French.\n\n"
            "Rules:\n"
            + _FR_SHARED_RULES + "\n"
            "- \"Gestell\" / \"Tischgestell\" → \"piètement\" or \"structure\" per context\n"
            "- \"Bezug\" → \"revêtement\" (NEVER \"housse\" for frame/structure components)\n"
            "- \"Korpus\" → \"caisson\"\n"
            "- \"Untergestell\" → \"structure inférieure\"\n"
            "- BHT / BxHxT / \"B x H x T\" → \"L x H x P\" (French dimension format)\n"
            "  Example: \"BHT: 100 x 80 x 45 cm\" → \"L x H x P : 100 x 80 x 45 cm\"\n"
            "- \"Artisan Eiche Dekor\" / \"Eiche Dekor\" → \"décor chêne artisan\" / \"décor chêne\"\n"
            "  The word 'décor' always comes FIRST: 'décor chêne artisan' NOT 'chêne artisan décor'\n"
            "- \"Grifflos\" → \"sans poignées\"\n"
            "- \"autark\" (kitchen) → \"équipée\"\n"
            "- Use professional French furniture terminology"
            f"{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "colorDetail":
        return (
            "You are a premium French copywriter for Home24 France furniture e-commerce.\n"
            "Translate the German color description to natural, attractive French.\n\n"
            "Rules:\n"
            "- All German words MUST be translated — zero German residue\n"
            "- Spaces around '/' in color combinations: 'Noir / Gris' NOT 'Noir/Gris'\n"
            "- French typography: space before and after ':'\n"
            "- dunkelgrau → gris foncé, hellgrau → gris clair\n"
            "- schwarz → noir, weiß → blanc, braun → brun, grau → gris\n"
            "- \"Schlamm\" / \"Schlammfarbe\" / mud-like colors → \"argile\" (NEVER 'boue')\n"
            "- \"Terra\" (color) → \"terracotta\"\n"
            "- \"Sand\" → \"sable\", \"Ton\" (clay color) → \"argile\"\n"
            "- \"Elfenbein\" → \"ivoire\", \"Puderrosa\" → \"rose poudré\"\n"
            "- \"Natur\" → \"naturel\", \"Blaugrün\" → \"bleu-vert\"\n"
            "- \"Multi\" → \"multicolore\" (or keep \"Multi\" if the brand uses it)\n"
            "- \"Artisan Eiche\" → \"chêne artisan\"\n"
            "- \"Artisan Eiche Dekor\" / \"Eiche Dekor\" → \"décor chêne artisan\" / \"décor chêne\"\n"
            "  'décor' always comes FIRST in the phrase\n"
            "- Preserve color codes and numbers exactly\n"
            "- Write elegant, commercial color names — not literal German"
            f"{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "textileCompositionCover1":
        return (
            "You are a precise translator for Home24 France product data.\n"
            "Translate German textile composition labels to French.\n\n"
            "Rules:\n"
            "- EU spacing format: 'X % material' (space before AND after %) — e.g. '100 % polyester'\n"
            "- Preserve the separator: '/' as in source\n"
            "- Fiber names in LOWERCASE\n"
            "- Polyester → polyester, Baumwolle → coton, Wolle → laine\n"
            "- Viskose → viscose, Polyamid → polyamide, Polypropylen → polypropylène\n"
            "- Modacryl → modacrylique, Kokos → coco, Gummi → caoutchouc\n"
            "- Leinen → lin, Seide → soie, Elasthan → élasthanne\n"
            "- Mikrofaser → microfibre, Acryl → acrylique, Leder → cuir\n"
            "- Preserve numbers exactly — NEVER alter percentages"
            f"{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "otherMeasurements":
        return (
            "You are a premium French copywriter for Home24 France furniture e-commerce.\n"
            "Translate the German measurement/dimension text to natural French.\n\n"
            "Rules:\n"
            "- BHT / BxHxT / \"B x H x T\" → \"L x H x P\" (French standard)\n"
            "  Example: \"BHT: 100 x 80 x 45 cm\" → \"L x H x P : 100 x 80 x 45 cm\"\n"
            "  Example: \"Tisch (BxHxT): 180 x 75 x 100 cm\" → \"Table (L x H x P) : 180 x 75 x 100 cm\"\n"
            "- French typography: ALWAYS a space before and after ':'\n"
            "- Spaces around 'x' in dimensions: '100 x 80 x 45' NOT '100x80x45'\n"
            "- Breite → largeur, Höhe → hauteur, Tiefe → profondeur, Länge → longueur\n"
            "- Preserve ALL numbers, units and symbols exactly — do not change values\n"
            "- Preserve <br> tags exactly"
            f"{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "deliveryScope":
        return (
            "You are a premium French copywriter for Home24 France furniture e-commerce.\n"
            "Translate the German delivery scope text to natural, elegant French.\n\n"
            "Rules:\n"
            + _FR_SHARED_RULES + "\n"
            "- \"inkl.\" / \"inklusive\" → \"inclus(e)\"\n"
            "- \"ohne Dekoration\" → \"sans décoration\"\n"
            "- Prefer \"composé de\" over \"contenant\" for set descriptions\n"
            "- When item names indicate multiple pieces, pluralise naturally:\n"
            "  e.g. '6 assiettes' not '6 assiette', 'Lot de 2 tables' not 'Lot de 2 table'\n"
            "- \"assiettes à manger\" is FORBIDDEN — use \"assiettes\" or \"assiettes de service\""
            f"{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    else:
        return (
            "You are a premium French copywriter for Home24 France furniture e-commerce.\n"
            "Translate the German text to natural, elegant French — real home24.fr wording.\n\n"
            "Rules:\n"
            + _FR_SHARED_RULES + "\n"
            "- BHT / BxHxT / 'B x H x T' → 'L x H x P'\n"
            "- \"Artisan Eiche Dekor\" → \"décor chêne artisan\" ('décor' FIRST)\n"
            "- Colors: 'Schlamm' / mud-like → 'argile' (NEVER 'boue')\n"
            "- 'Chaiselongue' → 'chaise longue' (always two words)"
            f"{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )


_NL_SHARED_RULES = (
    "- GSP / GSP-Blende / Geschirrspüler-Blende → \"vaatwasserpaneel\" (NEVER leave GSP untranslated)\n"
    "- BHT / BxHxT / \"B x H x T\" → \"B x H x D\"\n"
    "- Grifflos → \"greeploos\"\n"
    "- Unterflurauszug / Unterflurführung → \"onderliggende ladegeleider\"\n"
    "- Küchenzeile → \"keukenblok\", Einbauküche → \"inbouwkeuken\", Arbeitsplatte → \"werkblad\"\n"
    "- Spüle → \"spoelbak\", Unterschrank → \"onderkast\", Hängeschrank → \"hangkast\"\n"
    "- Oberschrank → \"bovenkast\", Hochschrank → \"hoge kast\", Apothekerschrank → \"apothekerskast\"\n"
    "- Einzelwaschtisch → \"enkele wastafel\", Doppelwaschtisch → \"dubbele wastafel\"\n"
    "- Waschtisch → \"wastafelmeubel\", Waschbecken → \"wastafel\"\n"
    "- Griff → \"greep\", Griffe → \"grepen\", Blende → \"frontpaneel\", Sockel → \"plint\"\n"
    "- Soft-Close / Softclose → \"soft-close\", Dämpfung → \"demping\"\n"
    "- Percentage: \"100%\" not \"100 %\"; lowercase after %: \"100% polyester\" not \"100% Polyester\"\n"
    "- Slash in color/material combos: NO spaces — \"zwart/grijs\" not \"zwart / grijs\"\n"
    "- Dekor terminology (CRITICAL — Home24 NL canonical):\n"
    "  • \"Eiche Artisan Dekor\" → \"Artisan eikenlook\"\n"
    "  • \"Eiche Viking Dekor\" → \"Viking eikenhouten look\"\n"
    "  • \"Eiche hell Dekor\" → \"lichte eikenhouten look\"\n"
    "  • \"Marmor Weiß Dekor\" → \"witte marmerlook\"\n"
    "  • \"Kernbuche Dekor\" → \"kernbeukenhouten look\"\n"
    "  • NEVER write \"eiken decor\" or \"Eiken decor\" — always use \"eikenlook\" or \"eikenhouten look\"\n"
    "  • The Dutch word is \"look\" (not \"decor\") for material finishes\n"
    "- Color-descriptor suffixes: \"Anthrazit\" → \"antracietkleurig\", \"Graphit\" → \"grafietkleurig\","
    " \"Silber\" → \"zilverkleurig\", \"Gold\" → \"goudkleurig\"\n"
    "- TV furniture: \"TV-Lowboard\" → \"Tv-meubel\", \"Fernsehsessel\" → \"tv-fauteuil\"\n"
    "- Seating: \"3-Sitzer\" → \"3-zits\", \"3-Sitzer Sofa\" → \"3-zitsbank\"\n"
    "- Lighting: \"Pendelleuchte\" → \"hanglamp\", \"Tischleuchte\" → \"tafellamp\","
    " \"Deckenleuchte\" → \"plafondlamp\", \"Stehleuchte\" → \"staande lamp\"\n"
    "- Bathroom: \"Badset\" → \"Badkamerset\"\n"
    "- Type/Typ: \"Typ A\" → \"type A\" (lowercase, no period)\n"
    "- Teilig: \"3-teilig\" → \"3-delig\"\n"
    "- Flammig: \"1-flammig\" → \"1-lichts\""
)


_NL_SYSTEM_PREAMBLE = (
    "You are a Home24 Netherlands localization specialist — NOT a generic translator.\n"
    "Your output must match real Home24 NL product copy style: native Dutch, commercial, compact.\n\n"
    "CRITICAL — NEVER do these:\n"
    "- Write \"eiken decor\" or \"decor eik\" → ALWAYS use \"eikenlook\" or \"eikenhouten look\"\n"
    "- Write \"TV lowboard\" or \"Televisiemeubel\" → ALWAYS \"Tv-meubel\"\n"
    "- Add spaces around slashes in color combos → \"zwart/grijs\" NOT \"zwart / grijs\"\n"
    "- Add space before % → \"95%\" NOT \"95 %\"\n"
    "- Leave \"Dekor\" untranslated → ALWAYS replace with \"look\"\n"
    "- Invent Dutch terminology → reuse TM vocabulary when provided\n"
    "- Write \"Anthrazit\" → ALWAYS \"antracietkleurig\"; \"Graphit\" → \"grafietkleurig\"\n\n"
)


def _build_nl_system_prompt(canonical: str, glossary_block: str, tm_guidance: str = "") -> str:
    """Dutch-specific system prompts — TM-guided, Home24 NL canonical."""
    tm_block = f"\n{tm_guidance}" if tm_guidance else ""

    if canonical == "name":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate the German product name to Dutch:\n"
            "- Maximum 40 characters — shorten ONLY at natural word boundaries\n"
            "- NEVER leave dangling adjectives: never end with: met, van, keramische, schuine, houten, geïntegreerde, verstelbare\n"
            "- If shortening needed: first remove 'met ...' phrase; if still long, drop whole optional clause; never cut mid-phrase\n"
            "- No commas, brackets, or parentheses\n"
            "- Natural, commercial Dutch product name\n"
            "- \"Sofa\" → \"bank\" / \"Sessel\" → \"fauteuil\" / \"Ecksofa\" → \"Hoekbank\"\n"
            "- \"Schlafsofa\" → \"slaapbank\" / \"Sitzer\" → \"zits\" / \"3-Sitzer Sofa\" → \"3-zitsbank\"\n"
            "- \"TV-Lowboard\" → \"Tv-meubel\" / \"Fernsehsessel\" → \"tv-fauteuil\"\n"
            "- \"Singleküche\" → \"Mini keuken\" (NEVER \"Enkele keuken\" or \"Eenpersoonskeuken\")\n"
            "- \"Pantryküche\" → \"Pantrykeuken\"\n"
            "- \"Küchenzeile\" / \"Küchenleerblock\" → \"Keukenblok\" (NEVER \"Keukenleerblok\")\n"
            "- \"Wohnlandschaft\" → \"Zithoek\"\n"
            "- \"Ecksofa\" → \"Hoekbank\" / \"Ottomane\" → \"ottomane\" (lowercase)\n"
            "- \"Pendelleuchte\" → \"hanglamp\" / \"Tischleuchte\" → \"tafellamp\"\n"
            "- \"Deckenleuchte\" → \"plafondlamp\" / \"Stehleuchte\" → \"staande lamp\"\n"
            "- \"Badset\" → \"Badkamerset\"\n"
            "- \"GSP-Blende\" → \"vaatwasserpaneel\" / \"Grifflos\" → \"greeploos\"\n"
            "- \"Kombi\" → \"combi\" / \"Variante\" → \"variant\"\n"
            "- Preserve model/collection names exactly\n"
            + _NL_SHARED_RULES
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "colorDetail":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate German color descriptions to natural Home24 NL Dutch:\n"
            "- Slash combos: NO spaces — \"zwart/grijs\" not \"zwart / grijs\"\n"
            "- Hyphen combos: hyphens — \"wit-grijs\" for contrast combos\n"
            "- dunkelgrau → donkergrijs / hellgrau → lichtgrijs\n"
            "- dunkelbraun → donkerbruin / hellbraun → lichtbruin\n"
            "- Schwarz → zwart / Weiß → wit / Grau → grijs / Braun → bruin\n"
            "- Blau → blauw / Grün → groen / Rot → rood / Gelb → geel\n"
            "- Anthrazit → antracietkleurig / Graphit → grafietkleurig\n"
            "- Silber → zilverkleurig / Gold → goudkleurig / Creme → crèmekleurig\n"
            "- Sandschwarz → zandzwart / Sandbeige → zandbeige\n"
            "- Dekor patterns: \"Eiche Artisan Dekor\" → \"Artisan eikenlook\","
            " \"Eiche hell Dekor\" → \"lichte eikenhouten look\"\n"
            "- Colors are lowercase after slash: \"zwart/grijs\" not \"Zwart/Grijs\"\n"
            "- Preserve color codes and numbers exactly"
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "materialDetail":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate the German material description to natural Home24 NL Dutch.\n\n"
            "Material rules:\n"
            "- \"Spanplatte, foliert\" → \"gefolieerde spaanplaat\" (adjective BEFORE noun)\n"
            "- \"Spanplatte, beschichtet\" → \"gecoate spaanplaat\"\n"
            "- \"MDF, beschichtet\" → \"gecoat MDF\" / \"MDF foliert\" → \"gefolieerd MDF\"\n"
            "- \"Metall, pulverbeschichtet\" → \"gepoedercoat metaal\"\n"
            "- Bezug → bekleding / Gestell → onderstel / Füße → poten\n"
            "- Korpus → romp / Schublade → lade / Schubladen → lades / Türen → deuren\n"
            "- Holzwerkstoff → houtmateriaal / Massivholz → massief hout\n"
            "- Eiche → eiken / Buche → beuk / Kiefer → grenen / Nussbaum → notelaar\n"
            "- Echtholzfurnier → fineer van echt hout / Furnier → fineer\n"
            "- Holz-Verbundstoff → hout-composietmateriaal\n"
            + _NL_SHARED_RULES + "\n"
            "- Preserve <br> tags exactly — NEVER replace with semicolons\n"
            "- NEVER use semicolons as property separators"
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "textileCompositionCover1":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate the German textile composition to Dutch:\n"
            "- Baumwolle → katoen / Polyester → polyester / Wolle → wol\n"
            "- Leinen → linnen / Viskose → viscose / Seide → zijde\n"
            "- Acryl → acryl / Elasthan → elastaan / Polyamid → polyamide\n"
            "- Jute → jute / Nylon → nylon\n"
            "- Füllung → vulling / Innenstoff → binnenstof\n"
            "- Percentages: no space before % → \"80% katoen\" not \"80 % Katoen\"\n"
            "- Comma separator in compositions: \"60% linnen, 40% katoen\"\n"
            "- Preserve all percentage numbers exactly"
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "deliveryScope":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate German delivery scope to natural Home24 NL Dutch:\n"
            "- Lieferumfang → leveringsomvang\n"
            "- inklusive/inkl. → inclusief/incl.\n"
            "- bestehend aus → bestaande uit\n"
            "- Set bestehend aus → set bestaande uit\n"
            "- ohne Dekoration → zonder decoratie\n"
            "- ohne Armatur → zonder armatuur\n"
            "- Loungeset bestehend aus: → Tuinset bestaande uit:\n"
            + _NL_SHARED_RULES + "\n"
            "- Natural Dutch e-commerce language\n"
            "- Preserve <br> tags exactly"
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "otherMeasurements":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate German measurements/dimensions to Dutch:\n"
            "- Maße/Abmessungen → afmetingen\n"
            "- Breite → breedte / Höhe → hoogte / Tiefe → diepte / Länge → lengte\n"
            "- BHT / BxHxT / \"B x H x T\" / \"B/H/T\" → \"BxHxD\" (no spaces — TM canonical)\n"
            "  Example: \"Schubkasten (BHT): 50x10x30 cm\" → \"lade (BxHxD): 50x10x30 cm\"\n"
            "- Preserve ALL numbers, units, and symbols exactly\n"
            "- Draagkracht for Belastbarkeit\n"
            "- Preserve <br> tags exactly"
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "qualityDetail":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate German quality details to natural Home24 NL Dutch:\n"
            "- Professional Dutch furniture/kitchen/bathroom e-commerce language\n"
            "- Fluent and commercial — sounds written by a native Dutch copywriter\n"
            + _NL_SHARED_RULES + "\n"
            "- Preserve <br> tags exactly as they appear\n"
            "- Do not invent product information"
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    elif canonical == "variantName":
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate German variant names to Home24 NL Dutch:\n"
            "- Natural, concise Dutch\n"
            "- GSP-Blende → \"vaatwasserpaneel\" / Grifflos → \"greeploos\"\n"
            "- Dekor patterns → eikenlook / eikenhouten look (see critical rules above)\n"
            "- Preserve model numbers and collection names exactly\n"
            "- Short and commercial — catalogue-ready"
            + _NL_SHARED_RULES
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )
    else:
        return (
            _NL_SYSTEM_PREAMBLE
            + "Translate the German text to natural Home24 NL Dutch:\n"
            "- Sounds written by a native Dutch furniture copywriter\n"
            "- Zero German residue\n"
            + _NL_SHARED_RULES + "\n"
            "- Preserve <br> tags exactly\n"
            "- Preserve numbers and dimensions exactly"
            + f"{tm_block}{glossary_block}\n"
            "Return ONLY the translated text, nothing else."
        )


def translate_batch(
    client,
    texts: list[str],
    canonical: str,
    token_counter: dict,
    glossary: dict,
    glossary_run_stats: dict,
    notify_fn=None,
    target_language: str = "French",
    product_type_hint: str = "",
) -> list[str]:
    if not texts:
        return []

    glossary_block = _glossary_prompt_block(glossary)
    n = len(texts)

    # Track glossary hits across all source texts in this batch
    all_hits: dict[str, int] = {}
    for text in texts:
        for term, count in count_glossary_hits(text, glossary).items():
            all_hits[term] = all_hits.get(term, 0) + count
    if all_hits:
        glossary_run_stats["total_hits"] = glossary_run_stats.get("total_hits", 0) + sum(all_hits.values())
        tc = glossary_run_stats.setdefault("term_counts", {})
        for term, count in all_hits.items():
            tc[term] = tc.get(term, 0) + count

    if target_language == "Dutch":
        if canonical == "name":
            batch_rules = (
                "Rules for each product name:\n"
                "- Maximum 40 characters, no commas, no brackets\n"
                "- ALL German words MUST be translated — zero German residue\n"
                "- Natural commercial Dutch furniture/kitchen/bathroom e-commerce\n"
                "- \"Sofa\"→\"bank\", \"Sessel\"→\"fauteuil\", "
                "\"Ecksofa\"→\"hoekbank\", \"Schlafsofa\"→\"slaapbank\", \"Sitzer\"→\"zits\"\n"
                "- \"Taschenfederkernmatratze\"→\"pocketveringmatras\"\n"
                "- \"GSP-Blende\"→\"vaatwasserpaneel\"\n"
                "- \"Einzelwaschtisch\"→\"enkele wastafel\", \"Doppelwaschtisch\"→\"dubbele wastafel\"\n"
                "- \"Grifflos\"→\"greeploos\"\n"
                "- Preserve model/collection names exactly (Asely, Arin, Bocca, Level36, etc.)"
            )
        elif canonical == "materialDetail":
            batch_rules = (
                "- Preserve <br> tags exactly — NEVER replace with semicolons\n"
                "- NEVER use semicolons (;) as property separators\n"
                "- Natural Dutch furniture/kitchen/bathroom terminology\n"
                "- Bezug→bekleding, Gestell→onderstel, Füße→poten\n"
                "- GSP-Blende→vaatwasserpaneel, BHT→\"B x H x D\", Grifflos→greeploos\n"
                "- Unterflurauszug→onderliggende ladegeleider\n"
                "- Küchenzeile→keukenblok, Arbeitsplatte→werkblad, Spüle→spoelbak\n"
                "- Unterschrank→onderkast, Hängeschrank→hangkast\n"
                "- Waschtisch→wastafelmeubel, Waschbecken→wastafel\n"
                "- Percentages: write \"80%\" not \"80 %\"; lowercase after %"
            )
        elif canonical == "qualityDetail":
            batch_rules = (
                "- Natural Dutch kitchen/bathroom/furniture e-commerce language\n"
                "- GSP-Blende→vaatwasserpaneel, BHT→\"B x H x D\", Grifflos→greeploos\n"
                "- Unterflurauszug→onderliggende ladegeleider\n"
                "- Preserve <br> tags exactly\n"
                "- Percentages: write \"80%\" not \"80 %\"; lowercase after %"
            )
        elif canonical == "deliveryScope":
            batch_rules = (
                "- Natural Dutch e-commerce language\n"
                "- bestehend aus→bestaande uit, ohne Dekoration→zonder decoratie\n"
                "- GSP-Blende→vaatwasserpaneel, Grifflos→greeploos\n"
                "- Preserve <br> tags exactly"
            )
        else:
            batch_rules = (
                "- Natural Dutch, not literal German\n"
                "- Remove all German traces\n"
                "- GSP-Blende→vaatwasserpaneel, BHT→\"B x H x D\", Grifflos→greeploos\n"
                "- Unterflurauszug→onderliggende ladegeleider\n"
                "- Preserve <br> tags exactly\n"
                "- Percentages: write \"80%\" not \"80 %\"; lowercase after %"
            )
        store_label = "Home24 Netherlands"
        target_label = "Dutch"
    else:
        if canonical == "name":
            batch_rules = (
                "Rules for each product name:\n"
                "- Maximum 40 characters, no commas, no brackets\n"
                "- ALL German words MUST be translated — zero German residue\n"
                "- Natural commercial French furniture e-commerce\n"
                "- \"Sofa\"→\"Canapé\", \"Sessel\"→\"Fauteuil\", "
                "\"Ecksofa\"→\"Canapé d'angle\", \"Sitzer\"→\"places\"\n"
                "- \"Loungeset\"→\"Salon de jardin\", \"Gartenessgruppe\"→\"Ensemble de jardin\"\n"
                "- \"Gartengruppe\"→\"Salon de jardin\", \"Sofaelement\"→\"Module de canapé\"\n"
                "- \"Taschenfederkernmatratze\"→\"Matelas ressorts ensachés\"\n"
                "- \"7-Zonen-Taschenfederkernmatratze\"→\"Matelas ressorts ensachés 7 zones\"\n"
                "- \"Matratze\"→\"Matelas\"\n"
                "- Preserve model/collection names exactly (Asely, Arin, Bocca, Vedene, Level36, etc.)"
            )
        elif canonical == "materialDetail":
            batch_rules = (
                "- Preserve <br> tags exactly — NEVER replace with semicolons\n"
                "- NEVER use semicolons (;) as property separators\n"
                "- ALL German words MUST be translated\n"
                "- \"pulverbeschichtet\"→\"thermolaqué\"\n"
                "- \"Geflecht\"/\"Polyrattan\"→\"résine tressée\"\n"
                "- \"Rattan\"→\"rotin\", \"Tischgestell\"→\"piètement de table\"\n"
                "- \"Bezug\"→\"revêtement\" (NEVER \"housse\" for frame components)\n"
                "- Natural French furniture/material terminology"
            )
        elif canonical == "deliveryScope":
            batch_rules = (
                "- ALL German words MUST be translated — zero German residue\n"
                "- \"Set bestehend aus\"→\"Ensemble composé de\"\n"
                "- \"bestehend aus\"→\"composé de\"\n"
                "- \"inkl.\"/\"inklusive\"→\"inclus(e)\"\n"
                "- \"ohne Dekoration\"→\"sans décoration\"\n"
                "- Prefer \"composé de\" over \"contenant\"\n"
                "- Preserve <br> tags exactly"
            )
        else:
            batch_rules = (
                "- ALL German words MUST be translated — zero German residue\n"
                "- Natural French, not literal German structure\n"
                "- \"pulverbeschichtet\"→\"thermolaqué\", \"Geflecht\"→\"résine tressée\"\n"
                "- \"bestehend aus\"→\"composé de\", \"ohne Dekoration\"→\"sans décoration\"\n"
                "- \"Kokosmatte\"→\"couche de coco\" (NEVER \"paillasson\")\n"
                "- \"Einseitige Kokosmatte\"→\"couche de coco sur une face\"\n"
                "- \"Doppeltuch\"→\"coutil double\"\n"
                "- \"Reißverschluss\"→\"fermeture éclair\"\n"
                "- \"4-seitiger Reißverschluss\"→\"fermeture éclair sur 4 côtés\"\n"
                "- \"Abnehmbarer Bezug\"→\"revêtement amovible\"\n"
                "- Preserve <br> tags exactly\n"
                "- Preserve numbers, dimensions and percentages exactly"
            )
        store_label = "Home24 France"
        target_label = "French"

    system_prompt = (
        f"You are a professional translator for {store_label} e-commerce.\n"
        f"Translate each German text to {target_label}.\n{batch_rules}{glossary_block}"
        f"{product_type_hint}\n\n"
        f"Return ONLY a valid JSON array of exactly {n} translated strings, "
        "in the same order as the input. No other text."
    )
    user_msg = f"Translate these {n} texts:\n{json.dumps(texts, ensure_ascii=False)}"

    for attempt in range(MAX_BATCH_RETRIES + 1):
        try:
            response = _api_call_with_retry(
                lambda: client.chat.completions.create(
                    model=OPENAI_MODEL,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user",   "content": user_msg},
                    ],
                    temperature=0.3,
                    max_tokens=min(4000, n * 200),
                    timeout=API_TIMEOUT_SECONDS,
                ),
                notify_fn=notify_fn,
            )
            if token_counter is not None and response.usage:
                token_counter["prompt_tokens"]     += response.usage.prompt_tokens
                token_counter["completion_tokens"] += response.usage.completion_tokens

            content = response.choices[0].message.content.strip()
            if not content.startswith("["):
                m = re.search(r'\[.*\]', content, re.DOTALL)
                content = m.group() if m else content

            translations = json.loads(content)
            if isinstance(translations, list) and len(translations) == n:
                out = [str(t).strip() for t in translations]
                if target_language == "Dutch":
                    out = [nl_post_process(t) for t in out]
                return out

        except Exception:
            pass

        if attempt < MAX_BATCH_RETRIES:
            continue
        break

    # Fallback: single-cell translation for each item
    out = _fallback_single_translations(client, texts, canonical, token_counter, glossary, notify_fn=notify_fn, target_language=target_language)
    if target_language == "Dutch":
        out = [nl_post_process(t) for t in out]
    return out


def _fallback_single_translations(
    client,
    texts: list[str],
    canonical: str,
    token_counter: dict,
    glossary: dict,
    notify_fn=None,
    target_language: str = "French",
) -> list[str]:
    glossary_block = _glossary_prompt_block(glossary)
    system_prompt  = _build_system_prompt(canonical, glossary_block, target_language)
    target_label   = "Dutch" if target_language == "Dutch" else "French"
    results        = []
    for text in texts:
        try:
            response = _api_call_with_retry(
                lambda t=text: client.chat.completions.create(
                    model=OPENAI_MODEL,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user",   "content": f"Translate to {target_label}:\n\n{t}"},
                    ],
                    temperature=0.3,
                    max_tokens=500,
                    timeout=API_TIMEOUT_SECONDS,
                ),
                notify_fn=notify_fn,
            )
            if token_counter is not None and response.usage:
                token_counter["prompt_tokens"]     += response.usage.prompt_tokens
                token_counter["completion_tokens"] += response.usage.completion_tokens
            results.append(response.choices[0].message.content.strip())
        except Exception:
            results.append(text)
    return results


def _run_batch_task(
    client,
    batch_id: int,
    batch_items: list,
    canonical: str,
    glossary: dict,
    retry_counter: list,
    retry_lock: threading.Lock,
    target_language: str = "French",
    product_type_hint: str = "",
) -> dict:
    """Worker: runs one translation batch in a thread. No shared-state mutations."""
    local_tokens  = {"prompt_tokens": 0, "completion_tokens": 0}
    local_glossary = {"total_hits": 0, "term_counts": {}}
    texts  = [item[4] for item in batch_items]
    failed = False
    start  = time.time()

    def _notify(_msg: str) -> None:
        with retry_lock:
            retry_counter[0] += 1

    try:
        translations = translate_batch(
            client, texts, canonical,
            local_tokens, glossary, local_glossary,
            notify_fn=_notify,
            target_language=target_language,
            product_type_hint=product_type_hint,
        )
    except Exception:
        failed       = True
        translations = list(texts)

    return {
        "batch_id":          batch_id,
        "batch_items":       batch_items,
        "translations":      translations,
        "prompt_tokens":     local_tokens["prompt_tokens"],
        "completion_tokens": local_tokens["completion_tokens"],
        "glossary_hits":     local_glossary.get("total_hits", 0),
        "glossary_terms":    local_glossary.get("term_counts", {}),
        "failed":            failed,
        "duration":          time.time() - start,
    }


def fix_german_residue(
    client,
    text: str,
    column_name: str,
    token_counter: dict | None = None,
    target_language: str = "French",
) -> str:
    if not text:
        return text

    # Step 1: Fast local term replacement — no API call needed
    locally_fixed = apply_furniture_terms(text, target_language)
    if locally_fixed != text:
        text = locally_fixed
        if not detect_german_residue(text, target_language):
            return text

    # Step 1b: Dutch NL post-processing pass (dekor, furniture, format)
    if target_language == "Dutch":
        text = nl_post_process(text, column_name)
        if not detect_german_residue(text, target_language):
            return text

    # Step 1c: French semantic normalization (no API)
    if target_language == "French":
        text = apply_french_semantic_normalization(text)
        if not detect_german_residue(text, target_language):
            return text

    if target_language == "Dutch":
        if column_name == "name":
            extra_rules = (
                "\n- Maximum 40 characters, no commas or brackets"
                "\n- \"Sofa\" → \"bank\" / \"Sessel\" → \"fauteuil\" / \"Sitzer\" → \"zits\""
                "\n- GSP-Blende → \"vaatwasserpaneel\""
                "\n- Einzelwaschtisch → \"enkele wastafel\" / Doppelwaschtisch → \"dubbele wastafel\""
                "\n- Grifflos → \"greeploos\""
            )
        elif column_name == "materialDetail":
            extra_rules = (
                "\n- Bezug → bekleding / Füße → poten / Gestell → onderstel"
                "\n- GSP-Blende → \"vaatwasserpaneel\" (NEVER leave GSP untranslated)"
                "\n- BHT / BxHxT / \"B x H x T\" → \"B x H x D\""
                "\n- Grifflos → \"greeploos\""
                "\n- Unterflurauszug → \"onderliggende ladegeleider\""
                "\n- Küchenzeile → \"keukenblok\" / Arbeitsplatte → \"werkblad\" / Spüle → \"spoelbak\""
                "\n- Waschtisch → \"wastafelmeubel\" / Waschbecken → \"wastafel\""
                "\n- Unterschrank → \"onderkast\" / Hängeschrank → \"hangkast\""
                "\n- Preserve <br> tags exactly — NEVER replace them with semicolons"
                "\n- NEVER use semicolons (;) as property separators"
                "\n- Percentages: \"100 %\" → \"100%\" (no space before %); lowercase after %"
            )
        elif column_name in ("qualityDetail", "deliveryScope"):
            extra_rules = (
                "\n- GSP-Blende → \"vaatwasserpaneel\""
                "\n- BHT / BxHxT → \"B x H x D\""
                "\n- Grifflos → \"greeploos\""
                "\n- Unterflurauszug → \"onderliggende ladegeleider\""
                "\n- Küchenzeile → \"keukenblok\" / Arbeitsplatte → \"werkblad\""
                "\n- Waschtisch → \"wastafelmeubel\" / Waschbecken → \"wastafel\""
                "\n- Preserve <br> tags exactly"
                "\n- Percentages: \"100 %\" → \"100%\"; lowercase after %"
            )
        else:
            extra_rules = (
                "\n- GSP-Blende → \"vaatwasserpaneel\""
                "\n- BHT / BxHxT → \"B x H x D\""
                "\n- Grifflos → \"greeploos\""
                "\n- Unterflurauszug → \"onderliggende ladegeleider\""
                "\n- Percentages: \"100 %\" → \"100%\"; lowercase after %"
            )
        fix_prompt = (
            f"This Dutch text still contains German words.\n"
            f"Rewrite it as clean, natural Dutch for Home24 Netherlands.\n"
            f"Replace ALL German words with Dutch equivalents.{extra_rules}\n\n"
            f"Text: {text}\n\n"
            f"Return ONLY the corrected Dutch text."
        )
        sys_msg = (
            "You are a professional Dutch editor for Home24 Netherlands e-commerce. "
            "You eliminate all German words from Dutch texts and rewrite them in natural, "
            "correct Dutch using proper kitchen and bathroom furniture terminology."
        )
    else:
        if column_name == "name":
            extra_rules = (
                "\n- Maximum 40 characters, no commas or brackets"
                "\n- \"Sofa\"→\"Canapé\" / \"Sessel\"→\"Fauteuil\" / \"Sitzer\"→\"places\""
                "\n- \"Loungeset\"→\"Salon de jardin\" / \"Gartenessgruppe\"→\"Ensemble de jardin\""
                "\n- \"Gartengruppe\"→\"Salon de jardin\" / \"Sofaelement\"→\"Module de canapé\""
                "\n- Preserve model names (Vedene, Arin, Bocca, etc.) exactly"
            )
        elif column_name == "materialDetail":
            extra_rules = (
                "\n- \"pulverbeschichtet\"→\"thermolaqué\""
                "\n- \"Geflecht\"/\"Polyrattan\"/\"Kunststoffgeflecht\"→\"résine tressée\""
                "\n- \"Rattan\"→\"rotin\" / \"Tischgestell\"→\"piètement de table\""
                "\n- \"Bezug\"→\"revêtement\" / \"Füße\"→\"pieds\" / \"Buche\"→\"hêtre\""
                "\n- Preserve <br> tags exactly — NEVER replace them with semicolons"
                "\n- NEVER use semicolons (;) as property separators"
            )
        elif column_name == "deliveryScope":
            extra_rules = (
                "\n- \"Set bestehend aus\"→\"Ensemble composé de\""
                "\n- \"bestehend aus\"→\"composé de\""
                "\n- \"ohne Dekoration\"→\"sans décoration\""
                "\n- \"inkl.\"/\"inklusive\"→\"inclus(e)\""
            )
        else:
            extra_rules = (
                "\n- \"pulverbeschichtet\"→\"thermolaqué\""
                "\n- \"Geflecht\"/\"Polyrattan\"→\"résine tressée\""
                "\n- \"bestehend aus\"→\"composé de\""
                "\n- \"ohne Dekoration\"→\"sans décoration\""
            )
        fix_prompt = (
            f"This French text still contains German words.\n"
            f"Rewrite it as clean, natural French for Home24 France.\n"
            f"Replace EVERY German word with the correct French equivalent.{extra_rules}\n\n"
            f"Text: {text}\n\n"
            f"Return ONLY the corrected French text."
        )
        sys_msg = (
            "You are a professional French editor for Home24 France e-commerce. "
            "You eliminate all German words from French texts and rewrite them in premium, natural French."
        )

    try:
        response = _api_call_with_retry(
            lambda: client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": sys_msg},
                    {"role": "user",   "content": fix_prompt},
                ],
                temperature=0.2,
                max_tokens=500,
                timeout=API_TIMEOUT_SECONDS,
            ),
        )
        if token_counter is not None and response.usage:
            token_counter["prompt_tokens"]     += response.usage.prompt_tokens
            token_counter["completion_tokens"] += response.usage.completion_tokens
        corrected = response.choices[0].message.content.strip()
        if column_name == "name":
            corrected = validate_product_name(corrected)
        return corrected
    except Exception:
        return text


# =============================================================================
# PREMIUM AI REFINEMENT
# =============================================================================

# Pattern: pure dimensions / percentages / numbers — never refine these
_REFINE_SKIP_PATTERN = re.compile(
    r'^[\d\s.,×xX%cm/mmkgcl×²³\-\+]+$'
)


def _should_refine(text: str, canonical: str) -> bool:
    """Return True if this translated cell is a good candidate for AI refinement."""
    if canonical not in REFINEMENT_COLUMNS:
        return False
    text = text.strip()
    if len(text) < REFINEMENT_MIN_CHARS:
        return False
    if _REFINE_SKIP_PATTERN.match(text):
        return False
    # Single word — likely a direct glossary match already correct
    if " " not in text:
        return False
    return True


def refine_batch(
    client,
    items: list[tuple[str, str]],
    token_counter: dict,
    target_language: str = "French",
) -> list[str]:
    """
    items: list of (translated_text, canonical_column_type)
    Sends a premium refinement pass to the AI.
    Returns refined texts in the same order; falls back to originals on failure.
    """
    if not items:
        return []

    n     = len(items)
    texts = [t for t, _ in items]

    if target_language == "Dutch":
        system_prompt = (
            "You are a premium Dutch copywriter for Home24 Netherlands furniture e-commerce.\n"
            "Improve the naturalness and professional tone of these Dutch product texts.\n\n"
            "Rules — follow every one:\n"
            "- Preserve the exact meaning — do NOT invent or remove product information\n"
            "- Improve Dutch fluency and furniture vocabulary\n"
            "- Avoid awkward literal translations from German patterns\n"
            "- Use established Dutch furniture/e-commerce terminology\n"
            "- Preserve ALL <br> tags exactly — do not add, remove, or move them\n"
            "- NEVER replace <br> tags with semicolons (;)\n"
            "- NEVER use semicolons as property separators\n"
            "- Do NOT modify numbers, dimensions, or percentages\n"
            "- Keep product names concise — do NOT make them longer\n"
            "- Do not exaggerate marketing claims\n"
            "- If the text is already natural and correct, return it unchanged\n\n"
            f"Return ONLY a valid JSON array of exactly {n} strings, in input order. No other text."
        )
        user_msg = f"Refine these {n} Dutch texts:\n{json.dumps(texts, ensure_ascii=False)}"
    else:
        system_prompt = (
            "You are a premium French copywriter for Home24 France furniture e-commerce.\n"
            "Improve the naturalness and professional tone of these French product texts.\n\n"
            "Rules — follow every one:\n"
            "- Preserve the exact meaning — do NOT invent or remove product information\n"
            "- Improve French fluency and furniture vocabulary\n"
            "- Avoid awkward literal translations from German patterns\n"
            "  Examples: 'décoré' → 'revêtu', 'revêtu d\\'un film décoratif' → 'revêtu de film mélaminé'\n"
            "- Use established French furniture/e-commerce terminology\n"
            "- Outdoor: 'thermolaqué' not 'poudré', 'résine tressée' not 'tressage plastique'\n"
            "- Delivery: 'ensemble composé de' not 'set contenant', 'composé de' not 'comprenant'\n"
            "- Frame: 'piètement' or 'structure' — NEVER 'housse' for frame components\n"
            "- If text says 'Structure' when it should be 'Revêtement', leave it — preserve meaning\n"
            "- Preserve ALL <br> tags exactly — do not add, remove, or move them\n"
            "- NEVER replace <br> tags with semicolons (;)\n"
            "- NEVER use semicolons as property separators\n"
            "- Do NOT modify numbers, dimensions, or percentages\n"
            "- Keep product names concise — do NOT make them longer\n"
            "- Do not exaggerate marketing claims\n"
            "- If the text is already natural and correct, return it unchanged\n\n"
            f"Return ONLY a valid JSON array of exactly {n} strings, in input order. No other text."
        )
        user_msg = f"Refine these {n} French texts:\n{json.dumps(texts, ensure_ascii=False)}"

    try:
        response = _api_call_with_retry(
            lambda: client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user",   "content": user_msg},
                ],
                temperature=0.35,
                max_tokens=min(4000, n * 220),
                timeout=API_TIMEOUT_SECONDS,
            )
        )
        if token_counter is not None and response.usage:
            token_counter["prompt_tokens"]     += response.usage.prompt_tokens
            token_counter["completion_tokens"] += response.usage.completion_tokens

        content = response.choices[0].message.content.strip()
        if not content.startswith("["):
            m = re.search(r'\[.*\]', content, re.DOTALL)
            content = m.group() if m else content

        refined = json.loads(content)
        if isinstance(refined, list) and len(refined) == n:
            return [str(r).strip() for r in refined]
    except Exception:
        pass

    return texts  # fallback — return originals unchanged


def _pipeline_status_html(steps: list[tuple[str, str]]) -> str:
    """Render a compact Quality Pipeline status panel.
    steps = [(label, status)] where status is "done" | "running" | "skipped" | "pending".
    """
    icons   = {"done": "✓", "running": "⟳", "skipped": "—", "pending": "○"}
    colours = {"done": "#22c55e", "running": "#f59e0b", "skipped": "#94a3b8", "pending": "#94a3b8"}
    rows = ""
    for label, status in steps:
        icon   = icons.get(status, "○")
        colour = colours.get(status, "#94a3b8")
        rows += (
            f'<div style="display:flex;align-items:center;gap:8px;padding:4px 0;">'
            f'<span style="color:{colour};font-weight:700;font-size:14px;width:16px">{icon}</span>'
            f'<span style="font-size:13px;color:{"#1e293b" if status=="done" else "#64748b"}">{label}</span>'
            f'</div>'
        )
    return (
        '<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;'
        'padding:12px 16px;margin-top:12px;">'
        '<div style="font-weight:700;font-size:12px;text-transform:uppercase;letter-spacing:.06em;'
        'color:#64748b;margin-bottom:8px;">Quality Pipeline</div>'
        + rows + '</div>'
    )


def _refinement_progress_html(sheet: str, done: int, total: int) -> str:
    pct = int((done / max(total, 1)) * 100)
    return f"""
    <div class="prog-shell">
        <div class="prog-head">
            <div>
                <div class="prog-phase">Phase 2.5 — Premium French Refinement</div>
                <div class="prog-sheet">Sheet: {sheet} · Elevating translation quality</div>
            </div>
            <span class="prog-badge"><span class="prog-badge-dot"></span>REFINING</span>
        </div>
        <div class="prog-track">
            <div class="prog-bar" style="width:{pct}%"></div>
        </div>
        <div class="prog-item">
            <div class="prog-item-dot"></div>
            <span class="prog-item-col">AI refinement pass</span>
            <span class="prog-item-row">{done} / {total} cells refined</span>
        </div>
        <div class="prog-stats">
            <div><span class="prog-stat-val">{done}</span><span class="prog-stat-lbl">Done</span></div>
            <div><span class="prog-stat-val">{total - done}</span><span class="prog-stat-lbl">Remaining</span></div>
            <div><span class="prog-stat-val">{pct}%</span><span class="prog-stat-lbl">Progress</span></div>
        </div>
    </div>
    """


# =============================================================================
# FRENCH CAPITALIZATION ENGINE
# =============================================================================

# Known product/collection names whose capitalisation must be preserved
_KNOWN_BRANDS: frozenset[str] = frozenset({
    "Arin", "Bocca", "Level36", "Sonoma", "Loft", "Scandi",
    "Asely", "Vedene", "Lano", "Bori", "Nalo", "Veda",
})

# Split on <br> tags (case-insensitive), preserving the tag in the result
_CAP_BR_RE = re.compile(r"(<br\s*/?>)", re.IGNORECASE)

# Capitalized word immediately following a percentage → needs lowercasing
_CAP_PERCENT_LOWER_RE = re.compile(
    r"(\d+\s*%)\s+([A-ZÉÀÈÙÂÊÎÔÛËÏÜÇŒÆ][a-zéàèùâêîôûëïüçœæA-ZÉÀÈÙÂÊÎÔÛËÏÜÇŒÆA-zA-Z]*)",
    re.UNICODE,
)

# Two words around a bare '/' with no surrounding spaces — structural separator
# Minimum 3 chars per side; matches any case so cap_first handles normalisation
_CAP_SLASH_RE = re.compile(
    r"([a-zA-Zéàèùâêîôûëïüçœæ][a-zA-Zéàèùâêîôûëïüçœæ'’\-]{2,})"
    r"/"
    r"([a-zA-Zéàèùâêîôûëïüçœæ][a-zA-Zéàèùâêîôûëïüçœæ'’\-]{2,})",
    re.UNICODE,
)


def _cap_first(s: str) -> str:
    """Capitalize the first alphabetic character of s only when s starts with a letter.

    Skips when the segment begins with a digit or symbol (e.g. "100% polyester",
    "60% coton/lin") to avoid re-capitalising words that follow percentages.
    """
    stripped = s.lstrip()
    if not stripped or not stripped[0].isalpha():
        return s
    for i, ch in enumerate(s):
        if ch.isalpha():
            return s[:i] + ch.upper() + s[i + 1:]
    return s


def _apply_slash_caps(segment: str) -> str:
    """Capitalise both sides of structural '/' separators within a text segment."""
    def repl(m: re.Match) -> str:
        # Skip if a percentage figure appears before this slash in the same segment
        # (material composition context: "60% polyester/coton")
        if re.search(r"\d+\s*%", segment[: m.start()]):
            return m.group(0)
        return _cap_first(m.group(1)) + "/" + _cap_first(m.group(2))

    return _CAP_SLASH_RE.sub(repl, segment)


def _apply_percent_lowercase(segment: str) -> str:
    """Lowercase a capitalised word that immediately follows a percentage figure."""
    return _CAP_PERCENT_LOWER_RE.sub(
        lambda m: m.group(1) + " " + m.group(2)[0].lower() + m.group(2)[1:],
        segment,
    )


def _restore_brand_caps(text: str) -> str:
    """Re-apply correct capitalisation for known product/brand names."""
    for brand in _KNOWN_BRANDS:
        text = re.sub(
            r"\b" + re.escape(brand) + r"\b",
            brand,
            text,
            flags=re.IGNORECASE | re.UNICODE,
        )
    return text


def apply_french_capitalization_rules(
    text: str, canonical: str, glossary: dict | None = None
) -> str:
    """Apply French e-commerce capitalisation rules to a translated cell value.

    Execution order per segment (between <br> tags):
      1. Structural '/' caps  — before start-cap so both sides get processed
      2. Percentage lowercase — "100% Polyester" → "100% polyester"
      3. Start-of-segment cap — always capitalise the first alphabetic character
      4. Known brand restore  — re-capitalise brand/model names anywhere in text

    Rules honoured:
    - Capitalise start of text + each <br>-delimited segment
    - Capitalise both words around a structural '/' (not inside % compositions)
    - Never capitalise after ':' (French convention)
    - Lowercase fibre/material words immediately following a percentage
    - Preserve HTML <br> tags, numbers, dimensions, and glossary terms
    - Restore known product/brand names to their canonical capitalisation
    """
    if not text or not text.strip():
        return text

    parts = _CAP_BR_RE.split(text)
    result: list[str] = []
    capitalize_next = True  # first segment always gets a capital start

    for part in parts:
        if _CAP_BR_RE.match(part):
            result.append("<br>")
            capitalize_next = True
            continue
        if not part:
            result.append(part)
            continue

        # Step 1: structural slash caps (before start-cap avoids missed right-side words)
        part = _apply_slash_caps(part)
        # Step 2: percentage lowercase
        part = _apply_percent_lowercase(part)
        # Step 3: capitalise start of segment
        if capitalize_next and part.strip():
            part = _cap_first(part)
            capitalize_next = False
        # Step 4: restore brand names
        part = _restore_brand_caps(part)
        result.append(part)

    return "".join(result)


# Dutch: remove space between number and % ("100 %" → "100%")
_NL_PCT_SPACE_RE = re.compile(r'(\d+)\s+%', re.UNICODE)


def apply_dutch_capitalization_rules(
    text: str, canonical: str, glossary: dict | None = None
) -> str:
    """Apply Dutch e-commerce formatting rules to a translated cell value.

    Per segment (split on <br>):
      1. Remove space before % ("100 %" → "100%")
      2. Lowercase capitalised word after % ("100% Polyester" → "100% polyester")
      3. Capitalise start of segment
      4. Restore known brand/model names
      5. Post-colon lowercase (Dutch rule: after ":", values are lowercase)
    """
    if not text or not text.strip():
        return text

    parts = _CAP_BR_RE.split(text)
    result: list[str] = []
    capitalize_next = True

    for part in parts:
        if _CAP_BR_RE.match(part):
            result.append("<br>")
            capitalize_next = True
            continue
        if not part:
            result.append(part)
            continue

        # Step 1: "100 %" → "100%"
        part = _NL_PCT_SPACE_RE.sub(r'\1%', part)
        # Step 2: "100% Polyester" → "100% polyester"
        part = _apply_percent_lowercase(part)
        # Step 3: capitalise start of segment
        if capitalize_next and part.strip():
            part = _cap_first(part)
            capitalize_next = False
        # Step 4: restore brand names
        part = _restore_brand_caps(part)
        # Step 5: Dutch rule — after ":" values are lowercase (colors, materials, etc.)
        part = apply_nl_post_colon_lowercase(part, canonical)
        result.append(part)

    return "".join(result)


# =============================================================================
# SEMICOLON → BR POST-PROCESSING
# =============================================================================

# Columns where the source uses <br> as a property separator.
# When the LLM replaces those <br> with " ; " I restore them — but ONLY when
# the source itself had <br>, so natural semicolons in source text are kept.
_SEMICOLON_BR_CANONICALS = frozenset({
    "materialDetail", "qualityDetail", "deliveryScope", "colorDetail",
})

_SEMICOLON_SEP_RE = re.compile(r"\s*;\s*", re.UNICODE)


# =============================================================================
# CSV EXPORT
# =============================================================================

_NAME_COLUMN_ALIASES = {
    "name", "productname", "product name", "produktname", "produkt name",
    "nom", "designation", "bezeichnung", "title", "titre",
}


def generate_csv_export(
    excel_bytes: bytes,
    sheet_name: str,
    original_filename: str,
    header_row: int = 1,
    force_exclude_header: str | None = None,
    target_language: str = "French",
) -> tuple[bytes | None, str | None, str | None]:
    """
    Reads the translated workbook from bytes, drops the name column, and
    returns (csv_bytes, csv_filename, excluded_col_name).
    Returns (None, None, None) if no name column is found and no forced header given.
    """
    import io
    from openpyxl import load_workbook as _load_wb

    wb = _load_wb(io.BytesIO(excel_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # Read header row
    headers = [
        (ws.cell(row=header_row, column=c).value or "")
        for c in range(1, ws.max_column + 1)
    ]

    # Find name column
    exclude_idx = None
    exclude_name = None

    if force_exclude_header is not None:
        for i, h in enumerate(headers):
            if str(h).strip() == force_exclude_header.strip():
                exclude_idx = i
                exclude_name = force_exclude_header
                break
    else:
        for i, h in enumerate(headers):
            canonical, _, _ = _classify_header(str(h))
            if canonical == "name":
                exclude_idx = i
                exclude_name = str(h).strip()
                break
        # fallback: alias check on raw header
        if exclude_idx is None:
            for i, h in enumerate(headers):
                norm = _normalize_col_header(str(h))
                if norm in _NAME_COLUMN_ALIASES or norm.replace(" ", "") in _NAME_COLUMN_ALIASES:
                    exclude_idx = i
                    exclude_name = str(h).strip()
                    break

    if exclude_idx is None:
        return None, None, None

    # Build CSV content
    lang_prefix = "NL" if target_language == "Dutch" else "FR"
    keep_cols = [i for i in range(len(headers)) if i != exclude_idx]
    base_name = original_filename.replace(".xlsx", "").replace(".xls", "")
    # Strip any existing language prefix before adding the correct one
    for pfx in ("FR-", "NL-"):
        if base_name.startswith(pfx):
            base_name = base_name[len(pfx):]
    csv_filename = f"{lang_prefix}-{base_name}.csv"

    lines: list[str] = []
    for row in ws.iter_rows(min_row=header_row, values_only=True):
        row_list = list(row)
        values = [str(row_list[i]) if row_list[i] is not None else "" for i in keep_cols]
        # Escape semicolons and quotes in cell values
        escaped = []
        for v in values:
            if ";" in v or '"' in v or "\n" in v:
                escaped.append('"' + v.replace('"', '""') + '"')
            else:
                escaped.append(v)
        lines.append(";".join(escaped))

    content = "\n".join(lines)
    csv_bytes = "﻿".encode("utf-8") + content.encode("utf-8")
    return csv_bytes, csv_filename, exclude_name


# =============================================================================
# EXCEL PROCESSING
# =============================================================================

def detect_target_sheet(sheet_names: list) -> str | None:
    for name in CANDIDATE_SHEETS:
        if name in sheet_names:
            return name
    if len(sheet_names) == 1:
        return sheet_names[0]
    return None


def _score_sheet_for_picking(ws, max_rows: int = 100) -> dict:
    """Lightweight 100-row scan for sheet selection scoring (one pass, read_only safe)."""
    row_buffer: dict[int, list[tuple[int, object]]] = {}
    non_empty_cells = 0
    real_max_col = 0

    try:
        # Force iter_rows to scan beyond a stale/corrupt <dimension> XML element.
        # In read_only mode the iterator caps at ws.max_row (from the dimension
        # tag). Files where that tag says "A1" instead of "A1:F264" would yield
        # only the first cell unless we supply an explicit large bound.
        _dim_rows = getattr(ws, "max_row", None) or 0
        _dim_cols = getattr(ws, "max_column", None) or 0
        _iter_max_row = max(_dim_rows + 1, max_rows + 1, 2000)
        _iter_max_col = max(_dim_cols + 1, 256)
        for row_tuple in ws.iter_rows(min_row=1, max_row=_iter_max_row, max_col=_iter_max_col):
            if not row_tuple:
                continue
            row_num = getattr(row_tuple[0], "row", None)
            if row_num is None:
                continue
            if row_num > max_rows:
                break
            for cell in row_tuple:
                val = getattr(cell, "value", None)
                if val is None or not str(val).strip():
                    continue
                c = getattr(cell, "column", None) or 0
                if not c:
                    continue
                non_empty_cells += 1
                if c > real_max_col:
                    real_max_col = c
                row_buffer.setdefault(row_num, []).append((c, val))
    except Exception:
        pass

    best_score = -1
    header_row = 1
    peek_headers: dict[str, int] = {}

    for row_num in sorted(row_buffer.keys()):
        score = 0
        candidate_headers: dict[str, int] = {}
        for col_idx, value in row_buffer[row_num]:
            text = str(value).strip()
            if not text:
                continue
            try:
                float(text)
                score -= 3
                continue
            except (ValueError, TypeError):
                pass
            candidate_headers[text] = col_idx
            norm = _normalize_col_header(text)
            norm_c = norm.replace(' ', '')
            if norm in HOME24_TRANSLATABLE_NORMALIZED or norm_c in HOME24_TRANSLATABLE_NORMALIZED:
                score += 10
            elif any(kw in norm for kw in HEADER_SCORE_KEYWORDS):
                score += 3
            if _is_protected(norm, text):
                score += 5
        if score > best_score:
            best_score = score
            header_row = row_num
            peek_headers = candidate_headers

    h24_headers: list[str] = []
    protected_headers: list[str] = []
    for raw_h in peek_headers:
        norm = _normalize_col_header(raw_h)
        norm_c = norm.replace(' ', '')
        if norm in HOME24_TRANSLATABLE_NORMALIZED or norm_c in HOME24_TRANSLATABLE_NORMALIZED:
            h24_headers.append(raw_h)
        elif _is_protected(norm, raw_h):
            protected_headers.append(raw_h)

    has_article_number = any(
        "articlenumber" in _normalize_col_header(h).replace(' ', '')
        or "artikelnummer" in _normalize_col_header(h).replace(' ', '')
        for h in peek_headers
    )

    selection_score = (
        len(h24_headers) * 10
        + len(protected_headers) * 5
        + min(non_empty_cells // 10, 50)
        + (8 if has_article_number else 0)
    )

    return {
        "selection_score":   selection_score,
        "h24_headers":       h24_headers,
        "protected_headers": protected_headers,
        "peek_headers":      peek_headers,
        "header_row":        header_row,
        "non_empty_cells":   non_empty_cells,
        "real_max_col":      real_max_col,
    }


def score_all_sheets(wb, max_rows: int = 100) -> dict[str, dict]:
    """Score every sheet in *wb* for translation suitability."""
    results: dict[str, dict] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        state = getattr(ws, "sheet_state", "visible")
        if state in ("hidden", "veryHidden"):
            results[name] = {
                "selection_score": -1000, "hidden": True,
                "h24_headers": [], "protected_headers": [],
                "peek_headers": {}, "header_row": 1,
                "non_empty_cells": 0, "real_max_col": 0,
            }
            continue
        score_dict = _score_sheet_for_picking(ws, max_rows=max_rows)
        score_dict["hidden"] = False
        if name in CANDIDATE_SHEETS:
            score_dict["selection_score"] += 2
        results[name] = score_dict
    return results


def pick_best_sheet(all_scores: dict[str, dict]) -> str:
    """Return the sheet name with the highest selection_score."""
    return max(all_scores, key=lambda n: all_scores[n]["selection_score"])


def scan_sheet(
    ws,
    header_scan_rows: int = 50,
    row_hard_limit: int = 5000,
) -> tuple[int, int, int, dict]:
    """
    Single-pass worksheet scan — safe in read_only mode (never iterates twice).

    Reads actual cell data from the XML stream; does NOT rely on ws.max_row or
    ws.max_column, which openpyxl can misreport when workbook metadata is stale.

    Returns:
        real_max_row  — last row that contains any non-empty cell
        real_max_col  — last column that contains any non-empty cell
        header_row    — 1-based row number of the best-scoring candidate header row
        headers       — {raw_header_text: col_index} for that header row
    """
    real_max_row = 0
    real_max_col = 0
    # Buffer the first header_scan_rows rows for scoring
    row_buffer: dict[int, list[tuple[int, object]]] = {}

    try:
        # In read_only mode, openpyxl caps iter_rows at ws.max_row (from the
        # <dimension> XML tag). Files with a stale tag (e.g. "A1" instead of
        # "A1:F264") would stop after the first row. Passing explicit bounds
        # that exceed the tag forces a full XML stream scan.
        _dim_rows = getattr(ws, "max_row", None) or 0
        _dim_cols = getattr(ws, "max_column", None) or 0
        _iter_max_row = max(_dim_rows + 1, row_hard_limit)
        _iter_max_col = max(_dim_cols + 1, 256)
        for row_tuple in ws.iter_rows(min_row=1, max_row=_iter_max_row, max_col=_iter_max_col):
            if not row_tuple:
                continue
            # openpyxl read_only mode yields EmptyCell for blank rows;
            # EmptyCell lacks .row in older versions — skip via getattr
            row_num = getattr(row_tuple[0], "row", None)
            if row_num is None:
                continue  # EmptyCell row — skip, don't break
            if row_num > row_hard_limit:
                break
            for cell in row_tuple:
                val = getattr(cell, "value", None)
                if val is None:
                    continue
                if not str(val).strip():
                    continue
                r = getattr(cell, "row", None) or 0
                c = getattr(cell, "column", None) or 0
                if not r or not c:
                    continue
                if r > real_max_row:
                    real_max_row = r
                if c > real_max_col:
                    real_max_col = c
                if r <= header_scan_rows:
                    row_buffer.setdefault(r, []).append((c, val))
    except Exception as _scan_exc:
        _DETECT_LOG.warning("scan_sheet: exception during row iteration — %s", _scan_exc)

    best_row    = 1
    best_score  = -1
    best_headers: dict[str, int] = {}

    for row_num in sorted(row_buffer.keys()):
        score   = 0
        headers: dict[str, int] = {}
        for col_idx, value in row_buffer[row_num]:
            raw  = str(value)
            text = raw.strip()
            if not text:
                continue
            try:
                float(raw)
                score -= 3
                continue
            except (ValueError, TypeError):
                pass
            headers[text] = col_idx
            score += 1
            norm = _normalize_col_header(text)
            if any(kw in norm for kw in HEADER_SCORE_KEYWORDS):
                score += 3
            # Home24 catalog headers get a strong bonus so they are never beaten
            # by a data row that happens to score well on keywords.
            if norm in HOME24_TRANSLATABLE_NORMALIZED or norm.replace(' ', '') in HOME24_TRANSLATABLE_NORMALIZED:
                score += 5
        _DETECT_LOG.debug("scan_sheet: row %d  score=%d  headers=%s", row_num, score, list(headers.keys()))
        if score > best_score:
            best_score   = score
            best_row     = row_num
            best_headers = headers

    _DETECT_LOG.debug(
        "scan_sheet: chosen header_row=%d  score=%d  real_max_row=%d  real_max_col=%d  headers=%s",
        best_row, best_score, real_max_row, real_max_col, list(best_headers.keys()),
    )
    return real_max_row, real_max_col, best_row, best_headers


def detect_header_row(worksheet, max_rows: int = 20) -> tuple[int, dict]:
    """
    Wrapper kept for backward compat. Prefer scan_sheet() for new callers.
    Uses scan_sheet internally so it inherits the real-range fix.
    """
    _, _, header_row, headers = scan_sheet(worksheet, header_scan_rows=max_rows)
    return header_row, headers


def detect_columns(worksheet) -> dict:
    """Backward-compatible wrapper — returns {raw_header: col_idx} from detected header row."""
    _, headers = detect_header_row(worksheet)
    return headers


def _is_protected(norm: str, raw: str) -> bool:
    """Return True if this column should never be translated."""
    # Exact match on fully-normalized, space-collapsed form
    norm_c = norm.replace(' ', '')
    if norm_c in PROTECTED_EXACT or norm in PROTECTED_EXACT:
        return True
    # Substring match for longer protected patterns
    if any(pk in norm_c for pk in PROTECTED_SUBSTRINGS):
        return True
    return False


def _classify_header(header: str) -> tuple[str | None, str, str]:
    """
    Classify a raw header into a canonical translation target.
    Returns (canonical_or_None, normalized_form, match_tier).
    """
    norm   = _normalize_col_header(header)
    norm_c = norm.replace(' ', '')          # collapsed form

    # Home24 fast-path — known catalog headers are accepted immediately.
    # Checked before T1/T2/T3 so variant-style files never fall through.
    if norm in HOME24_TRANSLATABLE_NORMALIZED:
        return HOME24_TRANSLATABLE_CANONICAL[norm], norm, "H24-exact"
    if norm_c in HOME24_TRANSLATABLE_NORMALIZED:
        return HOME24_TRANSLATABLE_CANONICAL[norm_c], norm, "H24-collapsed"

    # T1 — exact match (try both spaced and collapsed)
    for can, aliases in TRANSLATE_ALIASES_T1.items():
        if norm in aliases or norm_c in aliases:
            return can, norm, "T1-exact"

    # T2 — substring of normalized header
    for can, substrings in TRANSLATE_ALIASES_T2.items():
        if any(sub in norm for sub in substrings):
            return can, norm, "T2-substring"

    # T3 — word-set overlap (≥2 words must match)
    header_words = set(norm.split())
    best_can, best_score = None, 0
    for can, word_set in CANONICAL_WORD_SETS.items():
        score = len(header_words & word_set)
        if score >= 2 and score > best_score:
            best_can, best_score = can, score
    if best_can:
        return best_can, norm, f"T3-wordset({best_score})"

    return None, norm, "no-match"


def classify_columns(all_columns: dict) -> dict:
    """
    Classify raw {header: col_idx} into translatable / protected / ignored.
    Returns a dict with debug metadata for the column report.
    """
    to_translate:    dict[str, tuple] = {}
    protected:       dict[str, int]   = {}
    ignored:         dict[str, int]   = {}
    possible_missed: list[str]        = []
    normalized_map:  dict[str, str]   = {}
    ignored_reasons: dict[str, str]   = {}
    match_tiers:     dict[str, str]   = {}

    _DETECT_LOG.debug("classify_columns: evaluating %d headers: %s",
                      len(all_columns), list(all_columns.keys()))

    for header, col_idx in all_columns.items():
        norm = _normalize_col_header(header)
        normalized_map[header] = norm

        if _is_protected(norm, header):
            protected[header] = col_idx
            match_tiers[header] = "protected"
            _DETECT_LOG.debug("  col=%d  %-30r  norm=%-30r  → PROTECTED", col_idx, header, norm)
            continue

        canonical, _, tier = _classify_header(header)

        if canonical is not None:
            to_translate[header] = (col_idx, canonical)
            match_tiers[header] = tier
            _DETECT_LOG.debug("  col=%d  %-30r  norm=%-30r  → TRANSLATE as %-25r  tier=%s",
                              col_idx, header, norm, canonical, tier)
        else:
            ignored[header] = col_idx
            reason = "No alias / keyword match"
            ignored_reasons[header] = reason
            match_tiers[header] = "ignored"
            _DETECT_LOG.debug("  col=%d  %-30r  norm=%-30r  → IGNORED (%s)", col_idx, header, norm, reason)
            if any(ik in norm for ik in IMPORTANT_KEYWORDS):
                possible_missed.append(header)

    _DETECT_LOG.debug(
        "classify_columns result: translatable=%d  protected=%d  ignored=%d  possible_missed=%s",
        len(to_translate), len(protected), len(ignored), possible_missed,
    )

    return {
        "to_translate":    to_translate,
        "protected":       protected,
        "ignored":         ignored,
        "possible_missed": possible_missed,
        "normalized_map":  normalized_map,
        "ignored_reasons": ignored_reasons,
        "match_tiers":     match_tiers,
    }


def _progress_html(phase: str, sheet: str, col_header: str, row_index: int,
                   total_rows: int, pct: int, elapsed: float, eta: float,
                   translated: int, skipped: int, fixes: int) -> str:
    return f"""
    <div class="prog-shell">
        <div class="prog-head">
            <div>
                <div class="prog-phase">{phase}</div>
                <div class="prog-sheet">Sheet: {sheet}</div>
            </div>
            <span class="prog-badge">
                <span class="prog-badge-dot"></span>ACTIVE
            </span>
        </div>
        <div class="prog-track">
            <div class="prog-bar" style="width:{pct}%"></div>
        </div>
        <div class="prog-item">
            <div class="prog-item-dot"></div>
            <span class="prog-item-col">{col_header}</span>
            <span class="prog-item-row">row {row_index} / {total_rows}</span>
        </div>
        <div class="prog-stats">
            <div><span class="prog-stat-val">{translated}</span><span class="prog-stat-lbl">Translated</span></div>
            <div><span class="prog-stat-val">{skipped}</span><span class="prog-stat-lbl">Skipped</span></div>
            <div><span class="prog-stat-val">{fixes}</span><span class="prog-stat-lbl">Fixes</span></div>
            <div><span class="prog-stat-val">{format_time(elapsed)}</span><span class="prog-stat-lbl">Elapsed</span></div>
            <div><span class="prog-stat-val">~{format_time(eta)}</span><span class="prog-stat-lbl">Remaining</span></div>
            <div><span class="prog-stat-val">{pct}%</span><span class="prog-stat-lbl">Progress</span></div>
        </div>
    </div>
    """


def _batch_progress_html(phase: str, sheet: str, batch_num: int, batch_size: int,
                         total_api: int, api_done: int, elapsed: float, eta: float,
                         tm_hits: int, tm_misses: int, pct: int) -> str:
    return f"""
    <div class="prog-shell">
        <div class="prog-head">
            <div>
                <div class="prog-phase">{phase}</div>
                <div class="prog-sheet">Sheet: {sheet}</div>
            </div>
            <span class="prog-badge">
                <span class="prog-badge-dot"></span>ACTIVE
            </span>
        </div>
        <div class="prog-track">
            <div class="prog-bar" style="width:{pct}%"></div>
        </div>
        <div class="prog-item">
            <div class="prog-item-dot"></div>
            <span class="prog-item-col">Batch {batch_num}</span>
            <span class="prog-item-row">{api_done} / {total_api} cells</span>
        </div>
        <div class="prog-stats">
            <div><span class="prog-stat-val">{batch_num}</span><span class="prog-stat-lbl">Batches</span></div>
            <div><span class="prog-stat-val">{batch_size}</span><span class="prog-stat-lbl">Batch Size</span></div>
            <div><span class="prog-stat-val">{tm_hits}</span><span class="prog-stat-lbl">TM Hits</span></div>
            <div><span class="prog-stat-val">{tm_misses}</span><span class="prog-stat-lbl">API Queue</span></div>
            <div><span class="prog-stat-val">{format_time(elapsed)}</span><span class="prog-stat-lbl">Elapsed</span></div>
            <div><span class="prog-stat-val">~{format_time(eta)}</span><span class="prog-stat-lbl">Remaining</span></div>
            <div><span class="prog-stat-val">{pct}%</span><span class="prog-stat-lbl">Progress</span></div>
        </div>
    </div>
    """


def _parallel_progress_html(
    sheet: str,
    total_batches: int,
    completed: int,
    running: int,
    failed: int,
    cells_done: int,
    total_cells: int,
    tm_hits: int,
    elapsed: float,
    eta: float,
    pct: int,
) -> str:
    fail_str = f" · <span style='color:#ef4444;'>{failed} failed</span>" if failed else ""
    return f"""
    <div class="prog-shell">
        <div class="prog-head">
            <div>
                <div class="prog-phase">Parallel Batch Translation</div>
                <div class="prog-sheet">Sheet: {sheet} · {running} batch(es) running concurrently{fail_str}</div>
            </div>
            <span class="prog-badge"><span class="prog-badge-dot"></span>ACTIVE</span>
        </div>
        <div class="prog-track">
            <div class="prog-bar" style="width:{pct}%"></div>
        </div>
        <div class="prog-item">
            <div class="prog-item-dot"></div>
            <span class="prog-item-col">Processing {running} batch(es) in parallel</span>
            <span class="prog-item-row">{cells_done} / {total_cells} cells</span>
        </div>
        <div class="prog-stats">
            <div><span class="prog-stat-val">{completed}/{total_batches}</span><span class="prog-stat-lbl">Batches done</span></div>
            <div><span class="prog-stat-val">{running}</span><span class="prog-stat-lbl">Running</span></div>
            <div><span class="prog-stat-val">{tm_hits}</span><span class="prog-stat-lbl">TM Hits</span></div>
            <div><span class="prog-stat-val">{failed}</span><span class="prog-stat-lbl">Failed</span></div>
            <div><span class="prog-stat-val">{format_time(elapsed)}</span><span class="prog-stat-lbl">Elapsed</span></div>
            <div><span class="prog-stat-val">~{format_time(eta)}</span><span class="prog-stat-lbl">Remaining</span></div>
            <div><span class="prog-stat-val">{pct}%</span><span class="prog-stat-lbl">Progress</span></div>
        </div>
    </div>
    """


def process_excel_with_progress(
    uploaded_file,
    progress_bar,
    progress_container,
    sheet_name: str,
    column_classification: dict,
    batch_size: int = DEFAULT_BATCH_SIZE,
    highlight_review_warnings: bool = False,
    max_concurrent_batches: int = DEFAULT_MAX_CONCURRENT,
    header_row: int = 1,
    enable_refinement: bool = True,
    enable_consistency: bool = True,
    enable_final_qa: bool = True,
    target_language: str = "French",
) -> tuple[BytesIO, dict]:
    client   = get_openai_client()
    tm       = load_translation_memory()
    glossary = load_glossary(target_language)

    token_counter      = {"prompt_tokens": 0, "completion_tokens": 0}
    glossary_run_stats: dict = {"total_hits": 0, "term_counts": {}}

    stats = {
        "cells_translated":    0,
        "cells_skipped":       0,
        "residue_corrections": 0,
        "unresolved_warnings": 0,
        "warning_details":     [],
        "sheet_name":          sheet_name,
        "tm_hits":             0,
        "tm_misses":           0,
        "batch_count":         0,
        "avg_batch_size":      0.0,
        "glossary_hits":       0,
        "glossary_top_terms":  {},
        "api_calls_made":      0,
        "api_calls_reduced":   0,
        "review_items":        [],
        "all_warnings":        [],
        "review_count":        0,
        "retry_count":         0,
        "failed_cells":        0,
        "failed_batches":      0,
        "avg_batch_duration":  0.0,
        "max_concurrent_used": max_concurrent_batches,
        "critical_warnings":      0,
        "high_warnings":          0,
        "medium_warnings":        0,
        "low_warnings":           0,
        "quality_score":          100,
        "warning_categories":     {},
        "refinement_enabled":     enable_refinement,
        "cells_refined":          0,
        "refinement_api_calls":   0,
        "refinement_prompt_tokens": 0,
        "refinement_completion_tokens": 0,
        # Intelligence Engine (populated during processing)
        "semantic_tm_hits":       0,
        "duplicate_groups":       0,
        "duplicate_cells_saved":  0,
        "glossary_only_count":    0,
        "pattern_count":          0,
        "gpt_calls_avoided":      0,
        "detected_product_type":  "generic",
        "_glossary_suggestions":  [],
        "auto_learned_terms":       0,
        "furniture_term_fixes":     0,
        # Consistency pass
        "consistency_corrections":  0,
        "consistency_detected":     0,
        "terms_harmonized":         0,
        # Final QA
        "qa_issues_found":          0,
        # Localization quality engine
        "forbidden_corrections":        0,
        "dutch_contamination_fixes":    0,
        "context_reconstructions":      0,
        "corpus_matches":               0,
        # Pipeline tracking
        "pipeline_refinement":      enable_refinement,
        "pipeline_consistency":     enable_consistency,
        "pipeline_final_qa":        enable_final_qa,
    }

    _forbidden_patterns: list = []

    # ── Enterprise pipeline: get session-level consistency memory ────────────
    _file_key_for_wcm = f"{uploaded_file.name}_{getattr(uploaded_file, 'size', 0)}"
    _consistency_mem  = _get_consistency_memory(_file_key_for_wcm)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name

    try:
        workbook     = load_workbook(filename=tmp_path, data_only=False)
        worksheet    = workbook[sheet_name]
        to_translate = column_classification["to_translate"]

        if not to_translate:
            raise ValueError("No translatable columns found in the file.")

        # ws.max_row can still be wrong in some openpyxl versions even without read_only.
        # Compute a reliable upper bound by taking max(ws.max_row, last row with data).
        _ws_max_row = worksheet.max_row or 1
        data_start_row = header_row + 1
        total_rows     = max(0, _ws_max_row - header_row)
        start_time     = time.time()

        # ── Enterprise pipeline: large file mode auto-detection ───────────────
        _lf_config     = detect_large_file_mode(total_rows)
        _large_mode    = _lf_config.active
        # User-supplied batch_size is honoured unless large file mode overrides
        _eff_batch     = min(batch_size, _lf_config.batch_size) if _large_mode else batch_size
        _eff_concurrent = min(max_concurrent_batches, _lf_config.max_concurrent) if _large_mode else max_concurrent_batches
        # Debug metrics tracker
        _dbg_metrics   = SheetDebugMetrics(sheet_name)
        _dbg_metrics.large_file_mode = _large_mode
        stats["large_file_mode"] = _large_mode
        stats["effective_batch_size"] = _eff_batch

        # ── Phase 0: Pre-scan ─────────────────────────────────────────────────
        progress_bar.progress(0.02)
        progress_container.markdown(
            _batch_progress_html(
                "Scanning", sheet_name, 0, 0, 0, 0, 0.0, 0.0, 0, 0, 2
            ),
            unsafe_allow_html=True,
        )

        cells_queue: list[tuple] = []
        for row_num in range(data_start_row, _ws_max_row + 1):
            for col_header, (col_idx, canonical) in to_translate.items():
                cell = worksheet.cell(row=row_num, column=col_idx)
                raw  = cell.value
                if raw is None or str(raw).strip() == "":
                    stats["cells_skipped"] += 1
                else:
                    cells_queue.append((row_num, col_header, col_idx, canonical, str(raw).strip()))

        total_to_process = len(cells_queue)
        if total_to_process == 0:
            raise ValueError("All translatable cells are empty — nothing to translate.")

        # ── Context Reconstruction: build per-row context from all visible columns ──
        row_contexts: dict[int, dict] = {}
        if target_language == "French":
            all_col_indices = list(range(1, worksheet.max_column + 1))
            for row_num in range(data_start_row, _ws_max_row + 1):
                row_data: dict[str, str] = {}
                for ci in all_col_indices:
                    v = worksheet.cell(row=row_num, column=ci).value
                    if v:
                        row_data[str(ci)] = str(v)
                if row_data:
                    ctx = build_row_context(row_data)
                    if ctx["product_type"] != "generic":
                        row_contexts[row_num] = ctx
                        stats["context_reconstructions"] += 1

        # ── Intelligence: Product Type Detection ─────────────────────────────
        name_samples = [text for _, _, _, can, text in cells_queue if can == "name"][:20]
        ctx_samples  = [text for _, _, _, can, text in cells_queue if can == "materialDetail"][:5]
        product_type = detect_product_type(name_samples, ctx_samples)
        product_hint = get_product_type_hint(product_type, target_language)

        # Enrich with home24.fr corpus style examples (French)
        if target_language == "French":
            product_hint += get_corpus_style_hint(product_type, target_language)
            stats["corpus_matches"] += 1 if product_hint.strip() else 0

        # Enrich Dutch prompt with Trados TM terminology
        _nl_corpus = None
        if target_language == "Dutch":
            _nl_corpus = _get_nl_corpus_engine()
            if _nl_corpus:
                _src_sample = " ".join(
                    text for _, _, _, _, text in cells_queue
                    if text and len(text) > 4
                )[:800]
                _tm_terms = _nl_corpus.extract_terminology(_src_sample, max_terms=15)
                if _tm_terms:
                    product_hint += "\nHome24 NL TM terminology — reuse these exact Dutch terms:\n"
                    product_hint += "\n".join(
                        f'  • "{de}" → "{nl}"' for de, nl in _tm_terms[:12]
                    )
                stats["corpus_matches"] += 1 if _tm_terms else 0

        # TIE stats tracking
        tie_stats = {
            "semantic_tm_hits":      0,
            "duplicate_groups":      0,
            "duplicate_cells_saved": 0,
            "glossary_only_count":   0,
            "pattern_count":         0,
            "gpt_calls_avoided":     0,
            "detected_product_type": product_type,
        }

        # ── Phase 1: Translation Memory check ────────────────────────────────
        results: dict[tuple, str] = {}
        api_queue: list[tuple]    = []

        for row_num, col_header, col_idx, canonical, text in cells_queue:
            col_type = _tm_col_type(canonical)
            cached   = tm_get(tm, text, col_type, target_language)
            if cached is not None:
                results[(row_num, col_idx)] = cached
                stats["tm_hits"] += 1
            else:
                api_queue.append((row_num, col_header, col_idx, canonical, text))
                stats["tm_misses"] += 1

        # ── Intelligence: Duplicate Detection ────────────────────────────────
        unique_queue, dup_restore_map, dup_groups, cells_saved = dedup_api_queue(api_queue)
        tie_stats["duplicate_groups"]      = dup_groups
        tie_stats["duplicate_cells_saved"] = cells_saved
        tie_stats["gpt_calls_avoided"]    += cells_saved

        # ── Intelligence: Trados TM + Glossary-Only + Pattern + Semantic TM ────
        final_api_queue: list[tuple] = []
        for item in unique_queue:
            row_num, col_header, col_idx, canonical, text = item
            col_type = _tm_col_type(canonical)
            resolved = False

            # ── Dutch: Trados TM exact match (highest priority) ───────────────
            if not resolved and target_language == "Dutch" and _nl_corpus:
                exact_tr = _nl_corpus.exact_match(text)
                if exact_tr is not None:
                    exact_tr = nl_post_process(exact_tr, canonical)
                    results[(row_num, col_idx)] = exact_tr
                    tm_put(tm, text, exact_tr, col_type, target_language)
                    tie_stats["gpt_calls_avoided"]   += 1
                    tie_stats.setdefault("trados_exact", 0)
                    tie_stats["trados_exact"]        += 1
                    resolved = True

            # ── Dutch: segment-level TM match (split cell → match parts) ─────
            if not resolved and target_language == "Dutch" and _nl_corpus:
                seg_tr = _nl_corpus.segment_tm_match(text)
                if seg_tr is not None:
                    results[(row_num, col_idx)] = seg_tr
                    tm_put(tm, text, seg_tr, col_type, target_language)
                    tie_stats["gpt_calls_avoided"]         += 1
                    tie_stats.setdefault("trados_segment", 0)
                    tie_stats["trados_segment"]            += 1
                    resolved = True

            # Glossary-only resolution
            if not resolved:
                gl_tr = try_glossary_only(text, glossary, target_language)
                if gl_tr is not None:
                    if target_language == "Dutch":
                        gl_tr = nl_post_process(gl_tr, canonical)
                    results[(row_num, col_idx)] = gl_tr
                    tm_put(tm, text, gl_tr, col_type, target_language)
                    tie_stats["glossary_only_count"] += 1
                    tie_stats["gpt_calls_avoided"]   += 1
                    resolved = True

            # Pattern / rule-based translation
            if not resolved:
                pat_tr = try_pattern_translation(text, glossary, target_language)
                if pat_tr is not None:
                    results[(row_num, col_idx)] = pat_tr
                    tm_put(tm, text, pat_tr, col_type, target_language)
                    tie_stats["pattern_count"]      += 1
                    tie_stats["gpt_calls_avoided"]  += 1
                    resolved = True

            # ── Dutch: Trados TM fuzzy match (after cheap resolutions) ────────
            if not resolved and target_language == "Dutch" and _nl_corpus:
                fuzzy_result = _nl_corpus.fuzzy_match(text, threshold=0.88)
                if fuzzy_result is not None:
                    fuzzy_tr, _fscore = fuzzy_result
                    fuzzy_tr = nl_post_process(fuzzy_tr, canonical)
                    results[(row_num, col_idx)] = fuzzy_tr
                    tm_put(tm, text, fuzzy_tr, col_type, target_language)
                    tie_stats["gpt_calls_avoided"]   += 1
                    tie_stats.setdefault("trados_fuzzy", 0)
                    tie_stats["trados_fuzzy"]        += 1
                    resolved = True

            # Semantic TM match (app-level TM)
            if not resolved:
                sem = semantic_tm_match(tm, text, col_type, target_language)
                if sem is not None:
                    sem_tr, _score = sem
                    if target_language == "Dutch":
                        sem_tr = nl_post_process(sem_tr, canonical)
                    results[(row_num, col_idx)] = sem_tr
                    tm_put(tm, text, sem_tr, col_type, target_language)
                    tie_stats["semantic_tm_hits"]   += 1
                    tie_stats["gpt_calls_avoided"]  += 1
                    resolved = True

            if not resolved:
                final_api_queue.append(item)

        # Track which cells went through API translation (for targeted residue check).
        # Populated now with representatives; duplicates are added after Phase 2.
        api_translated_cells: set[tuple] = {
            (row_num, col_idx) for row_num, _, col_idx, _, _ in final_api_queue
        }

        # ── Phase 2: Parallel batch translation ──────────────────────────────
        # Enterprise pipeline: use hierarchical (cluster-first) batching for large files.
        # Flat batching is used for small files to preserve legacy behaviour.
        total_api_cells = len(final_api_queue)

        batch_list, _cluster_count = build_clustered_batches(
            final_api_queue,
            batch_size=_eff_batch,
            cluster_first=_large_mode,
            max_cluster_size=_lf_config.cluster_size,
        )
        _dbg_metrics.cluster_count = _cluster_count
        stats["cluster_count"] = _cluster_count

        total_batches  = len(batch_list)
        retry_counter  = [0]
        retry_lock     = threading.Lock()
        batch_results  = []
        completed_batches = 0
        failed_batches    = 0
        api_cells_done    = 0
        batch_durations   = []

        with ThreadPoolExecutor(max_workers=max(1, _eff_concurrent)) as executor:
            future_map = {
                executor.submit(
                    _run_batch_task,
                    client, bid, batch_items, canonical, glossary,
                    retry_counter, retry_lock, target_language, product_hint,
                ): bid
                for bid, (batch_items, canonical) in enumerate(batch_list)
            }

            for future in as_completed(future_map):
                result = future.result()
                batch_results.append(result)
                completed_batches += 1
                api_cells_done    += len(result["batch_items"])
                if result["failed"]:
                    failed_batches += 1
                batch_durations.append(result["duration"])

                # Accumulate counters on main thread (thread-safe)
                token_counter["prompt_tokens"]     += result["prompt_tokens"]
                token_counter["completion_tokens"] += result["completion_tokens"]
                glossary_run_stats["total_hits"]   += result["glossary_hits"]
                for term, cnt in result["glossary_terms"].items():
                    glossary_run_stats["term_counts"][term] = (
                        glossary_run_stats["term_counts"].get(term, 0) + cnt
                    )

                # Progress update (main thread only)
                pct_api = int((api_cells_done / max(total_api_cells, 1)) * 100)
                elapsed = time.time() - start_time
                rate    = api_cells_done / max(elapsed, 0.1)
                eta     = (total_api_cells - api_cells_done) / max(rate, 0.1)
                in_flight = max(0, min(max_concurrent_batches, total_batches - completed_batches))
                progress_bar.progress(0.05 + (api_cells_done / max(total_api_cells, 1)) * 0.60)
                progress_container.markdown(
                    _parallel_progress_html(
                        sheet_name, total_batches, completed_batches, in_flight,
                        failed_batches, api_cells_done, total_api_cells,
                        stats["tm_hits"], elapsed, eta, pct_api,
                    ),
                    unsafe_allow_html=True,
                )

        # Apply results sequentially (main thread — no concurrent Excel writes)
        _consistency_fixes_batch = 0
        for result in sorted(batch_results, key=lambda r: r["batch_id"]):
            for i, (row_num, col_header, col_idx, canonical, text) in enumerate(result["batch_items"]):
                tr = str(result["translations"][i]).strip() if i < len(result["translations"]) else text
                if canonical == "name":
                    tr = validate_product_name(tr)
                # Enterprise pipeline: enforce workbook-level consistency
                tr_enforced = _consistency_mem.enforce(text, canonical, tr)
                if tr_enforced != tr:
                    _consistency_fixes_batch += 1
                    tr = tr_enforced
                results[(row_num, col_idx)] = tr
                tm_put(tm, text, tr, _tm_col_type(canonical), target_language)
                stats["cells_translated"] += 1

        stats["consistency_memory_fixes"] = _consistency_fixes_batch

        # Restore duplicate cells — representative cells are now in results
        for (dup_row, dup_col), (rep_row, rep_col) in dup_restore_map.items():
            if (rep_row, rep_col) in results:
                results[(dup_row, dup_col)] = results[(rep_row, rep_col)]
                # Include restored duplicates in residue tracking
                if (rep_row, rep_col) in api_translated_cells:
                    api_translated_cells.add((dup_row, dup_col))

        # Restore duplicate cells into the translated count
        stats["cells_translated"] += len(dup_restore_map)

        # Count TM + intelligence hits as translated (exact TM)
        stats["cells_translated"] += stats["tm_hits"]
        # Add intelligence-resolved cells to translated count
        stats["cells_translated"] += (
            tie_stats["glossary_only_count"]
            + tie_stats["pattern_count"]
            + tie_stats["semantic_tm_hits"]
        )

        stats["batch_count"]       = total_batches
        stats["failed_batches"]    = failed_batches
        stats["avg_batch_duration"] = round(
            sum(batch_durations) / max(len(batch_durations), 1), 2
        )
        stats["max_concurrent_used"] = _eff_concurrent
        stats["avg_batch_size"] = (
            round(total_api_cells / max(total_batches, 1), 1)
            if total_batches > 0 else 0.0
        )

        # Merge TIE stats
        stats.update(tie_stats)

        # Enterprise pipeline: populate per-sheet debug metrics
        _dbg_metrics.rows            = total_rows
        _dbg_metrics.cells_total     = total_to_process
        _dbg_metrics.tm_hits         = stats["tm_hits"]
        _dbg_metrics.trados_exact    = tie_stats.get("trados_exact", 0)
        _dbg_metrics.trados_fuzzy    = tie_stats.get("trados_fuzzy", 0)
        _dbg_metrics.glossary_only   = tie_stats["glossary_only_count"]
        _dbg_metrics.pattern         = tie_stats["pattern_count"]
        _dbg_metrics.semantic_tm     = tie_stats["semantic_tm_hits"]
        _dbg_metrics.ai_cells        = total_api_cells
        _dbg_metrics.failed_batches  = failed_batches
        _dbg_metrics.processing_time = time.time() - start_time
        stats["_sheet_debug_metrics"] = _dbg_metrics.to_dict()

        # API call accounting: include all intelligence-layer savings
        stats["api_calls_made"]    = total_batches
        total_avoided = (
            stats["tm_hits"]
            + tie_stats["gpt_calls_avoided"]
        )
        stats["api_calls_reduced"] = max(total_to_process - total_batches - total_avoided, 0)

        # Glossary suggestions — extract unknown terms from source texts for admin review
        _src_texts_for_suggestions = [text for _, _, _, _, text in cells_queue]
        suggestions = extract_glossary_suggestions(
            _src_texts_for_suggestions, glossary, target_language,
            min_occurrences=2, max_results=20,
        )
        stats["_glossary_suggestions"] = suggestions

        # Persist TM
        tm["global_stats"]["total_hits"]            += stats["tm_hits"]
        tm["global_stats"]["total_misses"]          += stats["tm_misses"]
        tm["global_stats"]["total_api_calls_saved"] += stats["tm_hits"] + tie_stats["semantic_tm_hits"]
        save_translation_memory(tm)

        # Persist Glossary stats
        stats["glossary_hits"]      = glossary_run_stats.get("total_hits", 0)
        stats["glossary_top_terms"] = dict(
            sorted(
                glossary_run_stats.get("term_counts", {}).items(),
                key=lambda x: -x[1],
            )[:5]
        )
        update_glossary_stats(glossary, glossary_run_stats.get("term_counts", {}))
        save_glossary(glossary, target_language)

        # Build source lookup for quality analysis
        source_lookup: dict[tuple, tuple] = {
            (row_num, col_idx): (text, canonical)
            for row_num, col_header, col_idx, canonical, text in cells_queue
        }

        # ── Phase 2.5: Premium AI Refinement ─────────────────────────────────
        refine_token_c: dict = {"prompt_tokens": 0, "completion_tokens": 0}
        cells_refined  = 0
        refine_api_calls = 0

        if enable_refinement:
            refine_queue: list[tuple] = []
            for key, translation in results.items():
                _, src_canonical = source_lookup.get(key, ("", "other"))
                if _should_refine(translation, src_canonical):
                    refine_queue.append((key, translation, src_canonical))

            total_refine = len(refine_queue)
            if refine_queue:
                progress_bar.progress(0.65)
                progress_container.markdown(
                    _refinement_progress_html(sheet_name, 0, total_refine),
                    unsafe_allow_html=True,
                )

                for batch_start in range(0, total_refine, REFINEMENT_BATCH_SIZE):
                    batch = refine_queue[batch_start : batch_start + REFINEMENT_BATCH_SIZE]
                    items = [(text, canonical) for _, text, canonical in batch]
                    refined_texts = refine_batch(client, items, refine_token_c, target_language)
                    refine_api_calls += 1

                    for i, (key, orig_text, canonical) in enumerate(batch):
                        if i >= len(refined_texts):
                            continue
                        new_text = refined_texts[i].strip()
                        if not new_text or new_text == orig_text:
                            continue

                        # Safety: <br> count must be exactly preserved
                        if orig_text.count("<br>") != new_text.count("<br>"):
                            continue

                        # Safety: name column — never allow refinement to lengthen it
                        if canonical == "name":
                            new_text = validate_product_name(new_text)
                            if len(new_text) > max(len(orig_text), 40):
                                continue

                        results[key] = new_text
                        cells_refined += 1

                    done_refine = min(batch_start + len(batch), total_refine)
                    progress_bar.progress(
                        0.65 + (done_refine / max(total_refine, 1)) * 0.02
                    )
                    progress_container.markdown(
                        _refinement_progress_html(sheet_name, done_refine, total_refine),
                        unsafe_allow_html=True,
                    )

            # Merge refinement tokens into main counter
            token_counter["prompt_tokens"]     += refine_token_c["prompt_tokens"]
            token_counter["completion_tokens"] += refine_token_c["completion_tokens"]

        stats["cells_refined"]                  = cells_refined
        stats["refinement_api_calls"]           = refine_api_calls
        stats["refinement_prompt_tokens"]       = refine_token_c["prompt_tokens"]
        stats["refinement_completion_tokens"]   = refine_token_c["completion_tokens"]

        # ── Phase 2.6: Local furniture term pass (fast, no API) ───────────────
        furniture_fixes = 0
        for key in list(results.keys()):
            original_tr = results[key]
            fixed_tr = apply_furniture_terms(original_tr, target_language)
            if fixed_tr != original_tr:
                results[key] = fixed_tr
                furniture_fixes += 1
        stats["furniture_term_fixes"] = furniture_fixes

        # ── Auto-learn glossary from source texts ─────────────────────────────
        all_source_texts = [text for _, _, _, _, text in cells_queue]
        learnable = auto_learn_glossary_from_source(
            all_source_texts, glossary, target_language, min_occurrences=2
        )
        if learnable:
            for lt in learnable:
                glossary["terms"].setdefault(lt["source_term"], lt["target_term"])
            save_glossary(glossary, target_language)
            stats["auto_learned_terms"] = len(learnable)

        # ── Phase 2.7: Terminology Consistency Pass (fast, no API) ────────────
        if enable_consistency:
            progress_bar.progress(0.68)
            progress_container.markdown(
                '<div class="prog-shell"><div class="prog-head"><div>'
                '<div class="prog-phase">Phase 2.7 — Terminology Consistency</div>'
                f'<div class="prog-sheet">Sheet: {sheet_name} · Harmonizing recurring terms</div>'
                '</div><span class="prog-badge"><span class="prog-badge-dot"></span>CHECKING</span>'
                '</div></div>',
                unsafe_allow_html=True,
            )
            con_result = run_local_consistency_pass(results, source_lookup, glossary, target_language)
            stats["consistency_corrections"] = con_result["corrections"]
            stats["consistency_detected"]    = con_result["detected"]
            stats["terms_harmonized"]        = con_result["harmonized"]

        # ── Signal separation ─────────────────────────────────────────────────
        # Three independent signals — only the first two may affect Excel color:
        #   original_excel_highlights : source cell fills → always preserved
        #   review_warnings           : quality issues  → REVIEW_FILL only when
        #                               highlight_review_warnings=True (opt-in checkbox)
        #   glossary_hits             : analytics only  → NEVER affect cell color

        # Capture original fills from the source workbook before writing anything.
        original_excel_highlights: dict[tuple, object] = {}
        for (rn, ci) in results:
            src_cell = worksheet.cell(row=rn, column=ci)
            if src_cell.fill and getattr(src_cell.fill, "fill_type", None) not in (None, "none"):
                try:
                    original_excel_highlights[(rn, ci)] = copy.copy(src_cell.fill)
                except Exception:
                    pass

        all_warnings: list[dict] = []
        review_items: list[dict] = []
        cells_original_highlight = 0
        cells_review_highlighted  = 0

        col_header_map = {ci: h for h, (ci, _) in to_translate.items()}

        for (row_num, col_idx), translation in results.items():
            cell = worksheet.cell(row=row_num, column=col_idx)
            src_text, src_canonical = source_lookup.get((row_num, col_idx), (translation, "other"))

            # Source-aware semicolon fix: only replace " ; " with <br> when the
            # *source* text used <br> as its separator — meaning the LLM swapped
            # them.  If the source already had semicolons, keep them intact.
            if (
                src_canonical in _SEMICOLON_BR_CANONICALS
                and "<br>" in src_text.lower()
                and ";" in translation
            ):
                translation = _SEMICOLON_SEP_RE.sub("<br>", translation)

            cell.value = translation
            has_original = (row_num, col_idx) in original_excel_highlights
            col_header   = col_header_map.get(col_idx, str(col_idx))

            # Quality analysis — glossary_hits tracked separately, never affect cell color.
            issues = analyze_translation_quality(src_text, translation, src_canonical, glossary)

            ts = datetime.now().isoformat(timespec="seconds")
            has_critical_or_high = False

            for issue in issues:
                if issue["type"] == "german_residue":
                    continue  # deferred to post-fix warning_details for accuracy
                sev = issue.get("severity", SEVERITY_LOW)
                if sev in (SEVERITY_CRITICAL, SEVERITY_HIGH):
                    has_critical_or_high = True
                src_snippet = src_text[:120] + "…" if len(src_text) > 120 else src_text
                tr_snippet  = translation[:120] + "…" if len(translation) > 120 else translation
                all_warnings.append({
                    "severity":        sev,
                    "category":        issue.get("category", "Manual review recommended"),
                    "row":             row_num,
                    "column":          col_header,
                    "original_text":   src_snippet,
                    "translated_text": tr_snippet,
                    "reason":          issue.get("reason", ""),
                    "suggested_fix":   issue.get("suggested_fix", ""),
                    "timestamp":       ts,
                })

            if issues:
                review_items.append({
                    "row":         row_num,
                    "col_idx":     col_idx,
                    "translation": translation[:60] + "…" if len(translation) > 60 else translation,
                    "issues":      issues,
                })
                # Highlight only Critical/High when checkbox is enabled
                if highlight_review_warnings and has_critical_or_high:
                    cell.fill = REVIEW_FILL
                    cells_review_highlighted += 1
                elif has_original:
                    cell.fill = original_excel_highlights[(row_num, col_idx)]
                    cells_original_highlight += 1
            elif has_original:
                cell.fill = original_excel_highlights[(row_num, col_idx)]
                cells_original_highlight += 1

        # Load forbidden patterns for this language (fast DB read, cached in pipeline)
        _forbidden_lang = "FR" if target_language == "French" else "NL"
        _forbidden_patterns = db_load_forbidden_patterns(_forbidden_lang) if target_language == "French" else []

        # ── Phase 3: Residue check (fast multi-layer QA first, AI only for flagged) ──
        # Enterprise pipeline: pre-filter with qa_cell_needs_ai_fix (L1 + L2, no API).
        # Only cells flagged by L1/L2 proceed to the full local-fix + AI loop.
        # This avoids iterating thousands of clean cells on large files.
        residue_candidates_raw = {
            (rn, ci)
            for rn, ci in results
            if (rn, ci) in api_translated_cells
        }
        # Pre-filter: skip cells that are already clean (fast, no API)
        if _large_mode:
            residue_candidates = {
                (rn, ci)
                for rn, ci in residue_candidates_raw
                if qa_cell_needs_ai_fix(
                    str(results.get((rn, ci), "")),
                    _forbidden_patterns,
                )[0]
            }
        else:
            residue_candidates = residue_candidates_raw
        total_residue_cells = max(len(residue_candidates), 1)
        checked = 0

        col_canonical_map = {ci: can for _, (ci, can) in to_translate.items()}
        col_header_for_ci = {ci: h for h, (ci, _) in to_translate.items()}

        for (row_num, col_idx) in residue_candidates:
            checked += 1
            canonical  = col_canonical_map.get(col_idx, "other")
            col_header = col_header_for_ci.get(col_idx, str(col_idx))
            elapsed    = time.time() - start_time
            progress   = 0.70 + (checked / total_residue_cells) * 0.20
            pct        = int(progress * 100)

            progress_bar.progress(progress)
            progress_container.markdown(
                _progress_html(
                    "Phase 3 — Residue Check", sheet_name, col_header,
                    checked, total_residue_cells, pct, elapsed, 0,
                    stats["cells_translated"], stats["cells_skipped"],
                    stats["residue_corrections"],
                ),
                unsafe_allow_html=True,
            )

            cell = worksheet.cell(row=row_num, column=col_idx)
            if cell.value is None or str(cell.value).strip() == "":
                continue

            text = str(cell.value)

            # Step 1: Fast local furniture term replacement (no API)
            locally_fixed = apply_furniture_terms(text, target_language)
            if locally_fixed != text:
                text = locally_fixed
                cell.value = text
                stats["residue_corrections"] += 1

            # Step 1b: French semantic normalization (fixes literal German-pattern output)
            if target_language == "French":
                normalized = apply_french_semantic_normalization(text)
                if normalized != text:
                    text = normalized
                    cell.value = text

            # Step 1c: Context-aware terminology (uses row context if available)
            if target_language == "French" and row_num in row_contexts:
                ctx_fixed = apply_context_terminology_fr(text, row_contexts[row_num])
                if ctx_fixed != text:
                    text = ctx_fixed
                    cell.value = text

            # Step 1d: Forbidden pattern corrections (DB-backed)
            if target_language == "French" and _forbidden_patterns:
                fp_fixed, fp_count = apply_forbidden_patterns(text, _forbidden_patterns)
                if fp_count > 0:
                    text = fp_fixed
                    cell.value = text
                    stats["forbidden_corrections"] += fp_count

            # Step 1e: Dutch-in-French contamination fix (language isolation guard)
            if target_language == "French":
                df_fixed, df_count = apply_dutch_to_french_fixes(text)
                if df_count > 0:
                    text = df_fixed
                    cell.value = text
                    stats["dutch_contamination_fixes"] += df_count

            # Step 2: Quick residue scan — if clean, done
            detected = detect_german_residue(text, target_language)
            if not detected:
                continue

            # Step 3: AI fix for persistent residue (max 2 attempts, stronger prompt)
            for attempt in range(2):
                text = fix_german_residue(client, text, canonical, token_counter, target_language)
                stats["residue_corrections"] += 1
                detected = detect_german_residue(text, target_language)
                if not detected:
                    break

            if detected:
                stats["unresolved_warnings"] += 1
                stats["warning_details"].append({
                    "row":     row_num,
                    "column":  col_header,
                    "text":    text[:50] + "..." if len(text) > 50 else text,
                    "residue": detected[:3],
                })

            cell.value = text

        # ── Phase 4: Final verification pass ─────────────────────────────────
        progress_bar.progress(0.91)
        elapsed = time.time() - start_time
        progress_container.markdown(
            _progress_html(
                "Phase 4 — Final Verification", sheet_name, "all columns",
                total_rows, total_rows, 91, elapsed, 0,
                stats["cells_translated"], stats["cells_skipped"],
                stats["residue_corrections"],
            ),
            unsafe_allow_html=True,
        )

        for (row_num, col_idx) in residue_candidates:
            canonical  = col_canonical_map.get(col_idx, "other")
            col_header = col_header_for_ci.get(col_idx, str(col_idx))
            cell       = worksheet.cell(row=row_num, column=col_idx)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            text     = str(cell.value)
            detected = detect_german_residue(text, target_language)
            if detected:
                corrected = fix_german_residue(client, text, canonical, token_counter, target_language)
                stats["residue_corrections"] += 1
                if detect_german_residue(corrected, target_language):
                    already = any(
                        w["row"] == row_num and w["column"] == col_header
                        for w in stats["warning_details"]
                    )
                    if not already:
                        stats["unresolved_warnings"] += 1
                        stats["warning_details"].append({
                            "row":     row_num,
                            "column":  col_header,
                            "text":    corrected[:50] + "..." if len(corrected) > 50 else corrected,
                            "residue": detect_german_residue(corrected)[:3],
                        })
                cell.value = corrected

        # ── Merge confirmed unresolved residue into all_warnings ─────────────────
        ts_fin = datetime.now().isoformat(timespec="seconds")
        for wd in stats["warning_details"]:
            all_warnings.append({
                "severity":        SEVERITY_CRITICAL,
                "category":        "German residue",
                "row":             wd["row"],
                "column":          wd["column"],
                "original_text":   "",
                "translated_text": wd["text"],
                "reason":          f"Unresolved German words after 3 fix attempts: {', '.join(wd.get('residue', []))}",
                "suggested_fix":   "Retranslate manually — automated residue fixing failed",
                "timestamp":       ts_fin,
            })

        _crit = sum(1 for w in all_warnings if w["severity"] == SEVERITY_CRITICAL)
        _high = sum(1 for w in all_warnings if w["severity"] == SEVERITY_HIGH)
        _med  = sum(1 for w in all_warnings if w["severity"] == SEVERITY_MEDIUM)
        _low  = sum(1 for w in all_warnings if w["severity"] == SEVERITY_LOW)
        from collections import Counter
        _cat_counts = Counter(w["category"] for w in all_warnings)

        stats["all_warnings"]       = all_warnings
        stats["critical_warnings"]  = _crit
        stats["high_warnings"]      = _high
        stats["medium_warnings"]    = _med
        stats["low_warnings"]       = _low
        stats["quality_score"]      = compute_quality_score(all_warnings)
        stats["warning_categories"] = dict(_cat_counts)
        stats["review_count"]       = len(all_warnings)

        # ── Final capitalisation / formatting pass ────────────────────────────
        if target_language == "French":
            for row_num in range(data_start_row, worksheet.max_row + 1):
                for col_header, (col_idx, canonical) in to_translate.items():
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    if cell.value and str(cell.value).strip():
                        val = str(cell.value)
                        if _forbidden_patterns:
                            val, _fc = apply_forbidden_patterns(val, _forbidden_patterns)
                            if _fc > 0:
                                stats["forbidden_corrections"] += _fc
                        val = apply_french_semantic_normalization(val)
                        val = apply_french_capitalization_rules(val, canonical, glossary)
                        val = apply_french_typography_rules(val)
                        # Final Dutch-contamination sweep in formatting pass
                        val, _dc = apply_dutch_to_french_fixes(val)
                        if _dc > 0:
                            stats["dutch_contamination_fixes"] += _dc
                        cell.value = val
        elif target_language == "Dutch":
            for row_num in range(data_start_row, worksheet.max_row + 1):
                for col_header, (col_idx, canonical) in to_translate.items():
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    if cell.value and str(cell.value).strip():
                        cell.value = apply_dutch_capitalization_rules(
                            str(cell.value), canonical, glossary
                        )

        # ── Phase 5: Final QA — full-file local scan ──────────────────────────
        if enable_final_qa:
            progress_bar.progress(0.96)
            progress_container.markdown(
                '<div class="prog-shell"><div class="prog-head"><div>'
                '<div class="prog-phase">Phase 5 — Final QA</div>'
                f'<div class="prog-sheet">Sheet: {sheet_name} · Scanning all cells</div>'
                '</div><span class="prog-badge"><span class="prog-badge-dot"></span>QA</span>'
                '</div></div>',
                unsafe_allow_html=True,
            )
            qa_issues = 0
            ts_qa = datetime.now().isoformat(timespec="seconds")
            for row_num in range(data_start_row, worksheet.max_row + 1):
                for col_header, (col_idx, canonical) in to_translate.items():
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    if cell.value is None or str(cell.value).strip() == "":
                        # Empty cell that should have content (flag if source was non-empty)
                        src_text, _ = source_lookup.get((row_num, col_idx), ("", "other"))
                        if src_text and src_text.strip():
                            qa_issues += 1
                            all_warnings.append({
                                "severity":        SEVERITY_HIGH,
                                "category":        "Empty translation",
                                "row":             row_num,
                                "column":          col_header,
                                "original_text":   src_text[:120],
                                "translated_text": "",
                                "reason":          "Cell is empty after translation — source had content",
                                "suggested_fix":   "Translate manually",
                                "timestamp":       ts_qa,
                            })
                        continue
                    text = str(cell.value)

                    # Dutch-contamination auto-fix pass (French only) — catches anything
                    # that slipped through Phase 3 (e.g. source already contained Dutch)
                    if target_language == "French":
                        df_fixed, df_count = apply_dutch_to_french_fixes(text)
                        if df_count > 0:
                            text = df_fixed
                            cell.value = text
                            stats["dutch_contamination_fixes"] += df_count

                    # Check for residue that slipped past Phases 3 & 4
                    residue = detect_german_residue(text, target_language)
                    if residue:
                        already = any(
                            w["row"] == row_num and w["column"] == col_header
                            and w["category"] == "German residue"
                            for w in all_warnings
                        )
                        if not already:
                            qa_issues += 1
                            all_warnings.append({
                                "severity":        SEVERITY_CRITICAL,
                                "category":        "German residue",
                                "row":             row_num,
                                "column":          col_header,
                                "original_text":   "",
                                "translated_text": text[:120],
                                "reason":          f"German residue detected in Final QA: {', '.join(residue[:3])}",
                                "suggested_fix":   "Retranslate manually",
                                "timestamp":       ts_qa,
                            })

                    # Dutch-in-French contamination warning (for any Dutch that survived)
                    if target_language == "French":
                        dutch_words = detect_dutch_in_french(text)
                        if dutch_words:
                            qa_issues += 1
                            all_warnings.append({
                                "severity":        SEVERITY_HIGH,
                                "category":        "Dutch contamination",
                                "row":             row_num,
                                "column":          col_header,
                                "original_text":   "",
                                "translated_text": text[:120],
                                "reason":          f"Dutch words in French output: {', '.join(dutch_words[:3])}",
                                "suggested_fix":   "Check source file — may contain Dutch variant names",
                                "timestamp":       ts_qa,
                            })
            stats["qa_issues_found"] = qa_issues

            # Recompute warning totals with QA additions
            _crit = sum(1 for w in all_warnings if w["severity"] == SEVERITY_CRITICAL)
            _high = sum(1 for w in all_warnings if w["severity"] == SEVERITY_HIGH)
            _med  = sum(1 for w in all_warnings if w["severity"] == SEVERITY_MEDIUM)
            _low  = sum(1 for w in all_warnings if w["severity"] == SEVERITY_LOW)
            from collections import Counter as _Counter
            _cat_counts = _Counter(w["category"] for w in all_warnings)
            stats["all_warnings"]       = all_warnings
            stats["critical_warnings"]  = _crit
            stats["high_warnings"]      = _high
            stats["medium_warnings"]    = _med
            stats["low_warnings"]       = _low
            stats["quality_score"]      = compute_quality_score(all_warnings)
            stats["warning_categories"] = dict(_cat_counts)
            stats["review_count"]       = len(all_warnings)

        progress_bar.progress(1.0)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        workbook.close()

        elapsed_seconds = time.time() - start_time
        stats["total_time"]        = format_time(elapsed_seconds)
        stats["elapsed_seconds"]   = elapsed_seconds
        stats["prompt_tokens"]     = token_counter["prompt_tokens"]
        stats["completion_tokens"] = token_counter["completion_tokens"]

        total_tokens = token_counter["prompt_tokens"] + token_counter["completion_tokens"]
        stats["estimated_cost_usd"] = (
            round(
                token_counter["prompt_tokens"]     * _INPUT_COST_PER_TOKEN +
                token_counter["completion_tokens"] * _OUTPUT_COST_PER_TOKEN,
                4,
            ) if total_tokens > 0 else None
        )

        stats["quality_gate"] = {
            "no_residue":         stats["unresolved_warnings"] == 0,
            "protected_columns":  list(column_classification["protected"].keys()),
            "possible_missed":    column_classification["possible_missed"],
            "translated_columns": list(to_translate.keys()),
        }

        stats["review_items"] = review_items  # per-cell legacy list
        stats["retry_count"]  = retry_counter[0]

        # Audit counters for the export highlighting report
        stats["original_highlights_preserved"] = cells_original_highlight
        stats["review_highlights_applied"]      = cells_review_highlighted
        stats["highlight_review_warnings"]      = highlight_review_warnings

        return output, stats

    finally:
        os.unlink(tmp_path)


# =============================================================================
# REVIEW DASHBOARD
# =============================================================================

def render_review_dashboard(all_warnings: list, stats: dict, highlight_in_excel: bool):
    if not all_warnings:
        st.markdown("""
        <div class="alert alert-success">
            <span class="alert-icon">✓</span>
            <span><strong>No warnings detected.</strong> Translation passed all quality checks.</span>
        </div>
        """, unsafe_allow_html=True)
        return

    excel_note = (
        "Critical and High warnings are highlighted yellow in the downloaded Excel."
        if highlight_in_excel
        else "Warnings shown here only — Excel highlighting is off (checkbox unchecked)."
    )
    st.markdown(f"""
    <div class="alert alert-info">
        <span class="alert-icon">ℹ</span>
        <span>{excel_note} Glossary hits never create highlights.</span>
    </div>
    """, unsafe_allow_html=True)

    with st.expander(f"Warning details — {len(all_warnings)} warning(s)", expanded=True):
        c1, c2, c3, c4 = st.columns([1.4, 1.4, 2, 2])
        with c1:
            sev_filter = st.selectbox(
                "Severity", ["All"] + SEVERITY_ORDER, key="wd_sev"
            )
        with c2:
            col_opts   = ["All"] + sorted(set(w["column"] for w in all_warnings))
            col_filter = st.selectbox("Column", col_opts, key="wd_col")
        with c3:
            cat_opts   = ["All"] + sorted(set(w["category"] for w in all_warnings))
            cat_filter = st.selectbox("Category", cat_opts, key="wd_cat")
        with c4:
            search = st.text_input("Search in reason / text", placeholder="e.g. Bezug", key="wd_search")

        filtered = all_warnings
        if sev_filter != "All":
            filtered = [w for w in filtered if w["severity"] == sev_filter]
        if col_filter != "All":
            filtered = [w for w in filtered if w["column"] == col_filter]
        if cat_filter != "All":
            filtered = [w for w in filtered if w["category"] == cat_filter]
        if search:
            sl = search.lower()
            filtered = [
                w for w in filtered
                if sl in w.get("reason", "").lower()
                or sl in w.get("original_text", "").lower()
                or sl in w.get("translated_text", "").lower()
            ]

        st.caption(f"{len(filtered)} of {len(all_warnings)} warning(s) shown")

        if filtered:
            rows = [
                {
                    "Severity":    w["severity"],
                    "Category":    w["category"],
                    "Row":         w["row"],
                    "Column":      w["column"],
                    "Reason":      w["reason"],
                    "Original":    w.get("original_text", "")[:70],
                    "Translation": w.get("translated_text", "")[:70],
                    "Suggested Fix": w.get("suggested_fix", ""),
                }
                for w in filtered
            ]
            st.dataframe(rows, use_container_width=True, hide_index=True)

            csv_buf = StringIO()
            writer  = csv.DictWriter(csv_buf, fieldnames=[
                "severity", "category", "row", "column",
                "original_text", "translated_text", "reason", "suggested_fix",
            ])
            writer.writeheader()
            for w in filtered:
                writer.writerow({k: w.get(k, "") for k in writer.fieldnames})
            st.download_button(
                label="↓ Download warnings as CSV",
                data=csv_buf.getvalue().encode("utf-8"),
                file_name="translation_warnings.csv",
                mime="text/csv",
                use_container_width=True,
            )


# =============================================================================
# PAGE: LOGIN
# =============================================================================

def login_page():
    col_left, col_right = st.columns([9, 11], gap="small")

    with col_left:
        st.markdown("""
        <div style="
            background: linear-gradient(160deg, #0F3D9E 0%, #1A56D6 100%);
            min-height: 92vh;
            padding: 60px 52px;
            display: flex;
            flex-direction: column;
            margin: -2.5rem -1rem -4rem -3rem;
            position: relative;
            overflow: hidden;
        ">
            <div style="position:absolute;top:-110px;right:-110px;width:340px;height:340px;
                        border-radius:50%;background:rgba(255,255,255,0.06);pointer-events:none;"></div>
            <div style="position:absolute;bottom:-80px;left:-70px;width:240px;height:240px;
                        border-radius:50%;background:rgba(255,255,255,0.05);pointer-events:none;"></div>
            <div style="display:flex;align-items:center;gap:10px;font-size:15px;font-weight:700;
                        color:rgba(255,255,255,0.95);letter-spacing:-0.01em;position:relative;z-index:1;">
                <div style="width:9px;height:9px;border-radius:50%;background:#12A150;
                            box-shadow:0 0 0 3px rgba(18,161,80,0.28);flex-shrink:0;"></div>
                Home24 Localization
            </div>
            <div style="flex:1;display:flex;flex-direction:column;justify-content:center;
                        padding:52px 0;position:relative;z-index:1;">
                <div style="font-size:10.5px;font-weight:700;text-transform:uppercase;
                            letter-spacing:0.14em;color:rgba(255,255,255,0.45);margin-bottom:22px;">
                    Enterprise · Internal Platform
                </div>
                <h1 style="font-size:36px;font-weight:800;color:#ffffff;letter-spacing:-0.04em;
                           line-height:1.15;margin:0 0 20px;">
                    Translate product data at scale.
                </h1>
                <p style="font-size:15px;color:rgba(255,255,255,0.68);line-height:1.72;
                          font-weight:400;margin:0 0 44px;max-width:340px;">
                    AI-powered German to French and Dutch localization — built for Home24's e-commerce catalog.
                </p>
                <ul style="list-style:none;padding:0;margin:0;">
                    <li style="display:flex;align-items:center;gap:13px;font-size:13.5px;
                               color:rgba(255,255,255,0.78);margin-bottom:16px;font-weight:500;">
                        <span style="width:7px;height:7px;border-radius:50%;background:#12A150;
                                     flex-shrink:0;box-shadow:0 0 0 2px rgba(18,161,80,0.3);"></span>
                        Translation Memory with semantic matching
                    </li>
                    <li style="display:flex;align-items:center;gap:13px;font-size:13.5px;
                               color:rgba(255,255,255,0.78);margin-bottom:16px;font-weight:500;">
                        <span style="width:7px;height:7px;border-radius:50%;background:#12A150;
                                     flex-shrink:0;box-shadow:0 0 0 2px rgba(18,161,80,0.3);"></span>
                        Verified furniture &amp; material terminology
                    </li>
                    <li style="display:flex;align-items:center;gap:13px;font-size:13.5px;
                               color:rgba(255,255,255,0.78);margin-bottom:16px;font-weight:500;">
                        <span style="width:7px;height:7px;border-radius:50%;background:#12A150;
                                     flex-shrink:0;box-shadow:0 0 0 2px rgba(18,161,80,0.3);"></span>
                        Quality gate with automatic residue detection
                    </li>
                    <li style="display:flex;align-items:center;gap:13px;font-size:13.5px;
                               color:rgba(255,255,255,0.78);font-weight:500;">
                        <span style="width:7px;height:7px;border-radius:50%;background:#12A150;
                                     flex-shrink:0;box-shadow:0 0 0 2px rgba(18,161,80,0.3);"></span>
                        Batch processing with Excel export
                    </li>
                </ul>
            </div>
            <div style="font-size:11px;color:rgba(255,255,255,0.30);font-weight:500;
                        letter-spacing:0.04em;position:relative;z-index:1;">
                Home24 · Internal use only · v6.0
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_right:
        st.markdown("<div style='height:60px'></div>", unsafe_allow_html=True)
        _, form_col, _ = st.columns([1, 5, 1])
        with form_col:
            with st.form("login_form"):
                st.markdown("""
                <div style="margin-bottom:32px;">
                    <div class="login-form-title">Welcome back</div>
                    <div class="login-form-sub">Sign in to your localization workspace</div>
                </div>
                """, unsafe_allow_html=True)

                email_input    = st.text_input("Email address", placeholder="you@home24.de")
                password_input = st.text_input("Password", type="password", placeholder="••••••••")
                submitted      = st.form_submit_button("Sign in", use_container_width=True)

            st.markdown(
                '<p style="text-align:center;font-size:0.75rem;color:#9BA8BE;margin-top:12px;">'
                'Demo guest access available</p>',
                unsafe_allow_html=True,
            )
            st.markdown('<p class="login-footer">Home24 · Internal use only</p>', unsafe_allow_html=True)

            if submitted:
                _, stored_pw = _get_admin_credentials()
                guest_email, guest_pw = _get_guest_credentials()
                if not stored_pw and not guest_pw:
                    st.error("Credentials not configured. Check Streamlit secrets or your .env file.")
                else:
                    role = verify_credentials(email_input, password_input)
                    if role:
                        session_id = str(uuid.uuid4())
                        st.session_state["authenticated"]    = True
                        st.session_state["user_role"]        = role
                        st.session_state["user_email"]       = email_input.strip().lower()
                        st.session_state["session_id"]       = session_id
                        st.session_state["language_selected"] = False
                        db_log_login(email_input.strip().lower(), role, session_id)
                        st.rerun()
                    else:
                        st.error("Invalid email or password.")


# =============================================================================
# PAGE: TRANSLATOR
# =============================================================================

def translator_page():
    target_language = st.session_state.get("target_language", "French")
    lang_code       = "NL" if target_language == "Dutch" else "FR"
    lang_label      = "Dutch (NL)" if target_language == "Dutch" else "French (FR)"

    st.markdown(f"""
    <div class="tr-hero">
        <div class="tr-hero-tag">DE → {lang_code} · AI-Powered</div>
        <h1 class="tr-hero-title">AI Localization Platform</h1>
        <p class="tr-hero-sub">
            Translate German product data into {lang_label.split(" ")[0]} with verified
            e-commerce terminology, translation memory, and quality assurance.
        </p>
    </div>
    """, unsafe_allow_html=True)

    if not _get_api_key():
        st.markdown("""
        <div class="alert alert-warn">
            <span class="alert-icon">⚠</span>
            <div>
                <strong>OpenAI API key not configured.</strong><br>
                Add <code>OPENAI_API_KEY</code> to your <code>.env</code> file or Streamlit secrets.
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    # Jira-sourced file takes priority over manual upload
    _jira_src = st.session_state.get("_jira_upload")
    if _jira_src is not None:
        _jk_banner = st.session_state.get("_jira_ticket_key", "")
        _jk_summ   = st.session_state.get("_jira_ticket_summary", "")
        st.markdown(
            f'<div class="alert alert-info">'
            f'<span class="alert-icon">⬇</span>'
            f'<span>File from Jira <strong>{_jk_banner}</strong>'
            f'{": " + _jk_summ if _jk_summ else ""}'
            f' — <strong>{_jira_src.name}</strong></span>'
            f'</div>',
            unsafe_allow_html=True,
        )
        uploaded_file = _jira_src
        if st.button("Use a different file instead", key="jira_clear_src_btn"):
            for _jk in ["_jira_upload", "_jira_ticket_key", "_jira_ticket_summary",
                        "_jira_attachment_id", "_jira_attachment_filename"]:
                st.session_state.pop(_jk, None)
            st.rerun()
    else:
        uploaded_file = st.file_uploader(
            "Upload your German Excel file",
            type=["xlsx"],
            label_visibility="visible",
        )

    if uploaded_file is not None:
        # Clear stale results when the user switches to a different file
        if st.session_state.get("_tr_result_file") != uploaded_file.name:
            st.session_state.pop("_tr_result", None)
            for _stale_key in [k for k in st.session_state if k.startswith("_sheet_")]:
                del st.session_state[_stale_key]
            st.session_state["_tr_result_file"] = uploaded_file.name

        st.markdown(f'<div class="file-chip">📄 {uploaded_file.name}</div>', unsafe_allow_html=True)

        output_filename = f"{lang_code}-{uploaded_file.name}"
        _file_key       = f"{uploaded_file.name}_{uploaded_file.size}"

        # ── Score all sheets once per file, cache in session_state ────────────
        _scores_key = f"_sheet_scores_{_file_key}"
        if _scores_key not in st.session_state:
            _wb_score = load_workbook(BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
            st.session_state[_scores_key] = score_all_sheets(_wb_score)
            _wb_score.close()
        all_sheet_scores: dict[str, dict] = st.session_state[_scores_key]
        available_sheets = list(all_sheet_scores.keys())
        best_sheet       = pick_best_sheet(all_sheet_scores)

        # Confidence: best sheet must beat the next-best by at least 5 pts
        _other_scores   = [v["selection_score"] for k, v in all_sheet_scores.items() if k != best_sheet]
        _second_best    = max(_other_scores, default=-1000)
        _is_confident   = all_sheet_scores[best_sheet]["selection_score"] >= _second_best + 5

        # ── Sheet selector — always visible when >1 sheet, pre-set to best ───
        if len(available_sheets) == 1:
            selected_sheet = available_sheets[0]
            st.markdown(
                f'<div class="alert alert-info">'
                f'<span class="alert-icon">ℹ</span>'
                f'<span>Sheet: <strong>{selected_sheet}</strong></span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        else:
            _best_idx = available_sheets.index(best_sheet) if best_sheet in available_sheets else 0
            _select_label = (
                "Sheet to translate (auto-detected):" if _is_confident
                else "Multiple sheets found — please verify the correct sheet:"
            )
            selected_sheet = st.selectbox(
                _select_label,
                available_sheets,
                index=_best_idx,
                key="sheet_selector",
            )
            if _is_confident and selected_sheet == best_sheet:
                _pts = all_sheet_scores[best_sheet]["selection_score"]
                st.markdown(
                    f'<div class="alert alert-info">'
                    f'<span class="alert-icon">✓</span>'
                    f'<span>Auto-selected <strong>{selected_sheet}</strong> '
                    f'({_pts} pts). Change the selector above if needed.</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Full scan of selected sheet — cached per sheet ────────────────────
        _scan_cache_key = f"_sheet_scan_{_file_key}_{selected_sheet}"
        if _scan_cache_key not in st.session_state:
            _wb_full = load_workbook(BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
            _full_ws = _wb_full[selected_sheet]
            _fmr, _fmc, _fhr, _fph = scan_sheet(_full_ws)
            _opx_mr  = _full_ws.max_row
            _opx_mc  = _full_ws.max_column
            _wb_full.close()
            st.session_state[_scan_cache_key] = {
                "real_max_row":      _fmr,
                "real_max_col":      _fmc,
                "header_row":        _fhr,
                "peek_headers":      _fph,
                "openpyxl_max_row":  _opx_mr,
                "openpyxl_max_col":  _opx_mc,
            }
        _cached = st.session_state[_scan_cache_key]
        real_max_row = _cached["real_max_row"]
        real_max_col = _cached["real_max_col"]
        header_row   = _cached["header_row"]
        peek_headers = _cached["peek_headers"]

        _ws_info = {
            "openpyxl_max_row": _cached["openpyxl_max_row"],
            "openpyxl_max_col": _cached["openpyxl_max_col"],
            "real_max_row":     real_max_row,
            "real_max_col":     real_max_col,
            "columns_found":    len(peek_headers),
        }

        classification = classify_columns(peek_headers)
        classification["header_row"] = header_row
        classification["ws_info"]    = _ws_info

        # ── Column detection summary ──────────────────────────────────────────
        _to_tr  = classification.get("to_translate", {})
        _missed = classification.get("possible_missed", [])
        _role   = st.session_state.get("user_role", "")

        if _to_tr:
            _col_names = ", ".join(f"<strong>{h}</strong>" for h in _to_tr.keys())
            _hdr_note  = (
                f" (headers in row {header_row})" if header_row != 1 else ""
            )
            st.markdown(
                f'<div class="alert alert-info">'
                f'<span class="alert-icon">✓</span>'
                f'<span>{len(_to_tr)} column(s) ready{_hdr_note}: {_col_names}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            if _missed:
                st.markdown(
                    f'<div class="alert alert-warn">'
                    f'<span class="alert-icon">⚠</span>'
                    f'<span><strong>Possible missed columns:</strong> {", ".join(_missed)}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            if _role == "admin":
                with st.expander("Column detection details", expanded=False):
                    render_column_report(classification)
        else:
            render_column_report(classification)

        # ── Admin: multi-sheet diagnostics ────────────────────────────────────
        if _role == "admin" and len(available_sheets) > 1:
            with st.expander("Multi-sheet diagnostics", expanded=False):
                _diag_rows = []
                for _sn, _sd in all_sheet_scores.items():
                    _diag_rows.append({
                        "Sheet":        _sn,
                        "Score":        _sd["selection_score"],
                        "H24 headers":  len(_sd.get("h24_headers", [])),
                        "Protected":    len(_sd.get("protected_headers", [])),
                        "Non-empty cells (≤100 rows)": _sd.get("non_empty_cells", 0),
                        "Header row":   _sd.get("header_row", 1),
                        "Hidden":       _sd.get("hidden", False),
                        "Selected":     "✓" if _sn == selected_sheet else "",
                    })
                import pandas as _pd_diag
                st.dataframe(
                    _pd_diag.DataFrame(_diag_rows).sort_values("Score", ascending=False),
                    hide_index=True, use_container_width=True,
                )

        # ── Manual fallback when automatic detection found nothing ────────────
        if not classification["to_translate"]:
            st.markdown("""
            <div class="alert alert-warn" style="margin-top:0;">
                <span class="alert-icon">⚠</span>
                <div>
                    <strong>Automatic detection failed.</strong>
                    Please select the columns you want to translate below.
                    If the wrong sheet is selected, change it in the selector above.
                    Protected columns (articleNumber, SKU, ID) will never be translated.
                </div>
            </div>
            """, unsafe_allow_html=True)

            # Build candidate list — headers first, column letters as final fallback.
            candidates: list[str] = []
            candidate_col_map: dict[str, int] = {}
            if peek_headers:
                protected_keys = set(classification["protected"].keys())
                candidates = [h for h in peek_headers if h not in protected_keys]
                candidate_col_map = {h: peek_headers[h] for h in candidates}
            if not candidates:
                _letters  = list(string.ascii_uppercase)
                _extended = _letters + [f"A{l}" for l in _letters]
                col_count = max(real_max_col, 1)
                candidates = [_extended[i] for i in range(min(col_count, len(_extended)))]
                candidate_col_map = {lbl: (i + 1) for i, lbl in enumerate(candidates)}

            manual_cols = st.multiselect(
                "Select columns to translate:",
                options=candidates,
                key="manual_col_select",
                help="Choose each column header that contains German text to translate.",
            )

            if manual_cols:
                manual_to_translate = {}
                for h in manual_cols:
                    col_idx = candidate_col_map[h]
                    canonical, _, _ = _classify_header(h)
                    manual_to_translate[h] = (col_idx, canonical or "other")
                classification["to_translate"] = manual_to_translate

        # Advanced settings
        with st.expander("Advanced settings"):
            col_bs, col_cc = st.columns(2)
            with col_bs:
                batch_size = st.slider(
                    "Batch size (cells per request)",
                    min_value=5, max_value=30,
                    value=DEFAULT_BATCH_SIZE,
                    help="Cells grouped into one API request. Larger = fewer calls, bigger payloads.",
                )
            with col_cc:
                max_concurrent_batches = st.slider(
                    "Max concurrent batches",
                    min_value=1, max_value=5,
                    value=DEFAULT_MAX_CONCURRENT,
                    help=(
                        "How many batches are sent to OpenAI in parallel. "
                        "3 is a safe default; set to 1 for sequential (same as before)."
                    ),
                )
            st.markdown("---")
            enable_refinement = st.checkbox(
                f"Premium {target_language} Refinement (Pass 2)",
                value=True,
                key="enable_refinement",
                help=(
                    f"Runs a second AI pass on materialDetail, qualityDetail, name, deliveryScope, "
                    f"and variantName to produce more natural, premium {target_language} e-commerce copy. "
                    "Short texts, pure colors, and dimensions are skipped automatically. "
                    "Adds a small cost (~5–15% extra API tokens)."
                ),
            )
            enable_consistency = st.checkbox(
                "Terminology Consistency Pass (Pass 4)",
                value=True,
                key="enable_consistency",
                help=(
                    "Fast local pass — no API cost. Detects when the same German source term "
                    "was translated differently across the file and harmonizes all occurrences "
                    "to a single canonical translation. Also fixes known wrong AI variants "
                    "(e.g. 'rotin synthétique' → 'résine tressée')."
                ),
            )
            enable_final_qa = st.checkbox(
                "Final QA Scan (Pass 5)",
                value=True,
                key="enable_final_qa",
                help=(
                    "Fast local scan of every translated cell — no API cost. Flags empty cells "
                    "that had source content and any German residue that slipped through earlier "
                    "passes. Issues are added to Warning Details; download is never blocked."
                ),
            )
            highlight_warnings_in_excel = st.checkbox(
                "Highlight review warnings in exported Excel",
                value=False,
                key="highlight_warnings_excel",
                help=(
                    "When checked, cells flagged by Human Review Mode are highlighted "
                    "yellow in the exported file. Off by default — warnings appear in the "
                    "dashboard report only, and original source highlights are always preserved."
                ),
            )

        # ── Enterprise pipeline: Large File Mode indicator ───────────────────
        _preview_lf = detect_large_file_mode(
            max(real_max_row - header_row, 0)
        )
        if _preview_lf.active:
            _lf_batch_info = (
                f"Batch size reduced to {_preview_lf.batch_size} · "
                f"Semantic clustering active · "
                f"Concurrency capped at {_preview_lf.max_concurrent}"
            )
            st.markdown(
                f'<div class="alert alert-info">'
                f'<span class="alert-icon">⚡</span>'
                f'<span><strong>Large File Mode</strong> — {real_max_row - header_row} rows detected. '
                f'{_lf_batch_info}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

        st.markdown('<div class="section-label">Translate</div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="alert alert-warn" style="margin-bottom:16px;">
            <span class="alert-icon">⚠</span>
            <span>Keep this tab open while the translation is running.</span>
        </div>
        """, unsafe_allow_html=True)

        _, btn_col, _ = st.columns([1, 2, 1])
        with btn_col:
            translate_button = st.button(
                "Run Translation →", type="primary", use_container_width=True
            )

        if translate_button:
            progress_bar       = st.progress(0)
            progress_container = st.empty()

            try:
                output_bytes, stats = process_excel_with_progress(
                    uploaded_file, progress_bar, progress_container,
                    selected_sheet, classification,
                    batch_size=int(batch_size),
                    highlight_review_warnings=highlight_warnings_in_excel,
                    max_concurrent_batches=int(max_concurrent_batches),
                    header_row=header_row,
                    enable_refinement=enable_refinement,
                    enable_consistency=enable_consistency,
                    enable_final_qa=enable_final_qa,
                    target_language=target_language,
                )
                progress_container.empty()
                progress_bar.empty()

                # Generate CSV before saving history so the record is accurate
                _excel_data = output_bytes.getvalue()
                _csv_bytes, _csv_filename, _csv_removed_col = generate_csv_export(
                    _excel_data, selected_sheet, uploaded_file.name,
                    header_row=header_row, target_language=target_language,
                )

                job_id = str(uuid.uuid4())
                save_history_record({
                    "id":                        job_id,
                    "datetime":                  datetime.now().isoformat(timespec="seconds"),
                    "original_filename":         uploaded_file.name,
                    "output_filename":           output_filename,
                    "sheet_name":                selected_sheet,
                    "source_language":           "German",
                    "target_language":           target_language,
                    "output_prefix":             lang_code,
                    "user_email":                st.session_state.get("user_email", ""),
                    "user_role":                 st.session_state.get("user_role", ""),
                    "cells_translated":          stats["cells_translated"],
                    "cells_skipped":             stats["cells_skipped"],
                    "residue_corrections":       stats["residue_corrections"],
                    "unresolved_warnings":       stats["unresolved_warnings"],
                    "processing_time_seconds":   stats["elapsed_seconds"],
                    "processing_time_formatted": stats["total_time"],
                    "estimated_cost_usd":        stats.get("estimated_cost_usd"),
                    "prompt_tokens":             stats.get("prompt_tokens"),
                    "completion_tokens":         stats.get("completion_tokens"),
                    "tm_hits":                   stats.get("tm_hits", 0),
                    "tm_misses":                 stats.get("tm_misses", 0),
                    "batch_count":               stats.get("batch_count", 0),
                    "avg_batch_size":            stats.get("avg_batch_size", 0.0),
                    "api_calls_reduced":         stats.get("api_calls_reduced", 0),
                    "glossary_hits":             stats.get("glossary_hits", 0),
                    "review_count":              stats.get("review_count", 0),
                    "retry_count":               stats.get("retry_count", 0),
                    "critical_warnings":         stats.get("critical_warnings", 0),
                    "high_warnings":             stats.get("high_warnings", 0),
                    "medium_warnings":           stats.get("medium_warnings", 0),
                    "low_warnings":              stats.get("low_warnings", 0),
                    "total_warnings":            len(stats.get("all_warnings", [])),
                    "quality_score":             stats.get("quality_score", 100),
                    "warning_categories":        stats.get("warning_categories", {}),
                    "excel_exported":            1,
                    "csv_exported":              1 if _csv_bytes is not None else 0,
                    "csv_removed_column":        _csv_removed_col or "",
                    "csv_delimiter":             ";",
                    "csv_encoding":              "utf-8-sig",
                    # Intelligence Engine fields
                    "semantic_tm_hits":          stats.get("semantic_tm_hits", 0),
                    "duplicate_groups":          stats.get("duplicate_groups", 0),
                    "duplicate_cells_saved":     stats.get("duplicate_cells_saved", 0),
                    "glossary_only_count":       stats.get("glossary_only_count", 0),
                    "pattern_count":             stats.get("pattern_count", 0),
                    "gpt_calls_avoided":         stats.get("gpt_calls_avoided", 0),
                    "detected_product_type":     stats.get("detected_product_type", "generic"),
                    # Consistency + QA pass fields
                    "consistency_corrections":   stats.get("consistency_corrections", 0),
                    "consistency_detected":      stats.get("consistency_detected", 0),
                    "terms_harmonized":          stats.get("terms_harmonized", 0),
                    "qa_issues_found":           stats.get("qa_issues_found", 0),
                    # Language isolation
                    "dutch_contamination_fixes": stats.get("dutch_contamination_fixes", 0),
                })
                db_save_warnings(job_id, stats.get("all_warnings", []))

                # Persist glossary suggestions for admin review
                suggestions = stats.get("_glossary_suggestions", [])
                if suggestions:
                    db_save_glossary_suggestions(suggestions, job_id, target_language)

                # Store everything in session_state so results persist across reruns
                # (e.g. when a download button is clicked)
                st.session_state["_tr_result"] = {
                    "job_id":             job_id,
                    "stats":              stats,
                    "excel_data":         _excel_data,
                    "csv_bytes":          _csv_bytes,
                    "csv_filename":       _csv_filename,
                    "csv_removed_col":    _csv_removed_col,
                    "output_filename":    output_filename,
                    "highlight_in_excel": highlight_warnings_in_excel,
                    "selected_sheet":     selected_sheet,
                    "header_row":         header_row,
                    "orig_filename":      uploaded_file.name,
                }

            except ValueError as e:
                st.error(f"Error: {e}")
            except Exception as e:
                st.error(f"Unexpected error: {e}")

        # ── Render results (persists across reruns / download clicks) ──────────
        if "last_result" in st.session_state:
            st.session_state["_tr_result"] = st.session_state.pop("last_result")

        if st.session_state.get("_tr_result_file") == uploaded_file.name and "_tr_result" in st.session_state:
            _r   = st.session_state["_tr_result"]
            _s   = _r["stats"]
            _hix = _r["highlight_in_excel"]
            _ed  = _r["excel_data"]
            _cb  = _r["csv_bytes"]
            _cf  = _r["csv_filename"]
            _crc = _r["csv_removed_col"]
            _ofn = _r["output_filename"]
            _ssh = _r["selected_sheet"]
            _hdr = _r["header_row"]
            _ofnm= _r["orig_filename"]

            # ── Results ──
            st.markdown('<div class="section-label">Results</div>', unsafe_allow_html=True)
            render_stats(_s)

            # ── Quality Pipeline status ──
            _pipeline_steps = [
                ("Pass 1 — Initial Translation", "done"),
                (
                    f"Pass 2 — Premium Refinement"
                    + (f" ({_s.get('cells_refined', 0)} cells improved)" if _s.get("pipeline_refinement") else ""),
                    "done" if _s.get("pipeline_refinement") else "skipped",
                ),
                (
                    f"Pass 3 — Residue Check"
                    + (f" ({_s.get('residue_corrections', 0)} fix(es))" if _s.get("residue_corrections") else ""),
                    "done",
                ),
                (
                    f"Pass 4 — Consistency"
                    + (f" ({_s.get('consistency_corrections', 0)} harmonized)" if _s.get("pipeline_consistency") else ""),
                    "done" if _s.get("pipeline_consistency") else "skipped",
                ),
                (
                    f"Pass 5 — Final QA"
                    + (f" ({_s.get('qa_issues_found', 0)} issue(s))" if _s.get("pipeline_final_qa") else ""),
                    "done" if _s.get("pipeline_final_qa") else "skipped",
                ),
            ]
            st.markdown(_pipeline_status_html(_pipeline_steps), unsafe_allow_html=True)

            # ── Enterprise pipeline: Large File Mode badge ───────────────────
            if _s.get("large_file_mode"):
                _cm_fixes = _s.get("consistency_memory_fixes", 0)
                _clusters = _s.get("cluster_count", 0)
                _eff_bs   = _s.get("effective_batch_size", _s.get("avg_batch_size", "?"))
                _lf_badge_parts = [
                    f"Batch size: {_eff_bs}",
                    f"Clusters: {_clusters}" if _clusters else None,
                    f"Consistency enforced: {_cm_fixes} fix(es)" if _cm_fixes else None,
                ]
                _lf_text = " · ".join(p for p in _lf_badge_parts if p)
                st.markdown(
                    f'<div class="alert alert-info">'
                    f'<span class="alert-icon">⚡</span>'
                    f'<span><strong>Large File Mode was active.</strong> {_lf_text}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # ── Admin: per-sheet debug metrics ──────────────────────────────
            if _role == "admin" and _s.get("_sheet_debug_metrics"):
                with st.expander("Per-sheet pipeline metrics (admin)", expanded=False):
                    import pandas as _pd_dbg
                    _dbg_row = _s["_sheet_debug_metrics"]
                    st.dataframe(
                        _pd_dbg.DataFrame([_dbg_row]),
                        hide_index=True,
                        use_container_width=True,
                    )

            # ── Translation Intelligence ──
            st.markdown('<div class="section-label">Translation Intelligence</div>', unsafe_allow_html=True)
            render_intelligence_stats(_s)

            if _s.get("glossary_top_terms"):
                top_terms_str = " · ".join(
                    f"{de} → {DEFAULT_GLOSSARY_TERMS.get(de, '?')}"
                    for de in list(_s["glossary_top_terms"].keys())[:3]
                )
                st.markdown(f"""
                <div class="alert alert-info" style="margin-top:0;">
                    <span class="alert-icon">ℹ</span>
                    <span><strong>Top glossary terms used:</strong> {top_terms_str}</span>
                </div>
                """, unsafe_allow_html=True)

            # ── Quality Gate ──
            st.markdown('<div class="section-label">Quality Gate</div>', unsafe_allow_html=True)
            qg = _s.get("quality_gate", {})
            gate_rows = [
                ("✅" if qg.get("no_residue") else "⚠️",
                 "German residue",
                 "Clean" if qg.get("no_residue") else "Residue found — see warnings"),
                ("✅", "Row 1 (headers)", "Untouched"),
                ("✅" if qg.get("protected_columns") else "ℹ️",
                 "Protected columns",
                 ", ".join(qg.get("protected_columns", [])) or "None found"),
                ("⚠️" if qg.get("possible_missed") else "✅",
                 "Missed columns",
                 ", ".join(qg.get("possible_missed", [])) or "All matched"),
                ("✅", "Translation Memory",
                 f"{_s.get('tm_hits', 0)} hits / {_s.get('tm_misses', 0)} misses"),
                ("✅", "Batch processing",
                 f"{_s.get('batch_count', 0)} batches · avg {_s.get('avg_batch_size', 0)} cells/req"),
                (
                    "✅" if _s.get("refinement_enabled") else "ℹ️",
                    "Premium French Refinement",
                    (
                        f"{_s.get('cells_refined', 0)} cells improved "
                        f"({_s.get('refinement_api_calls', 0)} refinement batch(es))"
                        if _s.get("refinement_enabled")
                        else "Disabled"
                    ),
                ),
            ]
            qg_rows_html = "".join(
                f"""<div class="qg-row">
                    <span class="qg-icon">{icon}</span>
                    <span class="qg-label">{label}</span>
                    <span class="qg-value">{value}</span>
                </div>"""
                for icon, label, value in gate_rows
            )
            st.markdown(f'<div class="qg">{qg_rows_html}</div>', unsafe_allow_html=True)

            # ── Review Dashboard ──
            st.markdown('<div class="section-label">Review Dashboard</div>', unsafe_allow_html=True)
            render_review_dashboard(_s.get("all_warnings", []), _s, _hix)

            # ── Export Audit ──
            n_orig       = _s.get("original_highlights_preserved", 0)
            n_rev        = _s.get("review_highlights_applied", 0)
            n_gloss      = _s.get("glossary_hits", 0)
            n_total_warn = len(_s.get("all_warnings", []))
            rev_label = (
                f"{n_rev} cell(s) highlighted (Critical + High only)"
                if _hix
                else f"0 — checkbox off, {n_total_warn} warning(s) in dashboard only"
            )
            orig_label = f"{n_orig} cell(s) — source formatting preserved" if n_orig else "None in this file"
            audit_rows = [
                ("📋", "Original source highlights",    orig_label),
                ("🟡" if n_rev else "✅", "Review warnings in Excel", rev_label),
                ("✅", "Glossary hits in Excel",
                 f"{n_gloss} hit(s) tracked in analytics — no cell color applied"),
            ]
            audit_html = "".join(
                f"""<div class="qg-row">
                    <span class="qg-icon">{icon}</span>
                    <span class="qg-label">{label}</span>
                    <span class="qg-value">{value}</span>
                </div>"""
                for icon, label, value in audit_rows
            )
            st.markdown('<div class="section-label">Export Audit</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="qg">{audit_html}</div>', unsafe_allow_html=True)

            # ── Completion banner ──
            cost_str        = (
                f"${_s['estimated_cost_usd']:.4f}"
                if _s.get("estimated_cost_usd") is not None
                else "cost N/A"
            )
            processed_sheet = _s.get("sheet_name", "")
            n_warn_str      = f" · {n_total_warn} warning(s)" if n_total_warn else ""
            if not _s.get("all_warnings"):
                st.markdown(f"""
                <div class="success-banner">
                    <div class="success-banner-icon">✓</div>
                    <div>
                        <div class="success-banner-title">Translation complete</div>
                        <div class="success-banner-sub">
                            Sheet: {processed_sheet} · {_s["cells_translated"]} cells ·
                            {_s["total_time"]} · {cost_str}
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="warn-banner">
                    <div class="warn-banner-title">
                        Translation complete{n_warn_str}
                    </div>
                    <div class="warn-banner-sub">
                        Sheet: {processed_sheet} · {_s["total_time"]} · {cost_str}
                    </div>
                </div>
                """, unsafe_allow_html=True)

            # ── Download buttons (always available after translation) ──
            st.markdown("<br>", unsafe_allow_html=True)

            if _cb is not None:
                dl_left, dl_right = st.columns(2)
                with dl_left:
                    st.download_button(
                        label="↓ Download Excel",
                        data=_ed,
                        file_name=_ofn,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_excel_btn",
                    )
                with dl_right:
                    st.download_button(
                        label=f"↓ Download CSV (sans «{_crc}»)",
                        data=_cb,
                        file_name=_cf,
                        mime="text/csv",
                        use_container_width=True,
                        key="dl_csv_btn",
                    )
            else:
                # Name column not found — Excel only + manual CSV column picker
                _, dl_col, _ = st.columns([1, 2, 1])
                with dl_col:
                    st.download_button(
                        label="↓ Download Excel",
                        data=_ed,
                        file_name=_ofn,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_excel_only_btn",
                    )
                import io as _io
                import openpyxl as _openpyxl
                _wb_pk = _openpyxl.load_workbook(_io.BytesIO(_ed), data_only=True)
                _ws_pk = _wb_pk[_ssh] if _ssh in _wb_pk.sheetnames else _wb_pk.active
                _all_h = [
                    str(_ws_pk.cell(row=_hdr, column=c).value or "")
                    for c in range(1, _ws_pk.max_column + 1)
                    if _ws_pk.cell(row=_hdr, column=c).value is not None
                ]
                st.warning(
                    "CSV export: could not auto-detect the product name column. "
                    "Select it manually to generate the CSV."
                )
                if _all_h:
                    _chosen = st.selectbox(
                        "Column to exclude from CSV",
                        options=_all_h,
                        key="csv_col_manual_select",
                    )
                    if st.button("Generate CSV without selected column", key="csv_manual_gen_btn"):
                        _csv_b2, _csv_f2, _csv_rc2 = generate_csv_export(
                            _ed, _ssh, _ofnm, header_row=_hdr, force_exclude_header=_chosen,
                        )
                        if _csv_b2:
                            # Persist the generated CSV so it survives the next rerun
                            st.session_state["_tr_result"]["csv_bytes"]       = _csv_b2
                            st.session_state["_tr_result"]["csv_filename"]    = _csv_f2
                            st.session_state["_tr_result"]["csv_removed_col"] = _csv_rc2
                            st.rerun()

            # ── Jira upload section (only when file came from Jira) ───────────
            _jira_tk   = st.session_state.get("_jira_ticket_key", "")
            _jira_summ = st.session_state.get("_jira_ticket_summary", "")
            if _jira_tk:
                st.markdown("---")
                st.markdown(
                    '<div class="section-label">Upload to Jira</div>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f'<div class="alert alert-info">'
                    f'<span class="alert-icon">⬆</span>'
                    f'<span>Ticket: <strong>{_jira_tk}</strong>'
                    f'{" — " + _jira_summ if _jira_summ else ""}'
                    f'</span></div>',
                    unsafe_allow_html=True,
                )
                _add_jira_comment = st.checkbox(
                    "Add Jira comment after upload",
                    value=True,
                    key="jira_add_comment_cb",
                )
                _target_lang_j = st.session_state.get("target_language", "French")
                _lang_name_j   = "French" if _target_lang_j == "French" else "Dutch"

                _jcol1, _jcol2, _jcol3 = st.columns(3)

                def _do_jira_upload(
                    upload_xlsx: bool,
                    upload_csv: bool,
                    excel_data: bytes,
                    csv_bytes,
                    output_filename: str,
                    csv_filename: str,
                    ticket_key: str,
                    ticket_summary: str,
                    lang_name: str,
                    add_comment: bool,
                ) -> None:
                    _jc, _jerr = get_jira_client()
                    if not _jc:
                        st.error(f"Jira connection failed: {_jerr}")
                        return
                    _uploaded_files: list[str] = []
                    if upload_xlsx and excel_data:
                        _res = _jc.upload_attachment(ticket_key, output_filename, excel_data)
                        if _res["ok"]:
                            _uploaded_files.append(output_filename)
                        else:
                            st.error(f"XLSX upload failed: {_res['error']}")
                    if upload_csv and csv_bytes:
                        _res = _jc.upload_attachment(ticket_key, csv_filename, csv_bytes)
                        if _res["ok"]:
                            _uploaded_files.append(csv_filename)
                        else:
                            st.error(f"CSV upload failed: {_res['error']}")
                    if _uploaded_files:
                        st.success(f"Uploaded to {ticket_key}: {', '.join(_uploaded_files)}")
                        _now_j = datetime.now().isoformat(timespec="seconds")
                        _comment_ok = False
                        if add_comment:
                            _comment_text = (
                                f"AI localization files generated and attached.\n"
                                f"\n"
                                f"Target language: {lang_name}\n"
                                f"Generated files:\n"
                                + "\n".join(f"- {fn}" for fn in _uploaded_files)
                                + "\n\nGenerated by Home24 AI Localization Platform."
                            )
                            _cr = _jc.add_comment(ticket_key, _comment_text)
                            if _cr["ok"]:
                                st.success("Comment added to Jira ticket.")
                                _comment_ok = True
                            else:
                                st.warning(f"Comment failed: {_cr['error']}")
                        _r_id = st.session_state.get("_tr_result", {}).get("job_id", "")
                        if _r_id:
                            db_update_jira_metadata(
                                _r_id,
                                jira_ticket_key=ticket_key,
                                jira_ticket_summary=ticket_summary,
                                jira_attachment_filename=st.session_state.get(
                                    "_jira_attachment_filename", ""
                                ),
                                jira_attachment_id=st.session_state.get(
                                    "_jira_attachment_id", ""
                                ),
                                uploaded_to_jira=1,
                                jira_upload_time=_now_j,
                                jira_comment_added=1 if _comment_ok else 0,
                            )

                with _jcol1:
                    if st.button(
                        "⬆ Upload XLSX", key="jira_ul_xlsx", use_container_width=True
                    ):
                        _do_jira_upload(
                            True, False, _ed, _cb, _ofn, _cf,
                            _jira_tk, _jira_summ, _lang_name_j, _add_jira_comment,
                        )
                with _jcol2:
                    if st.button(
                        "⬆ Upload CSV", key="jira_ul_csv",
                        disabled=_cb is None, use_container_width=True,
                    ):
                        _do_jira_upload(
                            False, True, _ed, _cb, _ofn, _cf,
                            _jira_tk, _jira_summ, _lang_name_j, _add_jira_comment,
                        )
                with _jcol3:
                    if st.button(
                        "⬆ Upload both", key="jira_ul_both",
                        disabled=_cb is None, use_container_width=True,
                    ):
                        _do_jira_upload(
                            True, True, _ed, _cb, _ofn, _cf,
                            _jira_tk, _jira_summ, _lang_name_j, _add_jira_comment,
                        )

                # ── Optional status transition ────────────────────────────────
                if st.checkbox(
                    "Apply status transition",
                    value=False,
                    key="jira_trans_cb",
                ):
                    _jc_t, _jerr_t = get_jira_client()
                    if not _jc_t:
                        st.warning(f"Cannot load transitions: {_jerr_t}")
                    else:
                        _trans_list = _jc_t.get_transitions(_jira_tk)
                        if _trans_list:
                            _t_names     = [t["name"] for t in _trans_list]
                            _chosen_t_nm = st.selectbox(
                                "Transition to:", _t_names, key="jira_trans_sel"
                            )
                            _chosen_t_id = next(
                                (t["id"] for t in _trans_list if t["name"] == _chosen_t_nm),
                                None,
                            )
                            if st.button(
                                f"Apply: {_chosen_t_nm}", key="jira_apply_trans_btn"
                            ):
                                if _chosen_t_id:
                                    _tr_res = _jc_t.apply_transition(_jira_tk, _chosen_t_id)
                                    if _tr_res["ok"]:
                                        st.success(
                                            f"Ticket {_jira_tk} transitioned to: {_chosen_t_nm}"
                                        )
                                        _r_id_t = st.session_state.get(
                                            "_tr_result", {}
                                        ).get("job_id", "")
                                        if _r_id_t:
                                            db_update_jira_metadata(
                                                _r_id_t,
                                                jira_transition_applied=_chosen_t_nm,
                                            )
                                    else:
                                        st.error(f"Transition failed: {_tr_res['error']}")
                        else:
                            st.info("No transitions available for this ticket.")


# =============================================================================
# PAGE: TRANSLATION HISTORY
# =============================================================================

def history_page():
    user_email = st.session_state.get("user_email", "")
    user_role  = st.session_state.get("user_role", "")
    subtitle   = "All jobs (admin view)" if user_role == "admin" else "Your translation jobs — most recent first"
    render_page_header("Translation History", subtitle)

    history = db_load_history_for_user(user_email, user_role)

    st.markdown("""
    <div class="cloud-note">
        On Streamlit Community Cloud, history resets on each redeployment (ephemeral file system).
        This is expected for this version.
    </div>
    """, unsafe_allow_html=True)

    if not history:
        st.markdown("""
        <div class="history-empty">
            No translations yet.<br>
            <span class="history-empty-sub">
                Go to <strong>Translator</strong> to get started.
            </span>
        </div>
        """, unsafe_allow_html=True)
        return

    total_files  = len(history)
    total_cells  = sum(r.get("cells_translated", 0) for r in history)
    total_time_s = sum(r.get("processing_time_seconds", 0) for r in history)
    costs        = [r["estimated_cost_usd"] for r in history if r.get("estimated_cost_usd") is not None]
    total_cost   = sum(costs) if costs else None
    cost_display = f"${total_cost:.4f}" if total_cost is not None else "—"

    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-label">Files Translated</div>
            <div class="kpi-value accent">{total_files}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total Cells</div>
            <div class="kpi-value">{total_cells:,}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total Time</div>
            <div class="kpi-value success">{format_time(total_time_s)}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total Est. Cost</div>
            <div class="kpi-value warn">{cost_display}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-label">Job Log</div>', unsafe_allow_html=True)

    rows = []
    for r in history:
        dt  = r.get("datetime", "")[:16].replace("T", " ")
        c   = r.get("estimated_cost_usd")
        qs  = r.get("quality_score")
        row = {
            "Date / Time":     dt,
            "Lang":            r.get("output_prefix", "FR"),
            "File":            r.get("original_filename", ""),
            "Sheet":           r.get("sheet_name", ""),
            "Translated":      r.get("cells_translated", 0),
            "Score":           f"{qs}/100" if qs is not None else "—",
            "Critical":        r.get("critical_warnings", "—"),
            "High":            r.get("high_warnings", "—"),
            "Warnings (total)": r.get("total_warnings", r.get("unresolved_warnings", 0)),
            "Time":            r.get("processing_time_formatted", ""),
            "Est. Cost (USD)": f"${c:.4f}" if c is not None else "—",
        }
        if user_role == "admin":
            row["User"] = r.get("user_email", "")
        rows.append(row)

    st.dataframe(rows, use_container_width=True, hide_index=True)


# =============================================================================
# PAGE: ANALYTICS
# =============================================================================

def analytics_page():
    user_email = st.session_state.get("user_email", "")
    user_role  = st.session_state.get("user_role", "")
    subtitle   = "Aggregated statistics across all jobs" if user_role == "admin" else "Your translation statistics"
    render_page_header("Analytics", subtitle)

    history = db_load_history_for_user(user_email, user_role)

    if not history:
        st.markdown("""
        <div class="history-empty">
            No data yet.<br>
            <span class="history-empty-sub">
                Complete a translation to see analytics.
            </span>
        </div>
        """, unsafe_allow_html=True)
        return

    total_files      = len(history)
    total_translated = sum(r.get("cells_translated", 0) for r in history)
    total_skipped    = sum(r.get("cells_skipped", 0) for r in history)
    total_residue    = sum(r.get("residue_corrections", 0) for r in history)
    total_warnings   = sum(r.get("unresolved_warnings", 0) for r in history)
    total_time_s     = sum(r.get("processing_time_seconds", 0) for r in history)
    costs            = [r["estimated_cost_usd"] for r in history if r.get("estimated_cost_usd") is not None]
    total_cost       = sum(costs) if costs else None

    manual_time_s = total_translated * MANUAL_SECONDS_PER_CELL
    saved_s       = max(manual_time_s - total_time_s, 0)
    saved_h       = saved_s / 3600
    cost_display  = f"${total_cost:.4f}" if total_cost is not None else "—"

    # ── Hero: time saved ──
    if total_translated > 0:
        st.markdown(f"""
        <div class="hero-kpi">
            <p class="hero-kpi-value">{saved_h:.1f}h</p>
            <p class="hero-kpi-label">estimated time saved vs. manual translation</p>
            <p class="hero-kpi-sub">
                {total_translated:,} cells × {MANUAL_SECONDS_PER_CELL}s manual
                — {format_time(total_time_s)} actual
            </p>
        </div>
        """, unsafe_allow_html=True)

    # ── Volume ──
    st.markdown('<div class="section-label">Volume</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-label">Files Translated</div>
            <div class="kpi-value accent">{total_files}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Cells Translated</div>
            <div class="kpi-value">{total_translated:,}</div>
            <div class="kpi-sub">→ French (FR)</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Cells Skipped</div>
            <div class="kpi-value">{total_skipped:,}</div>
            <div class="kpi-sub">Empty or unchanged</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Unresolved Warnings</div>
            <div class="kpi-value warn">{total_warnings}</div>
            <div class="kpi-sub">Need manual review</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Performance ──
    st.markdown('<div class="section-label">Performance & Cost</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row-3">
        <div class="kpi">
            <div class="kpi-label">Total Processing Time</div>
            <div class="kpi-value">{format_time(total_time_s)}</div>
            <div class="kpi-sub">Actual machine time</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Est. Manual Time</div>
            <div class="kpi-value">{format_time(manual_time_s)}</div>
            <div class="kpi-sub">@ {MANUAL_SECONDS_PER_CELL}s per cell</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Estimated API Cost</div>
            <div class="kpi-value warn">{cost_display}</div>
            <div class="kpi-sub">GPT-4o-mini pricing</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Cost Dashboard ──
    total_prompt     = sum(r.get("prompt_tokens", 0) or 0 for r in history)
    total_completion = sum(r.get("completion_tokens", 0) or 0 for r in history)
    total_tokens_all = total_prompt + total_completion
    avg_cost_file    = round(total_cost / total_files, 4)     if total_cost and total_files > 0     else None
    avg_cost_cell    = round(total_cost / total_translated, 6) if total_cost and total_translated > 0 else None

    st.markdown('<div class="section-label">Cost Dashboard</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-label">Total Tokens Used</div>
            <div class="kpi-value accent">{total_tokens_all:,}</div>
            <div class="kpi-sub">All jobs combined</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Prompt Tokens</div>
            <div class="kpi-value">{total_prompt:,}</div>
            <div class="kpi-sub">${total_prompt * _INPUT_COST_PER_TOKEN:.4f} input cost</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Completion Tokens</div>
            <div class="kpi-value">{total_completion:,}</div>
            <div class="kpi-sub">${total_completion * _OUTPUT_COST_PER_TOKEN:.4f} output cost</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Avg Cost / File</div>
            <div class="kpi-value warn">{f"${avg_cost_file:.4f}" if avg_cost_file is not None else "—"}</div>
            <div class="kpi-sub">Per translation job</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    if avg_cost_cell is not None:
        st.markdown(f"""
        <div class="kpi-row-3">
            <div class="kpi">
                <div class="kpi-label">Avg Cost / Cell</div>
                <div class="kpi-value warn">${avg_cost_cell:.6f}</div>
                <div class="kpi-sub">Per translated cell</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Input Rate</div>
                <div class="kpi-value" style="font-size:18px;">$0.15 / 1M</div>
                <div class="kpi-sub">GPT-4o-mini prompt tokens</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Output Rate</div>
                <div class="kpi-value" style="font-size:18px;">$0.60 / 1M</div>
                <div class="kpi-sub">GPT-4o-mini completion tokens</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    cost_by_job = {
        f"#{i + 1}": r.get("estimated_cost_usd") or 0
        for i, r in enumerate(reversed(history))
        if r.get("estimated_cost_usd") is not None
    }
    if len(cost_by_job) > 1:
        st.markdown('<div class="section-label">Cost per Job</div>', unsafe_allow_html=True)
        st.bar_chart(cost_by_job)

    # ── Quality ──
    st.markdown('<div class="section-label">Quality</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row-3">
        <div class="kpi">
            <div class="kpi-label">Residue Auto-fixes</div>
            <div class="kpi-value success">{total_residue}</div>
            <div class="kpi-sub">German words corrected</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Source Language</div>
            <div class="kpi-value" style="font-size:20px;">German (DE)</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">FR Jobs</div>
            <div class="kpi-value accent">{sum(1 for r in history if r.get("target_language","French")=="French")}</div>
            <div class="kpi-sub">French translations</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">NL Jobs</div>
            <div class="kpi-value" style="color:#e8523a;">{sum(1 for r in history if r.get("target_language","")=="Dutch")}</div>
            <div class="kpi-sub">Dutch translations</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Quality Score Analytics ──
    scored_records = [r for r in history if r.get("quality_score") is not None]
    if scored_records:
        avg_score      = round(sum(r["quality_score"] for r in scored_records) / len(scored_records), 1)
        total_critical = sum(r.get("critical_warnings", 0) for r in history)
        total_high     = sum(r.get("high_warnings", 0) for r in history)
        total_medium   = sum(r.get("medium_warnings", 0) for r in history)
        total_low      = sum(r.get("low_warnings", 0) for r in history)
        files_critical = sum(1 for r in history if r.get("critical_warnings", 0) > 0)

        score_color = "#22C55E" if avg_score >= 85 else ("#F59E0B" if avg_score >= 70 else "#EF4444")
        st.markdown('<div class="section-label">Quality Score</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="kpi-row">
            <div class="kpi">
                <div class="kpi-label">Avg Quality Score</div>
                <div class="kpi-value" style="color:{score_color};">{avg_score}/100</div>
                <div class="kpi-sub">Across {len(scored_records)} job(s)</div>
            </div>
            <div class="kpi">
                <div class="kpi-label" style="color:#ef4444;">Critical Warnings</div>
                <div class="kpi-value" style="color:#ef4444;">{total_critical}</div>
                <div class="kpi-sub">{files_critical} file(s) affected</div>
            </div>
            <div class="kpi">
                <div class="kpi-label" style="color:#f59e0b;">High Warnings</div>
                <div class="kpi-value" style="color:#f59e0b;">{total_high}</div>
                <div class="kpi-sub">−5 pts each</div>
            </div>
            <div class="kpi">
                <div class="kpi-label" style="color:#ca8a04;">Medium / Low</div>
                <div class="kpi-value" style="color:#ca8a04;">{total_medium + total_low}</div>
                <div class="kpi-sub">Medium {total_medium} · Low {total_low}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Most common warning categories across all jobs
        from collections import Counter
        all_cats: Counter = Counter()
        for r in history:
            for cat, cnt in r.get("warning_categories", {}).items():
                all_cats[cat] += cnt
        if all_cats:
            top_cats = all_cats.most_common(5)
            chips = "".join(
                f'<span class="chip chip-muted">{cat} <span class="chip-arrow">· {n}×</span></span>'
                for cat, n in top_cats
            )
            st.markdown(f"""
            <div class="card" style="margin-top:0;">
                <div class="card-title">Most common warning categories</div>
                <div>{chips}</div>
            </div>
            """, unsafe_allow_html=True)

    # ── Translation Memory ──
    tm = load_translation_memory()
    tm_global      = tm.get("global_stats", {})
    tm_total_hits  = tm_global.get("total_hits", 0)
    tm_total_miss  = tm_global.get("total_misses", 0)
    tm_saved_calls = tm_global.get("total_api_calls_saved", 0)
    tm_entries     = len(tm.get("entries", {}))
    # Estimate cost saved: each TM hit avoided ~500 input + 100 output tokens
    tm_cost_saved  = round(
        tm_total_hits * (500 * _INPUT_COST_PER_TOKEN + 100 * _OUTPUT_COST_PER_TOKEN), 4
    )
    hit_rate = int(tm_total_hits / max(tm_total_hits + tm_total_miss, 1) * 100)

    st.markdown('<div class="section-label">Translation Memory</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-label">Memory Entries</div>
            <div class="kpi-value accent">{tm_entries:,}</div>
            <div class="kpi-sub">Unique source phrases</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Cache Hits</div>
            <div class="kpi-value success">{tm_total_hits:,}</div>
            <div class="kpi-sub">{hit_rate}% hit rate</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">API Calls Saved</div>
            <div class="kpi-value">{tm_saved_calls:,}</div>
            <div class="kpi-sub">via memory reuse</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Est. Cost Saved</div>
            <div class="kpi-value warn">${tm_cost_saved:.4f}</div>
            <div class="kpi-sub">from TM cache hits</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Batch Processing ──
    total_batches      = sum(r.get("batch_count", 0) for r in history)
    total_api_reduced  = sum(r.get("api_calls_reduced", 0) for r in history)
    avg_batch_sizes    = [r.get("avg_batch_size", 0) for r in history if r.get("avg_batch_size", 0) > 0]
    overall_avg_batch  = round(sum(avg_batch_sizes) / max(len(avg_batch_sizes), 1), 1)
    speed_multiplier   = round(overall_avg_batch, 1) if overall_avg_batch > 0 else 1.0

    st.markdown('<div class="section-label">Batch Processing</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row-3">
        <div class="kpi">
            <div class="kpi-label">Total API Batches</div>
            <div class="kpi-value accent">{total_batches:,}</div>
            <div class="kpi-sub">Requests sent to OpenAI</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Avg Batch Size</div>
            <div class="kpi-value">{overall_avg_batch}</div>
            <div class="kpi-sub">cells per request</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">API Calls Reduced</div>
            <div class="kpi-value success">{total_api_reduced:,}</div>
            <div class="kpi-sub">~{speed_multiplier}× fewer requests</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Glossary ──
    glossary      = load_glossary()
    gloss_stats   = glossary.get("stats", {})
    gloss_hits    = gloss_stats.get("total_hits", 0)
    gloss_tc      = gloss_stats.get("term_counts", {})
    gloss_terms   = len(glossary.get("terms", {}))
    top_terms     = sorted(gloss_tc.items(), key=lambda x: -x[1])[:5]

    st.markdown('<div class="section-label">Glossary</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row-3">
        <div class="kpi">
            <div class="kpi-label">Glossary Terms</div>
            <div class="kpi-value accent">{gloss_terms}</div>
            <div class="kpi-sub">DE→FR mappings defined</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total Glossary Hits</div>
            <div class="kpi-value success">{gloss_hits:,}</div>
            <div class="kpi-sub">Consistent terms enforced</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Top Term</div>
            <div class="kpi-value" style="font-size:18px;">
                {top_terms[0][0] if top_terms else "—"}
            </div>
            <div class="kpi-sub">
                {f"{top_terms[0][1]} uses" if top_terms else "No data yet"}
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if top_terms:
        top_html = "".join(
            f'<span class="chip chip-accent">{de} <span class="chip-arrow">→ {glossary["terms"].get(de,"?")}</span> · {n}×</span>'
            for de, n in top_terms
        )
        st.markdown(f"""
        <div class="card" style="margin-top:0;">
            <div class="card-title">Top glossary terms used</div>
            <div>{top_html}</div>
        </div>
        """, unsafe_allow_html=True)

    if total_cost is None:
        st.markdown("""
        <div class="alert alert-info" style="margin-top:16px;">
            <span class="alert-icon">ℹ</span>
            <span>
                <strong>API cost unavailable</strong> for jobs completed before token tracking was added.
                Future translations will include cost estimates.
            </span>
        </div>
        """, unsafe_allow_html=True)

    # ── Translation Intelligence Engine ──────────────────────────────────────
    total_sem_hits    = sum(r.get("semantic_tm_hits", 0)      for r in history)
    total_dup_groups  = sum(r.get("duplicate_groups", 0)       for r in history)
    total_dup_saved   = sum(r.get("duplicate_cells_saved", 0)  for r in history)
    total_gloss_only  = sum(r.get("glossary_only_count", 0)    for r in history)
    total_pattern     = sum(r.get("pattern_count", 0)          for r in history)
    total_avoided     = sum(r.get("gpt_calls_avoided", 0)      for r in history)

    # Product type distribution
    from collections import Counter as _Counter
    type_counts = _Counter(
        r.get("detected_product_type", "generic")
        for r in history
        if r.get("detected_product_type") and r.get("detected_product_type") != "generic"
    )

    if total_avoided > 0 or total_sem_hits > 0 or type_counts:
        st.markdown('<div class="section-label">Translation Intelligence Engine</div>', unsafe_allow_html=True)
        avoided_cost = round(
            total_avoided * (500 * _INPUT_COST_PER_TOKEN + 100 * _OUTPUT_COST_PER_TOKEN), 4
        )
        st.markdown(f"""
        <div class="kpi-row">
            <div class="kpi">
                <div class="kpi-label">GPT Calls Avoided</div>
                <div class="kpi-value success">{total_avoided:,}</div>
                <div class="kpi-sub">Est. saved ${avoided_cost:.4f}</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Duplicate Groups</div>
                <div class="kpi-value accent">{total_dup_groups:,}</div>
                <div class="kpi-sub">{total_dup_saved:,} cells deduplicated</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Glossary-Only</div>
                <div class="kpi-value">{total_gloss_only:,}</div>
                <div class="kpi-sub">No API needed</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Pattern Matches</div>
                <div class="kpi-value">{total_pattern:,}</div>
                <div class="kpi-sub">Dimensions / percentages</div>
            </div>
        </div>
        <div class="kpi-row-3">
            <div class="kpi">
                <div class="kpi-label">Semantic TM Hits</div>
                <div class="kpi-value accent">{total_sem_hits:,}</div>
                <div class="kpi-sub">Near-match reuse (≥88%)</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Top Product Category</div>
                <div class="kpi-value" style="font-size:18px;text-transform:capitalize;">
                    {type_counts.most_common(1)[0][0] if type_counts else "—"}
                </div>
                <div class="kpi-sub">Detected in source files</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Categories Detected</div>
                <div class="kpi-value">{len(type_counts)}</div>
                <div class="kpi-sub">Unique furniture types</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if type_counts:
            cat_chips = "".join(
                f'<span class="chip chip-accent" style="text-transform:capitalize;">{cat} · {n}×</span>'
                for cat, n in type_counts.most_common(8)
            )
            st.markdown(f"""
            <div class="card" style="margin-top:0;">
                <div class="card-title">Product categories translated</div>
                <div>{cat_chips}</div>
            </div>
            """, unsafe_allow_html=True)


# =============================================================================
# PAGE: GLOSSARY MANAGEMENT
# =============================================================================

def glossary_page():
    render_page_header(
        "Glossary Management",
        "Terminology enforced consistently across all translations",
    )

    # Language filter
    gloss_lang = st.radio(
        "Show glossary for:",
        ["French (FR)", "Dutch (NL)"],
        horizontal=True,
        key="gloss_lang_filter",
    )
    active_lang = "Dutch" if "Dutch" in gloss_lang else "French"

    glossary    = load_glossary(active_lang)
    terms       = glossary.get("terms", {})
    gloss_stats = glossary.get("stats", {})
    term_counts = gloss_stats.get("term_counts", {})

    # ── Stats ──
    total_hits  = gloss_stats.get("total_hits", 0)
    total_terms = len(terms)
    used_terms  = len([t for t in terms if t in term_counts])

    st.markdown(f"""
    <div class="kpi-row-3">
        <div class="kpi">
            <div class="kpi-label">Total Terms</div>
            <div class="kpi-value accent">{total_terms}</div>
            <div class="kpi-sub">Defined mappings</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Terms Used</div>
            <div class="kpi-value success">{used_terms}</div>
            <div class="kpi-sub">Matched in source texts</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total Hits</div>
            <div class="kpi-value">{total_hits:,}</div>
            <div class="kpi-sub">Across all translations</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    target_col_label = "Dutch" if active_lang == "Dutch" else "French"

    # ── Glossary table ──
    st.markdown(f'<div class="section-label">Term List — DE→{target_col_label}</div>', unsafe_allow_html=True)

    rows = []
    for de, target in sorted(terms.items()):
        rows.append({
            "German":             de,
            target_col_label:     target,
            "Times Used":         term_counts.get(de, 0),
        })

    st.dataframe(rows, use_container_width=True, hide_index=True)

    # ── Add new term ──
    st.markdown('<div class="section-label">Add / Update Term</div>', unsafe_allow_html=True)

    with st.form("add_term_form"):
        col_de, col_tgt = st.columns(2)
        with col_de:
            new_de = st.text_input("German term", placeholder="e.g. Kopfteil")
        with col_tgt:
            new_target = st.text_input(
                f"{target_col_label} translation",
                placeholder=("e.g. hoofdbord" if active_lang == "Dutch" else "e.g. Tête de lit"),
            )
        add_submitted = st.form_submit_button("Add term →", use_container_width=True)

    if add_submitted:
        new_de     = new_de.strip()
        new_target = new_target.strip()
        if new_de and new_target:
            glossary["terms"][new_de] = new_target
            save_glossary(glossary, active_lang)
            st.success(f"Added: **{new_de}** → **{new_target}** ({active_lang})")
            st.rerun()
        else:
            st.error("Both fields are required.")

    # ── Reset to defaults ──
    st.markdown('<div class="section-label">Reset</div>', unsafe_allow_html=True)
    default_terms = DEFAULT_NL_GLOSSARY_TERMS if active_lang == "Dutch" else DEFAULT_GLOSSARY_TERMS
    if st.button(f"Reset {target_col_label} glossary to defaults"):
        glossary["terms"] = default_terms.copy()
        save_glossary(glossary, active_lang)
        st.success(f"{target_col_label} glossary reset to defaults.")
        st.rerun()

    # ── Auto-learned furniture terms ─────────────────────────────────────────
    st.markdown('<div class="section-label">Auto-learned Terminology</div>', unsafe_allow_html=True)
    furniture_map = FURNITURE_TERM_MAP_FR if active_lang == "French" else FURNITURE_TERM_MAP_NL
    auto_in_glossary = {
        de: tr for de, tr in furniture_map.items()
        if de in terms
    }
    auto_missing = {
        de: tr for de, tr in furniture_map.items()
        if de not in terms
    }

    if auto_in_glossary:
        st.markdown(
            f'<div class="alert alert-success">'
            f'<span class="alert-icon">✓</span>'
            f'<span><strong>{len(auto_in_glossary)} furniture term(s) active</strong> — '
            f'learned automatically from translation jobs.</span></div>',
            unsafe_allow_html=True,
        )

    if auto_missing:
        st.markdown(
            f'<div class="alert alert-info">'
            f'<span class="alert-icon">ℹ</span>'
            f'<span>{len(auto_missing)} furniture term(s) will be added automatically '
            f'when they appear in source files (min. 2 occurrences).</span></div>',
            unsafe_allow_html=True,
        )
        if st.button(f"Add all {len(auto_missing)} furniture terms now", key="add_all_furniture"):
            for de, tr in auto_missing.items():
                glossary["terms"].setdefault(de, tr)
            save_glossary(glossary, active_lang)
            st.success(f"Added {len(auto_missing)} furniture terms to glossary.")
            st.rerun()

    # ── Unknown term suggestions (for terms not in furniture map) ────────────
    suggestions = db_load_glossary_suggestions(target_language=active_lang, status="pending")
    unknown_suggestions = [
        s for s in suggestions
        if s["term"] not in furniture_map
    ]

    if unknown_suggestions:
        st.markdown('<div class="section-label">Unknown Terms — Review Required</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="alert alert-warn"><span class="alert-icon">💡</span>'
            f'<span><strong>{len(unknown_suggestions)} unknown term(s)</strong> detected from recent jobs. '
            f'These are not in the furniture vocabulary — review and add if needed.</span></div>',
            unsafe_allow_html=True,
        )
        for s in unknown_suggestions:
            sid  = s["id"]
            term = s["term"]
            occ  = s["occurrences"]
            ctx  = s.get("example_context", "")[:80]
            col_a, col_b, col_c = st.columns([3, 1, 1])
            with col_a:
                proposed = st.text_input(
                    f"**{term}** ({occ}× in source):",
                    key=f"sug_input_{sid}",
                    placeholder=f"Type {target_col_label} translation…",
                    help=f"Context: {ctx}" if ctx else "",
                )
            with col_b:
                if st.button("Accept", key=f"sug_acc_{sid}", use_container_width=True):
                    if proposed.strip():
                        glossary["terms"][term] = proposed.strip()
                        save_glossary(glossary, active_lang)
                        db_update_suggestion_status(sid, "accepted")
                        st.success(f"Added **{term}** → **{proposed.strip()}**")
                        st.rerun()
                    else:
                        st.error("Enter a translation first.")
            with col_c:
                if st.button("Reject", key=f"sug_rej_{sid}", use_container_width=True):
                    db_update_suggestion_status(sid, "rejected")
                    st.rerun()


# =============================================================================
# PAGE: TRANSLATION MEMORY
# =============================================================================

def translation_memory_page():
    render_page_header(
        "Translation Memory",
        "Cached translations reused across jobs to reduce API calls and cost",
    )

    tm = load_translation_memory()
    entries     = tm.get("entries", {})
    gs          = tm.get("global_stats", {})
    total_hits  = gs.get("total_hits", 0)
    total_miss  = gs.get("total_misses", 0)
    saved_calls = gs.get("total_api_calls_saved", 0)
    hit_rate    = int(total_hits / max(total_hits + total_miss, 1) * 100)
    tm_cost_saved = round(
        total_hits * (_INPUT_COST_PER_TOKEN * 500 + _OUTPUT_COST_PER_TOKEN * 100), 4
    )

    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-label">Memory Entries</div>
            <div class="kpi-value accent">{len(entries):,}</div>
            <div class="kpi-sub">Unique cached phrases</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Cache Hits</div>
            <div class="kpi-value success">{total_hits:,}</div>
            <div class="kpi-sub">{hit_rate}% hit rate</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">API Calls Saved</div>
            <div class="kpi-value">{saved_calls:,}</div>
            <div class="kpi-sub">via memory reuse</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Est. Cost Saved</div>
            <div class="kpi-value warn">${tm_cost_saved:.4f}</div>
            <div class="kpi-sub">from TM cache hits</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if not entries:
        st.markdown("""
        <div class="history-empty">
            No translation memory entries yet.<br>
            <span class="history-empty-sub">
                Run a translation to start building the cache.
            </span>
        </div>
        """, unsafe_allow_html=True)
        return

    # Sort by hit count descending
    sorted_entries = sorted(entries.items(), key=lambda x: -x[1].get("hit_count", 0))

    st.markdown('<div class="section-label">Cached Translations</div>', unsafe_allow_html=True)

    rows_html = "".join(
        f"""<div class="qg-row">
            <span class="qg-label" style="width:40%;word-break:break-word;">{de}</span>
            <span class="qg-value" style="width:40%;word-break:break-word;">{val.get("translation","")}</span>
            <span style="font-size:11px;color:#94A3B8;min-width:60px;text-align:right;">
                {val.get("hit_count",0)}× &nbsp;·&nbsp; {val.get("col_type","other")}
            </span>
        </div>"""
        for de, val in sorted_entries[:200]
    )
    st.markdown(f'<div class="qg">{rows_html}</div>', unsafe_allow_html=True)

    if len(sorted_entries) > 200:
        st.markdown(
            f'<div style="font-size:12px;color:#94A3B8;margin-top:8px;">'
            f'Showing top 200 of {len(sorted_entries):,} entries.</div>',
            unsafe_allow_html=True,
        )

    # ── Trados NL TM Corpus ───────────────────────────────────────────────────
    role = st.session_state.get("user_role", "")
    if role == "admin":
        st.markdown("---")
        st.markdown('<div class="section-label">Dutch Trados TM Corpus (NL)</div>', unsafe_allow_html=True)

        trados_count = db_nl_trados_count()
        col_a, col_b = st.columns([2, 1])
        with col_a:
            if trados_count > 0:
                engine = _get_nl_corpus_engine()
                st.success(
                    f"Trados TM loaded: **{trados_count:,} entries** in DB"
                    + (f", **{len(engine):,} active** in engine" if engine else "")
                )
            else:
                st.info("No Dutch Trados TM imported yet. Upload the Trados XLSX export below.")

        with col_b:
            uploaded_tm = st.file_uploader(
                "Import Trados TM Export (XLSX)",
                type=["xlsx"],
                key="trados_tm_upload",
                help="Upload Translation_Memory_Export_NL.xlsx from Trados",
            )
            if uploaded_tm and st.button("Import TM", key="import_trados_btn", type="primary"):
                with st.spinner(f"Importing {uploaded_tm.name}…"):
                    try:
                        import tempfile, os
                        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                            tmp.write(uploaded_tm.getvalue())
                            tmp_path = tmp.name
                        entries = parse_trados_xlsx(tmp_path)
                        os.unlink(tmp_path)
                        n_imported = db_nl_trados_import(entries)
                        _reload_nl_corpus_engine()
                        st.success(f"Imported {n_imported:,} TM entries. Engine reloaded.")
                        st.rerun()
                    except Exception as exc:
                        st.error(f"Import failed: {exc}")


# =============================================================================
# PAGE: LANGUAGE SELECTION
# =============================================================================

def language_selection_page():
    st.markdown("""
    <div style="text-align:center;padding:72px 0 48px;animation:fadeUp 0.4s ease;">
        <div style="display:inline-flex;align-items:center;gap:8px;font-size:10.5px;
                    font-weight:700;text-transform:uppercase;letter-spacing:0.14em;
                    color:#9BA8BE;margin-bottom:28px;">
            <div style="width:6px;height:6px;border-radius:50%;background:#12A150;"></div>
            Home24 AI Localization
        </div>
        <h1 style="font-size:34px;font-weight:800;color:#0F3D9E;letter-spacing:-0.04em;
                   line-height:1.15;margin:0 0 14px;">Choose your target language</h1>
        <p style="font-size:15px;color:#6B7A99;font-weight:400;margin:0;">
            Select the language for this translation session
        </p>
    </div>
    """, unsafe_allow_html=True)

    _, col, _ = st.columns([1, 2.2, 1])
    with col:
        col_fr, col_nl = st.columns(2, gap="large")

        with col_fr:
            st.markdown("""
            <div style="background:#FFFFFF;border:1px solid #E8ECF2;border-radius:20px;
                        padding:36px 28px 28px;text-align:center;
                        box-shadow:0 2px 14px rgba(15,61,158,0.05);margin-bottom:14px;
                        transition:border-color 0.2s,box-shadow 0.2s;">
                <div style="font-size:44px;margin-bottom:18px;line-height:1;">🇫🇷</div>
                <div style="font-size:17px;font-weight:700;color:#1A2035;margin-bottom:6px;">French</div>
                <div style="font-size:12px;color:#9BA8BE;font-weight:500;">
                    German → French · FR prefix
                </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Select French", use_container_width=True, key="pick_fr"):
                st.session_state["target_language"]  = "French"
                st.session_state["language_selected"] = True
                st.rerun()

        with col_nl:
            st.markdown("""
            <div style="background:#FFFFFF;border:1px solid #E8ECF2;border-radius:20px;
                        padding:36px 28px 28px;text-align:center;
                        box-shadow:0 2px 14px rgba(15,61,158,0.05);margin-bottom:14px;
                        transition:border-color 0.2s,box-shadow 0.2s;">
                <div style="font-size:44px;margin-bottom:18px;line-height:1;">🇳🇱</div>
                <div style="font-size:17px;font-weight:700;color:#1A2035;margin-bottom:6px;">Dutch</div>
                <div style="font-size:12px;color:#9BA8BE;font-weight:500;">
                    German → Dutch · NL prefix
                </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Select Dutch", use_container_width=True, key="pick_nl"):
                st.session_state["target_language"]  = "Dutch"
                st.session_state["language_selected"] = True
                st.rerun()

    render_footer()


# =============================================================================
# PAGE: ADMIN DASHBOARD
# =============================================================================

def admin_dashboard_page():
    render_page_header(
        "Admin Dashboard",
        "Platform overview — metadata only, no file contents",
    )

    stats = db_get_admin_stats()

    # ── KPIs ──
    st.markdown('<div class="section-label">Platform Overview</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-label">Total Jobs</div>
            <div class="kpi-value accent">{stats["total_jobs"]}</div>
            <div class="kpi-sub">All translation jobs</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total Logins</div>
            <div class="kpi-value">{stats["total_logins"]}</div>
            <div class="kpi-sub">Across all users</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">FR Jobs</div>
            <div class="kpi-value accent">{stats["fr_jobs"]}</div>
            <div class="kpi-sub">French translations</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">NL Jobs</div>
            <div class="kpi-value" style="color:#D97706;">{stats["nl_jobs"]}</div>
            <div class="kpi-sub">Dutch translations</div>
        </div>
    </div>
    <div class="kpi-row-3">
        <div class="kpi">
            <div class="kpi-label">Total Cells</div>
            <div class="kpi-value">{stats["total_cells"]:,}</div>
            <div class="kpi-sub">Translated across all jobs</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total API Cost</div>
            <div class="kpi-value warn">${stats["total_cost"]:.4f}</div>
            <div class="kpi-sub">GPT-4o-mini estimated</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Active Users</div>
            <div class="kpi-value success">{len(stats["jobs_by_user"])}</div>
            <div class="kpi-sub">Users with jobs</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Jobs by user ──
    st.markdown('<div class="section-label">Jobs by User</div>', unsafe_allow_html=True)
    if stats["jobs_by_user"]:
        user_rows = []
        for u in stats["jobs_by_user"]:
            cost = u.get("cost") or 0.0
            user_rows.append({
                "User":         u.get("user_email", ""),
                "Jobs":         u.get("job_count", 0),
                "Cells":        u.get("cells") or 0,
                "Est. Cost":    f"${cost:.4f}" if cost else "—",
            })
        st.dataframe(user_rows, use_container_width=True, hide_index=True)
    else:
        st.info("No jobs recorded yet.")

    # ── Recent jobs (metadata only) ──
    st.markdown('<div class="section-label">Recent Translation Jobs</div>', unsafe_allow_html=True)
    all_jobs = db_load_history()
    if all_jobs:
        job_rows = []
        for r in all_jobs[:50]:
            dt = r.get("datetime", "")[:16].replace("T", " ")
            c  = r.get("estimated_cost_usd")
            job_rows.append({
                "Date":         dt,
                "User":         r.get("user_email", ""),
                "Lang":         r.get("output_prefix", "FR"),
                "File":         r.get("original_filename", ""),
                "Cells":        r.get("cells_translated", 0),
                "Warnings":     r.get("total_warnings", 0),
                "Time (s)":     round(r.get("processing_time_seconds", 0), 1),
                "Est. Cost":    f"${c:.4f}" if c is not None else "—",
            })
        st.dataframe(job_rows, use_container_width=True, hide_index=True)
    else:
        st.info("No jobs yet.")

    # ── Login activity ──
    st.markdown('<div class="section-label">Login Activity</div>', unsafe_allow_html=True)
    logins = db_get_login_activity()
    if logins:
        login_rows = []
        for l in logins[:100]:
            login_rows.append({
                "Login Time":  l.get("login_time", "")[:16].replace("T", " "),
                "User":        l.get("user_email", ""),
                "Role":        l.get("role", ""),
                "Last Seen":   (l.get("last_seen") or "")[:16].replace("T", " "),
            })
        st.dataframe(login_rows, use_container_width=True, hide_index=True)
    else:
        st.info("No login records yet.")

    st.markdown("""
    <div class="alert alert-info" style="margin-top:24px;">
        <span class="alert-icon">ℹ</span>
        <span><strong>Privacy note:</strong> File contents and translated data are never
        stored in this dashboard. Only metadata (filename, cell counts, timestamps) is shown.</span>
    </div>
    """, unsafe_allow_html=True)


# =============================================================================
# MAIN
# =============================================================================

def main():
    st.set_page_config(
        page_title="Home24 AI Localization",
        page_icon="🌿",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False
    if "user_role" not in st.session_state:
        st.session_state["user_role"] = ""
    if "user_email" not in st.session_state:
        st.session_state["user_email"] = ""
    if "language_selected" not in st.session_state:
        st.session_state["language_selected"] = False
    if "target_language" not in st.session_state:
        st.session_state["target_language"] = "French"
    if "db_initialized" not in st.session_state:
        init_db(
            default_glossary=DEFAULT_GLOSSARY_TERMS,
            default_nl_glossary=DEFAULT_NL_GLOSSARY_TERMS,
        )
        st.session_state["db_initialized"] = True

    inject_custom_css()

    if not st.session_state["authenticated"]:
        login_page()
        return

    if not st.session_state["language_selected"]:
        language_selection_page()
        return

    page = render_sidebar()

    if page == "Admin Dashboard":
        admin_dashboard_page()
    elif page == "Translator":
        translator_page()
    elif page == "Translation History":
        history_page()
    elif page == "Analytics":
        analytics_page()
    elif page == "Glossary":
        glossary_page()
    elif page == "Translation Memory":
        translation_memory_page()
    elif page == "Report an Issue":
        report_issue_page()
    elif page == "Issue Reports":
        admin_issue_reports_page()
    elif page == "Jira Tickets":
        jira_tickets_page()

    render_footer()


if __name__ == "__main__":
    main()
