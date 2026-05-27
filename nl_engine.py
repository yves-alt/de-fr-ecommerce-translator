"""
Home24 Dutch (NL) Localization Engine — Trados TM-powered.

Pipeline priority:
  1. Exact TM match        → EXACT_TM_MATCH
  2. Segment TM match      → SEGMENT_TM_MATCH
  3. Fuzzy TM match        → FUZZY_TM_MATCH
  4. Token fuzzy match     → TOKEN_FUZZY_MATCH
  5. Terminology injection → AI_GUIDED (constrains GPT output)
  6. Dekor normalization   → applied pre/post-AI
  7. Post-colon lowercase  → applied post-AI
  8. Forbidden patterns    → applied post-AI
  9. Format normalization  → applied post-AI
 10. QA validation         → applied post-AI
"""

import re
from collections import Counter
from difflib import SequenceMatcher


# =============================================================================
# CANONICAL TERMINOLOGY  — derived from Trados TM corpus analysis
# =============================================================================

# Dekor / finish patterns — longest first so multi-word phrases replace before substrings
NL_DEKOR_MAP: list[tuple[str, str]] = [
    # Multi-word compound dekor forms first (longest match wins)
    ("Eiche Sägerau Dekor",    "grof gezaagde eikenlook"),   # TM: exact
    ("Eiche Artisan Dekor",    "Artisan eikenlook"),          # TM: exact
    ("Eiche Viking Dekor",     "Viking eikenhouten look"),    # TM: exact
    ("Eiche hell Dekor",       "lichte eikenhouten look"),    # TM: exact
    ("Eiche dunkel Dekor",     "donkere eikenhouten look"),
    ("Eiche natur Dekor",      "natuurlijke eikenhouten look"),
    ("Zinneiche Dekor",        "tin-eikenhouten look"),
    ("Kernbuche Dekor",        "kernbeukenhouten look"),      # TM: exact
    ("Marmor Weiß Dekor",      "witte marmerlook"),           # TM: exact
    ("Marmor Schwarz Dekor",   "zwarte marmerlook"),
    ("Marmor Grau Dekor",      "grijze marmerlook"),
    ("Nussbaum Dekor",         "notenlook"),                  # TM-derived; never "Notelaar Dekor"
    ("Beton Dekor",            "betonlook"),
    ("Artisan Dekor",          "Artisan look"),
    # Sägerau without "Eiche" prefix
    ("Sägerau Dekor",          "grof gezaagde eikenlook"),
    ("Sägerau",                "grof gezaagd eiken"),
    # Single-word forms — must come after compounds
    ("Eiche Dekor",            "eikenlook"),                  # TM pattern
    ("Eiche",                  "eiken"),                      # raw wood, not dekor
    ("Dekor",                  "look"),                       # generic fallback — last
]

# Color map — full German → Dutch canonical
NL_COLOR_MAP: list[tuple[str, str]] = [
    # Compound shades first
    ("dunkelgrau",     "donkergrijs"),
    ("hellgrau",       "lichtgrijs"),
    ("dunkelbraun",    "donkerbruin"),
    ("hellbraun",      "lichtbruin"),
    ("dunkelblau",     "donkerblauw"),
    ("hellblau",       "lichtblauw"),
    ("dunkelgrün",     "donkergroen"),
    ("hellgrün",       "lichtgroen"),
    ("Sandanthrazit",  "zandantraciet"),
    ("Sandschwarz",    "zandzwart"),
    ("Sandbeige",      "zandbeige"),
    # Descriptor-style colors (always -kleurig in NL)
    ("Anthrazit",      "antracietkleurig"),
    ("Graphit",        "grafietkleurig"),
    ("Silber",         "zilverkleurig"),
    ("Gold",           "goudkleurig"),
    ("Creme",          "crèmekleurig"),
    ("Kupfer",         "koper"),
    # Basic colors
    ("Schwarz",        "zwart"),
    ("Weiß",           "wit"),
    ("Grau",           "grijs"),
    ("Braun",          "bruin"),
    ("Blau",           "blauw"),
    ("Grün",           "groen"),
    ("Rot",            "rood"),
    ("Gelb",           "geel"),
    ("Rosa",           "roze"),
    ("Orange",         "oranje"),
    ("Beige",          "beige"),
    ("Türkis",         "turquoise"),
    ("Lila",           "lila"),
    ("Violett",        "paars"),
    ("Pink",           "pink"),
    ("Natur",          "natuurlijk"),
    ("Matt",           "mat"),
]

# Canonical furniture/product name map — TM-verified (longest first)
NL_FURNITURE_CANONICAL: list[tuple[str, str]] = [
    # Kitchen — most specific forms first (compound forms before simple)
    ("Kochfeldunterkast",       "kookplaatonderkast"),  # German-Dutch hybrid → correct NL
    ("Kochfeldunterschrank",    "kookplaatonderkast"),
    ("Singleküche",             "mini keuken"),         # TM: "Singleküche Kinneloa" → "Mini keuken Kinneloa"
    ("Pantryküche",             "Pantrykeuken"),
    ("Küchenleerblock",         "Keukenblok"),          # Never "Keukenleerblok"
    ("Kücheninsel",             "kookeiland"),          # TM: exact (lowercase — Dutch convention)
    ("Autarke Küchenzeile",     "Keukenblok"),          # TM: "Autarke Küchenzeile" → "Keukenblok"
    ("Küchenzeile",             "Keukenblok"),          # TM: exact
    ("Einbauküche",             "Inbouwkeuken"),
    ("Kochfeld",                "kookplaat"),           # TM: Kochfeld = kookplaat
    ("Kochplatte",              "kookplaat"),           # keramische Kochplatte
    ("Dunstabzugshaube",        "afzuigkap"),
    ("Arbeitsplatte",           "werkblad"),            # Home24 NL standard
    ("Korpus",                  "Body"),                # Home24 NL: Body not Romp
    # Living room seating
    ("Wohnlandschaft",          "Zithoek"),             # TM: "Wohnlandschaft Fardah" → "Zithoek Fardah"
    ("Ecksofa",                 "Hoekbank"),            # TM: "Ecksofa BANDO" → "Hoekbank BANDO"
    ("Polstergarnitur",         "bankstellen"),
    ("Loungesessel",            "loungefauteuil"),      # TM: "Loungesessel" → "loungefauteuil"
    ("Bigsofa",                 "XXL-bank"),
    ("Big-Sofa",                "XXL-bank"),
    ("XXL Sessel",              "XXL-fauteuil"),
    ("Schlafsessel",            "slaapfauteuil"),
    ("Drehsessel",              "draaifauteuil"),
    ("Freischwinger",           "sledestoel"),
    ("Armlehnenstuhl",          "stoel met armleuningen"),
    ("Cocktailsessel",          "fauteuil"),
    ("Loungesofa",              "Loungebank"),          # TM: "Loungesofa Valera" → "Loungebank Valera"
    # Ottomane — always lowercase in NL
    ("Ottomane",                "ottomane"),            # TM: "mit Ottomane" → "met ottomane"
    ("Ottoman",                 "ottomane"),
    # TV furniture
    ("TV-Lowboard",             "Tv-meubel"),
    ("TV Lowboard",             "Tv-meubel"),
    ("Fernsehsessel",           "tv-fauteuil"),
    # Sideboard — TM: "Sideboard Weallup" → "Dressoir Weallup"
    ("Sideboard",               "Dressoir"),
    # Kommode — TM: "Kommode Weallup" → "Kast Weallup"
    ("Kommode",                 "Kast"),
    # Lighting — compound LED forms first
    ("LED-Pendelleuchte",       "LED-hanglamp"),
    ("LED-Deckenleuchte",       "LED-plafondlamp"),
    ("LED-Wandleuchte",         "LED-wandlamp"),
    ("LED-Tischleuchte",        "LED-tafellamp"),
    ("LED-Stehleuchte",         "LED-staande lamp"),
    ("LED-Deckenleuchten",      "LED-plafondlamp"),
    ("Pendelleuchte",           "hanglamp"),
    ("Tischleuchte",            "tafellamp"),
    ("Deckenleuchten",          "plafondlamp"),
    ("Deckenleuchte",           "plafondlamp"),
    ("Wandleuchte",             "wandlamp"),
    ("Stehleuchte",             "staande lamp"),
    # Bathroom
    ("Badezimmerset",           "Badkamerset"),
    ("Badset",                  "Badkamerset"),
    ("Waschbeckenunterschrank", "wastafelonderkast"),
    ("Spiegelschrank",          "spiegelkast"),
    ("Spiegelpaneel",           "spiegelpaneel"),
    # Textiles / decor
    ("Dekokissen",              "sierkussen"),
    ("Raffgardinenrollo",       "vouwgordijn"),
    ("Wanddekoration",          "wanddecoratie"),
    ("Wanduhr",                 "wandklok"),
    # Tableware
    ("Tellerset",               "bordenset"),
    ("Trinkhalm-Set",           "set rietjes"),
    ("Dekanter",                "karaf"),
    ("Tellerstand",             "bordenstandaard"),     # plate rack; never "Tellerstand" untranslated
    ("Tablett",                 "dienblad"),            # tray; never "Tablett" untranslated
    # Storage
    ("Steckregal",              "opbergrek"),
    # Outdoor
    ("Gartensitzgruppe",        "tuinset"),
    ("Gartenessgruppe",         "Outdoor-diningset"),   # TM: exact pattern
    # Bathroom / cleaning
    ("Badewannenmatte",         "badkuipmat"),          # bathtub mat ≠ badmat
    ("Badematte",               "badmat"),              # regular bath mat
    ("Duschmatte",              "douchemat"),           # shower mat; singular, not "Douchematt"
    # Iron board
    ("Bügelbrettbezug",         "strijkplankhoes"),     # never "strijkplankbekleding"
    ("Bügelbrett-Bezug",        "strijkplankhoes"),
    # Multi-color — one standard, no variants
    ("Mehrfarbig",              "meerdere kleuren"),    # never Multikleur/Multikleurig
    ("Multicolor",              "meerdere kleuren"),
    # Fireplace
    ("Kaminset",                "haardset"),
    ("Kamingarnitur",           "haardgarnituur"),
    # Valet stand
    ("Herrendiener",            "herenknecht"),         # never "Herenstandaard"
    # Column labels (appear as "Label:" prefixes in structured data)
    ("Absetzung",               "afwerking"),
    ("Füllung",                 "vulling"),
    ("Abdeckplatte",            "afdekplaat"),
    ("Oberplatte",              "bovenblad"),
    ("Kleiderstange",           "kledingroede"),
    ("Innenstoff",              "binnenstof"),
    ("Belastbarkeit",           "draagkracht"),
    ("Lieferumfang",            "leveringsomvang"),
    ("Inklusive",               "inclusief"),
    # Castors on furniture — lowercase, never "wielen" (vehicle wheels)
    ("Rollen",                  "rollen"),
    # Iron material — German Eisen → Dutch IJzer (capital IJ digraph mandatory)
    ("Eisen",                   "IJzer"),
    # Variant/combo labels in product names
    ("Variante",                "variant"),             # TM: "Variante A" → "variant A"
    ("Kombi",                   "combi"),               # TM: "Kombi A" → "combi A"
]

# Forbidden outputs → canonical replacements (detected post-AI, applied as correction)
NL_FORBIDDEN_REPLACEMENTS: list[tuple[str, str]] = [
    # Kitchen — German-Dutch hybrids and wrong forms
    ("Keukenleerblok",          "Keukenblok"),          # "Leerblok" is not Dutch
    ("Enkele keuken",           "mini keuken"),         # spec forbidden
    ("Eenpersoonskeuken",       "mini keuken"),         # spec forbidden
    ("Keukeninsel",             "kookeiland"),          # German-Dutch hybrid
    ("Kookfeld",                "kookplaat"),           # German-Dutch hybrid
    ("Keuken eiland",           "kookeiland"),          # wrong space
    ("Kookkom",                 "kookplaat"),           # wrong word entirely
    # Kitchen structure — "Body" not "Romp" in Home24 NL
    ("Romp & werkblad",         "Body & werkblad"),
    ("Romp & Werkblad",         "Body & werkblad"),
    ("Body & Werkblad",         "Body & werkblad"),     # capital W wrong after &
    # Sofa/seating — wrong forms
    ("Zits bank",               "zitsbank"),            # space is wrong
    ("Woonlandschap",           "Zithoek"),             # TM says Zithoek
    # Finish / dekor wrong forms
    ("Zaagruw Decor",           "grof gezaagde eikenlook"),
    ("Zaagruw decor",           "grof gezaagde eikenlook"),
    ("Zagerau",                 "grof gezaagde eikenlook"),
    ("Notelaar look",           "notenlook"),
    ("Notelaar dekor",          "notenlook"),
    ("Notelaar",                "notenlook"),           # wrong for Nussbaum Dekor context
    # TV furniture wrong forms
    ("TV lowboard",             "Tv-meubel"),
    ("TV-lowboard",             "Tv-meubel"),
    ("televisie meubel",        "Tv-meubel"),
    ("televisie-meubel",        "Tv-meubel"),
    ("televisiemeubel",         "Tv-meubel"),
    ("Televisiemeubel",         "Tv-meubel"),
    # Dekor anti-patterns
    ("eiken decor",             "eikenlook"),
    ("Eiken decor",             "eikenlook"),
    ("decor eik",               "eikenlook"),
    ("Decor eik",               "eikenlook"),
    ("eikenhouten decor",       "eikenhouten look"),
    # Multi-color wrong variants — only "meerdere kleuren" is allowed
    ("Multikleur",              "meerdere kleuren"),
    ("Multikleurig",            "meerdere kleuren"),
    # Bath/cleaning wrong forms
    ("Douchematt",              "douchemat"),           # typo (double t)
    ("Douchematten",            "douchemat"),           # wrong plural for singular source
    ("Strijkplankbekleding",    "strijkplankhoes"),     # wrong term
    # Valet stand — wrong term
    ("Herenstandaard",          "herenknecht"),
    # Rattan/weaving wrong forms
    ("synthetisch rotan",       "kunststof vlechtwerk"),
    ("kunststof rotan",         "kunststof vlechtwerk"),
    # Gepoedercoat wrong form
    ("poedergecoat",            "gepoedercoat"),
    # Incorrect set composition wording
    ("bestaand uit",            "bestaande uit"),
    # Greeploos wrong forms
    ("zonder greep",            "greeploos"),
    ("zonder grepen",           "greeploos"),
    # GSP wrong form — TM says vaatwasserpaneel, not vaatwasserfront
    ("vaatwasserfront",         "vaatwasserpaneel"),
    ("vaatwasser front",        "vaatwasserpaneel"),
    ("vaatwasser paneel",       "vaatwasserpaneel"),
    # Furniture castors — "wielen" (vehicle wheels) is wrong; use "rollen"
    ("Wielen",                  "rollen"),
    # "IJs" (ice) as translation of "Eisen" (iron) — semantic error
    ("IJs",                     "IJzer"),
    # Note: "Ijzer" → "IJzer" is handled by normalize_ij_capitalization, not here,
    # because a case-insensitive replacement would falsely match the correct "IJzer".
    # Typos
    ("opbervolume",             "opbergvolume"),
    # Nonsense word — replace with safest option
    ("vloerweefsel",            "vloerglijders"),
]


# =============================================================================
# FORMAT NORMALIZATION RULES  — derived from TM analysis
# =============================================================================

_NL_PCT_SPACE    = re.compile(r'(\d)\s+%')
_NL_SLASH_SPACES = re.compile(
    r'(?<=[a-zA-Zéàèùâêîôûëïüçœæ\d])\s+/\s+(?=[a-zA-Zéàèùâêîôûëïüçœæ\d])',
    re.UNICODE,
)
_NL_DEKOR_RESIDUE = re.compile(r'\bDekor\b', re.UNICODE)
# TM shows BxHxD (no spaces): "B/H/T:" → "BxHxD:", "BHT" → "BxHxD"
_NL_BHT    = re.compile(
    r'\bBHT\b|\bBxHxT\b|B\s*x\s*H\s*x\s*T\b|B\s*/\s*H\s*/\s*T\b',
    re.IGNORECASE,
)
_NL_TEILIG = re.compile(r'(\d+(?:[.,]\d+)?)-?teilig\b', re.IGNORECASE)
_NL_TYP    = re.compile(r'\bTyp\b\s+([A-Z]\b)', re.UNICODE)
_NL_SITZER = re.compile(r'(\d+(?:[.,]5)?)-?[Ss]itzer\b')
_NL_SITZER_SOFA = re.compile(r'(\d+(?:[.,]5)?)-?[Ss]itzer\s+[Ss]ofa\b')
_NL_FLAMMIG    = re.compile(r'(\d+)-flammig\b', re.IGNORECASE)
_NL_PCT_UPPER  = re.compile(r'(\d+%)\s+([A-ZÄÖÜ][a-zäöüß]+)')
_NL_INCL       = re.compile(r'\binklusive\b', re.IGNORECASE)
_NL_INKL       = re.compile(r'\binkl\.\b',    re.IGNORECASE)
_NL_OHNE_DEKO  = re.compile(r'\bohne\s+Dekoration\b', re.IGNORECASE)
_NL_BESTEHEND  = re.compile(r'\bbestehend\s+aus\b', re.IGNORECASE)
_NL_SET_BEST   = re.compile(r'\bSet\s+bestehend\s+aus\b', re.IGNORECASE)
# German "mit" (with) → Dutch "met" — unambiguously German, safe word-boundary replace
_NL_MIT        = re.compile(r'\bmit\b', re.UNICODE)
# "3-Zits" with capital Z → "3-zits"
_NL_ZITS_UPPER = re.compile(r'(\d+(?:[.,]5)?)-Zits\b', re.UNICODE)
# "2er-Set" → "set van 2" (German quantity compound → Dutch)
_NL_ER_SET = re.compile(r'\b(\d+)er-[Ss]et\b')


def apply_nl_format_normalization(text: str) -> str:
    """Apply Home24 NL formatting conventions derived from TM corpus."""
    if not text:
        return text
    # Percentage: remove space before %
    text = _NL_PCT_SPACE.sub(r'\1%', text)
    # Lowercase material name after percentage
    text = _NL_PCT_UPPER.sub(lambda m: f"{m.group(1)} {m.group(2).lower()}", text)
    # Slash: no spaces around (color/material combos)
    text = _NL_SLASH_SPACES.sub('/', text)
    # BHT / BxHxT / B/H/T → BxHxD (no spaces — TM canonical form)
    text = _NL_BHT.sub('BxHxD', text)
    # Residual "Dekor" → "look"
    text = _NL_DEKOR_RESIDUE.sub('look', text)
    # "3-teilig" → "3-delig"
    text = _NL_TEILIG.sub(lambda m: m.group(1).replace(',', '.') + '-delig', text)
    # "Typ A" → "type A"
    text = _NL_TYP.sub(lambda m: f"type {m.group(1)}", text)
    # "3-Sitzer Sofa" → "3-zitsbank" (combined)
    text = _NL_SITZER_SOFA.sub(lambda m: f"{m.group(1).replace(',', '.')}-zitsbank", text)
    # "3-Sitzer" → "3-zits"
    text = _NL_SITZER.sub(lambda m: f"{m.group(1).replace(',', '.')}-zits", text)
    # "3-Zits" (wrong capital) → "3-zits"
    text = _NL_ZITS_UPPER.sub(lambda m: m.group(1) + '-zits', text)
    # "1-flammig" → "1-lichts"
    text = _NL_FLAMMIG.sub(lambda m: f"{m.group(1)}-lichts", text)
    # German connector words that slip through
    text = _NL_SET_BEST.sub('set bestaande uit', text)
    text = _NL_BESTEHEND.sub('bestaande uit', text)
    text = _NL_INCL.sub('inclusief', text)
    text = _NL_INKL.sub('incl.', text)
    text = _NL_OHNE_DEKO.sub('zonder decoratie', text)
    # German "mit" → Dutch "met" (standalone word, always safe)
    text = _NL_MIT.sub('met', text)
    # "2er-Set" → "set van 2"
    text = _NL_ER_SET.sub(lambda m: f"set van {m.group(1)}", text)
    # Collapse extra spaces
    text = re.sub(r'  +', ' ', text)
    return text.strip()


def apply_nl_dekor_patterns(text: str) -> str:
    """Replace German Dekor compounds with canonical NL equivalents."""
    for de_form, nl_form in NL_DEKOR_MAP:
        if de_form.lower() not in text.lower():
            continue
        pat = re.compile(r'(?<!\w)' + re.escape(de_form) + r'(?!\w)', re.IGNORECASE | re.UNICODE)
        text = pat.sub(nl_form, text)
    return text


def apply_nl_color_normalization(text: str) -> str:
    """Replace residual German color words with canonical NL forms."""
    for de_color, nl_color in NL_COLOR_MAP:
        if de_color.lower() not in text.lower():
            continue
        pat = re.compile(r'(?<![a-zäöüß])' + re.escape(de_color) + r'(?![a-zäöüß])', re.IGNORECASE | re.UNICODE)
        text = pat.sub(nl_color, text)
    return text


def apply_nl_furniture_canonical(text: str) -> str:
    """Apply canonical NL furniture name mappings derived from TM."""
    for de_term, nl_term in NL_FURNITURE_CANONICAL:
        if de_term.lower() not in text.lower():
            continue
        pat = re.compile(r'(?<!\w)' + re.escape(de_term) + r'(?!\w)', re.IGNORECASE | re.UNICODE)
        text = pat.sub(nl_term, text)
    return text


def apply_nl_forbidden_patterns(text: str) -> tuple[str, int]:
    """Replace forbidden NL patterns with canonical equivalents."""
    corrections = 0
    for wrong, right in NL_FORBIDDEN_REPLACEMENTS:
        if wrong.lower() not in text.lower():
            continue
        pat = re.compile(r'(?<!\w)' + re.escape(wrong) + r'(?!\w)', re.IGNORECASE | re.UNICODE)
        new = pat.sub(right, text)
        if new != text:
            corrections += 1
            text = new
    return text, corrections


# =============================================================================
# POST-COLON LOWERCASE  — Dutch rule: after ":", values are lowercase
# =============================================================================

# Applies to structured attribute columns; skipped for product name column.
_NL_POST_COLON_SAFE_CANONICALS = frozenset({
    'colorDetail', 'materialDetail', 'textileCompositionCover1',
    'qualityDetail', 'otherMeasurements', 'deliveryScope', 'variantName',
})

# Matches ": " followed by a word starting with uppercase then lowercase letters.
# This avoids matching ALL-CAPS abbreviations (GSP, LED, MDF, etc.).
_NL_COLON_UPPER_RE = re.compile(r'(:\s+)([A-ZÜÖÄ][a-züöäß])', re.UNICODE)


def apply_nl_post_colon_lowercase(text: str, canonical: str = "") -> str:
    """
    Lowercase the first letter of the word immediately following ': '
    in attribute columns (colors, materials, etc.).
    Skipped for the product name column and ALL-CAPS terms.
    """
    if canonical == 'name' or not text:
        return text
    # Apply to all structural attribute columns (or if canonical is unspecified)
    if not canonical or canonical in _NL_POST_COLON_SAFE_CANONICALS:
        return _NL_COLON_UPPER_RE.sub(lambda m: m.group(1) + m.group(2).lower(), text)
    return text


# =============================================================================
# IJ DIGRAPH + CASE NORMALIZATION  — Dutch-specific rules
# =============================================================================

# "Ij..." at word boundary → "IJ..." (Dutch digraph always capitalised as a unit)
_NL_IJ_RE = re.compile(r'\bIj(?=[a-zA-ZäöüèéêëÄÖÜ])', re.UNICODE)

# First capitalised word after /, & or <br> in attribute columns → lowercase.
# The <br> pattern uses a negative lookahead to skip label words (e.g. "Poten:").
_NL_SLASH_UPPER_RE = re.compile(r'(?<=/)\s*([A-ZÜÖÄ][a-züöäß])', re.UNICODE)
_NL_AMP_UPPER_RE   = re.compile(r'(?<=&)\s+([A-ZÜÖÄ][a-züöäß])', re.UNICODE)
_NL_BR_UPPER_RE    = re.compile(
    r'<br>([A-ZÜÖÄ][a-züöäß][a-zA-ZüöäßÜÖÄ]*)(?!\w)(?!\s*:)',
    re.UNICODE,
)


def normalize_ij_capitalization(text: str) -> str:
    """Fix Dutch IJ digraph: 'Ij...' → 'IJ...' (e.g. Ijzer → IJzer, Ijsland → IJsland)."""
    return _NL_IJ_RE.sub('IJ', text)


def normalize_dutch_case(text: str, canonical: str = "") -> str:
    """
    Lowercase the first word after '/', '&' and '<br>' in attribute columns.
    Skipped for product name column and ALL-CAPS abbreviations.
    """
    if canonical == 'name' or not text:
        return text
    if not canonical or canonical in _NL_POST_COLON_SAFE_CANONICALS:
        text = _NL_SLASH_UPPER_RE.sub(lambda m: m.group(1).lower(), text)
        text = _NL_AMP_UPPER_RE.sub(lambda m: ' ' + m.group(1).lower(), text)
        text = _NL_BR_UPPER_RE.sub(lambda m: '<br>' + m.group(1)[0].lower() + m.group(1)[1:], text)
    return text


# =============================================================================
# CELL SEGMENTATION  — split cells into TM-matchable segments
# =============================================================================

_NL_SEG_BR = re.compile(r'(<br>)', re.IGNORECASE)
# Group 1: "Label:", Group 2: whitespace after colon, Group 3: value
_NL_LABEL_COLON = re.compile(r'^([^:]{1,40}:)(\s+)(.+)$', re.DOTALL)


def segment_cell_for_tm_matching(text: str) -> list[tuple[str, str]]:
    """
    Split a cell value into TM-matchable segments.
    Returns list of (segment_text, following_separator).
    The caller reassembles by concatenating segment_text + separator for each tuple.

    The colon-space separator is stored as the separator of the label segment so it
    is always preserved even if TM translation strips trailing whitespace.

    Example:
      "Bekleding: Zwart<br>Poten: Zwart"
      → [("Bekleding:", " "), ("Zwart", "<br>"), ("Poten:", " "), ("Zwart", "")]
    """
    if not text:
        return []

    segments: list[tuple[str, str]] = []

    # Split on <br> first — capturing group keeps the <br> in the result list
    br_parts = _NL_SEG_BR.split(text)

    i = 0
    while i < len(br_parts):
        part = br_parts[i]
        # <br> tag itself — skip (it is used as separator below)
        if _NL_SEG_BR.fullmatch(part):
            i += 1
            continue

        # Find the <br> separator following this part (if any)
        br_sep = ""
        if i + 1 < len(br_parts) and _NL_SEG_BR.fullmatch(br_parts[i + 1]):
            br_sep = "<br>"

        if not part.strip():
            segments.append((part, br_sep))
            i += 2 if br_sep else 1
            continue

        # Try to split on "Label: value" pattern
        m = _NL_LABEL_COLON.match(part.strip())
        if m:
            label_colon = m.group(1)   # "Bekleding:" (no trailing space)
            colon_space = m.group(2)   # " " (the space(s) after the colon)
            value       = m.group(3)   # "Zwart"
            # Store colon_space as separator so it survives TM translation of the label
            segments.append((label_colon, colon_space))
            segments.append((value, br_sep))
        else:
            segments.append((part.strip(), br_sep))

        i += 2 if br_sep else 1

    return segments


def reconstruct_from_segments(segments: list[tuple[str, str]]) -> str:
    """Reassemble segments (each tuple: translated_text, separator) into a full cell value."""
    return "".join(seg + sep for seg, sep in segments)


# =============================================================================
# SAFE PRODUCT NAME SHORTENER
# =============================================================================

# Words after which cutting is dangerous — the sentence is left incomplete
_NL_DANGLING_WORDS = frozenset({
    'met', 'van', 'voor', 'en', 'of', 'een', 'de', 'het', 'op', 'uit',
    'keramische', 'schuin', 'schuine', 'recht', 'rechte', 'houten', 'eikenhouten',
    'geïntegreerde', 'verstelbare', 'uitschuifbare', 'inklapbare',
    'ingebouwde', 'kleine', 'grote', 'massieve', 'massief',
    'gecoate', 'gepoedercoate', 'geslepen', 'gebogen',
})

# Connectors that introduce optional phrases — remove from here to end to shorten
_NL_OPTIONAL_CONNECTOR = re.compile(
    r'\s+(?:met|van|voor)\s+.+$',
    re.IGNORECASE | re.UNICODE,
)


def _strip_dangling_ending(name: str) -> str:
    """Remove trailing dangling adjective/connector words from a product name."""
    words = name.split()
    while words and words[-1].lower() in _NL_DANGLING_WORDS:
        words.pop()
    return ' '.join(words)


def safe_shorten_product_name_nl(name: str, max_chars: int = 40) -> str:
    """
    Shorten a Dutch product name to ≤ max_chars without leaving dangling adjectives.

    Strategy:
    0. Always strip dangling endings first (even if name is short).
    1. Return if already within limit after stripping.
    2. Find last safe word boundary within max_chars.
    3. Remove "met/van/voor ..." optional phrase if step 2 failed.
    4. Hard-truncate at word boundary as last resort.
    Never output names ending with: met, van, keramische, schuine, houten, etc.
    """
    if not name:
        return name

    # Always clean dangling endings regardless of length
    name = _strip_dangling_ending(name)

    if len(name) <= max_chars:
        return name

    words = name.split()

    # Pass 1: find last safe cut at ≤ max_chars
    running = ""
    last_safe = ""
    for i, word in enumerate(words):
        candidate = (running + " " + word).strip()
        if len(candidate) > max_chars:
            # Cannot add this word — cut before it
            if running and words[i - 1].lower() not in _NL_DANGLING_WORDS:
                last_safe = running
            break
        running = candidate
    else:
        # All words fit
        return running.strip()

    if last_safe:
        return last_safe.strip()

    # Pass 2: remove "met/van/voor ..." optional connector phrase, try again
    shortened = _NL_OPTIONAL_CONNECTOR.sub('', name).strip()
    if shortened != name and len(shortened) <= max_chars:
        return shortened

    if len(shortened) > max_chars:
        return safe_shorten_product_name_nl(shortened, max_chars)

    # Pass 3: hard cut at word boundary, then remove any dangling tail
    truncated = name[:max_chars].rsplit(' ', 1)[0].strip()
    trunc_words = truncated.split()
    while trunc_words and trunc_words[-1].lower() in _NL_DANGLING_WORDS:
        trunc_words.pop()
    return ' '.join(trunc_words) if trunc_words else name[:max_chars].strip()


# =============================================================================
# POST-PROCESSING PIPELINE
# =============================================================================

def nl_post_process(text: str, canonical: str = "") -> str:
    """
    Full Dutch post-processing pipeline (zero API calls).
    Applied after AI translation:
      dekor → colors → furniture → forbidden → format → IJ digraph →
      post-colon lowercase → extended case normalization.
    """
    text = apply_nl_dekor_patterns(text)
    text = apply_nl_color_normalization(text)
    text = apply_nl_furniture_canonical(text)
    text = apply_nl_forbidden_patterns(text)[0]
    text = nl_naturalness_rewrite(text)
    text = apply_nl_format_normalization(text)
    text = normalize_ij_capitalization(text)
    if canonical:
        text = apply_nl_post_colon_lowercase(text, canonical)
        text = normalize_dutch_case(text, canonical)
    return text


# =============================================================================
# GERMAN RESIDUE DETECTION (NL-specific)
# =============================================================================

_NL_GERMAN_RESIDUE_RE = re.compile(
    r'\b(?:'
    r'Dekor|Schrank|Tisch|Sofa|Sessel|Leuchte|Lampe|Schublade|Kommode|Bett|Stuhl|'
    r'Regal|Sideboard|Highboard|Spanplatte|Arbeitsplatte|Geschirrspüler|Grifflos|'
    r'Hängeschrank|Oberschrank|Unterschrank|Waschtisch|Waschbecken|Badezimmer|'
    r'Badset|Spiegel|Füße|Füsse|Korpus|Breite|Höhe|Tiefe|Länge|Maße|'
    r'Baumwolle|Polyester|Wolle|Leinen|Viskose|'
    r'Schwarz|Weiß|Grau|Braun|Blau|Grün|Rot|Anthrazit|Graphit|Silber|'
    r'Kochfeld|Kochplatte|Dunstabzugshaube|Kaminset|Herrendiener|'
    r'Tablett|Tellerstand|Bügelbrettbezug|Duschmatte|Badewannenmatte|'
    r'teilig|Sitzer|flammig|Fernsehsessel|Typ\b'
    r')\b',
    re.UNICODE,
)


def detect_nl_german_residue(text: str) -> list[str]:
    """Return list of German words found in NL output."""
    return _NL_GERMAN_RESIDUE_RE.findall(text)


# =============================================================================
# NL QA ENGINE
# =============================================================================

# Product name endings that are always wrong (dangling adjective/connector)
_NL_DANGLING_ENDING_RE = re.compile(
    r'\s+(?:met|van|voor|keramische|schuine|rechte|houten|eikenhouten|'
    r'geïntegreerde|verstelbare|en|of)\s*$',
    re.IGNORECASE | re.UNICODE,
)

# Post-colon uppercase in attribute cells (a word starting uppercase after ": ")
_NL_QA_COLON_UPPER_RE = re.compile(r':\s+[A-ZÜÖÄ][a-züöäß]', re.UNICODE)

# "3-Zits" with wrong capital
_NL_QA_ZITS_CAP_RE = re.compile(r'\d+-Zits\b', re.UNICODE)


# Specific forbidden strings that get their own critical QA entry
_NL_QA_CRITICAL_FORBIDDEN: list[tuple[str, str]] = [
    ("Douchematt",          "douchemat"),
    ("Strijkplankbekleding","strijkplankhoes"),
    ("Keukeninsel",         "kookeiland"),
    ("Kookfeld",            "kookplaat"),
    ("Keuken eiland",       "kookeiland"),
    ("Kookkom",             "kookplaat"),
    ("Zaagruw Decor",       "grof gezaagde eikenlook"),
    ("Zagerau",             "grof gezaagde eikenlook"),
    ("opbervolume",         "opbergvolume"),
    ("vloerweefsel",        "vloerglijders"),
    ("Herenstandaard",      "herenknecht"),
    # IJ digraph enforcement — case-SENSITIVE (tuple[str, str, bool])
    ("Ijzer",               "IJzer",    True),  # wrong IJ casing; "IJzer" must not be flagged
    # Semantic error: "IJs" (ice) for source "Eisen" (iron) — case-sensitive
    ("IJs",                 "IJzer",    True),  # completely wrong translation
    # Furniture castors — "wielen" is for vehicles, not furniture
    ("Wielen",              "rollen"),
]

# Untranslated German terms that should always be caught
_NL_QA_UNTRANSLATED: list[tuple[str, str]] = [
    ("Tablett",             "dienblad"),
    ("Tellerstand",         "bordenstandaard"),
    ("Kaminset",            "haardset"),
]


def nl_qa_check(translation: str, canonical: str = "") -> list[dict]:
    """
    Run QA checks on a Dutch translation.
    Returns list of issues [{severity, category, message}].
    """
    issues = []
    t_lower = translation.lower()

    residues = detect_nl_german_residue(translation)
    if residues:
        issues.append({
            "severity": "High",
            "category": "German residue",
            "message":  f"German words in NL output: {', '.join(sorted(set(residues)))}",
        })

    if re.search(r'\bDekor\b', translation, re.UNICODE):
        issues.append({
            "severity": "High",
            "category": "Untranslated term",
            "message":  '"Dekor" not localized → should be "look"',
        })

    if re.search(r'\d\s+%', translation):
        issues.append({
            "severity": "Low",
            "category": "Formatting",
            "message":  'Space before % — should be "100%" not "100 %"',
        })

    if re.search(r'\bBHT\b', translation, re.IGNORECASE):
        issues.append({
            "severity": "Medium",
            "category": "Formatting",
            "message":  "BHT not converted to BxHxD",
        })

    if _NL_QA_ZITS_CAP_RE.search(translation):
        issues.append({
            "severity": "High",
            "category": "Capitalization",
            "message":  '"3-Zits" has wrong capital — should be "3-zits"',
        })

    if canonical != 'name' and _NL_QA_COLON_UPPER_RE.search(translation):
        issues.append({
            "severity": "Medium",
            "category": "Capitalization",
            "message":  "Uppercase word after colon — should be lowercase (color/material)",
        })

    if canonical == 'name' and _NL_DANGLING_ENDING_RE.search(translation):
        issues.append({
            "severity": "High",
            "category": "Product name",
            "message":  "Product name ends with a dangling connector/adjective",
        })

    # Critical forbidden patterns — each gets its own entry.
    # Tuple is (wrong, correct) or (wrong, correct, case_sensitive).
    for entry in _NL_QA_CRITICAL_FORBIDDEN:
        wrong, correct = entry[0], entry[1]
        case_sensitive = entry[2] if len(entry) > 2 else False
        matched = (wrong in translation) if case_sensitive else (wrong.lower() in t_lower)
        if matched:
            issues.append({
                "severity": "Critical",
                "category": "Forbidden pattern",
                "message":  f'"{wrong}" is forbidden → use "{correct}"',
            })

    # Untranslated German terms
    for term, correct in _NL_QA_UNTRANSLATED:
        if re.search(r'(?<!\w)' + re.escape(term) + r'(?!\w)', translation, re.IGNORECASE):
            issues.append({
                "severity": "High",
                "category": "Untranslated term",
                "message":  f'"{term}" not translated → should be "{correct}"',
            })

    # General forbidden replacements (any remaining)
    for wrong, _ in NL_FORBIDDEN_REPLACEMENTS:
        if wrong.lower() in t_lower:
            issues.append({
                "severity": "High",
                "category": "Forbidden pattern",
                "message":  f'Forbidden pattern detected: "{wrong}"',
            })
            break  # one report per cell is enough

    return issues


# =============================================================================
# HYBRID LANGUAGE DETECTOR
# =============================================================================

# Known German-Dutch hybrid words — partial translations AI is prone to generating
_NL_HYBRID_PATTERNS: list[tuple[str, str]] = [
    ("Keukeninsel",             "kookeiland"),          # NL prefix + DE root
    ("Kookfeld",                "kookplaat"),           # NL prefix + DE root
    ("Keukenleerblok",          "Keukenblok"),          # NL + non-existent hybrid
    ("Kookplaatunterkast",      "kookplaatonderkast"),  # NL prefix + DE suffix
    ("Kookplaatonderschrank",   "kookplaatonderkast"),  # NL + DE hybrid
    ("Enkeleküche",             "mini keuken"),         # NL adjective + DE noun
    ("Einzelküche",             "mini keuken"),         # alternative DE form
    ("Ijzer Dekor",             "grof gezaagde eikenlook"),  # wrong IJ context
]


def detect_hybrid_language_words(text: str) -> list[tuple[str, str]]:
    """
    Detect partial translations (German-Dutch hybrids) in NL output.
    Returns list of (wrong_form, canonical_nl) pairs found in the text.
    """
    found: list[tuple[str, str]] = []
    t_lower = text.lower()
    for wrong, correct in _NL_HYBRID_PATTERNS:
        if wrong.lower() in t_lower:
            found.append((wrong, correct))
    return found


# =============================================================================
# DUTCH NATURALNESS REWRITER  — fixes common unnatural AI phrasing
# =============================================================================

_NL_NAT_RULES: list[tuple[re.Pattern, str]] = [
    # "is gemaakt van" → "van" (AI verbosity in attribute descriptions)
    (re.compile(r'\bis gemaakt van\b', re.IGNORECASE | re.UNICODE), 'van'),
    # "die bestaat uit" → "bestaande uit" (grammatically more natural in NL)
    (re.compile(r'\bdie bestaat uit\b', re.IGNORECASE | re.UNICODE), 'bestaande uit'),
    # "wordt geleverd met" → "inclusief" (shorter, standard Home24 NL wording)
    (re.compile(r'\bwordt geleverd met\b', re.IGNORECASE | re.UNICODE), 'inclusief'),
]


def nl_naturalness_rewrite(text: str) -> str:
    """Fix common unnatural Dutch constructions produced by AI translation."""
    for pat, replacement in _NL_NAT_RULES:
        text = pat.sub(replacement, text)
    return text


# =============================================================================
# NORMALIZATION HELPER
# =============================================================================

def _normalize(text: str) -> str:
    """Normalize for TM lookup: strip, lowercase, collapse whitespace."""
    return re.sub(r'\s+', ' ', str(text).strip().lower())


# =============================================================================
# WORKBOOK CONSISTENCY MEMORY
# =============================================================================

class DutchWorkbookConsistencyMemory:
    """
    Tracks source → target choices within one workbook translation session.
    Ensures the same German source always receives the same Dutch output.
    Stores optional metadata (origin, confidence, hit count) per entry.
    """

    def __init__(self) -> None:
        self._memory: dict[str, str] = {}
        self._meta: dict[str, dict] = {}

    def get(self, source: str) -> str | None:
        key = _normalize(source)
        if key in self._memory:
            if key in self._meta:
                self._meta[key]["count"] += 1
        return self._memory.get(key)

    def put(self, source: str, target: str, *, origin: str = "", confidence: float = 1.0) -> None:
        """Record source→target. First write wins; subsequent calls increment hit count."""
        key = _normalize(source)
        if key not in self._memory:
            self._memory[key] = target
            self._meta[key] = {"origin": origin, "confidence": confidence, "count": 1}

    def get_meta(self, source: str) -> dict | None:
        """Return metadata dict for a source term, or None if not recorded."""
        return self._meta.get(_normalize(source))

    def get_or_default(self, source: str, fallback: str, *, origin: str = "", confidence: float = 1.0) -> str:
        existing = self.get(source)
        if existing is not None:
            return existing
        self.put(source, fallback, origin=origin, confidence=confidence)
        return fallback

    def clear(self) -> None:
        self._memory.clear()
        self._meta.clear()

    def __len__(self) -> int:
        return len(self._memory)


# =============================================================================
# HOME24 DUTCH CORPUS ENGINE
# =============================================================================

class Home24DutchCorpusEngine:
    """
    TM-powered Dutch localization engine.
    Loads ~38k Trados TM entries and provides:
      - Exact match
      - Segment-level match (split cell → match parts → reconstruct)
      - Fuzzy match (SequenceMatcher, wider candidate scan)
      - Terminology extraction for AI prompt guidance
    """

    def __init__(self, entries: list[dict]) -> None:
        """
        entries: list of {"source": str, "target": str, "usage_count": int}
        """
        self._exact: dict[str, str] = {}
        raw: list[tuple[str, str, int]] = []

        for e in entries:
            src = str(e.get("source", "")).strip()
            tgt = str(e.get("target", "")).strip()
            if not src or not tgt or src == tgt:
                continue
            key = _normalize(src)
            self._exact[key] = tgt
            raw.append((key, tgt, e.get("usage_count", 0)))

        # Sort: longer entries first, then by usage count (for fuzzy scan quality)
        raw.sort(key=lambda x: (-len(x[0]), -x[2]))
        self._entries: list[tuple[str, str]] = [(k, t) for k, t, _ in raw]

        # Short TM segments (≤5 tokens) → NL term for terminology injection
        self._term_index: dict[str, str] = {}
        term_candidates: dict[str, list[str]] = {}
        for src_norm, tgt, usage in raw:
            token_count = src_norm.count(' ') + 1
            if 1 <= token_count <= 5:
                term_candidates.setdefault(src_norm, []).append(tgt)
        for src_key, targets in term_candidates.items():
            self._term_index[src_key] = Counter(targets).most_common(1)[0][0]

    # ── Exact match ──────────────────────────────────────────────────────────

    def exact_match(self, source: str) -> str | None:
        """Return exact TM match or None."""
        return self._exact.get(_normalize(source))

    # ── Segment-level match ──────────────────────────────────────────────────

    def segment_tm_match(
        self,
        source: str,
        min_coverage: float = 0.6,
        seg_fuzzy_threshold: float = 0.82,
    ) -> str | None:
        """
        Split source into segments (by <br> and label: value patterns),
        TM-match each segment individually, then reconstruct.
        Returns assembled translation if ≥ min_coverage of non-empty segments matched.
        """
        segs = segment_cell_for_tm_matching(source)
        if not segs:
            return None

        translated: list[tuple[str, str]] = []
        match_count = 0
        total_content_segs = 0

        for seg_text, sep in segs:
            stripped = seg_text.strip()
            if not stripped:
                translated.append((seg_text, sep))
                continue

            total_content_segs += 1

            # Try exact TM match
            exact = self.exact_match(stripped)
            if exact:
                translated.append((exact, sep))
                match_count += 1
                continue

            # Try content rules (dekor, color, furniture — actual term translation)
            # Only count as a match if these content rules changed the text, NOT
            # if only format normalization changed it (format-only ≠ content translated).
            content_processed = apply_nl_dekor_patterns(stripped)
            content_processed = apply_nl_color_normalization(content_processed)
            content_processed = apply_nl_furniture_canonical(content_processed)
            if content_processed != stripped:
                # Apply full formatting on top
                content_processed = apply_nl_format_normalization(content_processed)
                translated.append((content_processed, sep))
                match_count += 1
                continue

            # Try fuzzy TM match on the segment
            fuzzy = self.fuzzy_match(stripped, threshold=seg_fuzzy_threshold, max_candidates=300)
            if fuzzy:
                translated.append((fuzzy[0], sep))
                match_count += 1
                continue

            # Apply format normalization even if no content match (don't count toward coverage)
            fmt_only = apply_nl_format_normalization(stripped)
            translated.append((fmt_only, sep))

        if total_content_segs == 0:
            return None
        if match_count / total_content_segs < min_coverage:
            return None

        result = reconstruct_from_segments(translated)
        result = nl_post_process(result.strip())

        # Reject if significant German residue remains — let GPT handle it instead
        residues = detect_nl_german_residue(result)
        # Allow up to 1 residue word for short segments (labels like "Stuhl"), but
        # reject if multiple distinct German content words survived in the output
        unique_residues = set(w.lower() for w in residues)
        if len(unique_residues) > 1:
            return None

        return result

    # ── Fuzzy match ──────────────────────────────────────────────────────────

    def fuzzy_match(
        self,
        source: str,
        threshold: float = 0.82,
        max_candidates: int = 1200,
    ) -> tuple[str, float] | None:
        """
        Best fuzzy TM match via sequence similarity.
        Returns (translation, score) or None.
        Increased max_candidates (was 600) for better recall on 38k entries.
        """
        src_norm = _normalize(source)
        slen = len(src_norm)
        if slen < 6 or slen > 220:
            return None

        best_score = 0.0
        best_tgt: str | None = None
        checked = 0

        for tm_src, tm_tgt in self._entries:
            # Quick length gate — skip if >50% length difference
            if abs(len(tm_src) - slen) > slen * 0.50 + 6:
                continue
            ratio = SequenceMatcher(None, src_norm, tm_src, autojunk=False).ratio()
            if ratio > best_score:
                best_score = ratio
                best_tgt = tm_tgt
            checked += 1
            if checked >= max_candidates:
                break

        if best_score >= threshold and best_tgt:
            return best_tgt, round(best_score, 3)
        return None

    # ── Terminology extraction ────────────────────────────────────────────────

    def extract_terminology(self, source: str, max_terms: int = 10) -> list[tuple[str, str]]:
        """
        Find TM short-segment matches that appear in source text.
        Returns list of (de_term, nl_term) pairs for AI prompt injection.
        """
        results: list[tuple[str, str]] = []
        src_lower = source.lower()
        seen_nl: set[str] = set()

        for tm_src, nl_tgt in self._term_index.items():
            if tm_src in src_lower and nl_tgt not in seen_nl:
                results.append((tm_src, nl_tgt))
                seen_nl.add(nl_tgt)
                if len(results) >= max_terms:
                    break

        return results

    # ── AI prompt guidance ────────────────────────────────────────────────────

    def get_prompt_guidance(self, source: str) -> str:
        """
        Return TM-derived guidance block for injection into the AI system prompt.
        Exact match → strong instruction. Fuzzy + terms → softer guidance.
        """
        exact = self.exact_match(source)
        if exact:
            return (
                f"\nTM EXACT MATCH — reproduce this wording exactly:\n"
                f"  DE: {source}\n"
                f"  NL: {exact}"
            )

        lines: list[str] = []

        fuzzy = self.fuzzy_match(source, threshold=0.85)
        if fuzzy:
            tgt, score = fuzzy
            lines.append(
                f"\nTM FUZZY MATCH ({score:.0%}) — strongly reuse this NL wording:\n"
                f"  NL reference: {tgt}"
            )

        terms = self.extract_terminology(source, max_terms=8)
        if terms:
            lines.append("\nTM terminology — use these exact NL equivalents:")
            for de, nl in terms:
                lines.append(f'  • "{de}" → "{nl}"')

        return "\n".join(lines)

    # ── Confidence classification ─────────────────────────────────────────────

    def confidence_level(self, source: str) -> str:
        if self.exact_match(source):
            return "EXACT_TM_MATCH"
        if self.fuzzy_match(source, threshold=0.85):
            return "FUZZY_TM_MATCH"
        if self.extract_terminology(source):
            return "AI_GUIDED"
        return "LOW_CONFIDENCE"

    def __len__(self) -> int:
        return len(self._entries)


# =============================================================================
# EXCEL IMPORT HELPER
# =============================================================================

def parse_trados_xlsx(path: str) -> list[dict]:
    """
    Parse Trados TM Export Excel file.
    Returns list of {"source": str, "target": str, "usage_count": int}.
    Skips identical source/target pairs (metadata fields, untranslated rows).
    Uses explicit iter_rows bounds to bypass stale <dimension> XML tags.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Translation Units"]
    except Exception as exc:
        raise ValueError(f"Cannot read TM file: {exc}") from exc

    # Bypass stale <dimension> tag — same fix as in scan_sheet()
    _dim_rows = getattr(ws, "max_row", None) or 0
    _dim_cols = getattr(ws, "max_column", None) or 0
    _iter_max_row = max(_dim_rows + 1, 50000)
    _iter_max_col = max(_dim_cols + 1, 10)

    entries: list[dict] = []
    header_skipped = False

    for row in ws.iter_rows(
        min_row=1,
        max_row=_iter_max_row,
        max_col=_iter_max_col,
        values_only=True,
    ):
        if not row or row[0] is None:
            continue
        if not header_skipped:
            header_skipped = True
            continue  # skip header row

        src   = row[1] if len(row) > 1 else None
        tgt   = row[2] if len(row) > 2 else None
        usage = row[7] if len(row) > 7 else 0

        src = str(src).strip() if src is not None else ""
        tgt = str(tgt).strip() if tgt is not None else ""
        if not src or not tgt or src == tgt:
            continue

        entries.append({
            "source":      src,
            "target":      tgt,
            "usage_count": int(usage) if usage and str(usage).isdigit() else 0,
        })

    wb.close()
    return entries
