"""
SQLite database backend for the Home24 Localization Platform.
Supports DE→FR and DE→NL translation memory, multilingual glossary,
user roles, and login activity tracking.
"""

import json
import sqlite3
import uuid
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

DB_PATH = Path(__file__).parent / "localization_platform.db"

# Base schema — uses CREATE TABLE IF NOT EXISTS so it is safe on existing DBs.
# Migration functions below handle upgrades from older schema versions.
_SCHEMA = """
PRAGMA journal_mode=WAL;

CREATE TABLE IF NOT EXISTS translation_jobs (
    id                        TEXT PRIMARY KEY,
    datetime                  TEXT NOT NULL,
    original_filename         TEXT,
    output_filename           TEXT,
    sheet_name                TEXT,
    source_language           TEXT DEFAULT 'German',
    target_language           TEXT DEFAULT 'French',
    output_prefix             TEXT DEFAULT 'FR',
    user_email                TEXT DEFAULT '',
    user_role                 TEXT DEFAULT '',
    cells_translated          INTEGER DEFAULT 0,
    cells_skipped             INTEGER DEFAULT 0,
    residue_corrections       INTEGER DEFAULT 0,
    unresolved_warnings       INTEGER DEFAULT 0,
    processing_time_seconds   REAL DEFAULT 0,
    processing_time_formatted TEXT,
    estimated_cost_usd        REAL,
    prompt_tokens             INTEGER,
    completion_tokens         INTEGER,
    tm_hits                   INTEGER DEFAULT 0,
    tm_misses                 INTEGER DEFAULT 0,
    batch_count               INTEGER DEFAULT 0,
    avg_batch_size            REAL DEFAULT 0,
    api_calls_reduced         INTEGER DEFAULT 0,
    glossary_hits             INTEGER DEFAULT 0,
    review_count              INTEGER DEFAULT 0,
    retry_count               INTEGER DEFAULT 0,
    critical_warnings         INTEGER DEFAULT 0,
    high_warnings             INTEGER DEFAULT 0,
    medium_warnings           INTEGER DEFAULT 0,
    low_warnings              INTEGER DEFAULT 0,
    total_warnings            INTEGER DEFAULT 0,
    quality_score             INTEGER DEFAULT 100,
    warning_categories        TEXT DEFAULT '{}',
    excel_exported            INTEGER DEFAULT 1,
    csv_exported              INTEGER DEFAULT 0,
    csv_removed_column        TEXT DEFAULT '',
    csv_delimiter             TEXT DEFAULT ';',
    csv_encoding              TEXT DEFAULT 'utf-8-sig',
    jira_ticket_key           TEXT DEFAULT '',
    jira_ticket_summary       TEXT DEFAULT '',
    jira_attachment_filename  TEXT DEFAULT '',
    jira_attachment_id        TEXT DEFAULT '',
    uploaded_to_jira          INTEGER DEFAULT 0,
    jira_upload_time          TEXT DEFAULT '',
    jira_comment_added        INTEGER DEFAULT 0,
    jira_transition_applied   TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS review_warnings (
    id              TEXT PRIMARY KEY,
    job_id          TEXT NOT NULL REFERENCES translation_jobs(id),
    severity        TEXT NOT NULL,
    category        TEXT,
    row             INTEGER,
    col_name        TEXT,
    source_text     TEXT,
    translated_text TEXT,
    reason          TEXT,
    suggested_fix   TEXT,
    timestamp       TEXT
);

CREATE TABLE IF NOT EXISTS translation_memory (
    tm_key      TEXT PRIMARY KEY,
    translation TEXT NOT NULL,
    col_type    TEXT,
    created_at  TEXT,
    hit_count   INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS glossary_terms (
    source_term     TEXT NOT NULL,
    target_language TEXT NOT NULL DEFAULT 'French',
    target_term     TEXT NOT NULL,
    hit_count       INTEGER DEFAULT 0,
    updated_at      TEXT,
    PRIMARY KEY (source_term, target_language)
);

CREATE TABLE IF NOT EXISTS user_logins (
    id            TEXT PRIMARY KEY,
    user_email    TEXT NOT NULL,
    role          TEXT NOT NULL,
    login_time    TEXT NOT NULL,
    last_seen     TEXT,
    session_id    TEXT,
    login_success INTEGER DEFAULT 1,
    created_at    TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS app_metrics (
    key   TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS glossary_suggestions (
    id              TEXT PRIMARY KEY,
    term            TEXT NOT NULL,
    target_language TEXT NOT NULL DEFAULT 'French',
    occurrences     INTEGER DEFAULT 1,
    example_context TEXT,
    status          TEXT DEFAULT 'pending',
    job_id          TEXT,
    created_at      TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS home24_corpus_entries (
    id                TEXT PRIMARY KEY,
    source_language   TEXT NOT NULL DEFAULT 'DE',
    target_language   TEXT NOT NULL DEFAULT 'FR',
    source_text       TEXT NOT NULL,
    translated_text   TEXT NOT NULL,
    category          TEXT,
    product_type      TEXT,
    terminology_tags  TEXT,
    confidence_score  REAL DEFAULT 1.0,
    approved          INTEGER DEFAULT 1,
    created_at        TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS forbidden_translation_patterns (
    id               TEXT PRIMARY KEY,
    forbidden_text   TEXT NOT NULL,
    replacement_text TEXT NOT NULL,
    category         TEXT,
    severity         TEXT DEFAULT 'HIGH',
    language         TEXT NOT NULL DEFAULT 'FR',
    notes            TEXT,
    auto_replace     INTEGER DEFAULT 1,
    created_at       TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS issue_reports (
    id                  TEXT PRIMARY KEY,
    user_email          TEXT NOT NULL DEFAULT '',
    user_role           TEXT NOT NULL DEFAULT '',
    category            TEXT NOT NULL,
    severity            TEXT NOT NULL DEFAULT 'Medium',
    target_language     TEXT NOT NULL DEFAULT 'Not language-related',
    filename            TEXT DEFAULT '',
    row_reference       TEXT DEFAULT '',
    column_reference    TEXT DEFAULT '',
    description         TEXT NOT NULL,
    expected_correction TEXT DEFAULT '',
    status              TEXT NOT NULL DEFAULT 'open',
    created_at          TEXT NOT NULL
);
"""


@contextmanager
def _db():
    conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# =============================================================================
# INIT & MIGRATION
# =============================================================================

def init_db(
    default_glossary: dict | None = None,
    default_nl_glossary: dict | None = None,
) -> None:
    with _db() as conn:
        conn.executescript(_SCHEMA)
    _ensure_v2_migration()
    _ensure_v3_migration()
    _ensure_v4_migration()
    _ensure_v5_migration()
    _ensure_v6_migration()
    _ensure_v7_migration()
    _ensure_v8_migration()
    _ensure_v9_migration()
    _ensure_v10_migration()
    _ensure_v11_migration()
    _ensure_v12_migration()
    _migrate_json_if_needed()
    if default_glossary:
        _seed_glossary_if_empty(default_glossary, "French")
    if default_nl_glossary:
        _seed_glossary_if_empty(default_nl_glossary, "Dutch")


def _get_schema_version() -> int:
    try:
        with _db() as conn:
            row = conn.execute(
                "SELECT value FROM app_metrics WHERE key='schema_version'"
            ).fetchone()
            return int(row[0]) if row else 0
    except Exception:
        return 0


def _set_schema_version(v: int) -> None:
    with _db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO app_metrics(key,value) VALUES('schema_version',?)",
            (str(v),),
        )


def _table_columns(conn, table: str) -> set[str]:
    return {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}


def _ensure_v2_migration() -> None:
    """Upgrade schema from v1 (FR-only) to v2 (multilingual)."""
    if _get_schema_version() >= 2:
        return
    try:
        _migrate_glossary_to_multilingual()
        _ensure_jobs_columns()
        _ensure_user_logins_table()
        _migrate_tm_keys_to_language_prefix()
        _set_schema_version(2)
    except Exception:
        pass


def _migrate_glossary_to_multilingual() -> None:
    """Convert old glossary_terms(de_term PK, fr_term) → new composite PK schema."""
    try:
        with _db() as conn:
            cols = _table_columns(conn, "glossary_terms")
            if "de_term" not in cols:
                return  # Already new schema or table doesn't exist
            conn.execute(
                "ALTER TABLE glossary_terms RENAME TO _glossary_terms_v1"
            )
            conn.execute("""
                CREATE TABLE IF NOT EXISTS glossary_terms (
                    source_term     TEXT NOT NULL,
                    target_language TEXT NOT NULL DEFAULT 'French',
                    target_term     TEXT NOT NULL,
                    hit_count       INTEGER DEFAULT 0,
                    updated_at      TEXT,
                    PRIMARY KEY (source_term, target_language)
                )
            """)
            conn.execute("""
                INSERT OR IGNORE INTO glossary_terms
                    (source_term, target_language, target_term, hit_count, updated_at)
                SELECT de_term, 'French', fr_term, hit_count, updated_at
                FROM _glossary_terms_v1
            """)
            conn.execute("DROP TABLE _glossary_terms_v1")
    except Exception:
        pass


def _ensure_jobs_columns() -> None:
    """Add v2 columns to translation_jobs if they are missing."""
    new_cols = [
        ("output_prefix", "TEXT DEFAULT 'FR'"),
        ("user_email",    "TEXT DEFAULT ''"),
        ("user_role",     "TEXT DEFAULT ''"),
        ("excel_exported",     "INTEGER DEFAULT 1"),
        ("csv_exported",       "INTEGER DEFAULT 0"),
        ("csv_removed_column", "TEXT DEFAULT ''"),
        ("csv_delimiter",      "TEXT DEFAULT ';'"),
        ("csv_encoding",       "TEXT DEFAULT 'utf-8-sig'"),
    ]
    try:
        with _db() as conn:
            existing = _table_columns(conn, "translation_jobs")
            for col_name, col_def in new_cols:
                if col_name not in existing:
                    conn.execute(
                        f"ALTER TABLE translation_jobs ADD COLUMN {col_name} {col_def}"
                    )
    except Exception:
        pass


def _ensure_user_logins_table() -> None:
    try:
        with _db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS user_logins (
                    id            TEXT PRIMARY KEY,
                    user_email    TEXT NOT NULL,
                    role          TEXT NOT NULL,
                    login_time    TEXT NOT NULL,
                    last_seen     TEXT,
                    session_id    TEXT,
                    login_success INTEGER DEFAULT 1,
                    created_at    TEXT NOT NULL
                )
            """)
    except Exception:
        pass


def _migrate_tm_keys_to_language_prefix() -> None:
    """Prefix all existing TM keys with 'fr:' (treat legacy entries as French)."""
    try:
        with _db() as conn:
            conn.execute("""
                UPDATE translation_memory
                SET tm_key = 'fr:' || tm_key
                WHERE tm_key NOT LIKE 'fr:%'
                  AND tm_key NOT LIKE 'nl:%'
            """)
    except Exception:
        pass


def _ensure_v5_migration() -> None:
    """Add consistency and QA stats columns to translation_jobs."""
    if _get_schema_version() >= 5:
        return
    try:
        new_cols = [
            ("consistency_corrections", "INTEGER DEFAULT 0"),
            ("consistency_detected",    "INTEGER DEFAULT 0"),
            ("terms_harmonized",        "INTEGER DEFAULT 0"),
            ("qa_issues_found",         "INTEGER DEFAULT 0"),
        ]
        with _db() as conn:
            existing = _table_columns(conn, "translation_jobs")
            for col_name, col_def in new_cols:
                if col_name not in existing:
                    conn.execute(
                        f"ALTER TABLE translation_jobs ADD COLUMN {col_name} {col_def}"
                    )
        _set_schema_version(5)
    except Exception:
        pass


def _ensure_v4_migration() -> None:
    """Add auto_learned_terms and furniture_term_fixes columns to translation_jobs."""
    if _get_schema_version() >= 4:
        return
    try:
        new_cols = [
            ("auto_learned_terms",   "INTEGER DEFAULT 0"),
            ("furniture_term_fixes", "INTEGER DEFAULT 0"),
        ]
        with _db() as conn:
            existing = _table_columns(conn, "translation_jobs")
            for col_name, col_def in new_cols:
                if col_name not in existing:
                    conn.execute(
                        f"ALTER TABLE translation_jobs ADD COLUMN {col_name} {col_def}"
                    )
        _set_schema_version(4)
    except Exception:
        pass


def _ensure_v3_migration() -> None:
    """Add Intelligence Engine columns to translation_jobs and glossary_suggestions table."""
    if _get_schema_version() >= 3:
        return
    try:
        new_cols = [
            ("semantic_tm_hits",       "INTEGER DEFAULT 0"),
            ("duplicate_groups",       "INTEGER DEFAULT 0"),
            ("duplicate_cells_saved",  "INTEGER DEFAULT 0"),
            ("glossary_only_count",    "INTEGER DEFAULT 0"),
            ("pattern_count",          "INTEGER DEFAULT 0"),
            ("gpt_calls_avoided",      "INTEGER DEFAULT 0"),
            ("detected_product_type",  "TEXT DEFAULT 'generic'"),
        ]
        with _db() as conn:
            existing = _table_columns(conn, "translation_jobs")
            for col_name, col_def in new_cols:
                if col_name not in existing:
                    conn.execute(
                        f"ALTER TABLE translation_jobs ADD COLUMN {col_name} {col_def}"
                    )
            # Ensure glossary_suggestions table exists (may not have been created on old DBs)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS glossary_suggestions (
                    id              TEXT PRIMARY KEY,
                    term            TEXT NOT NULL,
                    target_language TEXT NOT NULL DEFAULT 'French',
                    occurrences     INTEGER DEFAULT 1,
                    example_context TEXT,
                    status          TEXT DEFAULT 'pending',
                    job_id          TEXT,
                    created_at      TEXT NOT NULL
                )
            """)
        _set_schema_version(3)
    except Exception:
        pass


# =============================================================================
# HISTORY
# =============================================================================

def db_load_history() -> list:
    """Return all jobs — for admin use."""
    return _load_jobs_where()


def db_load_history_for_user(user_email: str, role: str) -> list:
    """Return jobs scoped to the user (admin sees all)."""
    if role == "admin":
        return _load_jobs_where()
    return _load_jobs_where("user_email = ?", (user_email.strip().lower(),))


def _load_jobs_where(condition: str = "", params: tuple = ()) -> list:
    try:
        sql = "SELECT * FROM translation_jobs"
        if condition:
            sql += f" WHERE {condition}"
        sql += " ORDER BY datetime DESC"
        with _db() as conn:
            rows = conn.execute(sql, params).fetchall()
        result = []
        for row in rows:
            rec = dict(row)
            if rec.get("warning_categories"):
                try:
                    rec["warning_categories"] = json.loads(rec["warning_categories"])
                except (json.JSONDecodeError, TypeError):
                    rec["warning_categories"] = {}
            result.append(rec)
        return result
    except Exception:
        return []


def db_save_history_record(record: dict) -> None:
    wc = record.get("warning_categories", {})
    wc_json = json.dumps(wc if isinstance(wc, dict) else {}, ensure_ascii=False)

    # Derive output_prefix from target_language when not explicitly set
    target_lang = record.get("target_language", "French")
    output_prefix = record.get("output_prefix", "NL" if target_lang == "Dutch" else "FR")

    params = (
        record.get("id") or str(uuid.uuid4()),
        record.get("datetime", ""),
        record.get("original_filename"),
        record.get("output_filename"),
        record.get("sheet_name"),
        record.get("source_language", "German"),
        target_lang,
        output_prefix,
        record.get("user_email", ""),
        record.get("user_role", ""),
        record.get("cells_translated", 0),
        record.get("cells_skipped", 0),
        record.get("residue_corrections", 0),
        record.get("unresolved_warnings", 0),
        record.get("processing_time_seconds", 0),
        record.get("processing_time_formatted"),
        record.get("estimated_cost_usd"),
        record.get("prompt_tokens"),
        record.get("completion_tokens"),
        record.get("tm_hits", 0),
        record.get("tm_misses", 0),
        record.get("batch_count", 0),
        record.get("avg_batch_size", 0.0),
        record.get("api_calls_reduced", 0),
        record.get("glossary_hits", 0),
        record.get("review_count", 0),
        record.get("retry_count", 0),
        record.get("critical_warnings", 0),
        record.get("high_warnings", 0),
        record.get("medium_warnings", 0),
        record.get("low_warnings", 0),
        record.get("total_warnings", 0),
        record.get("quality_score", 100),
        wc_json,
        record.get("excel_exported", 1),
        record.get("csv_exported", 0),
        record.get("csv_removed_column", ""),
        record.get("csv_delimiter", ";"),
        record.get("csv_encoding", "utf-8-sig"),
        # Intelligence Engine fields (v3)
        record.get("semantic_tm_hits", 0),
        record.get("duplicate_groups", 0),
        record.get("duplicate_cells_saved", 0),
        record.get("glossary_only_count", 0),
        record.get("pattern_count", 0),
        record.get("gpt_calls_avoided", 0),
        record.get("detected_product_type", "generic"),
        record.get("dutch_contamination_fixes", 0),
    )

    try:
        with _db() as conn:
            conn.execute(
                """
                INSERT OR IGNORE INTO translation_jobs (
                    id, datetime, original_filename, output_filename,
                    sheet_name, source_language, target_language, output_prefix,
                    user_email, user_role,
                    cells_translated, cells_skipped, residue_corrections,
                    unresolved_warnings, processing_time_seconds,
                    processing_time_formatted, estimated_cost_usd,
                    prompt_tokens, completion_tokens,
                    tm_hits, tm_misses, batch_count, avg_batch_size,
                    api_calls_reduced, glossary_hits, review_count, retry_count,
                    critical_warnings, high_warnings, medium_warnings, low_warnings,
                    total_warnings, quality_score, warning_categories,
                    excel_exported, csv_exported, csv_removed_column,
                    csv_delimiter, csv_encoding,
                    semantic_tm_hits, duplicate_groups, duplicate_cells_saved,
                    glossary_only_count, pattern_count, gpt_calls_avoided,
                    detected_product_type, dutch_contamination_fixes
                ) VALUES (
                    ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                    ?,?,?,?,?,?,?,?
                )
                """,
                params,
            )
    except Exception:
        pass


def db_save_warnings(job_id: str, warnings: list) -> None:
    if not warnings:
        return
    rows = []
    for w in warnings:
        rows.append((
            str(uuid.uuid4()),
            job_id,
            w.get("severity", "Low"),
            w.get("category"),
            w.get("row"),
            w.get("column"),
            w.get("source_text"),
            w.get("translated_text"),
            w.get("reason"),
            w.get("suggested_fix"),
            w.get("timestamp"),
        ))
    try:
        with _db() as conn:
            conn.executemany(
                """
                INSERT OR IGNORE INTO review_warnings
                    (id, job_id, severity, category, row, col_name,
                     source_text, translated_text, reason, suggested_fix, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows,
            )
    except Exception:
        pass


# =============================================================================
# TRANSLATION MEMORY
# =============================================================================

def db_load_translation_memory() -> dict:
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT tm_key, translation, col_type, created_at, hit_count "
                "FROM translation_memory"
            ).fetchall()
            metrics = conn.execute(
                "SELECT key, value FROM app_metrics WHERE key IN "
                "('tm_total_hits','tm_total_misses','tm_api_calls_saved')"
            ).fetchall()

        entries = {}
        for row in rows:
            entries[row["tm_key"]] = {
                "translation": row["translation"],
                "col_type":    row["col_type"],
                "created_at":  row["created_at"],
                "hit_count":   row["hit_count"],
            }

        m = {r["key"]: int(r["value"]) for r in metrics}
        return {
            "entries": entries,
            "global_stats": {
                "total_hits":            m.get("tm_total_hits", 0),
                "total_misses":          m.get("tm_total_misses", 0),
                "total_api_calls_saved": m.get("tm_api_calls_saved", 0),
            },
        }
    except Exception:
        return {
            "entries": {},
            "global_stats": {
                "total_hits":            0,
                "total_misses":          0,
                "total_api_calls_saved": 0,
            },
        }


def db_save_translation_memory(tm: dict) -> None:
    entries = tm.get("entries", {})
    gs = tm.get("global_stats", {})

    rows = []
    now = datetime.now().isoformat(timespec="seconds")
    for key, val in entries.items():
        rows.append((
            key,
            val.get("translation", ""),
            val.get("col_type", "other"),
            val.get("created_at", now),
            val.get("hit_count", 0),
        ))

    try:
        with _db() as conn:
            conn.executemany(
                """
                INSERT INTO translation_memory (tm_key, translation, col_type, created_at, hit_count)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(tm_key) DO UPDATE SET
                    translation = excluded.translation,
                    hit_count   = excluded.hit_count
                """,
                rows,
            )
            for k, v in [
                ("tm_total_hits",       gs.get("total_hits", 0)),
                ("tm_total_misses",     gs.get("total_misses", 0)),
                ("tm_api_calls_saved",  gs.get("total_api_calls_saved", 0)),
            ]:
                conn.execute(
                    "INSERT INTO app_metrics(key, value) VALUES(?,?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (k, str(v)),
                )
    except Exception:
        pass


# =============================================================================
# GLOSSARY — multilingual
# =============================================================================

def db_load_glossary(target_language: str = "French") -> dict | None:
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT source_term, target_term, hit_count "
                "FROM glossary_terms "
                "WHERE target_language = ? "
                "  AND COALESCE(active, 1) = 1 "
                "  AND COALESCE(is_deleted, 0) = 0 "
                "ORDER BY rowid",
                (target_language,),
            ).fetchall()

        if not rows:
            return None

        terms = {row["source_term"]: row["target_term"] for row in rows}
        total_hits = sum(row["hit_count"] for row in rows)
        term_counts = {
            row["source_term"]: row["hit_count"]
            for row in rows
            if row["hit_count"] > 0
        }
        return {
            "terms":           terms,
            "target_language": target_language,
            "stats":           {"total_hits": total_hits, "term_counts": term_counts},
        }
    except Exception:
        return None


def db_save_glossary(glossary: dict, target_language: str = "French") -> None:
    terms = glossary.get("terms", {})
    stats = glossary.get("stats", {})
    term_counts = stats.get("term_counts", {})
    now = datetime.now().isoformat(timespec="seconds")

    rows = []
    for source, target in terms.items():
        rows.append((
            source,
            target_language,
            target,
            term_counts.get(source, 0),
            now,
        ))

    try:
        with _db() as conn:
            conn.executemany(
                """
                INSERT INTO glossary_terms
                    (source_term, target_language, target_term, hit_count, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(source_term, target_language) DO UPDATE SET
                    target_term = excluded.target_term,
                    hit_count   = excluded.hit_count,
                    updated_at  = excluded.updated_at
                """,
                rows,
            )
    except Exception:
        pass


# =============================================================================
# LOGIN ACTIVITY
# =============================================================================

def db_log_login(user_email: str, role: str, session_id: str) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                """
                INSERT INTO user_logins
                    (id, user_email, role, login_time, last_seen, session_id, login_success, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 1, ?)
                """,
                (str(uuid.uuid4()), user_email.strip().lower(), role, now, now, session_id, now),
            )
    except Exception:
        pass


def db_update_last_seen(user_email: str, session_id: str) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                "UPDATE user_logins SET last_seen = ? WHERE session_id = ?",
                (now, session_id),
            )
    except Exception:
        pass


def db_get_login_activity() -> list:
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT * FROM user_logins ORDER BY login_time DESC"
            ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


# =============================================================================
# ADMIN STATS
# =============================================================================

def db_get_admin_stats() -> dict:
    try:
        with _db() as conn:
            total_jobs = conn.execute("SELECT COUNT(*) FROM translation_jobs").fetchone()[0]
            total_logins = conn.execute("SELECT COUNT(*) FROM user_logins").fetchone()[0]
            total_cost = conn.execute(
                "SELECT SUM(estimated_cost_usd) FROM translation_jobs"
            ).fetchone()[0]
            total_cells = conn.execute(
                "SELECT SUM(cells_translated) FROM translation_jobs"
            ).fetchone()[0]
            fr_jobs = conn.execute(
                "SELECT COUNT(*) FROM translation_jobs WHERE target_language='French'"
            ).fetchone()[0]
            nl_jobs = conn.execute(
                "SELECT COUNT(*) FROM translation_jobs WHERE target_language='Dutch'"
            ).fetchone()[0]
            jobs_by_user = conn.execute(
                """
                SELECT user_email, COUNT(*) as job_count,
                       SUM(cells_translated) as cells,
                       SUM(estimated_cost_usd) as cost
                FROM translation_jobs
                GROUP BY user_email
                ORDER BY job_count DESC
                """
            ).fetchall()
        return {
            "total_jobs":    total_jobs,
            "total_logins":  total_logins,
            "total_cost":    total_cost or 0.0,
            "total_cells":   total_cells or 0,
            "fr_jobs":       fr_jobs,
            "nl_jobs":       nl_jobs,
            "jobs_by_user":  [dict(r) for r in jobs_by_user],
        }
    except Exception:
        return {
            "total_jobs": 0, "total_logins": 0, "total_cost": 0.0,
            "total_cells": 0, "fr_jobs": 0, "nl_jobs": 0, "jobs_by_user": [],
        }


# =============================================================================
# STATUS
# =============================================================================

def db_get_status() -> dict:
    try:
        with _db() as conn:
            jobs     = conn.execute("SELECT COUNT(*) FROM translation_jobs").fetchone()[0]
            tm       = conn.execute("SELECT COUNT(*) FROM translation_memory").fetchone()[0]
            glossary = conn.execute("SELECT COUNT(*) FROM glossary_terms").fetchone()[0]
        return {
            "connected":      True,
            "jobs":           jobs,
            "tm_entries":     tm,
            "glossary_terms": glossary,
        }
    except Exception:
        return {"connected": False, "jobs": 0, "tm_entries": 0, "glossary_terms": 0}


# =============================================================================
# V6 MIGRATION — Corpus & Forbidden Patterns
# =============================================================================

def _ensure_v6_migration() -> None:
    """Add home24_corpus_entries and forbidden_translation_patterns tables."""
    if _get_schema_version() >= 6:
        return
    try:
        with _db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS home24_corpus_entries (
                    id                TEXT PRIMARY KEY,
                    source_language   TEXT NOT NULL DEFAULT 'DE',
                    target_language   TEXT NOT NULL DEFAULT 'FR',
                    source_text       TEXT NOT NULL,
                    translated_text   TEXT NOT NULL,
                    category          TEXT,
                    product_type      TEXT,
                    terminology_tags  TEXT,
                    confidence_score  REAL DEFAULT 1.0,
                    approved          INTEGER DEFAULT 1,
                    created_at        TEXT NOT NULL
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS forbidden_translation_patterns (
                    id               TEXT PRIMARY KEY,
                    forbidden_text   TEXT NOT NULL,
                    replacement_text TEXT NOT NULL,
                    category         TEXT,
                    severity         TEXT DEFAULT 'HIGH',
                    language         TEXT NOT NULL DEFAULT 'FR',
                    notes            TEXT,
                    auto_replace     INTEGER DEFAULT 1,
                    created_at       TEXT NOT NULL
                )
            """)
        _seed_forbidden_patterns_if_empty()
        _seed_corpus_if_empty()
        _set_schema_version(6)
    except Exception:
        pass


_INITIAL_FORBIDDEN_FR = [
    ("cuisine autonome",          "cuisine équipée",           "kitchen",   "CRITICAL", "Literal translation of 'autark'"),
    ("chaiselongue",              "chaise longue",             "seating",   "CRITICAL", "Must always be two words"),
    ("assiettes à manger",        "assiettes",                 "tableware", "CRITICAL", "Not natural French e-commerce"),
    ("assiette à manger",         "assiette",                  "tableware", "CRITICAL", "Not natural French e-commerce"),
    ("boue",                      "argile",                    "colors",    "HIGH",     "Unattractive color name in product copy"),
    ("table sans revêtement",     "table",                     "tables",    "HIGH",     "Unclear literal phrase from 'ohne Bezug'"),
    ("revêtu de poudre",          "thermolaqué",               "materials", "CRITICAL", "Wrong powder coating term"),
    ("revêtu par poudre",         "thermolaqué",               "materials", "CRITICAL", "Wrong powder coating term"),
    ("revêtement de poudre",      "thermolaqué",               "materials", "CRITICAL", "Wrong powder coating term"),
    ("laqué par poudre",          "thermolaqué",               "materials", "CRITICAL", "Wrong powder coating term"),
    ("traitement en poudre",      "thermolaqué",               "materials", "HIGH",     "Wrong powder coating term"),
    ("thermopoudré",              "thermolaqué",               "materials", "CRITICAL", "Incorrect portmanteau — not used in FR"),
    ("set consistant de",         "ensemble composé de",       "sets",      "HIGH",     "Literal German sentence structure"),
    ("ensemble consistant de",    "ensemble composé de",       "sets",      "HIGH",     "Literal German sentence structure"),
    ("paillasson en coco",        "couche de coco",            "mattress",  "CRITICAL", "Wrong translation of Kokosmatte"),
    ("paillasson de coco",        "couche de coco",            "mattress",  "CRITICAL", "Wrong translation of Kokosmatte"),
    ("nattes en coco",            "couche de coco",            "mattress",  "HIGH",     "Wrong translation of Kokosmatte"),
    ("décoration non comprise",   "accessoires non inclus",    "delivery",  "HIGH",     "Literal translation"),
    ("chêne artisan décor",       "décor chêne artisan",       "materials", "HIGH",     "Wrong word order for home24.fr"),
    ("artisan chêne décor",       "décor chêne artisan",       "materials", "HIGH",     "Wrong word order for home24.fr"),
    ("chêne décor artisan",       "décor chêne artisan",       "materials", "HIGH",     "Wrong word order for home24.fr"),
    ("table à manger extensible", "table extensible",          "tables",    "MEDIUM",   "Prefer shorter home24.fr form"),
    ("cuisine autonome complète", "cuisine équipée complète",  "kitchen",   "CRITICAL", "Literal translation of 'autark'"),
    ("pouf de pieds",             "repose-pieds",              "seating",   "MEDIUM",   "Wrong term in lounge context"),
    ("support de pieds",          "repose-pieds",              "seating",   "MEDIUM",   "Wrong term in lounge context"),
    ("housse amovible",           "revêtement amovible",       "materials", "MEDIUM",   "Use 'revêtement' for structured covers"),
    ("sans revêtement",           "sans finition",             "materials", "LOW",      "Clarify meaning before using"),
    ("meuble de télévision",      "meuble TV",                 "furniture", "MEDIUM",   "Prefer commercial short form"),
    ("armoire de garde-robe",     "armoire",                   "furniture", "LOW",      "Redundant in French"),
    ("structure en métal",        "structure métal",           "materials", "LOW",      "home24.fr drops 'en' in attribute lists"),
]


def _seed_forbidden_patterns_if_empty() -> None:
    try:
        with _db() as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM forbidden_translation_patterns WHERE language='FR'"
            ).fetchone()[0]
            if count > 0:
                return
        now = datetime.now().isoformat(timespec="seconds")
        rows = []
        for forbidden_text, replacement_text, category, severity, notes in _INITIAL_FORBIDDEN_FR:
            rows.append((
                str(uuid.uuid4()),
                forbidden_text,
                replacement_text,
                category,
                severity,
                "FR",
                notes,
                1,
                now,
            ))
        with _db() as conn:
            conn.executemany(
                """INSERT OR IGNORE INTO forbidden_translation_patterns
                   (id, forbidden_text, replacement_text, category, severity, language, notes, auto_replace, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                rows,
            )
    except Exception:
        pass


_INITIAL_CORPUS_FR = [
    # (source_text, translated_text, category, product_type, terminology_tags)
    ("Geschirrservice 18-teilig Nature Collection", "Service de vaisselle Nature Collection — 18 pièces", "tableware", "tableware", "service,vaisselle,collection"),
    ("Assiette de service 6er Set", "Assiettes de service — lot de 6", "tableware", "tableware", "assiettes,lot,service"),
    ("Gartenlounge Set 5-teilig", "Salon de jardin — 5 pièces", "outdoor", "outdoor", "salon,jardin,ensemble"),
    ("Loungeset bestehend aus Sofa, 2 Sesseln und Tisch", "Salon de jardin composé de : 1 canapé, 2 fauteuils et 1 table basse", "outdoor", "outdoor", "lounge,composé,ensemble"),
    ("Chaiselongue Gartenliege", "Chaise longue de jardin", "outdoor", "outdoor", "chaise,longue,jardin"),
    ("Fußhocker passend zum Loungesessel", "Repose-pieds assorti au fauteuil lounge", "outdoor", "outdoor", "repose-pieds,fauteuil,lounge"),
    ("Artisan Eiche Dekor Küchenzeile", "Cuisine décor chêne artisan", "kitchen", "kitchen", "décor,chêne,artisan,cuisine"),
    ("Einbauküche autark", "Cuisine équipée", "kitchen", "kitchen", "cuisine,équipée"),
    ("Hängeschrank Meuble haut", "Meuble haut de cuisine", "kitchen", "kitchen", "meuble,haut"),
    ("Taschenfederkernmatratze 7 Zonen", "Matelas à ressorts ensachés 7 zones", "mattress", "mattress", "matelas,ressorts,zones"),
    ("Einseitige Kokosmatte Matratze", "Matelas avec couche de coco sur une face", "mattress", "mattress", "coco,matelas"),
    ("Schlafsofa mit Bettkasten", "Canapé convertible avec coffre de rangement", "seating", "seating", "canapé,convertible,rangement"),
    ("Esstisch ausziehbar Artisan Eiche", "Table extensible décor chêne artisan", "tables", "table", "table,extensible,chêne"),
    ("Meuble TV Lowboard mit LED", "Meuble TV bas lumineux", "furniture", "generic", "meuble,TV,lumineux"),
    ("Badezimmer Waschtischunterschrank", "Meuble sous-vasque", "bathroom", "bathroom", "meuble,vasque,salle"),
    ("Set bestehend aus 6 Teilen", "Ensemble composé de 6 éléments", "sets", "generic", "ensemble,composé,éléments"),
    ("pulverbeschichtet Aluminium", "aluminium thermolaqué", "materials", "outdoor", "thermolaqué,aluminium"),
    ("Bezug abnehmbar Matratze", "Revêtement amovible", "mattress", "mattress", "revêtement,amovible"),
    ("Gestell Metall schwarz", "Structure métal noir", "materials", "generic", "structure,métal,noir"),
    ("Füße Massivholz Eiche", "Pieds en chêne massif", "materials", "generic", "pieds,chêne,massif"),
]


def _seed_corpus_if_empty() -> None:
    try:
        with _db() as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM home24_corpus_entries WHERE target_language='FR'"
            ).fetchone()[0]
            if count > 0:
                return
        now = datetime.now().isoformat(timespec="seconds")
        rows = []
        for source_text, translated_text, category, product_type, terminology_tags in _INITIAL_CORPUS_FR:
            rows.append((
                str(uuid.uuid4()),
                "DE",
                "FR",
                source_text,
                translated_text,
                category,
                product_type,
                terminology_tags,
                1.0,
                1,
                now,
            ))
        with _db() as conn:
            conn.executemany(
                """INSERT OR IGNORE INTO home24_corpus_entries
                   (id, source_language, target_language, source_text, translated_text,
                    category, product_type, terminology_tags, confidence_score, approved, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                rows,
            )
    except Exception:
        pass


# =============================================================================
# SEEDING
# =============================================================================

def _seed_glossary_if_empty(terms: dict, target_language: str) -> None:
    try:
        with _db() as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM glossary_terms WHERE target_language = ?",
                (target_language,),
            ).fetchone()[0]
            if count > 0:
                return

        now = datetime.now().isoformat(timespec="seconds")
        rows = [(de, target_language, target, 0, now) for de, target in terms.items()]
        with _db() as conn:
            conn.executemany(
                """
                INSERT OR IGNORE INTO glossary_terms
                    (source_term, target_language, target_term, hit_count, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                rows,
            )
    except Exception:
        pass


# =============================================================================
# JSON MIGRATION (one-time, from legacy JSON files)
# =============================================================================

def _migrate_json_if_needed() -> None:
    try:
        with _db() as conn:
            row = conn.execute(
                "SELECT value FROM app_metrics WHERE key='migration_done'"
            ).fetchone()
            if row:
                return

        _run_json_migration()

        with _db() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO app_metrics(key,value) VALUES('migration_done','1')"
            )
    except Exception:
        pass


def _run_json_migration() -> None:
    base = DB_PATH.parent

    history_file = base / "translation_history.json"
    if history_file.exists():
        try:
            import json as _json
            with open(history_file, encoding="utf-8") as f:
                records = _json.load(f)
            for rec in records:
                if "id" not in rec:
                    rec["id"] = str(uuid.uuid4())
                db_save_history_record(rec)
        except Exception:
            pass

    tm_file = base / "translation_memory.json"
    if tm_file.exists():
        try:
            import json as _json
            with open(tm_file, encoding="utf-8") as f:
                tm = _json.load(f)
            if "entries" in tm:
                db_save_translation_memory(tm)
        except Exception:
            pass

    glossary_file = base / "glossary.json"
    if glossary_file.exists():
        try:
            import json as _json
            with open(glossary_file, encoding="utf-8") as f:
                glossary = _json.load(f)
            if "terms" in glossary:
                db_save_glossary(glossary, "French")
        except Exception:
            pass


# =============================================================================
# GLOSSARY SUGGESTIONS
# =============================================================================

def db_save_glossary_suggestions(
    suggestions: list[dict],
    job_id: str,
    target_language: str = "French",
) -> None:
    """Persist new glossary suggestions (pending review). Skips duplicates."""
    if not suggestions:
        return
    now = datetime.now().isoformat(timespec="seconds")
    rows = []
    for s in suggestions:
        rows.append((
            str(uuid.uuid4()),
            s.get("term", ""),
            target_language,
            s.get("occurrences", 1),
            s.get("example_context", ""),
            "pending",
            job_id,
            now,
        ))
    try:
        with _db() as conn:
            # Only insert if the term+language doesn't already have a pending suggestion
            for row in rows:
                existing = conn.execute(
                    "SELECT id FROM glossary_suggestions "
                    "WHERE term=? AND target_language=? AND status='pending'",
                    (row[1], row[2]),
                ).fetchone()
                if not existing:
                    conn.execute(
                        """
                        INSERT OR IGNORE INTO glossary_suggestions
                            (id, term, target_language, occurrences, example_context,
                             status, job_id, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        row,
                    )
    except Exception:
        pass


def db_load_glossary_suggestions(
    target_language: str | None = None,
    status: str = "pending",
) -> list[dict]:
    """Load glossary suggestions, optionally filtered by language and status."""
    try:
        with _db() as conn:
            if target_language:
                rows = conn.execute(
                    "SELECT * FROM glossary_suggestions "
                    "WHERE target_language=? AND status=? "
                    "ORDER BY occurrences DESC, created_at DESC",
                    (target_language, status),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM glossary_suggestions "
                    "WHERE status=? ORDER BY occurrences DESC, created_at DESC",
                    (status,),
                ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_update_suggestion_status(suggestion_id: str, status: str) -> None:
    """Set status to 'accepted' or 'rejected' for a suggestion."""
    try:
        with _db() as conn:
            conn.execute(
                "UPDATE glossary_suggestions SET status=? WHERE id=?",
                (status, suggestion_id),
            )
    except Exception:
        pass


# =============================================================================
# CORPUS & FORBIDDEN PATTERNS — CRUD
# =============================================================================

def db_load_forbidden_patterns(language: str = "FR") -> list[dict]:
    """Load all active forbidden patterns for a language."""
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT * FROM forbidden_translation_patterns "
                "WHERE language=? AND auto_replace=1 "
                "ORDER BY CASE severity WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'MEDIUM' THEN 3 ELSE 4 END ASC, "
                "LENGTH(forbidden_text) DESC",
                (language,),
            ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_save_forbidden_pattern(
    forbidden_text: str,
    replacement_text: str,
    category: str = "",
    severity: str = "HIGH",
    language: str = "FR",
    notes: str = "",
) -> bool:
    """Add a new forbidden pattern. Returns True on success."""
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                """INSERT OR IGNORE INTO forbidden_translation_patterns
                   (id, forbidden_text, replacement_text, category, severity, language, notes, auto_replace, created_at)
                   VALUES (?,?,?,?,?,?,?,1,?)""",
                (str(uuid.uuid4()), forbidden_text.strip(), replacement_text.strip(),
                 category, severity, language, notes, now),
            )
        return True
    except Exception:
        return False


def db_delete_forbidden_pattern(pattern_id: str) -> bool:
    try:
        with _db() as conn:
            conn.execute("DELETE FROM forbidden_translation_patterns WHERE id=?", (pattern_id,))
        return True
    except Exception:
        return False


def db_load_corpus_entries(
    product_type: str | None = None,
    language: str = "FR",
    limit: int = 10,
) -> list[dict]:
    """Load corpus entries, optionally filtered by product type."""
    try:
        with _db() as conn:
            if product_type and product_type != "generic":
                rows = conn.execute(
                    "SELECT * FROM home24_corpus_entries "
                    "WHERE target_language=? AND approved=1 AND product_type=? "
                    "ORDER BY confidence_score DESC LIMIT ?",
                    (language, product_type, limit),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM home24_corpus_entries "
                    "WHERE target_language=? AND approved=1 "
                    "ORDER BY confidence_score DESC LIMIT ?",
                    (language, limit),
                ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_add_corpus_entry(
    source_text: str,
    translated_text: str,
    category: str = "",
    product_type: str = "generic",
    terminology_tags: str = "",
    confidence_score: float = 1.0,
    source_language: str = "DE",
    target_language: str = "FR",
) -> bool:
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                """INSERT OR IGNORE INTO home24_corpus_entries
                   (id, source_language, target_language, source_text, translated_text,
                    category, product_type, terminology_tags, confidence_score, approved, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,1,?)""",
                (str(uuid.uuid4()), source_language, target_language,
                 source_text.strip(), translated_text.strip(),
                 category, product_type, terminology_tags, confidence_score, now),
            )
        return True
    except Exception:
        return False


def db_get_corpus_count() -> dict:
    """Return corpus entry counts by language."""
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT target_language, COUNT(*) as cnt FROM home24_corpus_entries GROUP BY target_language"
            ).fetchall()
        return {r["target_language"]: r["cnt"] for r in rows}
    except Exception:
        return {}


def db_get_forbidden_count() -> dict:
    """Return forbidden pattern counts by language."""
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT language, COUNT(*) as cnt FROM forbidden_translation_patterns GROUP BY language"
            ).fetchall()
        return {r["language"]: r["cnt"] for r in rows}
    except Exception:
        return {}


# =============================================================================
# V7 MIGRATION — Issue Reports
# =============================================================================

def _ensure_v7_migration() -> None:
    """Add issue_reports table (in-app bug / translation issue reporting)."""
    if _get_schema_version() >= 7:
        return
    try:
        with _db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS issue_reports (
                    id                  TEXT PRIMARY KEY,
                    user_email          TEXT NOT NULL DEFAULT '',
                    user_role           TEXT NOT NULL DEFAULT '',
                    category            TEXT NOT NULL,
                    severity            TEXT NOT NULL DEFAULT 'Medium',
                    target_language     TEXT NOT NULL DEFAULT 'Not language-related',
                    filename            TEXT DEFAULT '',
                    row_reference       TEXT DEFAULT '',
                    column_reference    TEXT DEFAULT '',
                    description         TEXT NOT NULL,
                    expected_correction TEXT DEFAULT '',
                    status              TEXT NOT NULL DEFAULT 'open',
                    created_at          TEXT NOT NULL
                )
            """)
        _set_schema_version(7)
    except Exception:
        pass


# =============================================================================
# ISSUE REPORTS — CRUD
# =============================================================================

def db_save_issue_report(report: dict) -> str:
    """Persist a new issue report. Returns the new report id."""
    report_id = str(uuid.uuid4())
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                """
                INSERT INTO issue_reports
                    (id, user_email, user_role, category, severity, target_language,
                     filename, row_reference, column_reference, description,
                     expected_correction, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'open', ?)
                """,
                (
                    report_id,
                    report.get("user_email", ""),
                    report.get("user_role", ""),
                    report.get("category", "Other"),
                    report.get("severity", "Medium"),
                    report.get("target_language", "Not language-related"),
                    report.get("filename", ""),
                    report.get("row_reference", ""),
                    report.get("column_reference", ""),
                    report.get("description", ""),
                    report.get("expected_correction", ""),
                    now,
                ),
            )
    except Exception:
        pass
    return report_id


def db_load_issue_reports(
    status_filter: str | None = None,
    severity_filter: str | None = None,
    category_filter: str | None = None,
    lang_filter: str | None = None,
) -> list[dict]:
    """Load issue reports with optional filters."""
    try:
        clauses: list[str] = []
        params: list = []
        if status_filter and status_filter != "All":
            clauses.append("status = ?")
            params.append(status_filter)
        if severity_filter and severity_filter != "All":
            clauses.append("severity = ?")
            params.append(severity_filter)
        if category_filter and category_filter != "All":
            clauses.append("category = ?")
            params.append(category_filter)
        if lang_filter and lang_filter != "All":
            clauses.append("target_language = ?")
            params.append(lang_filter)
        sql = "SELECT * FROM issue_reports"
        if clauses:
            sql += " WHERE " + " AND ".join(clauses)
        sql += " ORDER BY created_at DESC"
        with _db() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_update_issue_report_status(report_id: str, status: str) -> None:
    """Update the status of an issue report."""
    try:
        with _db() as conn:
            conn.execute(
                "UPDATE issue_reports SET status = ? WHERE id = ?",
                (status, report_id),
            )
    except Exception:
        pass


def db_get_issue_report_counts() -> dict:
    """Return counts grouped by status."""
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT status, COUNT(*) as cnt FROM issue_reports GROUP BY status"
            ).fetchall()
        return {r["status"]: r["cnt"] for r in rows}
    except Exception:
        return {}


# =============================================================================
# V8 MIGRATION — Jira integration columns
# =============================================================================

def _ensure_v8_migration() -> None:
    """Add Jira tracking columns to translation_jobs."""
    if _get_schema_version() >= 8:
        return
    try:
        new_cols = [
            ("jira_ticket_key",          "TEXT DEFAULT ''"),
            ("jira_ticket_summary",      "TEXT DEFAULT ''"),
            ("jira_attachment_filename", "TEXT DEFAULT ''"),
            ("jira_attachment_id",       "TEXT DEFAULT ''"),
            ("uploaded_to_jira",         "INTEGER DEFAULT 0"),
            ("jira_upload_time",         "TEXT DEFAULT ''"),
            ("jira_comment_added",       "INTEGER DEFAULT 0"),
            ("jira_transition_applied",  "TEXT DEFAULT ''"),
        ]
        with _db() as conn:
            existing = _table_columns(conn, "translation_jobs")
            for col_name, col_def in new_cols:
                if col_name not in existing:
                    conn.execute(
                        f"ALTER TABLE translation_jobs ADD COLUMN {col_name} {col_def}"
                    )
        _set_schema_version(8)
    except Exception:
        pass


# =============================================================================
# JIRA METADATA — CRUD
# =============================================================================

def db_update_jira_metadata(
    job_id: str,
    jira_ticket_key: str = "",
    jira_ticket_summary: str = "",
    jira_attachment_filename: str = "",
    jira_attachment_id: str = "",
    uploaded_to_jira: int = 0,
    jira_upload_time: str = "",
    jira_comment_added: int = 0,
    jira_transition_applied: str = "",
) -> None:
    """Persist Jira workflow metadata for a translation job."""
    try:
        with _db() as conn:
            conn.execute(
                """
                UPDATE translation_jobs SET
                    jira_ticket_key          = COALESCE(NULLIF(?, ''), jira_ticket_key),
                    jira_ticket_summary      = COALESCE(NULLIF(?, ''), jira_ticket_summary),
                    jira_attachment_filename = COALESCE(NULLIF(?, ''), jira_attachment_filename),
                    jira_attachment_id       = COALESCE(NULLIF(?, ''), jira_attachment_id),
                    uploaded_to_jira         = MAX(uploaded_to_jira, ?),
                    jira_upload_time         = COALESCE(NULLIF(?, ''), jira_upload_time),
                    jira_comment_added       = MAX(jira_comment_added, ?),
                    jira_transition_applied  = COALESCE(NULLIF(?, ''), jira_transition_applied)
                WHERE id = ?
                """,
                (
                    jira_ticket_key,
                    jira_ticket_summary,
                    jira_attachment_filename,
                    jira_attachment_id,
                    uploaded_to_jira,
                    jira_upload_time,
                    jira_comment_added,
                    jira_transition_applied,
                    job_id,
                ),
            )
    except Exception:
        pass


def db_get_jira_stats() -> dict:
    """Return Jira workflow statistics for the admin dashboard."""
    try:
        with _db() as conn:
            total_from_jira = conn.execute(
                "SELECT COUNT(*) FROM translation_jobs WHERE jira_ticket_key != ''"
            ).fetchone()[0]
            total_uploaded = conn.execute(
                "SELECT COUNT(*) FROM translation_jobs WHERE uploaded_to_jira = 1"
            ).fetchone()[0]
            recent_rows = conn.execute(
                """
                SELECT jira_ticket_key, jira_ticket_summary,
                       original_filename, output_filename,
                       target_language, user_email,
                       uploaded_to_jira, jira_upload_time,
                       jira_comment_added, jira_transition_applied,
                       datetime
                FROM translation_jobs
                WHERE jira_ticket_key != ''
                ORDER BY datetime DESC
                LIMIT 50
                """
            ).fetchall()
        return {
            "total_from_jira": total_from_jira,
            "total_uploaded":  total_uploaded,
            "recent":          [dict(r) for r in recent_rows],
        }
    except Exception:
        return {"total_from_jira": 0, "total_uploaded": 0, "recent": []}


# =============================================================================
# V9 MIGRATION — NL Trados TM corpus table
# =============================================================================

def _ensure_v10_migration() -> None:
    """Add dutch_contamination_fixes column to translation_jobs (language isolation tracking)."""
    if _get_schema_version() >= 10:
        return
    try:
        with _db() as conn:
            existing = _table_columns(conn, "translation_jobs")
            if "dutch_contamination_fixes" not in existing:
                conn.execute(
                    "ALTER TABLE translation_jobs ADD COLUMN dutch_contamination_fixes INTEGER DEFAULT 0"
                )
        _set_schema_version(10)
    except Exception:
        pass


def _ensure_v9_migration() -> None:
    """Create nl_trados_tm table for the Dutch Home24 Trados corpus."""
    if _get_schema_version() >= 9:
        return
    try:
        with _db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS nl_trados_tm (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_de    TEXT NOT NULL,
                    target_nl    TEXT NOT NULL,
                    source_norm  TEXT NOT NULL,
                    usage_count  INTEGER DEFAULT 0,
                    imported_at  TEXT NOT NULL
                )
            """)
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_nl_trados_norm ON nl_trados_tm(source_norm)"
            )
        _set_schema_version(9)
    except Exception:
        pass


# =============================================================================
# NL TRADOS TM — CRUD
# =============================================================================

def db_nl_trados_import(entries: list[dict]) -> int:
    """
    Bulk-replace the nl_trados_tm table with new entries.
    entries: list of {"source": str, "target": str, "usage_count": int}
    Returns count of imported rows.
    """
    import re as _re
    now = datetime.utcnow().isoformat()

    def _norm(t: str) -> str:
        return _re.sub(r'\s+', ' ', str(t).strip().lower())

    rows = []
    for e in entries:
        src = str(e.get("source", "")).strip()
        tgt = str(e.get("target", "")).strip()
        if not src or not tgt or src == tgt:
            continue
        rows.append((src, tgt, _norm(src), int(e.get("usage_count", 0)), now))

    try:
        with _db() as conn:
            conn.execute("DELETE FROM nl_trados_tm")
            conn.executemany(
                "INSERT INTO nl_trados_tm(source_de, target_nl, source_norm, usage_count, imported_at)"
                " VALUES (?, ?, ?, ?, ?)",
                rows,
            )
        return len(rows)
    except Exception:
        return 0


def db_nl_trados_load_all() -> list[dict]:
    """Load all NL Trados TM entries from DB for engine initialization."""
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT source_de, target_nl, usage_count FROM nl_trados_tm ORDER BY usage_count DESC"
            ).fetchall()
        return [{"source": r[0], "target": r[1], "usage_count": r[2]} for r in rows]
    except Exception:
        return []


def db_nl_trados_count() -> int:
    """Return number of TM entries stored in the DB."""
    try:
        with _db() as conn:
            return conn.execute("SELECT COUNT(*) FROM nl_trados_tm").fetchone()[0]
    except Exception:
        return 0


# =============================================================================
# V11 MIGRATION — Full glossary management schema
# =============================================================================

def _ensure_v11_migration() -> None:
    """Extend glossary_terms with management fields and create audit log table."""
    if _get_schema_version() >= 11:
        return
    try:
        new_cols = [
            ("source_language",       "TEXT DEFAULT 'German'"),
            ("category",              "TEXT DEFAULT ''"),
            ("source_type",           "TEXT DEFAULT 'MANUAL'"),
            ("confidence",            "REAL DEFAULT 1.0"),
            ("priority",              "TEXT DEFAULT 'NORMAL'"),
            ("active",                "INTEGER DEFAULT 1"),
            ("is_deleted",            "INTEGER DEFAULT 0"),
            ("last_used_at",          "TEXT DEFAULT ''"),
            ("created_by",            "TEXT DEFAULT ''"),
            ("created_at",            "TEXT DEFAULT ''"),
            ("notes",                 "TEXT DEFAULT ''"),
            ("forbidden_alternatives","TEXT DEFAULT ''"),
            ("synonyms",              "TEXT DEFAULT ''"),
        ]
        with _db() as conn:
            existing = _table_columns(conn, "glossary_terms")
            for col_name, col_def in new_cols:
                if col_name not in existing:
                    conn.execute(
                        f"ALTER TABLE glossary_terms ADD COLUMN {col_name} {col_def}"
                    )
            # Backfill created_at from updated_at for existing terms
            conn.execute(
                "UPDATE glossary_terms SET created_at = updated_at "
                "WHERE created_at = '' AND updated_at IS NOT NULL AND updated_at != ''"
            )
            # Indexes
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_glossary_lang_active "
                "ON glossary_terms(target_language, active, is_deleted)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_glossary_source_type "
                "ON glossary_terms(source_type, target_language)"
            )
            # Audit log table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS glossary_audit_log (
                    id           TEXT PRIMARY KEY,
                    source_term  TEXT NOT NULL,
                    target_language TEXT NOT NULL,
                    action       TEXT NOT NULL,
                    old_value    TEXT DEFAULT '',
                    new_value    TEXT DEFAULT '',
                    changed_by   TEXT DEFAULT '',
                    timestamp    TEXT NOT NULL
                )
            """)
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_gloss_audit_ts "
                "ON glossary_audit_log(timestamp DESC)"
            )
        _set_schema_version(11)
    except Exception:
        pass


# =============================================================================
# GLOSSARY MANAGEMENT — full CRUD
# =============================================================================

_SOURCE_TYPE_LABELS = {
    "MANUAL":            "👤 Manual",
    "TM_IMPORTED":       "📚 TM Import",
    "AI_GENERATED":      "🤖 AI",
    "AUTO_LEARNED":      "⚡ Auto-learned",
    "CORPUS_EXTRACTED":  "🔍 Corpus",
}

_PRIORITY_LABELS = {
    "CRITICAL": "🔴 CRITICAL",
    "HIGH":     "🟠 HIGH",
    "NORMAL":   "🟢 NORMAL",
    "LOW":      "⚪ LOW",
}


def db_glossary_get_all(
    target_language: str | None = None,
    include_deleted: bool = False,
    include_inactive: bool = True,
) -> list[dict]:
    """Return all glossary terms with full management metadata."""
    try:
        clauses = []
        params: list = []
        if target_language:
            clauses.append("target_language = ?")
            params.append(target_language)
        if not include_deleted:
            clauses.append("COALESCE(is_deleted, 0) = 0")
        if not include_inactive:
            clauses.append("COALESCE(active, 1) = 1")
        sql = (
            "SELECT source_term, target_language, target_term, "
            "COALESCE(source_language,'German') AS source_language, "
            "COALESCE(category,'') AS category, "
            "COALESCE(source_type,'MANUAL') AS source_type, "
            "COALESCE(confidence,1.0) AS confidence, "
            "COALESCE(priority,'NORMAL') AS priority, "
            "COALESCE(active,1) AS active, "
            "COALESCE(is_deleted,0) AS is_deleted, "
            "COALESCE(hit_count,0) AS hit_count, "
            "COALESCE(last_used_at,'') AS last_used_at, "
            "COALESCE(created_by,'') AS created_by, "
            "COALESCE(created_at, updated_at, '') AS created_at, "
            "COALESCE(updated_at,'') AS updated_at, "
            "COALESCE(notes,'') AS notes, "
            "COALESCE(synonyms,'') AS synonyms, "
            "COALESCE(forbidden_alternatives,'') AS forbidden_alternatives "
            "FROM glossary_terms"
        )
        if clauses:
            sql += " WHERE " + " AND ".join(clauses)
        sql += " ORDER BY source_term COLLATE NOCASE"
        with _db() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_glossary_add_term(
    source_term: str,
    target_term: str,
    target_language: str = "French",
    source_language: str = "German",
    category: str = "",
    source_type: str = "MANUAL",
    confidence: float = 1.0,
    priority: str = "NORMAL",
    notes: str = "",
    synonyms: str = "",
    forbidden_alternatives: str = "",
    created_by: str = "",
) -> bool:
    """Insert a new glossary term with full metadata. Returns False if it already exists."""
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                """
                INSERT OR IGNORE INTO glossary_terms
                    (source_term, target_language, target_term,
                     source_language, category, source_type,
                     confidence, priority, active, is_deleted,
                     hit_count, last_used_at, created_by, created_at, updated_at,
                     notes, synonyms, forbidden_alternatives)
                VALUES (?,?,?,?,?,?,?,?,1,0,0,'',?,?,?,?,?,?)
                """,
                (
                    source_term.strip(), target_language, target_term.strip(),
                    source_language, category, source_type,
                    confidence, priority,
                    created_by, now, now,
                    notes, synonyms, forbidden_alternatives,
                ),
            )
        db_log_glossary_audit(source_term, target_language, "CREATED", "", target_term, created_by)
        return True
    except Exception:
        return False


def db_glossary_update_term(
    source_term: str,
    target_language: str,
    updates: dict,
    edited_by: str = "",
) -> bool:
    """Update editable fields of a glossary term."""
    allowed = {
        "target_term", "category", "notes", "synonyms",
        "forbidden_alternatives", "priority", "confidence",
        "source_type",
    }
    fields = {k: v for k, v in updates.items() if k in allowed}
    if not fields:
        return False
    now = datetime.now().isoformat(timespec="seconds")
    fields["updated_at"] = now
    set_clause = ", ".join(f"{k} = ?" for k in fields)
    params = list(fields.values()) + [source_term, target_language]
    try:
        with _db() as conn:
            conn.execute(
                f"UPDATE glossary_terms SET {set_clause} "
                "WHERE source_term = ? AND target_language = ?",
                params,
            )
        old_target = updates.get("_old_target_term", "")
        new_target = updates.get("target_term", "")
        db_log_glossary_audit(
            source_term, target_language, "UPDATED",
            old_target, new_target, edited_by,
        )
        return True
    except Exception:
        return False


def db_glossary_soft_delete(
    source_term: str,
    target_language: str,
    deleted_by: str = "",
) -> bool:
    """Mark a term as deleted (soft delete)."""
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                "UPDATE glossary_terms SET is_deleted=1, active=0, updated_at=? "
                "WHERE source_term=? AND target_language=?",
                (now, source_term, target_language),
            )
        db_log_glossary_audit(source_term, target_language, "DELETED", "active", "deleted", deleted_by)
        return True
    except Exception:
        return False


def db_glossary_restore(
    source_term: str,
    target_language: str,
    restored_by: str = "",
) -> bool:
    """Restore a soft-deleted term."""
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                "UPDATE glossary_terms SET is_deleted=0, active=1, updated_at=? "
                "WHERE source_term=? AND target_language=?",
                (now, source_term, target_language),
            )
        db_log_glossary_audit(source_term, target_language, "RESTORED", "deleted", "active", restored_by)
        return True
    except Exception:
        return False


def db_glossary_toggle_active(
    source_term: str,
    target_language: str,
    user: str = "",
) -> bool:
    """Toggle active flag. Returns the new active state."""
    try:
        with _db() as conn:
            row = conn.execute(
                "SELECT COALESCE(active, 1) FROM glossary_terms "
                "WHERE source_term=? AND target_language=?",
                (source_term, target_language),
            ).fetchone()
            if not row:
                return False
            new_active = 0 if row[0] else 1
            now = datetime.now().isoformat(timespec="seconds")
            conn.execute(
                "UPDATE glossary_terms SET active=?, updated_at=? "
                "WHERE source_term=? AND target_language=?",
                (new_active, now, source_term, target_language),
            )
        action = "ENABLED" if new_active else "DISABLED"
        db_log_glossary_audit(source_term, target_language, action, "", str(new_active), user)
        return bool(new_active)
    except Exception:
        return False


def db_glossary_increment_usage(source_term: str, target_language: str) -> None:
    """Increment hit_count and update last_used_at. Called from translation engine."""
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                "UPDATE glossary_terms SET hit_count = hit_count + 1, last_used_at = ? "
                "WHERE source_term = ? AND target_language = ?",
                (now, source_term, target_language),
            )
    except Exception:
        pass


def db_log_glossary_audit(
    source_term: str,
    target_language: str,
    action: str,
    old_value: str,
    new_value: str,
    user: str,
) -> None:
    """Write an entry to the glossary audit log."""
    now = datetime.now().isoformat(timespec="seconds")
    try:
        with _db() as conn:
            conn.execute(
                """
                INSERT INTO glossary_audit_log
                    (id, source_term, target_language, action, old_value, new_value, changed_by, timestamp)
                VALUES (?,?,?,?,?,?,?,?)
                """,
                (str(uuid.uuid4()), source_term, target_language, action,
                 old_value or "", new_value or "", user or "", now),
            )
    except Exception:
        pass


def db_get_glossary_audit_log(limit: int = 200) -> list[dict]:
    """Return the most recent audit log entries."""
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT * FROM glossary_audit_log ORDER BY timestamp DESC LIMIT ?",
                (limit,),
            ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_glossary_bulk_import(
    rows: list[dict],
    created_by: str = "",
    overwrite: bool = False,
) -> tuple[int, int]:
    """
    Bulk-import glossary terms from a list of dicts.
    Expected keys: source_term, target_term, target_language, category, source_type, priority.
    Returns (imported_count, skipped_count).
    """
    imported = 0
    skipped = 0
    now = datetime.now().isoformat(timespec="seconds")
    for row in rows:
        src  = str(row.get("source_term", "")).strip()
        tgt  = str(row.get("target_term", "")).strip()
        lang = str(row.get("target_language", "French")).strip()
        if not src or not tgt:
            skipped += 1
            continue
        cat       = str(row.get("category", "")).strip()
        src_type  = str(row.get("source_type", "MANUAL")).strip().upper()
        priority  = str(row.get("priority", "NORMAL")).strip().upper()
        notes_val = str(row.get("notes", "")).strip()
        if src_type not in ("MANUAL", "TM_IMPORTED", "AI_GENERATED", "AUTO_LEARNED", "CORPUS_EXTRACTED"):
            src_type = "MANUAL"
        if priority not in ("CRITICAL", "HIGH", "NORMAL", "LOW"):
            priority = "NORMAL"
        try:
            with _db() as conn:
                if overwrite:
                    conn.execute(
                        """
                        INSERT INTO glossary_terms
                            (source_term, target_language, target_term, category,
                             source_type, priority, active, is_deleted, hit_count,
                             created_by, created_at, updated_at, notes)
                        VALUES (?,?,?,?,?,?,1,0,0,?,?,?,?)
                        ON CONFLICT(source_term, target_language) DO UPDATE SET
                            target_term  = excluded.target_term,
                            category     = excluded.category,
                            source_type  = excluded.source_type,
                            priority     = excluded.priority,
                            notes        = excluded.notes,
                            updated_at   = excluded.updated_at
                        """,
                        (src, lang, tgt, cat, src_type, priority,
                         created_by, now, now, notes_val),
                    )
                else:
                    conn.execute(
                        """
                        INSERT OR IGNORE INTO glossary_terms
                            (source_term, target_language, target_term, category,
                             source_type, priority, active, is_deleted, hit_count,
                             created_by, created_at, updated_at, notes)
                        VALUES (?,?,?,?,?,?,1,0,0,?,?,?,?)
                        """,
                        (src, lang, tgt, cat, src_type, priority,
                         created_by, now, now, notes_val),
                    )
            imported += 1
        except Exception:
            skipped += 1
    return imported, skipped


def db_glossary_get_stats(target_language: str | None = None) -> dict:
    """Return summary statistics for the glossary management dashboard."""
    try:
        lang_clause = "WHERE target_language = ?" if target_language else ""
        params = (target_language,) if target_language else ()
        with _db() as conn:
            total = conn.execute(
                f"SELECT COUNT(*) FROM glossary_terms {lang_clause}", params
            ).fetchone()[0]
            active = conn.execute(
                f"SELECT COUNT(*) FROM glossary_terms "
                f"{'WHERE' if lang_clause else 'WHERE'} "
                f"{'target_language=? AND ' if target_language else ''}"
                f"COALESCE(active,1)=1 AND COALESCE(is_deleted,0)=0",
                params,
            ).fetchone()[0]
            inactive = conn.execute(
                f"SELECT COUNT(*) FROM glossary_terms "
                f"WHERE {'target_language=? AND ' if target_language else ''}"
                f"COALESCE(active,1)=0 AND COALESCE(is_deleted,0)=0",
                params,
            ).fetchone()[0]
            deleted = conn.execute(
                f"SELECT COUNT(*) FROM glossary_terms "
                f"WHERE {'target_language=? AND ' if target_language else ''}"
                f"COALESCE(is_deleted,0)=1",
                params,
            ).fetchone()[0]
            total_hits = conn.execute(
                f"SELECT COALESCE(SUM(hit_count),0) FROM glossary_terms {lang_clause}", params
            ).fetchone()[0]
            critical = conn.execute(
                f"SELECT COUNT(*) FROM glossary_terms "
                f"WHERE {'target_language=? AND ' if target_language else ''}"
                f"priority='CRITICAL' AND COALESCE(is_deleted,0)=0",
                params,
            ).fetchone()[0]
        return {
            "total": total,
            "active": active,
            "inactive": inactive,
            "deleted": deleted,
            "total_hits": total_hits,
            "critical": critical,
        }
    except Exception:
        return {"total": 0, "active": 0, "inactive": 0, "deleted": 0, "total_hits": 0, "critical": 0}


# =============================================================================
# V12 MIGRATION — Dutch glossary dedicated table
# =============================================================================

def _ensure_v12_migration() -> None:
    """Create dutch_glossary_terms table with full-text and fuzzy lookup support."""
    if _get_schema_version() >= 12:
        return
    try:
        with _db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS dutch_glossary_terms (
                    id              TEXT PRIMARY KEY,
                    source_term_de  TEXT NOT NULL,
                    target_term_nl  TEXT NOT NULL,
                    normalized_source TEXT NOT NULL,
                    normalized_target TEXT NOT NULL,
                    category        TEXT DEFAULT '',
                    frequency       INTEGER DEFAULT 0,
                    confidence      REAL DEFAULT 1.0,
                    source_type     TEXT DEFAULT 'MANUAL',
                    active          INTEGER DEFAULT 1,
                    created_at      TEXT NOT NULL,
                    updated_at      TEXT NOT NULL,
                    UNIQUE(source_term_de)
                )
            """)
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_dutch_gloss_source "
                "ON dutch_glossary_terms(normalized_source)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_dutch_gloss_active "
                "ON dutch_glossary_terms(active, source_type)"
            )
        _set_schema_version(12)
    except Exception:
        pass


# =============================================================================
# DUTCH GLOSSARY — CRUD
# =============================================================================

def _normalize_gloss(text: str) -> str:
    """Normalize for Dutch glossary lookup: lower, strip, collapse spaces."""
    import re as _re
    return _re.sub(r'\s+', ' ', str(text).strip().lower())


def db_dutch_glossary_load_all(active_only: bool = True) -> list[dict]:
    """Load all Dutch glossary terms. Optionally filter to active only."""
    try:
        with _db() as conn:
            if active_only:
                rows = conn.execute(
                    "SELECT * FROM dutch_glossary_terms WHERE active=1 "
                    "ORDER BY source_term_de COLLATE NOCASE"
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM dutch_glossary_terms "
                    "ORDER BY source_term_de COLLATE NOCASE"
                ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_dutch_glossary_lookup(source_term: str) -> str | None:
    """Exact match lookup by normalized_source. Returns Dutch target or None."""
    norm = _normalize_gloss(source_term)
    try:
        with _db() as conn:
            row = conn.execute(
                "SELECT target_term_nl FROM dutch_glossary_terms "
                "WHERE normalized_source = ? AND active = 1",
                (norm,),
            ).fetchone()
        return row[0] if row else None
    except Exception:
        return None


def db_dutch_glossary_fuzzy_lookup(source_term: str, threshold: float = 0.85) -> str | None:
    """
    Fuzzy match using SequenceMatcher against all active normalized sources.
    Returns best Dutch target above the threshold or None.
    """
    from difflib import SequenceMatcher
    norm = _normalize_gloss(source_term)
    if not norm:
        return None
    try:
        with _db() as conn:
            rows = conn.execute(
                "SELECT normalized_source, target_term_nl FROM dutch_glossary_terms WHERE active=1"
            ).fetchall()
        best_score = 0.0
        best_tgt: str | None = None
        for row in rows:
            score = SequenceMatcher(None, norm, row[0], autojunk=False).ratio()
            if score > best_score:
                best_score = score
                best_tgt = row[1]
        if best_score >= threshold and best_tgt:
            return best_tgt
        return None
    except Exception:
        return None


def db_dutch_glossary_import_excel(path: str) -> tuple[int, int]:
    """
    Parse DE→NL glossary Excel (sheet: "Clean Technical Glossary").
    Column A: German term, Column B: Dutch term.
    Strips whitespace, skips empty/identical/digit-only/punctuation-only rows.
    Returns (imported, skipped).
    """
    import re as _re
    try:
        import openpyxl
    except ImportError:
        return 0, 0

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Clean Technical Glossary"]
    except Exception:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return 0, 0

    rows_to_import: list[tuple[str, str]] = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if not row:
            continue
        de = str(row[0]).strip() if row[0] is not None else ""
        nl = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not de or not nl:
            continue
        if de == nl:
            continue
        if _re.match(r'^\d+$', de):
            continue
        if _re.match(r'^[^\w]+$', de):
            continue
        rows_to_import.append((de, nl))

    try:
        wb.close()
    except Exception:
        pass

    return db_dutch_glossary_bulk_import([
        {
            "source_term_de": de,
            "target_term_nl": nl,
            "source_type":    "IMPORTED_EXCEL",
            "frequency":      1,
            "confidence":     0.9,
        }
        for de, nl in rows_to_import
    ])


def db_dutch_glossary_bulk_import(rows: list[dict]) -> tuple[int, int]:
    """
    Bulk-import Dutch glossary rows.
    Each row: {source_term_de, target_term_nl, source_type, frequency, confidence, category}.
    Uses ON CONFLICT(source_term_de) DO UPDATE if new confidence is higher.
    Returns (imported, skipped).
    """
    imported = 0
    skipped = 0
    now = datetime.now().isoformat(timespec="seconds")
    for row in rows:
        de  = str(row.get("source_term_de", "")).strip()
        nl  = str(row.get("target_term_nl", "")).strip()
        if not de or not nl:
            skipped += 1
            continue
        norm_de = _normalize_gloss(de)
        norm_nl = _normalize_gloss(nl)
        src_type = str(row.get("source_type", "MANUAL")).upper()
        freq     = int(row.get("frequency", 0))
        conf     = float(row.get("confidence", 1.0))
        category = str(row.get("category", "")).strip()
        try:
            with _db() as conn:
                conn.execute(
                    """
                    INSERT INTO dutch_glossary_terms
                        (id, source_term_de, target_term_nl,
                         normalized_source, normalized_target,
                         category, frequency, confidence, source_type,
                         active, created_at, updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,1,?,?)
                    ON CONFLICT(source_term_de) DO UPDATE SET
                        target_term_nl    = CASE WHEN excluded.confidence > confidence
                                                 THEN excluded.target_term_nl
                                                 ELSE target_term_nl END,
                        normalized_target = CASE WHEN excluded.confidence > confidence
                                                 THEN excluded.normalized_target
                                                 ELSE normalized_target END,
                        confidence        = MAX(excluded.confidence, confidence),
                        frequency         = frequency + excluded.frequency,
                        updated_at        = excluded.updated_at
                    """,
                    (str(uuid.uuid4()), de, nl, norm_de, norm_nl,
                     category, freq, conf, src_type, now, now),
                )
            imported += 1
        except Exception:
            skipped += 1
    return imported, skipped


def db_dutch_glossary_add_term(
    source_term_de: str,
    target_term_nl: str,
    category: str = "",
    source_type: str = "MANUAL",
    confidence: float = 1.0,
) -> bool:
    """Add a single Dutch glossary term. Returns True on success."""
    now = datetime.now().isoformat(timespec="seconds")
    de  = source_term_de.strip()
    nl  = target_term_nl.strip()
    if not de or not nl:
        return False
    try:
        with _db() as conn:
            conn.execute(
                """
                INSERT OR IGNORE INTO dutch_glossary_terms
                    (id, source_term_de, target_term_nl,
                     normalized_source, normalized_target,
                     category, frequency, confidence, source_type,
                     active, created_at, updated_at)
                VALUES (?,?,?,?,?,?,0,?,?,1,?,?)
                """,
                (str(uuid.uuid4()), de, nl,
                 _normalize_gloss(de), _normalize_gloss(nl),
                 category, confidence, source_type.upper(), now, now),
            )
        return True
    except Exception:
        return False


def db_dutch_glossary_get_stats() -> dict:
    """Return summary statistics for the Dutch glossary."""
    try:
        with _db() as conn:
            total = conn.execute(
                "SELECT COUNT(*) FROM dutch_glossary_terms"
            ).fetchone()[0]
            active = conn.execute(
                "SELECT COUNT(*) FROM dutch_glossary_terms WHERE active=1"
            ).fetchone()[0]
            by_type = conn.execute(
                "SELECT source_type, COUNT(*) as cnt "
                "FROM dutch_glossary_terms GROUP BY source_type"
            ).fetchall()
            total_freq = conn.execute(
                "SELECT COALESCE(SUM(frequency),0) FROM dutch_glossary_terms"
            ).fetchone()[0]
        return {
            "total":      total,
            "active":     active,
            "inactive":   total - active,
            "total_freq": total_freq,
            "by_type":    {r["source_type"]: r["cnt"] for r in by_type},
        }
    except Exception:
        return {"total": 0, "active": 0, "inactive": 0, "total_freq": 0, "by_type": {}}
